"""Export system: schedule, client, book, renewals, LLM context dumps."""

from __future__ import annotations

import csv
import io
import json
import sqlite3
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from policydb import config as cfg
from policydb.analysis import build_program_audit
from policydb.display import fmt_currency, fmt_limit, fmt_pct, fmt_days

# Fields excluded from all client-facing (schedule) exports
_INTERNAL_FIELDS = {
    "commission_rate", "commission_amount", "prior_premium", "rate_change",
    "renewal_status", "placement_colleague", "underwriter_name",
    "underwriter_contact", "notes", "account_exec",
}

TODAY = date.today().isoformat()


def _schedule_rows_for_client(conn: sqlite3.Connection, client_id: int) -> list[sqlite3.Row]:
    return conn.execute(
        """SELECT * FROM v_schedule WHERE client_name = (SELECT name FROM clients WHERE id = ?)""",
        (client_id,),
    ).fetchall()


def _row_to_dict(row: sqlite3.Row, exclude: set[str] | None = None) -> dict:
    d = dict(row)
    if exclude:
        for k in exclude:
            d.pop(k, None)
    return d


# ─── SCHEDULE OF INSURANCE ───────────────────────────────────────────────────

def export_schedule_md(conn: sqlite3.Connection, client_id: int, client_name: str) -> str:
    rows = _schedule_rows_for_client(conn, client_id)
    account_exec = cfg.get("default_account_exec", "Grant")
    total = sum(r["Premium"] or 0 for r in rows)

    lines = [
        "# Schedule of Insurance",
        "",
        f"**Insured:** {client_name}",
        f"**Prepared:** {TODAY}",
        f"**Prepared by:** {account_exec}, Marsh",
        "",
        "| Line of Business | Carrier | Policy # | Effective | Expiration | Premium | Limit | Deductible | Form | Layer | Comments |",
        "|------------------|---------|----------|-----------|------------|---------|-------|------------|------|-------|----------|",
    ]

    for r in rows:
        pnum = r["Policy Number"] or ""
        form = r["Form"] or ""
        layer = r["Layer"] or "Primary"
        comments = (r["Comments"] or "").replace("|", "\\|")
        lines.append(
            f"| {r['Line of Business']} | {r['Carrier']} | {pnum} | {r['Effective']} | {r['Expiration']}"
            f" | {fmt_currency(r['Premium'])} | {fmt_limit(r['Limit'])} | {fmt_limit(r['Deductible'])}"
            f" | {form} | {layer} | {comments} |"
        )

    lines += [
        "",
        f"**Total Annual Premium:** {fmt_currency(total)}",
        f"**Policies:** {len(rows)}",
    ]
    return "\n".join(lines)


def export_schedule_csv(conn: sqlite3.Connection, client_id: int) -> str:
    rows = _schedule_rows_for_client(conn, client_id)
    if not rows:
        return ""
    buf = io.StringIO()
    cols = [k for k in rows[0].keys() if k != "client_name"]
    writer = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow({k: r[k] for k in cols})
    return buf.getvalue()


def export_schedule_json(conn: sqlite3.Connection, client_id: int, client_name: str) -> str:
    rows = _schedule_rows_for_client(conn, client_id)
    policies = []
    for r in rows:
        d = dict(r)
        d.pop("client_name", None)
        policies.append(d)
    return json.dumps({
        "client": client_name,
        "prepared": TODAY,
        "total_premium": sum(r["Premium"] or 0 for r in rows),
        "policies": policies,
    }, indent=2, default=str)


# ─── LLM CLIENT EXPORT ───────────────────────────────────────────────────────

def export_llm_client_md(conn: sqlite3.Connection, client_id: int) -> str:
    from collections import defaultdict
    from policydb.analysis import layer_notation as _ln

    client = conn.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    summary = conn.execute("SELECT * FROM v_client_summary WHERE id = ?", (client_id,)).fetchone()
    policies = conn.execute(
        "SELECT * FROM v_policy_status WHERE client_id = ? ORDER BY policy_type, layer_position",
        (client_id,),
    ).fetchall()
    pipeline = conn.execute(
        "SELECT * FROM v_renewal_pipeline WHERE client_name = ? ORDER BY expiration_date",
        (client["name"],),
    ).fetchall()
    activities = conn.execute(
        """SELECT a.*, c.name AS client_name, p.policy_uid, p.policy_type AS policy_type_ref
           FROM activity_log a
           JOIN clients c ON a.client_id = c.id
           LEFT JOIN policies p ON a.policy_id = p.id
           WHERE a.client_id = ? AND a.activity_date >= date('now', '-180 days')
           ORDER BY a.activity_date DESC""",
        (client_id,),
    ).fetchall()
    overdue = conn.execute(
        "SELECT * FROM v_overdue_followups WHERE client_name = ?",
        (client["name"],),
    ).fetchall()
    upcoming_followups = conn.execute(
        """SELECT a.id, a.activity_date, a.activity_type, a.subject, a.follow_up_date,
                  a.contact_person, p.policy_uid
           FROM activity_log a
           LEFT JOIN policies p ON a.policy_id = p.id
           WHERE a.client_id = ? AND a.follow_up_date >= date('now') AND a.follow_up_done = 0
           ORDER BY a.follow_up_date ASC""",
        (client_id,),
    ).fetchall()
    history = conn.execute(
        "SELECT * FROM premium_history WHERE client_id = ? ORDER BY policy_type, term_effective DESC",
        (client_id,),
    ).fetchall()
    project_notes_rows = conn.execute(
        "SELECT name AS project_name, notes FROM projects WHERE client_id = ? AND notes != '' ORDER BY name",
        (client_id,),
    ).fetchall()
    contacts = conn.execute(
        "SELECT * FROM client_contacts WHERE client_id = ? ORDER BY is_primary DESC, name",
        (client_id,),
    ).fetchall()
    scratchpad_row = conn.execute(
        "SELECT content FROM client_scratchpad WHERE client_id = ?", (client_id,)
    ).fetchone()
    audit = build_program_audit(conn, client_id)

    policy_contacts_rows = conn.execute(
        """SELECT pc.policy_id, pc.name, pc.email, pc.phone, pc.role, p.policy_uid
           FROM policy_contacts pc JOIN policies p ON pc.policy_id = p.id
           WHERE p.client_id = ? ORDER BY pc.name""",
        (client_id,),
    ).fetchall()

    client_dict = dict(client)
    s = summary
    total_premium = s["total_premium"] if s else 0
    total_commission = s["total_commission"] if s else 0
    total_fees = s["total_fees"] if s else 0
    total_revenue = s["total_revenue"] if s else 0
    next_pol = pipeline[0] if pipeline else None

    # Build lookup maps
    project_notes_map = {r["project_name"]: r["notes"] for r in project_notes_rows}

    # Policy contacts grouped by policy_id
    from collections import defaultdict as _dd
    policy_contacts_map: dict = _dd(list)
    for pc in policy_contacts_rows:
        policy_contacts_map[pc["policy_id"]].append(dict(pc))

    # Primary location address per project (from most recent policy)
    project_addresses: dict[str, str] = {}
    for p in sorted([dict(r) for r in policies], key=lambda x: x.get("id", 0)):
        proj = (p.get("project_name") or "").strip()
        if proj not in project_addresses:
            parts = [p.get("exposure_address"), p.get("exposure_city"),
                     p.get("exposure_state"), p.get("exposure_zip")]
            addr = ", ".join(x for x in parts if x)
            if addr:
                project_addresses[proj] = addr

    # Tower layers grouped by project then tower_group
    tower_by_project: dict = defaultdict(lambda: defaultdict(list))
    for p_dict in [dict(p) for p in policies]:
        tg = p_dict.get("tower_group")
        if tg:
            proj = (p_dict.get("project_name") or "").strip()
            tower_by_project[proj][tg].append(p_dict)

    # Group policies by project
    policy_groups: dict[str, list] = defaultdict(list)
    for p in policies:
        proj = (dict(p).get("project_name") or "").strip()
        policy_groups[proj].append(dict(p))

    # Sort projects: named A-Z, blank ("Corporate / Standalone") last
    sorted_projects = sorted(policy_groups.keys(), key=lambda x: "\xff" if not x else x.lower())

    # ─── YAML frontmatter ────────────────────────────────────────────────────
    lines = [
        "---",
        "export_type: client_program_summary",
        f'client: "{client["name"]}"',
        f'industry: "{client["industry_segment"]}"',
        f'account_executive: "{client["account_exec"]}"',
        f'export_date: "{TODAY}"',
        f"total_policies: {len(policies)}",
        f"total_annual_premium: {fmt_currency(total_premium)}",
        f"estimated_annual_revenue: {fmt_currency(total_revenue)}",
    ]
    if next_pol:
        lines.append(
            f'next_renewal: "{next_pol["expiration_date"]} ({next_pol["days_to_renewal"]}d'
            f' — {next_pol["policy_uid"]}, {next_pol["policy_type"]})"'
        )
    lines += ["---", ""]

    # ─── Title + business description ────────────────────────────────────────
    lines += [f"# Client Program Summary: {client['name']}", ""]

    if client_dict.get("business_description"):
        lines += [client_dict["business_description"], ""]

    meta = f"**Industry:** {client['industry_segment']}"
    if client_dict.get("cn_number"):
        meta += f" | **CN:** {client['cn_number']}"
    meta += f" | **Onboarded:** {client['date_onboarded']} | **Account Executive:** {client['account_exec']}"
    if client_dict.get("client_since"):
        meta += f" | **Client Since:** {client_dict['client_since']}"
    if client_dict.get("website"):
        meta += f" | **Web:** {client_dict['website']}"
    if client_dict.get("preferred_contact_method"):
        meta += f" | **Preferred Contact:** {client_dict['preferred_contact_method']}"
    if client_dict.get("referral_source"):
        meta += f" | **Source:** {client_dict['referral_source']}"
    lines += [meta, ""]

    if client_dict.get("renewal_month"):
        import calendar
        month_name = calendar.month_name[client_dict["renewal_month"]]
        lines += [f"*Typical renewal month: {month_name}*", ""]

    # Contacts
    if contacts:
        lines += ["### Contacts", ""]
        for c in contacts:
            c = dict(c)
            primary_marker = " ★" if c.get("is_primary") else ""
            entry = f"- **{c['name']}{primary_marker}**"
            if c.get("title"):
                entry += f", {c['title']}"
            if c.get("role"):
                entry += f" ({c['role']})"
            if c.get("contact_type") == "internal":
                if c.get("assignment"):
                    entry += f" — {c['assignment']}"
                else:
                    entry += " [internal]"
            if c.get("email"):
                entry += f" — {c['email']}"
            if c.get("phone"):
                entry += f" · {c['phone']}"
            lines.append(entry)
        lines.append("")
    elif client_dict.get("primary_contact"):
        contact_line = f"- **{client['primary_contact']}**"
        if client_dict.get("contact_email"):
            contact_line += f" — {client['contact_email']}"
        if client_dict.get("contact_phone"):
            contact_line += f" · {client['contact_phone']}"
        lines += ["### Contacts", "", contact_line, ""]

    if client_dict.get("notes"):
        lines += ["### Account Notes (Legacy)", "", client_dict["notes"], ""]

    scratchpad_content = (scratchpad_row["content"] if scratchpad_row else "").strip()
    if scratchpad_content:
        lines += ["### Working Notes", "", scratchpad_content, ""]

    # ─── Program Overview ────────────────────────────────────────────────────
    rev_detail = ""
    if total_commission and total_fees:
        rev_detail = f" ({fmt_currency(total_commission)} commission + {fmt_currency(total_fees)} flat fee)"
    elif total_commission:
        rev_detail = " (commission)"
    elif total_fees:
        rev_detail = " (flat fee)"

    lines += [
        "## Program Overview",
        "",
        f"- **Total annual premium:** {fmt_currency(total_premium)}",
        f"- **Estimated annual revenue:** {fmt_currency(total_revenue)}{rev_detail}",
        f"- **Active policies:** {len(policies)} ({audit['standalone_count']} standalone)",
        f"- **Coverage lines:** {', '.join(audit['coverage_lines'])}",
        f"- **Carriers:** {', '.join(audit['carriers'])}",
        "",
    ]
    if next_pol:
        lines += [
            f"Next renewal: **{next_pol['policy_type']}** ({next_pol['policy_uid']}) with "
            f"{next_pol['carrier']} — expires {next_pol['expiration_date']}, "
            f"{next_pol['days_to_renewal']} days out.",
            "",
        ]

    # ─── Renewal Pipeline ────────────────────────────────────────────────────
    if pipeline:
        lines += [
            "## Renewal Pipeline (Next 180 Days)",
            "",
            "| Policy | Line of Business | Carrier | Expires | Days | Premium | Urgency | Status | Colleague |",
            "|--------|------------------|---------|---------|------|---------|---------|--------|-----------|",
        ]
        for r in pipeline:
            lines.append(
                f"| {r['policy_uid']} | {r['policy_type']} | {r['carrier']}"
                f" | {r['expiration_date']} | {r['days_to_renewal']}"
                f" | {fmt_currency(r['premium'])} | {r['urgency']}"
                f" | {r['renewal_status']} | {r['placement_colleague'] or '—'} |"
            )
        lines.append("")

    # ─── Insurance Program (grouped by project) ───────────────────────────
    lines += ["## Insurance Program", ""]

    for proj in sorted_projects:
        group = policy_groups[proj]
        display_name = proj if proj else "Corporate / Standalone"
        lines += [f"### {display_name}", ""]

        addr = project_addresses.get(proj, "")
        if addr:
            lines += [f"**Location:** {addr}", ""]

        note = project_notes_map.get(proj, "")
        if note:
            lines += [note, ""]

        # Policy summary table
        lines += [
            "| UID | Line of Business | Carrier | Policy # | Effective | Expires | Premium | Limit | Ded. | Status |",
            "|-----|------------------|---------|----------|-----------|---------|---------|-------|------|--------|",
        ]
        for p in group:
            lines.append(
                f"| {p['policy_uid']} | {p['policy_type']} | {p['carrier']}"
                f" | {p.get('policy_number') or '—'} | {p['effective_date']} | {p['expiration_date']}"
                f" | {fmt_currency(p['premium'])} | {fmt_limit(p.get('limit_amount'))}"
                f" | {fmt_limit(p.get('deductible'))} | {p['renewal_status']} |"
            )
        lines.append("")

        # Per-policy narrative detail (description, placement, team, internal notes)
        for p in group:
            details = []
            if p.get("first_named_insured") and p["first_named_insured"] != client["name"]:
                details.append(f"First Named Insured: {p['first_named_insured']}")
            if p.get("access_point"):
                details.append(f"Access Point: {p['access_point']}")
            if p.get("description"):
                details.append(p["description"])
            # Policy team contacts (new system)
            team = policy_contacts_map.get(p.get("id", 0), [])
            if team:
                team_parts = []
                for tm in team:
                    part = tm["name"]
                    if tm.get("role"):
                        part += f" ({tm['role']})"
                    if tm.get("email"):
                        part += f" <{tm['email']}>"
                    team_parts.append(part)
                details.append("Team: " + "; ".join(team_parts))
            elif p.get("placement_colleague"):
                col = f"Placement: {p['placement_colleague']}"
                if p.get("placement_colleague_email"):
                    col += f" ({p['placement_colleague_email']})"
                details.append(col)
            if p.get("underwriter_name"):
                uw = f"Underwriter: {p['underwriter_name']}"
                if p.get("underwriter_contact"):
                    uw += f" ({p['underwriter_contact']})"
                details.append(uw)
            if p.get("notes"):
                details.append(f"Internal notes: {p['notes']}")
            if details:
                lines.append(f"**{p['policy_uid']} — {p['policy_type']}:** " + " | ".join(details))
        has_details = any(
            p.get("description") or p.get("placement_colleague") or p.get("notes")
            or p.get("underwriter_name") or p.get("first_named_insured")
            or p.get("access_point") or policy_contacts_map.get(p.get("id", 0))
            for p in group
        )
        if has_details:
            lines.append("")

        # Tower structure for this project (inline)
        proj_towers = tower_by_project.get(proj, {})
        if proj_towers:
            for tg_name in sorted(proj_towers.keys()):
                layers = sorted(
                    proj_towers[tg_name],
                    key=lambda lp: lp.get("attachment_point") or 0,
                    reverse=True,
                )
                total_tower_premium = sum(lp.get("premium") or 0 for lp in layers)
                lines += [
                    f"**Tower: {tg_name}** ({fmt_currency(total_tower_premium)} total, {len(layers)} carrier{'s' if len(layers) != 1 else ''})",
                    "",
                    "| Layer | Carrier | Limit | Premium | Colleague |",
                    "|-------|---------|-------|---------|-----------|",
                ]
                for lp in layers:
                    notation = _ln(lp.get("limit_amount"), lp.get("attachment_point"), lp.get("participation_of"))
                    lines.append(
                        f"| {notation or lp.get('layer_position', 'Primary')}"
                        f" | {lp['carrier']} | {fmt_limit(lp.get('limit_amount'))}"
                        f" | {fmt_currency(lp['premium'])} | {lp.get('placement_colleague') or '—'} |"
                    )
                lines.append("")

    # ─── Premium History ──────────────────────────────────────────────────────
    if history:
        lines += ["## Premium History", ""]
        current_type = None
        for r in history:
            if r["policy_type"] != current_type:
                current_type = r["policy_type"]
                lines += [
                    f"### {current_type}",
                    "",
                    "| Term | Carrier | Premium | Limit | Notes |",
                    "|------|---------|---------|-------|-------|",
                ]
            lines.append(
                f"| {r['term_effective']} → {r['term_expiration']}"
                f" | {r['carrier'] or '—'} | {fmt_currency(r['premium'])}"
                f" | {fmt_limit(r['limit_amount'])} | {r['notes'] or '—'} |"
            )
        lines.append("")

    # ─── Account Activity ────────────────────────────────────────────────────
    if activities:
        lines += [
            "## Account Activity (Last 180 Days)",
            "",
            "| Date | Type | Policy | Contact | Subject | Follow-Up |",
            "|------|------|--------|---------|---------|-----------|",
        ]
        detail_blocks = []
        for a in activities:
            subject = (a["subject"] or "").replace("|", "\\|")
            policy_ref = (a["policy_uid"] or "")
            lines.append(
                f"| {a['activity_date']} | {a['activity_type']}"
                f" | {policy_ref or '—'} | {a['contact_person'] or '—'} | {subject}"
                f" | {a['follow_up_date'] or '—'} |"
            )
            if a["details"] and a["details"].strip():
                detail_blocks.append((a["activity_date"], a["subject"], a["details"].strip(), policy_ref))
        lines.append("")
        if detail_blocks:
            lines += ["### Activity Notes", ""]
            for date, subject, details, policy_ref in detail_blocks:
                header = f"**{date} — {subject}**"
                if policy_ref:
                    header += f" [{policy_ref}]"
                lines.append(header)
                lines.append(f"{details}")
                lines.append("")

    # ─── Opportunities ────────────────────────────────────────────────────────
    opportunities = [p for p in [dict(r) for r in policies] if p.get("is_opportunity")]
    if opportunities:
        lines += ["## Opportunities / New Business", "", "| UID | Line | Status | Target Effective | Premium |",
                  "|-----|------|--------|-----------------|---------|"]
        for op in opportunities:
            lines.append(
                f"| {op['policy_uid']} | {op['policy_type']} | {op.get('opportunity_status') or '—'}"
                f" | {op.get('target_effective_date') or '—'} | {fmt_currency(op['premium'])} |"
            )
        lines.append("")

    # ─── Open Follow-Ups ─────────────────────────────────────────────────────
    if overdue or upcoming_followups:
        lines += ["## Open Follow-Ups", ""]
        if overdue:
            lines += ["**Overdue:**", ""]
            for o in overdue:
                lines.append(
                    f"- {o['follow_up_date']} ({o['days_overdue']}d overdue) — "
                    f"{o['activity_type']}: {o['subject']}"
                )
            lines.append("")
        if upcoming_followups:
            lines += ["**Upcoming:**", ""]
            for u in upcoming_followups:
                u = dict(u)
                entry = f"- {u['follow_up_date']} — {u['activity_type']}: {u['subject']}"
                if u.get("contact_person"):
                    entry += f" (w/ {u['contact_person']})"
                if u.get("policy_uid"):
                    entry += f" [{u['policy_uid']}]"
                lines.append(entry)
            lines.append("")

    return "\n".join(lines)


def export_llm_client_json(conn: sqlite3.Connection, client_id: int) -> str:
    client = conn.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    summary = conn.execute("SELECT * FROM v_client_summary WHERE id = ?", (client_id,)).fetchone()
    policies = conn.execute(
        "SELECT * FROM v_policy_status WHERE client_id = ?", (client_id,)
    ).fetchall()
    activities = conn.execute(
        """SELECT a.* FROM activity_log a WHERE a.client_id = ?
           AND a.activity_date >= date('now', '-90 days')
           ORDER BY a.activity_date DESC""",
        (client_id,),
    ).fetchall()
    overdue = conn.execute(
        "SELECT * FROM v_overdue_followups WHERE client_name = ?", (client["name"],)
    ).fetchall()
    upcoming_fups = conn.execute(
        """SELECT a.id, a.activity_date, a.activity_type, a.subject, a.follow_up_date,
                  a.contact_person, p.policy_uid
           FROM activity_log a
           LEFT JOIN policies p ON a.policy_id = p.id
           WHERE a.client_id = ? AND a.follow_up_date >= date('now') AND a.follow_up_done = 0
           ORDER BY a.follow_up_date ASC""",
        (client_id,),
    ).fetchall()
    history = conn.execute(
        "SELECT * FROM premium_history WHERE client_id = ? ORDER BY policy_type, term_effective DESC",
        (client_id,),
    ).fetchall()
    proj_notes = conn.execute(
        "SELECT name AS project_name, notes FROM projects WHERE client_id = ? AND notes != '' ORDER BY name",
        (client_id,),
    ).fetchall()
    contacts = conn.execute(
        "SELECT name, title, email, phone, role, assignment, contact_type, notes, is_primary FROM client_contacts WHERE client_id = ? ORDER BY is_primary DESC, name",
        (client_id,),
    ).fetchall()
    policy_contacts = conn.execute(
        """SELECT pc.policy_id, pc.name, pc.email, pc.phone, pc.role, p.policy_uid
           FROM policy_contacts pc JOIN policies p ON pc.policy_id = p.id
           WHERE p.client_id = ? ORDER BY p.policy_uid, pc.name""",
        (client_id,),
    ).fetchall()
    scratchpad = conn.execute(
        "SELECT content, updated_at FROM client_scratchpad WHERE client_id = ?", (client_id,)
    ).fetchone()
    audit = build_program_audit(conn, client_id)

    # Group policy_contacts by policy_uid
    from collections import defaultdict as _dd2
    pc_by_uid: dict = _dd2(list)
    for pc in policy_contacts:
        pc_by_uid[pc["policy_uid"]].append(dict(pc))

    data = {
        "metadata": {
            "export_type": "client_program",
            "date": TODAY,
            "client": client["name"],
        },
        "client": dict(client),
        "summary": dict(summary) if summary else {},
        "contacts": [dict(c) for c in contacts],
        "working_notes": scratchpad["content"] if scratchpad else "",
        "project_notes": [dict(r) for r in proj_notes],
        "policies": [
            {
                **dict(p),
                "team": pc_by_uid.get(p["policy_uid"], []),
                "computed": {
                    "days_to_renewal": p["days_to_renewal"],
                    "urgency": p["urgency"],
                    "commission_amount": p["commission_amount"],
                    "rate_change": p["rate_change"],
                },
            }
            for p in policies
        ],
        "coverage_analysis": {
            "gaps": audit["gap_observations"],
            "tower_count": audit["tower_count"],
            "standalone_count": audit["standalone_count"],
            "duplicate_count": audit["duplicate_count"],
        },
        "premium_history": [dict(r) for r in history],
        "activities": [dict(a) for a in activities],
        "overdue_followups": [dict(o) for o in overdue],
        "upcoming_followups": [dict(u) for u in upcoming_fups],
    }
    return json.dumps(data, indent=2, default=str)


# ─── LLM BOOK EXPORT ─────────────────────────────────────────────────────────

def export_llm_book_md(conn: sqlite3.Connection) -> str:
    clients = conn.execute("SELECT * FROM v_client_summary ORDER BY name").fetchall()
    all_policies = conn.execute("SELECT * FROM v_policy_status").fetchall()
    pipeline = conn.execute(
        "SELECT * FROM v_renewal_pipeline ORDER BY expiration_date"
    ).fetchall()
    overdue = conn.execute("SELECT * FROM v_overdue_followups").fetchall()

    total_premium = sum(c["total_premium"] for c in clients)
    total_commission = sum(c["total_commission"] for c in clients)

    urgency_counts: dict[str, dict] = {}
    for p in all_policies:
        u = p["urgency"]
        if u not in urgency_counts:
            urgency_counts[u] = {"count": 0, "premium": 0}
        urgency_counts[u]["count"] += 1
        urgency_counts[u]["premium"] += p["premium"] or 0

    lines = [
        "---",
        "export_type: book_of_business",
        f'account_executive: "{cfg.get("default_account_exec", "Grant")}"',
        f'export_date: "{TODAY}"',
        f"total_clients: {len(clients)}",
        f"total_policies: {len(all_policies)}",
        f"total_premium: {fmt_currency(total_premium)}",
        f"total_commission: {fmt_currency(total_commission)}",
        f"urgent_renewals: {urgency_counts.get('URGENT', {}).get('count', 0)}",
        "---",
        "",
        "# Book of Business Summary",
        "",
        "## Key Metrics",
        f"- Clients: {len(clients)}",
        f"- Policies: {len(all_policies)}",
        f"- Total premium: {fmt_currency(total_premium)}",
        f"- Total commission: {fmt_currency(total_commission)}",
        "",
        "## Urgency Summary",
        "",
    ]
    for urgency_level in ["EXPIRED", "URGENT", "WARNING", "UPCOMING", "OK"]:
        data = urgency_counts.get(urgency_level, {"count": 0, "premium": 0})
        lines.append(f"- {urgency_level}: {data['count']} policies, {fmt_currency(data['premium'])}")
    lines.append("")

    # Client summaries
    lines += ["## Client Summaries", ""]
    lines += [
        "| Client | Segment | Policies | Premium | Next Renewal | Flags |",
        "|--------|---------|----------|---------|--------------|-------|",
    ]
    for c in clients:
        flags = []
        if c["next_renewal_days"] is not None and c["next_renewal_days"] <= 90:
            flags.append("URGENT RENEWAL")
        if c["activity_last_90d"] == 0:
            flags.append("NO RECENT ACTIVITY")
        lines.append(
            f"| {c['name']} | {c['industry_segment']}"
            f" | {c['total_policies']} | {fmt_currency(c['total_premium'])}"
            f" | {fmt_days(c['next_renewal_days'])} | {', '.join(flags) or '—'} |"
        )
    lines.append("")

    # Stale renewals
    stale = [
        p for p in pipeline
        if p["renewal_status"] == "Not Started"
    ]
    if stale:
        lines += ["## Stale Renewals (Not Started — Within 180d)", ""]
        for p in stale:
            lines.append(f"- {p['client_name']} / {p['policy_type']} ({p['policy_uid']}) — expires {p['expiration_date']}, {p['days_to_renewal']}d")
        lines.append("")

    # Overdue follow-ups
    if overdue:
        lines += ["## Overdue Follow-Ups", ""]
        for o in overdue:
            lines.append(f"- {o['client_name']}: {o['subject']} (due {o['follow_up_date']}, {o['days_overdue']}d overdue)")
        lines.append("")

    return "\n".join(lines)


def export_llm_book_json(conn: sqlite3.Connection) -> str:
    clients = conn.execute("SELECT * FROM v_client_summary ORDER BY name").fetchall()
    all_policies = conn.execute("SELECT * FROM v_policy_status").fetchall()
    pipeline = conn.execute("SELECT * FROM v_renewal_pipeline ORDER BY expiration_date").fetchall()
    overdue = conn.execute("SELECT * FROM v_overdue_followups").fetchall()

    return json.dumps({
        "metadata": {"export_type": "book_of_business", "date": TODAY},
        "clients": [dict(c) for c in clients],
        "policies": [dict(p) for p in all_policies],
        "renewal_pipeline": [dict(p) for p in pipeline],
        "overdue_followups": [dict(o) for o in overdue],
    }, indent=2, default=str)


# ─── CLIENT EXPORT ────────────────────────────────────────────────────────────

def export_client_md(conn: sqlite3.Connection, client_id: int) -> str:
    """Full client profile + policies in Markdown."""
    return export_llm_client_md(conn, client_id)


def export_client_json(conn: sqlite3.Connection, client_id: int) -> str:
    return export_llm_client_json(conn, client_id)


def export_client_csv(conn: sqlite3.Connection, client_id: int) -> str:
    rows = conn.execute(
        "SELECT * FROM v_policy_status WHERE client_id = ? ORDER BY policy_type",
        (client_id,),
    ).fetchall()
    if not rows:
        return ""
    buf = io.StringIO()
    cols = list(rows[0].keys())
    writer = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow(dict(r))
    return buf.getvalue()


# ─── RENEWAL EXPORT ───────────────────────────────────────────────────────────

def export_renewals_md(conn: sqlite3.Connection, window_days: int = 180) -> str:
    rows = conn.execute(
        """SELECT * FROM v_renewal_pipeline WHERE days_to_renewal <= ?
           ORDER BY expiration_date""",
        (window_days,),
    ).fetchall()

    # Build policy_contacts map keyed by policy_uid
    pc_rows = conn.execute(
        """SELECT p.policy_uid, pc.name, pc.role, pc.email
           FROM policy_contacts pc JOIN policies p ON pc.policy_id = p.id
           WHERE p.policy_uid IN (SELECT policy_uid FROM v_renewal_pipeline WHERE days_to_renewal <= ?)
           ORDER BY pc.name""",
        (window_days,),
    ).fetchall()
    from collections import defaultdict as _ddr
    pc_map: dict = _ddr(list)
    for pc in pc_rows:
        pc_map[pc["policy_uid"]].append(pc["name"])

    total = sum(r["premium"] or 0 for r in rows)
    lines = [
        f"# Renewal Pipeline — Next {window_days} Days",
        "",
        f"**As of:** {TODAY}  ",
        f"**Policies:** {len(rows)}  ",
        f"**Premium at Risk:** {fmt_currency(total)}",
        "",
        "| UID | Client | Line | Carrier | Expires | Days | Urgency | Premium | Status | Team |",
        "|-----|--------|------|---------|---------|------|---------|---------|--------|------|",
    ]
    for r in rows:
        team = pc_map.get(r["policy_uid"])
        team_str = "; ".join(team) if team else (r["placement_colleague"] or "—")
        lines.append(
            f"| {r['policy_uid']} | {r['client_name']} | {r['policy_type']}"
            f" | {r['carrier']} | {r['expiration_date']} | {r['days_to_renewal']}"
            f" | {r['urgency']} | {fmt_currency(r['premium'])}"
            f" | {r['renewal_status']} | {team_str} |"
        )
    return "\n".join(lines)


def export_renewals_json(conn: sqlite3.Connection, window_days: int = 180) -> str:
    rows = conn.execute(
        """SELECT * FROM v_renewal_pipeline WHERE days_to_renewal <= ?
           ORDER BY expiration_date""",
        (window_days,),
    ).fetchall()
    return json.dumps({
        "window_days": window_days,
        "export_date": TODAY,
        "renewals": [dict(r) for r in rows],
    }, indent=2, default=str)


def export_renewals_csv(conn: sqlite3.Connection, window_days: int = 180) -> str:
    rows = conn.execute(
        """SELECT * FROM v_renewal_pipeline WHERE days_to_renewal <= ?
           ORDER BY expiration_date""",
        (window_days,),
    ).fetchall()
    if not rows:
        return ""
    buf = io.StringIO()
    cols = list(rows[0].keys())
    writer = csv.DictWriter(buf, fieldnames=cols)
    writer.writeheader()
    for r in rows:
        writer.writerow(dict(r))
    return buf.getvalue()


# ─── XLSX HELPERS ────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_CURRENCY_FMT = '"$"#,##0.00'
_CURRENCY_COLS = {
    "Premium", "Limit", "Deductible", "premium", "limit_amount", "deductible",
    "prior_premium", "commission_amount", "exposure_amount",
}


def _write_sheet(wb: Workbook, title: str, rows: list) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["No data"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row[k] for k in headers])

    # Currency formatting
    for col_idx, col_name in enumerate(headers, 1):
        if col_name in _CURRENCY_COLS:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = _CURRENCY_FMT

    # Auto-size columns
    for col_idx, col_name in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *(len(str(row[col_name] or "")) for row in rows),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_schedule_xlsx(conn: sqlite3.Connection, client_id: int, client_name: str) -> bytes:
    rows = _schedule_rows_for_client(conn, client_id)
    # Strip client_name column — already in the filename
    cleaned = [{k: r[k] for k in r.keys() if k != "client_name"} for r in rows]
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Schedule of Insurance", cleaned)
    # Summary row
    ws = wb["Schedule of Insurance"]
    ws.append([])
    total = sum(r["Premium"] or 0 for r in rows)
    ws.append(["Total Annual Premium", fmt_currency(total)])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    # Project Notes sheet (if any)
    notes_rows = conn.execute(
        "SELECT name AS \"Project / Location\", notes AS \"Notes\" FROM projects WHERE client_id = ? AND notes != '' ORDER BY name",
        (client_id,),
    ).fetchall()
    if notes_rows:
        _write_sheet(wb, "Project Notes", [dict(r) for r in notes_rows])

    return _wb_to_bytes(wb)


def export_client_xlsx(conn: sqlite3.Connection, client_id: int) -> bytes:
    policies = conn.execute(
        "SELECT * FROM v_policy_status WHERE client_id = ? ORDER BY policy_type, layer_position",
        (client_id,),
    ).fetchall()
    history = conn.execute(
        "SELECT * FROM premium_history WHERE client_id = ? ORDER BY policy_type, term_effective DESC",
        (client_id,),
    ).fetchall()
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Policies", [dict(r) for r in policies])
    _write_sheet(wb, "Premium History", [dict(r) for r in history])
    return _wb_to_bytes(wb)


def export_full_xlsx(conn: sqlite3.Connection, client_id: int, client_name: str) -> bytes:
    """Full internal data export: all policy fields + contacts + notes + activities."""
    policies = conn.execute(
        """SELECT policy_uid, policy_type, carrier, policy_number,
                  effective_date, expiration_date, premium, limit_amount, deductible,
                  description, coverage_form, layer_position, tower_group, is_standalone,
                  project_name, renewal_status, commission_rate, commission_amount,
                  prior_premium, rate_change,
                  placement_colleague, placement_colleague_email,
                  underwriter_name, underwriter_contact,
                  follow_up_date, attachment_point, participation_of,
                  exposure_basis, exposure_amount, exposure_unit,
                  exposure_address, exposure_city, exposure_state, exposure_zip,
                  notes, account_exec, urgency, days_to_renewal,
                  first_named_insured, access_point
           FROM v_policy_status WHERE client_id = ?
           ORDER BY project_name, policy_type, layer_position""",
        (client_id,),
    ).fetchall()

    contacts = conn.execute(
        "SELECT name, title, role, assignment, contact_type, email, phone, notes, is_primary, created_at FROM client_contacts WHERE client_id = ? ORDER BY is_primary DESC, name",
        (client_id,),
    ).fetchall()

    policy_team = conn.execute(
        """SELECT p.policy_uid, p.policy_type, pc.name, pc.role, pc.email, pc.phone
           FROM policy_contacts pc JOIN policies p ON pc.policy_id = p.id
           WHERE p.client_id = ? ORDER BY p.policy_uid, pc.name""",
        (client_id,),
    ).fetchall()

    project_notes = conn.execute(
        """SELECT name AS project_name, notes, created_at, updated_at
           FROM projects WHERE client_id = ? ORDER BY name""",
        (client_id,),
    ).fetchall()

    activities = conn.execute(
        """SELECT activity_date, activity_type, contact_person, subject, details,
                  follow_up_date, follow_up_done, account_exec, created_at
           FROM activity_log WHERE client_id = ?
           ORDER BY activity_date DESC""",
        (client_id,),
    ).fetchall()

    history = conn.execute(
        "SELECT policy_type, term_effective, term_expiration, carrier, premium, limit_amount, deductible, notes FROM premium_history WHERE client_id = ? ORDER BY policy_type, term_effective DESC",
        (client_id,),
    ).fetchall()

    scratchpad = conn.execute(
        "SELECT content, updated_at FROM client_scratchpad WHERE client_id = ?",
        (client_id,),
    ).fetchone()

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Policies (Full)", [dict(r) for r in policies])
    _write_sheet(wb, "Contacts", [dict(r) for r in contacts])
    _write_sheet(wb, "Policy Team", [dict(r) for r in policy_team])
    _write_sheet(wb, "Project Notes", [dict(r) for r in project_notes])
    _write_sheet(wb, "Activities", [dict(r) for r in activities])
    _write_sheet(wb, "Premium History", [dict(r) for r in history])

    # Working Notes as a simple text sheet
    ws = wb.create_sheet("Working Notes")
    ws.append(["Working Notes"])
    ws["A1"].font = Font(bold=True)
    if scratchpad and scratchpad["content"]:
        for line in scratchpad["content"].splitlines():
            ws.append([line])
        ws.append([])
        ws.append([f"Last updated: {scratchpad['updated_at']}"])
    else:
        ws.append(["(no working notes)"])
    ws.column_dimensions["A"].width = 80

    return _wb_to_bytes(wb)


def export_renewals_xlsx(conn: sqlite3.Connection, window_days: int = 180) -> bytes:
    rows = conn.execute(
        """SELECT * FROM v_renewal_pipeline WHERE days_to_renewal <= ?
           ORDER BY expiration_date""",
        (window_days,),
    ).fetchall()
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, f"Renewals — Next {window_days}d", [dict(r) for r in rows])
    return _wb_to_bytes(wb)


# ─── SAVE HELPER ─────────────────────────────────────────────────────────────

def save_export(content: str, filename: str) -> Path:
    exports_dir = Path(cfg.get("export_dir", str(Path.home() / ".policydb" / "exports")))
    exports_dir.mkdir(parents=True, exist_ok=True)
    out = exports_dir / filename
    out.write_text(content)
    return out


def save_export_bytes(content: bytes, filename: str) -> Path:
    exports_dir = Path(cfg.get("export_dir", str(Path.home() / ".policydb" / "exports")))
    exports_dir.mkdir(parents=True, exist_ok=True)
    out = exports_dir / filename
    out.write_bytes(content)
    return out
