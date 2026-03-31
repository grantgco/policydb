#!/usr/bin/env python3
"""
Generate a sample XLSX export showing the standard PolicyDB/Marsh brand styling.

Run:  python templates/sample-export.xlsx.py
Out:  templates/sample-export.xlsx

This demonstrates the brand palette, fonts, alternating rows, currency formatting,
compliance status fills, and column width patterns used across all PolicyDB exports.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─── MARSH BRAND PALETTE ────────────────────────────────────────────────────
# These constants are defined in src/policydb/exporter.py and must stay in sync.

HEADER_FILL = PatternFill("solid", fgColor="003865")       # Navy #003865
HEADER_FONT = Font(name="Noto Sans", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Noto Sans", size=11, color="3D3C37")  # Charcoal #3D3C37
ALT_ROW_FILL = PatternFill("solid", fgColor="F7F3EE")       # Warm cream #F7F3EE
BORDER_COLOR = "B9B6B1"                                      # Warm gray #B9B6B1
THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

CURRENCY_FMT = '"$"#,##0.00'

# ─── COMPLIANCE STATUS FILLS ────────────────────────────────────────────────
# Used on the Compliance Matrix sheet for status-based cell coloring.

COMPLIANCE_FILLS = {
    "Compliant": PatternFill("solid", fgColor="C6EFCE"),     # Light green
    "Gap": PatternFill("solid", fgColor="FFC7CE"),           # Light red
    "Partial": PatternFill("solid", fgColor="FFEB9C"),       # Light yellow
    "N/A": PatternFill("solid", fgColor="D9D2E9"),           # Light purple
    "Needs Review": PatternFill("solid", fgColor="D9D9D9"),  # Light gray
    "Waived": PatternFill("solid", fgColor="D9D9D9"),        # Light gray
}

BOLD = Font(bold=True)
BOLD_WHITE = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, horizontal="center", vertical="center")
WRAP_RIGHT = Alignment(wrap_text=True, horizontal="right")


def build_sample_workbook() -> Workbook:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ── Sheet 1: Policy Schedule (standard data table) ───────────────────
    ws = wb.create_sheet("Policy Schedule")
    headers = [
        "Line of Business", "Carrier", "Policy #", "Effective",
        "Expiration", "Premium", "Limit", "Deductible", "Description",
    ]
    currency_cols = {"Premium", "Limit", "Deductible"}

    sample_rows = [
        ["General Liability", "Zurich via Marsh", "GL-2026-001", "01/01/2026",
         "01/01/2027", 125000, 1000000, 10000, "Commercial General Liability"],
        ["Property", "Chubb", "PR-2026-042", "03/15/2026",
         "03/15/2027", 87500, 5000000, 25000, "All-Risk Property"],
        ["Workers Compensation", "Hartford", "WC-2026-010", "01/01/2026",
         "01/01/2027", 210000, 1000000, 0, "Statutory WC + Employers Liability"],
        ["Commercial Auto", "Travelers", "CA-2026-005", "06/01/2026",
         "06/01/2027", 45000, 1000000, 5000, "Fleet Auto — 12 vehicles"],
        ["Umbrella", "AIG via Marsh", "UMB-2026-003", "01/01/2026",
         "01/01/2027", 62000, 10000000, 0, "Follow-form Umbrella"],
        ["Professional Liability", "Berkley", "PL-2026-008", "04/01/2026",
         "04/01/2027", 38000, 2000000, 15000, "E&O / Professional Liability"],
    ]

    # Header row
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for row_data in sample_rows:
        ws.append(row_data)

    for col_idx, col_name in enumerate(headers, 1):
        is_currency = col_name in currency_cols
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_RIGHT if is_currency else WRAP
            if is_currency:
                cell.number_format = CURRENCY_FMT
            if (row_idx - 2) % 2 == 1:
                cell.fill = ALT_ROW_FILL

    # Column widths
    col_widths = [22, 24, 18, 14, 14, 16, 16, 16, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Compliance Matrix (status-colored cells) ────────────────
    ws2 = wb.create_sheet("Compliance Matrix")
    matrix_headers = ["Coverage Line", "Location A", "Location B", "Location C"]
    ws2.append(matrix_headers)
    for cell in ws2[1]:
        cell.font = BOLD_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = WRAP_CENTER

    matrix_data = [
        ["General Liability", "Compliant", "Gap", "Compliant"],
        ["Property", "Compliant", "Compliant", "Partial"],
        ["Workers Compensation", "Compliant", "Compliant", "Compliant"],
        ["Commercial Auto", "Gap", "Needs Review", "N/A"],
        ["Umbrella / Excess", "Partial", "Compliant", "Waived"],
        ["Builders Risk", "N/A", "Gap", "Needs Review"],
    ]

    for row_data in matrix_data:
        ws2.append(row_data)

    for row_idx in range(2, ws2.max_row + 1):
        # Coverage name column
        ws2.cell(row=row_idx, column=1).font = DATA_FONT
        ws2.cell(row=row_idx, column=1).border = THIN_BORDER
        ws2.cell(row=row_idx, column=1).alignment = WRAP
        # Status columns with compliance fills
        for col_idx in range(2, 5):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_CENTER
            status = cell.value
            if status in COMPLIANCE_FILLS:
                cell.fill = COMPLIANCE_FILLS[status]

    ws2.column_dimensions["A"].width = 30
    for col in ["B", "C", "D"]:
        ws2.column_dimensions[col].width = 18

    # ── Sheet 3: Executive Summary (label-value layout) ──────────────────
    ws3 = wb.create_sheet("Executive Summary")
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 50

    summary_data = [
        ("Client", "Acme Construction Corp"),
        ("Account #", "CN123456789"),
        ("Account Executive", "Grant Greeson"),
        ("Industry", "Construction — General Contractor"),
        ("Total Policies", "6"),
        ("Total Premium", "$567,500"),
        ("Renewal Window", "180 days"),
        ("Next Expiration", "01/01/2027"),
        ("Compliance Status", "83% Compliant"),
        ("Data Health Score", "92 / 100"),
    ]

    for label, value in summary_data:
        ws3.append([label, value])

    for row_idx in range(1, ws3.max_row + 1):
        ws3.cell(row=row_idx, column=1).font = Font(name="Noto Sans", bold=True, size=11, color="003865")
        ws3.cell(row=row_idx, column=2).font = DATA_FONT

    return wb


if __name__ == "__main__":
    import os
    wb = build_sample_workbook()
    out = os.path.join(os.path.dirname(__file__), "sample-export.xlsx")
    wb.save(out)
    print(f"Saved: {out}")
