"""Client routes."""

from __future__ import annotations

import logging
logger = logging.getLogger("policydb.web.routes.clients")

from fastapi import APIRouter, Depends, Form, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates

from babel.dates import format_datetime as babel_format_datetime
from collections import Counter
from datetime import datetime, timedelta

from policydb import config as cfg
from policydb.data_health import score_client
from policydb.utils import clean_email, format_fein, format_phone, normalize_client_name, format_city, format_state, format_zip
from policydb.queries import (
    get_activities,
    get_all_clients,
    get_client_by_id,
    get_client_contacts,
    get_client_summary,
    get_client_total_hours,
    get_or_create_contact,
    assign_contact_to_client,
    assign_contact_to_policy,
    remove_contact_from_client,
    set_primary_contact,
    get_linked_group_for_client,
    get_linked_group_overview,
    get_policies_for_client,
    get_saved_notes,
    get_saved_notes_for_client_timeline,
    save_note,
    delete_saved_note,
    full_text_search,
    create_linked_group,
    add_client_to_group,
    remove_client_from_group,
    update_linked_group,
    delete_linked_group,
    get_client_exposures,
    get_exposure_years,
    get_distinct_custom_exposure_types,
    get_exposure_observations,
    get_exposure_by_id,
    attach_open_issues,
)
from policydb.web.app import get_db, templates

router = APIRouter(prefix="/clients")


def _pinned_notes_for_page(conn, scope, scope_id, client_id=None):
    from policydb.web.routes.pinned_notes import get_pinned_notes_with_cascade
    return get_pinned_notes_with_cascade(conn, scope, scope_id, client_id=client_id)


def _get_all_client_contact_orgs(conn):
    """Get all distinct organization values from contacts."""
    rows = conn.execute(
        "SELECT DISTINCT organization FROM contacts WHERE organization IS NOT NULL AND organization != '' ORDER BY organization"
    ).fetchall()
    return [r["organization"] for r in rows]


def _find_similar_clients(conn, name: str, threshold: int = 85) -> list[dict]:
    """Find existing clients with names similar to the given name using fuzzy matching."""
    from rapidfuzz import fuzz
    normalized = normalize_client_name(name)
    existing = conn.execute(
        "SELECT id, name, industry_segment FROM clients WHERE archived = 0"
    ).fetchall()
    matches = []
    for r in existing:
        score = fuzz.WRatio(normalized, r["name"])
        if score >= threshold:
            matches.append({"id": r["id"], "name": r["name"],
                           "industry": r["industry_segment"], "score": round(score)})
    return sorted(matches, key=lambda x: -x["score"])


_CLIENT_SORT_FIELDS = {
    "name", "industry_segment", "total_policies", "total_premium",
    "total_revenue", "next_renewal_days", "activity_last_90d",
    "last_activity_date",
}


def _relative_days(delta: int) -> str:
    """Return a human-friendly relative time string for a number of days ago."""
    if delta == 0:
        return "Today"
    if delta == 1:
        return "1d ago"
    if delta < 7:
        return f"{delta}d ago"
    if delta < 14:
        return "1w ago"
    if delta < 21:
        return "2w ago"
    if delta < 28:
        return "3w ago"
    if delta < 60:
        return "1mo ago"
    if delta < 90:
        return "2mo ago"
    return None  # caller will format as month+day


def _enrich_last_activity(clients: list[dict]) -> None:
    """Add last_activity_ago and last_activity_urgency to each client dict."""
    from datetime import date as _date
    today = _date.today()
    for c in clients:
        raw = c.get("last_activity_date")
        if raw:
            d = _date.fromisoformat(raw)
            delta = (today - d).days
            rel = _relative_days(delta)
            if rel is None:
                # >90 days — show abbreviated date like "Mar 12"
                c["last_activity_ago"] = d.strftime("%b %-d")
            else:
                c["last_activity_ago"] = rel
            if delta < 14:
                c["last_activity_urgency"] = "green"
            elif delta < 60:
                c["last_activity_urgency"] = "neutral"
            elif delta <= 90:
                c["last_activity_urgency"] = "amber"
            else:
                c["last_activity_urgency"] = "red"
        else:
            c["last_activity_ago"] = "Never"
            c["last_activity_urgency"] = "neutral"


def _apply_client_filters(clients, segment="", urgent="", inactive="", prospect=""):
    if segment:
        clients = [c for c in clients if c["industry_segment"] == segment]
    if urgent:
        clients = [c for c in clients if (c.get("next_renewal_days") or 999) <= 90]
    if inactive:
        clients = [c for c in clients if (c.get("activity_last_90d") or 0) == 0]
    if prospect:
        clients = [c for c in clients if c.get("is_prospect")]
    return clients


def _sort_clients(clients, sort="name", dir="asc"):
    field = sort if sort in _CLIENT_SORT_FIELDS else "name"
    reverse = dir == "desc"
    clients.sort(
        key=lambda c: (c.get(field) is None, c.get(field) if c.get(field) is not None else ""),
        reverse=reverse,
    )
    return clients


def _get_us_states():
    from policydb.web.routes.policies import US_STATES
    return US_STATES


def _get_project_locations(conn, client_id: int) -> list[dict]:
    """Load all location-type projects with policy/opportunity counts, premium, and revenue."""
    rows = conn.execute("""
        SELECT p.id, p.name, p.address, p.city, p.state, p.zip, p.notes,
               (SELECT COUNT(*) FROM policies pol
                WHERE pol.project_id = p.id AND pol.archived = 0) AS total_coverages,
               (SELECT COUNT(*) FROM policies pol
                WHERE pol.project_id = p.id AND pol.archived = 0
                AND (pol.is_opportunity = 0 OR pol.is_opportunity IS NULL)) AS placed_coverages,
               (SELECT COALESCE(SUM(pol.premium), 0) FROM policies pol
                WHERE pol.project_id = p.id AND pol.archived = 0) AS total_premium,
               (SELECT COALESCE(SUM(CASE WHEN pol.commission_rate > 0
                THEN pol.premium * pol.commission_rate ELSE 0 END), 0)
                FROM policies pol
                WHERE pol.project_id = p.id AND pol.archived = 0) AS total_revenue
        FROM projects p
        WHERE p.client_id = ? AND (p.project_type = 'Location' OR p.project_type IS NULL)
        ORDER BY p.name
    """, (client_id,)).fetchall()
    return [dict(r) for r in rows]


def _get_project_pipeline(conn, client_id: int) -> list[dict]:
    """Load all non-location projects with computed coverage stats."""
    projects = conn.execute("""
        SELECT p.*,
               (SELECT COUNT(*) FROM policies pol
                WHERE pol.project_id = p.id AND pol.archived = 0) AS total_coverages,
               (SELECT COUNT(*) FROM policies pol
                WHERE pol.project_id = p.id AND pol.archived = 0
                AND (pol.is_opportunity = 0 OR pol.is_opportunity IS NULL)) AS placed_coverages,
               (SELECT COALESCE(SUM(pol.premium), 0) FROM policies pol
                WHERE pol.project_id = p.id AND pol.archived = 0) AS total_premium,
               (SELECT COALESCE(SUM(CASE WHEN pol.commission_rate > 0
                THEN pol.premium * pol.commission_rate ELSE 0 END), 0)
                FROM policies pol
                WHERE pol.project_id = p.id AND pol.archived = 0) AS total_revenue
        FROM projects p
        WHERE p.client_id = ? AND p.project_type != 'Location'
        ORDER BY p.insurance_needed_by, p.start_date, p.name
    """, (client_id,)).fetchall()
    return [dict(r) for r in projects]


def _build_timeline_data(projects: list[dict]) -> list[dict]:
    """Build percentage-based timeline bar data for the project pipeline.

    Returns a list of dicts (one per project with a start or completion date),
    each with: name, left_pct, width_pct, color, ins_marker_pct (or None).
    Returns empty list when fewer than 2 projects have dates.
    """
    from datetime import date as _date_type

    def _parse(ds: str | None) -> _date_type | None:
        if not ds:
            return None
        try:
            return _date_type.fromisoformat(str(ds))
        except (ValueError, TypeError):
            return None

    dated = [
        p for p in projects
        if _parse(p.get("start_date")) or _parse(p.get("target_completion"))
    ]
    if len(dated) < 2:
        return []

    all_dates: list[_date_type] = []
    for p in dated:
        for field in ("start_date", "target_completion", "insurance_needed_by"):
            d = _parse(p.get(field))
            if d:
                all_dates.append(d)

    d_min = min(all_dates)
    d_max = max(all_dates)
    total_days = max((d_max - d_min).days, 1)

    status_colors: dict[str, str] = {
        "Upcoming": "bg-gray-300",
        "Quoting": "bg-blue-400",
        "Bound": "bg-green-400",
        "Active": "bg-green-400",
        "Complete": "bg-gray-200",
    }

    result = []
    for p in dated:
        d_start = _parse(p.get("start_date"))
        d_end = _parse(p.get("target_completion"))
        # Use whichever end we have
        if not d_start and not d_end:
            continue
        if not d_start:
            d_start = d_end
        if not d_end:
            d_end = d_start

        left_pct = round((d_start - d_min).days / total_days * 100, 2)
        width_pct = max(round((d_end - d_start).days / total_days * 100, 2), 1.5)
        color = status_colors.get(p.get("status", ""), "bg-gray-300")

        ins_marker_pct = None
        d_ins = _parse(p.get("insurance_needed_by"))
        if d_ins:
            ins_marker_pct = round((d_ins - d_min).days / total_days * 100, 2)

        result.append({
            "name": p.get("name", ""),
            "left_pct": left_pct,
            "width_pct": width_pct,
            "color": color,
            "ins_marker_pct": ins_marker_pct,
        })

    return result


@router.get("", response_class=HTMLResponse)
def client_list(
    request: Request,
    q: str = "",
    segment: str = "",
    urgent: str = "",
    inactive: str = "",
    prospect: str = "",
    sort: str = "name",
    dir: str = "asc",
    conn=Depends(get_db),
):
    clients = [dict(r) for r in get_all_clients(conn)]

    # Book summary — computed from ALL clients before filtering
    book_stats = {
        "total_clients": len(clients),
        "total_premium": sum(c.get("total_premium") or 0 for c in clients),
        "total_revenue": sum(c.get("total_revenue") or 0 for c in clients),
        "opportunity_count": sum(c.get("opportunity_count") or 0 for c in clients),
        "opportunity_premium": sum(c.get("opportunity_premium") or 0 for c in clients),
        "opportunity_revenue": sum(c.get("opportunity_revenue") or 0 for c in clients),
    }
    segment_counts = Counter(c.get("industry_segment", "Other") or "Other" for c in clients)
    top_segments = segment_counts.most_common(4)
    top_set = set(k for k, v in top_segments)
    other_count = sum(v for k, v in segment_counts.items() if k not in top_set)
    book_stats["segments"] = top_segments
    if other_count > 0 and len(segment_counts) > 4:
        book_stats["other_count"] = other_count

    clients = _apply_client_filters(clients, segment, urgent, inactive, prospect)
    clients = _sort_clients(clients, sort, dir)
    archived_clients = [dict(r) for r in conn.execute(
        """SELECT c.id, c.name, c.industry_segment,
                  COUNT(p.id) AS policy_count
           FROM clients c
           LEFT JOIN policies p ON p.client_id = c.id
           WHERE c.archived = 1
           GROUP BY c.id
           ORDER BY c.name""",
    ).fetchall()]
    linked_client_ids = {r["client_id"] for r in conn.execute(
        "SELECT client_id FROM client_group_members"
    ).fetchall()}
    # Build group membership map: client_id → group_id
    _group_rows = conn.execute(
        """SELECT gm.client_id, gm.group_id, cg.label, c.name AS client_name
           FROM client_group_members gm
           JOIN client_groups cg ON gm.group_id = cg.id
           JOIN clients c ON gm.client_id = c.id
           ORDER BY gm.group_id, c.name"""
    ).fetchall()
    client_group_map = {}  # client_id → group_id
    group_labels = {}  # group_id → label
    group_member_names = {}  # group_id → [client_name, ...]
    for gr in _group_rows:
        client_group_map[gr["client_id"]] = gr["group_id"]
        if gr["group_id"] not in group_labels:
            group_labels[gr["group_id"]] = gr["label"]
        group_member_names.setdefault(gr["group_id"], []).append(gr["client_name"])

    # Re-order clients so grouped ones appear together
    grouped_ids_seen = set()
    ordered_clients = []
    for c in clients:
        if c["id"] in grouped_ids_seen:
            continue
        gid = client_group_map.get(c["id"])
        if gid:
            # Find all clients in this group and insert them together
            group_members = [gc for gc in clients if client_group_map.get(gc["id"]) == gid]
            for gm in group_members:
                if gm["id"] not in grouped_ids_seen:
                    grouped_ids_seen.add(gm["id"])
                    ordered_clients.append(gm)
        else:
            ordered_clients.append(c)
    clients = ordered_clients
    _enrich_last_activity(clients)
    for c in clients:
        score_client(conn, c, include_staleness=False)

    return templates.TemplateResponse("clients/list.html", {
        "request": request,
        "active": "clients",
        "clients": clients,
        "q": q,
        "segment": segment,
        "urgent": urgent,
        "inactive": inactive,
        "prospect": prospect,
        "sort": sort if sort in _CLIENT_SORT_FIELDS else "name",
        "dir": dir,
        "industry_segments": cfg.get("industry_segments", []),
        "archived_clients": archived_clients,
        "linked_client_ids": linked_client_ids,
        "client_group_map": client_group_map,
        "group_labels": group_labels,
        "group_member_names": group_member_names,
        "book_stats": book_stats,
    })


@router.get("/search", response_class=HTMLResponse)
def client_search(
    request: Request,
    q: str = "",
    segment: str = "",
    urgent: str = "",
    inactive: str = "",
    prospect: str = "",
    sort: str = "name",
    dir: str = "asc",
    conn=Depends(get_db),
):
    """HTMX partial: filtered client table rows."""
    if q.strip():
        raw = full_text_search(conn, q.strip())
        client_ids = {r["id"] for r in raw["clients"]}
        all_clients = [dict(r) for r in get_all_clients(conn)]
        clients = [c for c in all_clients if c["id"] in client_ids or q.lower() in c["name"].lower()]
    else:
        clients = [dict(r) for r in get_all_clients(conn)]
    clients = _apply_client_filters(clients, segment, urgent, inactive, prospect)
    clients = _sort_clients(clients, sort, dir)
    _enrich_last_activity(clients)
    return templates.TemplateResponse("clients/_table_rows.html", {
        "request": request,
        "clients": clients,
    })


@router.get("/new", response_class=HTMLResponse)
def client_new_form(request: Request):
    return templates.TemplateResponse("clients/edit.html", {
        "request": request,
        "active": "clients",
        "client": None,
        "industry_segments": cfg.get("industry_segments"),
    })


@router.post("/new")
def client_new_post(
    request: Request,
    name: str = Form(...),
    industry_segment: str = Form(...),
    cn_number: str = Form(""),
    is_prospect: str = Form(""),
    primary_contact: str = Form(""),
    contact_email: str = Form(""),
    contact_phone: str = Form(""),
    contact_mobile: str = Form(""),
    address: str = Form(""),
    notes: str = Form(""),
    broker_fee: str = Form(""),
    business_description: str = Form(""),
    website: str = Form(""),
    renewal_month: str = Form(""),
    client_since: str = Form(""),
    preferred_contact_method: str = Form(""),
    referral_source: str = Form(""),
    latitude: str = Form(""),
    longitude: str = Form(""),
    conn=Depends(get_db),
):
    def _float(v):
        try:
            return float(v) if str(v).strip() else None
        except ValueError:
            return None

    def _int(v):
        try:
            return int(v) if str(v).strip() else None
        except ValueError:
            return None

    account_exec = cfg.get("default_account_exec", "Grant")
    name = normalize_client_name(name) if name else name

    # Duplicate detection: warn if a similar client already exists (unless ?force=1)
    force = request.query_params.get("force", "")
    if not force:
        dupes = _find_similar_clients(conn, name)
        if dupes:
            return templates.TemplateResponse("clients/edit.html", {
                "request": request,
                "active": "clients",
                "client": None,
                "industry_segments": cfg.get("industry_segments"),
                "duplicate_warning": dupes,
                # Pre-fill the form so the user doesn't have to retype
                "prefill": {
                    "name": name,
                    "industry_segment": industry_segment,
                    "cn_number": cn_number,
                    "is_prospect": is_prospect,
                    "primary_contact": primary_contact,
                    "contact_email": contact_email,
                    "contact_phone": contact_phone,
                    "contact_mobile": contact_mobile,
                    "address": address,
                    "notes": notes,
                    "broker_fee": broker_fee,
                    "business_description": business_description,
                    "website": website,
                    "renewal_month": renewal_month,
                    "client_since": client_since,
                    "preferred_contact_method": preferred_contact_method,
                    "referral_source": referral_source,
                    "latitude": latitude,
                    "longitude": longitude,
                },
            })

    cursor = conn.execute(
        """INSERT INTO clients (name, industry_segment, cn_number, is_prospect, primary_contact, contact_email,
           contact_phone, contact_mobile, address, notes, account_exec, broker_fee, business_description,
           website, renewal_month, client_since, preferred_contact_method, referral_source,
           latitude, longitude)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (name, industry_segment, cn_number.strip() or None, 1 if is_prospect else 0,
         primary_contact or None, clean_email(contact_email) or None,
         format_phone(contact_phone) or None, format_phone(contact_mobile) or None,
         address or None, notes or None, account_exec,
         _float(broker_fee), business_description or None,
         website or None, _int(renewal_month), client_since or None,
         preferred_contact_method or None, referral_source or None,
         _float(latitude), _float(longitude)),
    )
    conn.commit()
    logger.info("Client %d created: %s", cursor.lastrowid, name)
    return RedirectResponse(f"/clients/{cursor.lastrowid}", status_code=303)


# ─── CLIENT SPREADSHEET VIEW ────────────────────────────────────────────────


@router.get("/spreadsheet", response_class=HTMLResponse)
def client_spreadsheet(request: Request, conn=Depends(get_db)):
    """Full-book editable client spreadsheet view using Tabulator."""
    from policydb.queries import get_all_clients_for_grid

    rows = get_all_clients_for_grid(conn)

    industry_segments = cfg.get("industry_segments", [])
    risk_levels = cfg.get("relationship_risk_levels", ["None", "Low", "Medium", "High"])
    service_models = cfg.get("service_model_options", ["Standard", "High-touch", "White-glove"])

    columns = [
        {"field": "name", "title": "Client", "width": 200,
         "editor": "input", "headerFilter": "input", "_format": "link"},
        {"field": "cn_number", "title": "Account #", "width": 120,
         "editor": "input", "headerFilter": "input"},
        {"field": "industry_segment", "title": "Industry", "width": 160,
         "editor": "list", "editorParams": {"values": industry_segments, "autocomplete": True, "freetext": True, "listOnEmpty": True},
         "headerFilter": "input"},
        {"field": "account_exec", "title": "Account Exec", "width": 130,
         "editor": "input", "headerFilter": "input"},
        {"field": "date_onboarded", "title": "Onboarded", "width": 115,
         "editor": "date", "_format": "date"},
        {"field": "website", "title": "Website", "width": 160,
         "editor": "input"},
        {"field": "fein", "title": "FEIN", "width": 110,
         "editor": "input"},
        {"field": "broker_fee", "title": "Broker Fee", "width": 110,
         "editor": "number", "editorParams": {"selectContents": True},
         "_format": "currency"},
        {"field": "hourly_rate", "title": "Hourly Rate", "width": 100,
         "editor": "number", "editorParams": {"selectContents": True},
         "_format": "currency"},
        {"field": "follow_up_date", "title": "Follow-Up", "width": 115,
         "editor": "date", "_format": "date"},
        {"field": "relationship_risk", "title": "Risk Level", "width": 110,
         "editor": "list", "editorParams": {"values": risk_levels, "autocomplete": True, "freetext": False, "listOnEmpty": True},
         "headerFilter": "list", "headerFilterParams": {"values": {s: s for s in risk_levels}, "clearable": True}},
        {"field": "service_model", "title": "Service Model", "width": 120,
         "editor": "list", "editorParams": {"values": service_models, "autocomplete": True, "freetext": False, "listOnEmpty": True},
         "headerFilter": "list", "headerFilterParams": {"values": {s: s for s in service_models}, "clearable": True}},
        {"field": "stewardship_date", "title": "Stewardship", "width": 115,
         "editor": "date", "_format": "date"},
        {"field": "renewal_strategy", "title": "Renewal Strategy", "width": 180,
         "editor": "input"},
        {"field": "growth_opportunities", "title": "Growth Opps", "width": 180,
         "editor": "input"},
        {"field": "account_priorities", "title": "Priorities", "width": 160,
         "editor": "input"},
        {"field": "business_description", "title": "Business Desc", "width": 200,
         "editor": "input"},
        {"field": "notes", "title": "Notes", "width": 180,
         "editor": "input"},
        # Read-only aggregates
        {"field": "total_policies", "title": "Policies", "width": 80,
         "hozAlign": "right", "headerHozAlign": "right"},
        {"field": "total_premium", "title": "Total Premium", "width": 120,
         "_format": "currency"},
        {"field": "total_revenue", "title": "Revenue", "width": 110,
         "_format": "currency"},
        {"field": "next_renewal_days", "title": "Next Renewal", "width": 105,
         "hozAlign": "right", "headerHozAlign": "right"},
    ]

    return templates.TemplateResponse("clients/spreadsheet.html", {
        "request": request,
        "active": "client-spreadsheet",
        "rows": rows,
        "columns": columns,
    })


@router.get("/spreadsheet/export")
def client_spreadsheet_export(request: Request, conn=Depends(get_db)):
    """Export client spreadsheet as branded XLSX."""
    from policydb.exporter import _write_sheet, _wb_to_bytes
    from openpyxl import Workbook
    from policydb.queries import get_all_clients_for_grid

    rows = get_all_clients_for_grid(conn)

    for key, val in request.query_params.items():
        if key.startswith("filter_") and val:
            field = key[7:]
            val_lower = val.lower()
            rows = [r for r in rows if val_lower in str(r.get(field, "") or "").lower()]

    sort_field = request.query_params.get("sort_field")
    sort_dir = request.query_params.get("sort_dir", "asc")
    if sort_field and rows:
        reverse = sort_dir.lower() == "desc"
        rows.sort(key=lambda r: (r.get(sort_field) is None, r.get(sort_field, "")), reverse=reverse)

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Client Spreadsheet", rows, wrap_text=False)
    content = _wb_to_bytes(wb)

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="client_spreadsheet.xlsx"'},
    )


@router.post("/quick-add", response_class=JSONResponse)
async def client_quick_add(request: Request, conn=Depends(get_db)):
    """Create a minimal client record for spreadsheet add-row."""
    body = await request.json()
    name = (body.get("name") or "").strip()
    if not name:
        return JSONResponse({"ok": False, "error": "Name required"}, status_code=400)

    existing = conn.execute("SELECT id FROM clients WHERE LOWER(name) = LOWER(?)", (name,)).fetchone()
    if existing:
        return JSONResponse({"ok": False, "error": "Client already exists"}, status_code=400)

    account_exec = cfg.get("default_account_exec", "")
    conn.execute(
        "INSERT INTO clients (name, industry_segment, account_exec) VALUES (?, '', ?)",
        (name, account_exec),
    )
    conn.commit()
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    row = conn.execute(
        """SELECT c.id, c.name, c.cn_number, c.industry_segment, c.account_exec,
                  c.date_onboarded, c.website, c.fein, c.broker_fee, c.hourly_rate,
                  c.follow_up_date, c.relationship_risk, c.service_model,
                  c.business_description, c.notes, c.stewardship_date,
                  c.renewal_strategy, c.growth_opportunities, c.account_priorities,
                  0 AS total_policies, 0 AS total_premium, 0 AS total_revenue,
                  NULL AS next_renewal_days
           FROM clients c WHERE c.id = ?""",
        (new_id,),
    ).fetchone()

    logger.info("Quick-add client %d: %s", new_id, name)
    return JSONResponse({"ok": True, "row": dict(row)})


@router.get("/{client_id}/tab/activity", response_class=HTMLResponse)
def client_tab_activity(request: Request, client_id: int, conn=Depends(get_db)):
    """Client Activity tab — activity log, quick-log, escalate."""
    client = get_client_by_id(conn, client_id, include_archived=True)
    if not client:
        return HTMLResponse("Not found", status_code=404)
    activities = [dict(a) for a in get_activities(conn, client_id=client_id, days=90)]
    from policydb.web.routes.activities import _attach_pc_emails
    _attach_pc_emails(conn, activities)
    _today_iso = datetime.now().strftime("%Y-%m-%d")
    overdue_followups = sorted(
        [a for a in activities if a.get("follow_up_date") and not a.get("follow_up_done") and a["follow_up_date"] < _today_iso],
        key=lambda a: a["follow_up_date"],
    )
    upcoming_followups = sorted(
        [a for a in activities if a.get("follow_up_date") and not a.get("follow_up_done") and a["follow_up_date"] >= _today_iso],
        key=lambda a: a["follow_up_date"],
    )
    _week_cutoff = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")
    due_soon = [a for a in upcoming_followups if a["follow_up_date"] <= _week_cutoff]
    later_followups = [a for a in upcoming_followups if a["follow_up_date"] > _week_cutoff]
    history = [a for a in activities if not (a.get("follow_up_date") and not a.get("follow_up_done"))]

    quick_log_templates = cfg.get("quick_log_templates", [])
    try:
        primary_contact_name = client["primary_contact"] or ""
    except (KeyError, IndexError):
        primary_contact_name = ""

    # Next renewal label
    _active = [dict(p) for p in get_policies_for_client(conn, client_id) if not p["is_opportunity"]]
    _today = datetime.now().strftime("%Y-%m-%d")
    _sorted = sorted([p for p in _active if p.get("expiration_date")], key=lambda p: p["expiration_date"])
    _future = [p for p in _sorted if p["expiration_date"] >= _today]
    next_renewal_label = (_future[0].get("policy_type") or "renewal") if _future else ((_sorted[-1].get("policy_type") or "renewal") if _sorted else "renewal")

    return templates.TemplateResponse("clients/_tab_activity.html", {
        "request": request,
        "client": client,
        "overdue_followups": overdue_followups,
        "due_soon": due_soon,
        "later_followups": later_followups,
        "history": history,
        "activity_types": cfg.get("activity_types", []),
        "quick_log_templates": quick_log_templates,
        "primary_contact_name": primary_contact_name,
        "next_renewal_label": next_renewal_label,
        "issue_severities": cfg.get("issue_severities", []),
    })


@router.get("/{client_id}/tab/overview", response_class=HTMLResponse)
def client_tab_overview(request: Request, client_id: int, conn=Depends(get_db)):
    client = get_client_by_id(conn, client_id, include_archived=True)
    if not client:
        return HTMLResponse("Not found", status_code=404)
    activities = [dict(a) for a in get_activities(conn, client_id=client_id, days=90)]
    from policydb.web.routes.activities import _attach_pc_emails
    _attach_pc_emails(conn, activities)
    # Split into 3 groups: overdue follow-ups, upcoming follow-ups, history
    _today_iso = datetime.now().strftime("%Y-%m-%d")
    overdue_followups = sorted(
        [a for a in activities if a.get("follow_up_date") and not a.get("follow_up_done") and a["follow_up_date"] < _today_iso],
        key=lambda a: a["follow_up_date"],
    )
    upcoming_followups = sorted(
        [a for a in activities if a.get("follow_up_date") and not a.get("follow_up_done") and a["follow_up_date"] >= _today_iso],
        key=lambda a: a["follow_up_date"],
    )
    _week_cutoff = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")
    due_soon = [a for a in upcoming_followups if a["follow_up_date"] <= _week_cutoff]
    later_followups = [a for a in upcoming_followups if a["follow_up_date"] > _week_cutoff]
    history = [a for a in activities if not (a.get("follow_up_date") and not a.get("follow_up_done"))]

    # Linked accounts
    linked_group = get_linked_group_for_client(conn, client_id)

    # Scratchpad
    _scratch = conn.execute("SELECT content, updated_at FROM client_scratchpad WHERE client_id=?", (client_id,)).fetchone()

    # Pulse data
    from policydb.web.routes.policies import _attach_milestone_progress
    _active_policies = [dict(p) for p in get_policies_for_client(conn, client_id) if not p["is_opportunity"]]
    _attach_milestone_progress(conn, _active_policies)
    pulse_milestone_done = sum(p.get("milestone_done", 0) for p in _active_policies)
    pulse_milestone_total = sum(p.get("milestone_total", 0) for p in _active_policies)
    pulse_overdue = conn.execute(
        """SELECT COUNT(*) FROM (
             SELECT 1 FROM activity_log WHERE client_id=? AND follow_up_done=0 AND follow_up_date < date('now')
             UNION ALL
             SELECT 1 FROM policies WHERE client_id=? AND archived=0 AND (is_opportunity=0 OR is_opportunity IS NULL)
               AND follow_up_date IS NOT NULL AND follow_up_date < date('now')
           )""", (client_id, client_id)
    ).fetchone()[0]
    # Next upcoming follow-up date (earliest pending across activities + policies)
    pulse_next_followup = conn.execute(
        """SELECT MIN(fu) FROM (
             SELECT MIN(follow_up_date) AS fu FROM activity_log
             WHERE client_id=? AND follow_up_done=0 AND follow_up_date >= date('now')
             UNION ALL
             SELECT MIN(follow_up_date) AS fu FROM policies
             WHERE client_id=? AND archived=0 AND (is_opportunity=0 OR is_opportunity IS NULL)
               AND follow_up_date IS NOT NULL AND follow_up_date >= date('now')
           )""", (client_id, client_id)
    ).fetchone()[0]

    _risks_for_pulse = conn.execute("SELECT severity FROM client_risks WHERE client_id=?", (client_id,)).fetchall()
    pulse_high_risks = sum(1 for r in _risks_for_pulse if r["severity"] in ("High", "Critical"))

    # Recent activity for pulse
    _today = datetime.now().strftime("%Y-%m-%d")
    pulse_recent = [dict(a) | {"_type": "activity"} for a in activities[:5]]
    _saved_notes = get_saved_notes_for_client_timeline(conn, client_id)
    for sn in _saved_notes[:5]:
        pulse_recent.append(dict(sn) | {"_type": "note"})
    pulse_recent.sort(key=lambda x: x.get("activity_date") or x.get("created_at") or "", reverse=True)
    pulse_recent = pulse_recent[:5]

    summary = get_client_summary(conn, client_id)

    # Quick-log template data
    quick_log_templates = cfg.get("quick_log_templates", [])
    try:
        primary_contact_name = client["primary_contact"] or ""
    except (KeyError, IndexError):
        primary_contact_name = ""
    # Determine next renewal label (policy type of next-expiring active policy)
    _next_renewal_label = ""
    _sorted_active = sorted(
        [p for p in _active_policies if p.get("expiration_date")],
        key=lambda p: p["expiration_date"],
    )
    _future = [p for p in _sorted_active if p["expiration_date"] >= _today]
    if _future:
        _next_renewal_label = _future[0].get("policy_type") or "renewal"
    elif _sorted_active:
        _next_renewal_label = _sorted_active[-1].get("policy_type") or "renewal"
    else:
        _next_renewal_label = "renewal"
    issue_severities = cfg.get("issue_severities", [])

    # ── What's Next card: single most urgent action item ─────────────
    from datetime import date as _wn_date
    _wn_today = _wn_date.today()
    _wn_today_str = _wn_today.isoformat()
    _wn_tomorrow_str = (_wn_today + timedelta(days=1)).isoformat()
    whats_next = None

    # Priority 1: Issues with breached SLA
    _sla_breached = conn.execute("""
        SELECT id, subject, issue_uid, issue_severity, issue_status, issue_sla_days,
               activity_date, client_id,
               CAST(julianday('now') - julianday(activity_date) AS INTEGER) AS age_days
        FROM activity_log
        WHERE client_id = ? AND item_kind = 'issue' AND issue_id IS NULL
          AND merged_into_id IS NULL
          AND (issue_status IS NULL OR issue_status NOT IN ('Resolved', 'Closed'))
          AND issue_sla_days > 0
          AND CAST(julianday('now') - julianday(activity_date) AS INTEGER) > issue_sla_days
        ORDER BY
          CASE issue_severity WHEN 'Critical' THEN 0 WHEN 'High' THEN 1 WHEN 'Normal' THEN 2 ELSE 3 END,
          (CAST(julianday('now') - julianday(activity_date) AS INTEGER) - issue_sla_days) DESC
        LIMIT 1
    """, [client_id]).fetchone()
    if _sla_breached:
        whats_next = {
            "kind": "issue_sla_breached",
            "id": _sla_breached["id"],
            "uid": _sla_breached["issue_uid"],
            "subject": _sla_breached["subject"],
            "severity": _sla_breached["issue_severity"],
            "age_days": _sla_breached["age_days"],
            "sla_days": _sla_breached["issue_sla_days"],
            "overdue_days": _sla_breached["age_days"] - _sla_breached["issue_sla_days"],
        }

    # Priority 2: Overdue follow-ups
    if not whats_next:
        _overdue = conn.execute("""
            SELECT id, subject, activity_type, follow_up_date, contact_person, policy_id,
                   CAST(julianday('now') - julianday(follow_up_date) AS INTEGER) AS overdue_days
            FROM activity_log
            WHERE client_id = ? AND follow_up_date < ? AND (follow_up_done = 0 OR follow_up_done IS NULL)
              AND (item_kind IS NULL OR item_kind = 'followup')
            ORDER BY follow_up_date ASC
            LIMIT 1
        """, [client_id, _wn_today_str]).fetchone()
        if _overdue:
            whats_next = {
                "kind": "overdue_followup",
                "id": _overdue["id"],
                "subject": _overdue["subject"],
                "type": _overdue["activity_type"],
                "follow_up_date": _overdue["follow_up_date"],
                "overdue_days": _overdue["overdue_days"],
                "contact": _overdue["contact_person"],
            }

    # Priority 3: Issues approaching SLA (within 1 day)
    if not whats_next:
        _approaching = conn.execute("""
            SELECT id, subject, issue_uid, issue_severity, issue_sla_days,
                   activity_date,
                   CAST(julianday('now') - julianday(activity_date) AS INTEGER) AS age_days
            FROM activity_log
            WHERE client_id = ? AND item_kind = 'issue' AND issue_id IS NULL
              AND merged_into_id IS NULL
              AND (issue_status IS NULL OR issue_status NOT IN ('Resolved', 'Closed'))
              AND issue_sla_days > 0
              AND CAST(julianday('now') - julianday(activity_date) AS INTEGER) >= (issue_sla_days - 1)
              AND CAST(julianday('now') - julianday(activity_date) AS INTEGER) <= issue_sla_days
            ORDER BY issue_sla_days - CAST(julianday('now') - julianday(activity_date) AS INTEGER)
            LIMIT 1
        """, [client_id]).fetchone()
        if _approaching:
            whats_next = {
                "kind": "issue_approaching_sla",
                "id": _approaching["id"],
                "uid": _approaching["issue_uid"],
                "subject": _approaching["subject"],
                "severity": _approaching["issue_severity"],
                "remaining_days": _approaching["issue_sla_days"] - _approaching["age_days"],
            }

    # Priority 4: Follow-ups due today or tomorrow
    if not whats_next:
        _due_soon_wn = conn.execute("""
            SELECT id, subject, activity_type, follow_up_date, contact_person
            FROM activity_log
            WHERE client_id = ? AND follow_up_date BETWEEN ? AND ?
              AND (follow_up_done = 0 OR follow_up_done IS NULL)
              AND (item_kind IS NULL OR item_kind = 'followup')
            ORDER BY follow_up_date ASC
            LIMIT 1
        """, [client_id, _wn_today_str, _wn_tomorrow_str]).fetchone()
        if _due_soon_wn:
            _is_today = _due_soon_wn["follow_up_date"] == _wn_today_str
            whats_next = {
                "kind": "due_followup",
                "id": _due_soon_wn["id"],
                "subject": _due_soon_wn["subject"],
                "type": _due_soon_wn["activity_type"],
                "due_label": "Due today" if _is_today else "Due tomorrow",
                "contact": _due_soon_wn["contact_person"],
            }

    # Priority 5: Next expiring policy (within 120 days)
    if not whats_next:
        _next_exp = conn.execute("""
            SELECT policy_uid, policy_type, carrier, expiration_date,
                   CAST(julianday(expiration_date) - julianday('now') AS INTEGER) AS days_to_exp
            FROM policies
            WHERE client_id = ? AND archived = 0
              AND (is_opportunity = 0 OR is_opportunity IS NULL)
              AND expiration_date >= ?
            ORDER BY expiration_date ASC
            LIMIT 1
        """, [client_id, _wn_today_str]).fetchone()
        if _next_exp and _next_exp["days_to_exp"] <= 120:
            whats_next = {
                "kind": "policy_expiry",
                "uid": _next_exp["policy_uid"],
                "policy_type": _next_exp["policy_type"],
                "carrier": _next_exp["carrier"],
                "days_to_exp": _next_exp["days_to_exp"],
                "expiration_date": _next_exp["expiration_date"],
            }

    # Priority 6: All clear
    if not whats_next:
        whats_next = {"kind": "all_clear"}

    # Anomalies for this client
    try:
        from policydb.anomaly_engine import get_anomalies_for_client
        client_anomalies = get_anomalies_for_client(conn, client_id)
    except Exception:
        client_anomalies = []

    # Last activity days for all-clear state
    _last_act = conn.execute(
        "SELECT MAX(activity_date) AS d FROM activity_log WHERE client_id = ?",
        [client_id],
    ).fetchone()
    _last_activity_days = None
    if _last_act and _last_act["d"]:
        try:
            _last_activity_days = (_wn_today - _wn_date.fromisoformat(_last_act["d"])).days
        except (ValueError, TypeError):
            pass

    return templates.TemplateResponse("clients/_tab_overview.html", {
        "request": request,
        "client": dict(client),
        "summary": dict(summary) if summary else {},
        "activities": activities,
        "overdue_followups": overdue_followups,
        "upcoming_followups": upcoming_followups,
        "due_soon": due_soon,
        "later_followups": later_followups,
        "history": history,
        "activity_types": cfg.get("activity_types"),
        "dispositions": cfg.get("follow_up_dispositions", []),
        "quick_log_templates": quick_log_templates,
        "primary_contact_name": primary_contact_name,
        "next_renewal_label": _next_renewal_label,
        "issue_severities": issue_severities,
        "all_clients": [{"id": client_id, "name": dict(client)["name"]}],
        "linked_group": linked_group,
        "linked_relationships": cfg.get("linked_account_relationships", []),
        "client_scratchpad": _scratch["content"] if _scratch else "",
        "client_scratchpad_updated": _scratch["updated_at"] if _scratch else "",
        "client_saved_notes": get_saved_notes(conn, "client", client_id),
        "pulse_overdue": pulse_overdue,
        "pulse_next_followup": pulse_next_followup,
        "pulse_milestone_done": pulse_milestone_done,
        "pulse_milestone_total": pulse_milestone_total,
        "pulse_high_risks": pulse_high_risks,
        "pulse_recent": pulse_recent,
        "open_issues": [dict(r) for r in conn.execute("""
            SELECT a.id, a.issue_uid, a.subject, a.issue_status, a.issue_severity,
                   a.activity_date, a.issue_sla_days,
                   julianday(date('now')) - julianday(a.activity_date) AS days_open
            FROM activity_log a
            WHERE a.client_id = ? AND a.item_kind = 'issue' AND a.issue_id IS NULL
              AND a.merged_into_id IS NULL
              AND (a.issue_status IS NULL OR a.issue_status NOT IN ('Resolved', 'Closed'))
        """, (client_id,)).fetchall()],
        "today": _today,
        "today_iso": _today,
        "locations": _get_project_locations(conn, client_id),
        "unassigned_count": conn.execute(
            """SELECT COUNT(*) FROM policies
               WHERE client_id=? AND archived=0
               AND (project_id IS NULL OR project_id=0)
               AND (is_opportunity=0 OR is_opportunity IS NULL)""",
            (client_id,),
        ).fetchone()[0],
        "whats_next": whats_next,
        "last_activity_days": _last_activity_days,
        "account_priority_options": cfg.get("account_priority_options", []),
        "relationship_risk_levels": cfg.get("relationship_risk_levels", []),
        "service_model_options": cfg.get("service_model_options", []),
        "client_anomalies": client_anomalies,
    })


@router.get("/{client_id}/tab/policies", response_class=HTMLResponse)
def client_tab_policies(request: Request, client_id: int, conn=Depends(get_db)):
    from collections import defaultdict
    client = get_client_by_id(conn, client_id, include_archived=True)
    if not client:
        return HTMLResponse("Not found", status_code=404)

    all_policies = [dict(p) for p in get_policies_for_client(conn, client_id)]
    opportunities = [p for p in all_policies if p.get("is_opportunity")]
    policies = [p for p in all_policies if not p.get("is_opportunity")]

    from policydb.web.routes.policies import _attach_milestone_progress, _attach_readiness_score
    from policydb.queries import attach_renewal_issues
    policies = _attach_readiness_score(conn, _attach_milestone_progress(conn, policies))
    attach_renewal_issues(conn, policies)

    # Attach sub-coverages for ghost row display on package policies
    from policydb.queries import get_sub_coverages_full_by_policy_id
    _pol_ids = [p["id"] for p in policies if p.get("id")]
    _sub_cov_map = get_sub_coverages_full_by_policy_id(conn, _pol_ids) if _pol_ids else {}
    for p in policies:
        p["sub_coverages"] = _sub_cov_map.get(p["id"], [])

    # Attach team contacts to opportunities
    if opportunities:
        opp_ids = [o["id"] for o in opportunities]
        _pc_ph = ",".join("?" * len(opp_ids))
        _opc = conn.execute(
            f"SELECT cpa.policy_id, co.name, co.email, co.phone, cpa.role, co.organization "  # noqa: S608
            f"FROM contact_policy_assignments cpa JOIN contacts co ON cpa.contact_id = co.id "
            f"WHERE cpa.policy_id IN ({_pc_ph}) ORDER BY cpa.id", opp_ids
        ).fetchall()
        _opc_map: dict = {}
        for _c in _opc:
            _opc_map.setdefault(_c["policy_id"], []).append(dict(_c))
        for o in opportunities:
            o["team"] = _opc_map.get(o["id"], [])

    # Group policies by project_name
    def _proj_key(name):
        if not name:
            return ""
        return " ".join(name.strip().split()).lower()

    groups: dict = defaultdict(list)
    group_display: dict = {}
    for p in policies:
        raw = (p.get("project_name") or "").strip()
        key = _proj_key(raw)
        groups[key].append(p)
        if key and key not in group_display:
            group_display[key] = raw

    policy_groups = sorted(
        [(group_display.get(k, ""), v) for k, v in groups.items()],
        key=lambda x: ("\xff" if not x[0] else x[0].lower()),
    )

    # Tower visuals
    tower_by_project: dict = defaultdict(lambda: defaultdict(list))
    for p in policies:
        tg = p.get("tower_group")
        if tg:
            proj = (p.get("project_name") or "").strip() or "Corporate / Standalone"
            tower_by_project[proj][tg].append(p)

    def _tower_sort_key(lp):
        att = lp.get("attachment_point")
        if att is not None:
            return (float(att), 0)
        pos = lp.get("layer_position") or "Primary"
        try:
            return (-1, int(pos))
        except (ValueError, TypeError):
            return (-1, 0)

    def _attach_ground_up(layers):
        sorted_layers = sorted(layers, key=_tower_sort_key)
        running = 0.0
        for lp in sorted_layers:
            lim = float(lp.get("limit_amount") or 0)
            att = lp.get("attachment_point")
            part = lp.get("participation_of")
            if att is not None and float(att) >= 0:
                layer_size = float(part) if part else lim
                lp["ground_up"] = float(att) + layer_size
            else:
                running += lim
                lp["ground_up"] = running
        return sorted_layers

    tower_groups = {
        proj: {tg: _attach_ground_up(layers) for tg, layers in sorted(tgs.items())}
        for proj, tgs in sorted(tower_by_project.items(),
            key=lambda x: ("\xff" if x[0] == "Corporate / Standalone" else x[0].lower()))
    }

    def _build_tower_visuals_tab(tg_dict):
        from policydb.analysis import layer_notation as _ln
        visuals = {}
        for proj, tgs in tg_dict.items():
            visuals[proj] = {}
            for tg_name, layers in tgs.items():
                if not layers:
                    continue
                total_gu = max(float(l.get("ground_up") or 0) for l in layers)
                if total_gu == 0:
                    continue
                grouped: dict = {}
                for l in layers:
                    att = l.get("attachment_point")
                    key = str(float(att)) if att is not None else f"pos-{l.get('layer_position', 'Primary')}"
                    grouped.setdefault(key, []).append(l)
                visual_layers = []
                for gkey, carriers in grouped.items():
                    parts = [float(c["participation_of"]) for c in carriers if c.get("participation_of")]
                    carrier_limits = [float(c.get("limit_amount") or 0) for c in carriers]
                    full_limit = max(parts) if parts else (sum(carrier_limits) if len(carriers) > 1 else (carrier_limits[0] if carrier_limits else 0))
                    flex = max(full_limit / 1_000_000, 0.5)
                    att_val = carriers[0].get("attachment_point")
                    gu_val = max(float(c.get("ground_up") or 0) for c in carriers)
                    carrier_data = []
                    for c in carriers:
                        climit = float(c.get("limit_amount") or 0)
                        fill_pct = round(climit / full_limit * 100, 1) if full_limit > 0 else 100
                        carrier_data.append({
                            "carrier": c.get("carrier", ""), "limit": climit, "fill_pct": fill_pct,
                            "policy_uid": c.get("policy_uid", ""), "policy_type": c.get("policy_type", ""),
                            "premium": c.get("premium") or 0,
                            "notation": _ln(c.get("limit_amount"), c.get("attachment_point"), c.get("participation_of")) or "",
                        })
                    is_shared = len(carriers) > 1 or any(c.get("participation_of") for c in carriers)
                    total_fill = sum(cd["fill_pct"] for cd in carrier_data)
                    open_pct = round(100 - total_fill, 1) if is_shared and total_fill < 100 else 0
                    open_amount = full_limit - sum(cd["limit"] for cd in carrier_data) if open_pct > 0 else 0
                    visual_layers.append({
                        "attachment": float(att_val) if att_val is not None else None,
                        "full_limit": full_limit, "flex": flex, "ground_up": gu_val,
                        "carriers": carrier_data, "total_premium": sum(cd["premium"] for cd in carrier_data),
                        "is_shared": is_shared, "open_pct": open_pct, "open_amount": open_amount,
                    })
                visual_layers.sort(key=lambda vl: vl["attachment"] if vl["attachment"] is not None else -1)
                visuals[proj][tg_name] = {
                    "layers": visual_layers, "total_ground_up": total_gu,
                    "total_premium": sum(vl["total_premium"] for vl in visual_layers),
                    "carrier_count": sum(len(vl["carriers"]) for vl in visual_layers),
                }
        return visuals

    tower_visuals = _build_tower_visuals_tab(tower_groups) if tower_groups else {}

    # Archived policies
    archived_policies = [dict(r) for r in conn.execute(
        """SELECT policy_uid, policy_type, carrier, effective_date, expiration_date,
                  premium, policy_number, project_name
           FROM policies WHERE client_id = ? AND archived = 1 ORDER BY expiration_date DESC""",
        (client_id,),
    ).fetchall()]

    # Programs v2 (from standalone programs table)
    from policydb.queries import get_programs_for_client
    programs_v2 = get_programs_for_client(conn, client_id)

    # Legacy programs removed — programs now come from standalone programs table (programs_v2)
    programs = []
    _program_linked_ids = set()

    # Project notes & addresses
    notes_rows = conn.execute("SELECT id, LOWER(TRIM(name)) AS key, name, notes FROM projects WHERE client_id = ?", (client_id,)).fetchall()
    project_notes = {r["key"]: r["notes"] for r in notes_rows}
    project_ids = {r["key"]: r["id"] for r in notes_rows}
    project_addresses: dict = {}
    for p in sorted(policies, key=lambda x: x.get("id", 0), reverse=True):
        key = _proj_key(p.get("project_name"))
        if key and key not in project_addresses:
            project_addresses[key] = {
                "exposure_address": p.get("exposure_address") or "",
                "exposure_city": p.get("exposure_city") or "",
                "exposure_state": p.get("exposure_state") or "",
                "exposure_zip": p.get("exposure_zip") or "",
            }

    from policydb.queries import get_schematic_completeness
    schematic_completeness = get_schematic_completeness(conn, client_id)
    # Build lookup by tower_group for template badge rendering
    completeness_by_tg = {c["tower_group"]: c for c in schematic_completeness}

    # Renewal pipeline mini-view — policies expiring in next 120 days
    from collections import OrderedDict
    _renewal_statuses = cfg.get("renewal_statuses", [])
    from policydb.queries import get_program_pipeline
    renewal_pipeline_policies = conn.execute(
        """SELECT policy_uid, policy_type, carrier, expiration_date, renewal_status,
                  project_name,
                  CAST(julianday(expiration_date) - julianday('now') AS INTEGER) AS days_to_exp
           FROM policies
           WHERE client_id = ? AND archived = 0
             AND (is_opportunity = 0 OR is_opportunity IS NULL)
             AND expiration_date >= date('now')
             AND CAST(julianday(expiration_date) - julianday('now') AS INTEGER) <= 120
             AND (program_id IS NULL OR NOT EXISTS (
                   SELECT 1 FROM programs pg WHERE pg.id = policies.program_id AND pg.archived = 0
                 ))
           ORDER BY expiration_date ASC""",
        [client_id],
    ).fetchall()
    program_pipeline = get_program_pipeline(conn, client_id=client_id, window_days=120)
    pipeline_by_status: OrderedDict = OrderedDict()
    for _rs in _renewal_statuses:
        pipeline_by_status[_rs] = []
    for _rp in renewal_pipeline_policies:
        _s = _rp["renewal_status"] or (_renewal_statuses[0] if _renewal_statuses else "Unknown")
        if _s not in pipeline_by_status:
            pipeline_by_status[_s] = []
        pipeline_by_status[_s].append(dict(_rp))
    # Merge programs into status columns alongside regular policies
    for pgm in program_pipeline:
        _s = pgm["renewal_status"] or (_renewal_statuses[0] if _renewal_statuses else "Unknown")
        if _s not in pipeline_by_status:
            pipeline_by_status[_s] = []
        pipeline_by_status[_s].append({**dict(pgm), "is_program": True})

    # ── Summary stats for policies header bar ──
    _total_premium = sum(float(p.get("premium") or 0) for p in policies)
    _expiring_30 = sum(1 for p in policies if p.get("days_to_renewal") is not None and 0 <= p["days_to_renewal"] <= 30)
    _expiring_60 = sum(1 for p in policies if p.get("days_to_renewal") is not None and 31 <= p["days_to_renewal"] <= 60)
    _not_started_90 = sum(
        1 for p in policies
        if p.get("days_to_renewal") is not None and 0 <= p["days_to_renewal"] <= 90
        and (p.get("renewal_status") or "").lower() in ("not started", "")
    )
    policy_summary = {
        "total": len(policies),
        "premium": _total_premium,
        "expiring_30": _expiring_30,
        "expiring_60": _expiring_60,
        "not_started_90": _not_started_90,
    }

    return templates.TemplateResponse("clients/_tab_policies.html", {
        "request": request,
        "client": dict(client),
        "policy_groups": policy_groups,
        "policy_summary": policy_summary,
        "pipeline_by_status": pipeline_by_status,
        "program_pipeline": program_pipeline,
        "has_renewal_pipeline": bool(renewal_pipeline_policies) or bool(program_pipeline),
        "tower_visuals": tower_visuals,
        "completeness_by_tg": completeness_by_tg,
        "renewal_statuses": cfg.get("renewal_statuses"),
        "programs": programs,
        "programs_v2": programs_v2,
        "program_linked_uids": _program_linked_ids,
        "archived_policies": archived_policies,
        "project_notes": project_notes,
        "project_ids": project_ids,
        "project_addresses": project_addresses,
        "opportunities": opportunities,
        "bundles": _get_request_bundles(conn, client_id),
        "today_iso": datetime.now().strftime("%Y-%m-%d"),
        "pipeline_projects": _get_project_pipeline(conn, client_id),
        "location_projects": _get_project_locations(conn, client_id),
        "project_stages": cfg.get("project_stages", []),
        "project_types": cfg.get("project_types", []),
        "policy_types": cfg.get("policy_types", []),
        "timeline_data": _build_timeline_data(_get_project_pipeline(conn, client_id)),
        "us_states": _get_us_states(),
    })


@router.get("/{client_id}/tab/contacts", response_class=HTMLResponse)
def client_tab_contacts(request: Request, client_id: int, add_contact: str = "", conn=Depends(get_db)):
    client = get_client_by_id(conn, client_id, include_archived=True)
    if not client:
        return HTMLResponse("Not found", status_code=404)
    import json as _json

    contacts = get_client_contacts(conn, client_id, contact_type='client')
    team_contacts = get_client_contacts(conn, client_id, contact_type='internal')
    external_contacts = get_client_contacts(conn, client_id, contact_type='external')

    # Placement colleagues — include archived/lost policies, tagged accordingly
    _pc_rows = conn.execute(
        """SELECT co.id, co.name, co.email, co.phone, co.mobile,
                  cpa.role, cpa.title, co.organization,
                  GROUP_CONCAT(p.policy_type, ', ') AS policy_types,
                  MAX(p.archived) AS has_archived,
                  MIN(p.archived) AS all_archived
           FROM contact_policy_assignments cpa
           JOIN contacts co ON cpa.contact_id = co.id
           JOIN policies p ON cpa.policy_id = p.id
           WHERE p.client_id = ? AND cpa.is_placement_colleague = 1
           GROUP BY co.id ORDER BY LOWER(co.name)""",
        (client_id,),
    ).fetchall()
    placement_colleagues = [
        dict(r) | {
            "organization": r["organization"] or "",
            "is_lost_only": bool(r["all_archived"]),
        }
        for r in _pc_rows
    ]

    # --- Key Contacts summary strip ---
    key_contacts = []

    # Primary client contact
    _primary = conn.execute(
        """SELECT c.name, c.email, c.phone, c.mobile, cca.role, cca.title
           FROM contacts c
           JOIN contact_client_assignments cca ON c.id = cca.contact_id
           WHERE cca.client_id = ? AND cca.contact_type = 'client' AND cca.is_primary = 1
           LIMIT 1""",
        (client_id,),
    ).fetchone()
    key_contacts.append({"label": "Primary Contact", "contact": dict(_primary) if _primary else None})

    # Account Manager (team / internal lead)
    _team_lead = conn.execute(
        """SELECT c.name, c.email, c.phone, c.mobile, cca.role, cca.title
           FROM contacts c
           JOIN contact_client_assignments cca ON c.id = cca.contact_id
           WHERE cca.client_id = ? AND cca.contact_type = 'internal' AND cca.is_primary = 1
           LIMIT 1""",
        (client_id,),
    ).fetchone()
    if not _team_lead:
        _team_lead = conn.execute(
            """SELECT c.name, c.email, c.phone, c.mobile, cca.role, cca.title
               FROM contacts c
               JOIN contact_client_assignments cca ON c.id = cca.contact_id
               WHERE cca.client_id = ? AND cca.contact_type = 'internal'
               ORDER BY cca.id LIMIT 1""",
            (client_id,),
        ).fetchone()
    key_contacts.append({"label": "Account Manager", "contact": dict(_team_lead) if _team_lead else None})

    # Key Underwriter (from policy assignments, most recent active policy)
    _underwriter = conn.execute(
        """SELECT c.name, c.email, c.phone, c.mobile, cpa.role, cpa.title
           FROM contacts c
           JOIN contact_policy_assignments cpa ON c.id = cpa.contact_id
           JOIN policies p ON p.id = cpa.policy_id
           WHERE p.client_id = ? AND p.archived = 0
             AND LOWER(cpa.role) LIKE '%underwriter%'
           ORDER BY p.expiration_date DESC
           LIMIT 1""",
        (client_id,),
    ).fetchone()
    key_contacts.append({"label": "Key Underwriter", "contact": dict(_underwriter) if _underwriter else None})

    # Billing contact (client assignment with billing role)
    _billing = conn.execute(
        """SELECT c.name, c.email, c.phone, c.mobile, cca.role, cca.title
           FROM contacts c
           JOIN contact_client_assignments cca ON c.id = cca.contact_id
           WHERE cca.client_id = ? AND LOWER(cca.role) LIKE '%billing%'
           LIMIT 1""",
        (client_id,),
    ).fetchone()
    key_contacts.append({"label": "Billing Contact", "contact": dict(_billing) if _billing else None})

    from policydb.email_templates import client_context as _client_ctx, render_tokens as _render_tokens
    _mail_ctx = _client_ctx(conn, client_id)
    mailto_subject = _render_tokens(cfg.get("email_subject_client", "Re: {{client_name}}"), _mail_ctx)

    _ac_rows = conn.execute(
        """SELECT co.name, MAX(co.email) AS email, MAX(co.phone) AS phone, MAX(co.mobile) AS mobile,
                  MAX(COALESCE(cca.title, cpa.title)) AS title, MAX(COALESCE(cca.role, cpa.role)) AS role
           FROM contacts co
           LEFT JOIN contact_client_assignments cca ON co.id = cca.contact_id AND cca.client_id = ?
           LEFT JOIN contact_policy_assignments cpa ON co.id = cpa.contact_id
           WHERE co.name IS NOT NULL AND co.name != '' GROUP BY co.id ORDER BY co.name""",
        (client_id,),
    ).fetchall()
    all_contacts_json = _json.dumps({r["name"]: {"email": r["email"] or "", "phone": r["phone"] or "", "mobile": r["mobile"] or "", "title": r["title"] or "", "role": r["role"] or ""} for r in _ac_rows})

    all_internal_contacts_json = _json.dumps({
        r["name"]: {"title": r["title"] or "", "email": r["email"] or "", "phone": r["phone"] or "", "mobile": r["mobile"] or "", "role": r["role"] or ""}
        for r in conn.execute(
            """SELECT co.name, MAX(cca.title) AS title, MAX(co.email) AS email,
                      MAX(co.phone) AS phone, MAX(co.mobile) AS mobile, MAX(cca.role) AS role
               FROM contacts co JOIN contact_client_assignments cca ON co.id = cca.contact_id
               WHERE cca.contact_type='internal' AND co.name IS NOT NULL AND co.name != ''
               GROUP BY LOWER(TRIM(co.name)) ORDER BY co.name"""
        ).fetchall()
    })

    return templates.TemplateResponse("clients/_tab_contacts.html", {
        "request": request,
        "client": dict(client),
        "contacts": contacts,
        "team_contacts": team_contacts,
        "external_contacts": external_contacts,
        "key_contacts": key_contacts,
        "billing_accounts": [dict(r) for r in conn.execute(
            "SELECT * FROM billing_accounts WHERE client_id=? ORDER BY is_master DESC, billing_id", (client_id,)
        ).fetchall()],
        "placement_colleagues": placement_colleagues,
        "mailto_subject": mailto_subject,
        "all_contacts_json": all_contacts_json,
        "all_internal_contacts_json": all_internal_contacts_json,
        "add_contact": add_contact,
        "contact_roles": cfg.get("contact_roles", []),
        "team_assignments": cfg.get("team_assignments", []),
        "all_orgs": _get_all_client_contact_orgs(conn),
    })


@router.get("/{client_id}/tab/risk", response_class=HTMLResponse)
def client_tab_risk(request: Request, client_id: int, conn=Depends(get_db)):
    client = get_client_by_id(conn, client_id, include_archived=True)
    if not client:
        return HTMLResponse("Not found", status_code=404)

    # Risks
    _risk_rows = conn.execute(
        """SELECT cr.*, p.policy_type AS linked_policy_type, p.carrier AS linked_carrier
           FROM client_risks cr LEFT JOIN policies p ON cr.policy_uid = p.policy_uid
           WHERE cr.client_id = ? ORDER BY
             CASE cr.severity WHEN 'Critical' THEN 0 WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 ELSE 3 END,
             cr.category""",
        (client_id,),
    ).fetchall()
    risks = [dict(r) for r in _risk_rows]
    for risk in risks:
        risk["coverage_lines"] = [dict(r) for r in conn.execute(
            "SELECT * FROM risk_coverage_lines WHERE risk_id=?", (risk["id"],)
        ).fetchall()]
        risk["controls"] = [dict(r) for r in conn.execute(
            "SELECT * FROM risk_controls WHERE risk_id=? ORDER BY id", (risk["id"],)
        ).fetchall()]

    # Policy UID options for linking
    policy_rows = conn.execute(
        "SELECT policy_uid, policy_type FROM policies WHERE client_id=? AND archived=0 ORDER BY policy_type",
        (client_id,),
    ).fetchall()
    policy_uid_options = [{"uid": r["policy_uid"], "label": f"{r['policy_uid']} — {r['policy_type']}"} for r in policy_rows]

    # Risk review prompts — industry-aware, coverage-gap-aware guided questions
    risk_prompts = []
    try:
        from policydb.compliance import get_risk_review_prompts
        cfg_prompts = cfg.get("risk_review_prompts", [])
        if cfg_prompts:
            risk_prompts = get_risk_review_prompts(
                client=dict(client),
                locations=[],
                policies=[dict(r) for r in policy_rows],
                cfg_prompts=cfg_prompts,
            )
    except Exception:
        pass

    return templates.TemplateResponse("clients/_tab_risk.html", {
        "request": request,
        "client": dict(client),
        "risks": risks,
        "risk_summary": _compute_risk_summary(risks),
        "risk_categories": cfg.get("risk_categories", []),
        "risk_severities": cfg.get("risk_severities", []),
        "risk_sources": cfg.get("risk_sources", []),
        "risk_control_types": cfg.get("risk_control_types", []),
        "risk_control_statuses": cfg.get("risk_control_statuses", []),
        "risk_adequacy_levels": cfg.get("risk_adequacy_levels", []),
        "policy_types": cfg.get("policy_types", []),
        "policy_uid_options": policy_uid_options,
        "bundles": _get_request_bundles(conn, client_id),
        "today_iso": datetime.now().strftime("%Y-%m-%d"),
        "risk_prompts": risk_prompts,
    })


@router.get("/{client_id}/tab/issues", response_class=HTMLResponse)
def client_tab_issues(request: Request, client_id: int, conn=Depends(get_db)):
    """Issues tab — open issues + resolved history with activity count and hours."""
    client = get_client_by_id(conn, client_id, include_archived=True)
    if not client:
        return HTMLResponse("Not found", status_code=404)
    client_dict = dict(client)

    all_issues = [dict(r) for r in conn.execute("""
        SELECT a.id, a.issue_uid, a.subject, a.issue_status, a.issue_severity,
               a.issue_sla_days, a.resolution_type, a.resolved_date, a.activity_date,
               a.is_renewal_issue,
               CAST(julianday('now') - julianday(a.activity_date) AS INTEGER) AS days_open,
               p.policy_uid, p.policy_type,
               (SELECT COUNT(*) FROM activity_log sub WHERE sub.issue_id = a.id) AS activity_count,
               (SELECT COALESCE(SUM(sub.duration_hours), 0) FROM activity_log sub
                WHERE sub.issue_id = a.id AND sub.duration_hours IS NOT NULL) AS total_hours
        FROM activity_log a
        LEFT JOIN policies p ON a.policy_id = p.id
        WHERE a.client_id = ? AND a.item_kind = 'issue' AND a.issue_id IS NULL
          AND a.merged_into_id IS NULL
        ORDER BY CASE WHEN a.issue_status IN ('Resolved','Closed') THEN 1 ELSE 0 END,
                 CASE a.issue_severity WHEN 'Critical' THEN 0 WHEN 'High' THEN 1 WHEN 'Normal' THEN 2 ELSE 3 END,
                 a.activity_date DESC
    """, (client_id,)).fetchall()]

    open_issues = [i for i in all_issues if i.get("issue_status") not in ("Resolved", "Closed")]
    resolved_issues = [i for i in all_issues if i.get("issue_status") in ("Resolved", "Closed")]

    return templates.TemplateResponse("clients/_tab_issues.html", {
        "request": request,
        "client": client_dict,
        "open_issues": open_issues,
        "resolved_issues": resolved_issues,
        "issue_severities": cfg.get("issue_severities", []),
        "all_clients": [dict(r) for r in conn.execute("SELECT id, name FROM clients WHERE archived = 0 ORDER BY name").fetchall()],
    })


@router.get("/{client_id}/tab/files", response_class=HTMLResponse)
def client_tab_files(
    request: Request,
    client_id: int,
    view: str = "grouped",
    q: str = "",
    rt: str = "",
    conn=Depends(get_db),
):
    """Files tab — rollup of all attachments across client + child records."""
    client = get_client_by_id(conn, client_id, include_archived=True)
    if not client:
        return HTMLResponse("Not found", status_code=404)
    client_dict = dict(client)

    from policydb.devonthink import is_devonthink_available

    # Rollup query: all attachments linked to this client or its child records
    rows = conn.execute("""
        SELECT a.*, ra.id AS link_id, ra.record_type, ra.record_id, ra.sort_order,
          CASE ra.record_type
            WHEN 'client' THEN 'Client-level'
            WHEN 'policy' THEN (SELECT policy_uid || ' — ' || COALESCE(policy_type, '') FROM policies WHERE id = ra.record_id)
            WHEN 'activity' THEN (SELECT COALESCE(activity_type, '') || ': ' || COALESCE(subject, '') FROM activity_log WHERE id = ra.record_id)
            WHEN 'project' THEN (SELECT name FROM projects WHERE id = ra.record_id)
            WHEN 'rfi_bundle' THEN (SELECT COALESCE(rfi_uid, '') || CASE WHEN rfi_uid IS NOT NULL AND rfi_uid <> '' THEN ' — ' ELSE '' END || COALESCE(title, 'RFI Bundle') FROM client_request_bundles WHERE id = ra.record_id)
            WHEN 'rfi_item' THEN (SELECT COALESCE(cri.description, 'RFI Item') FROM client_request_items cri WHERE cri.id = ra.record_id)
          END AS source_label,
          CASE ra.record_type
            WHEN 'rfi_bundle' THEN ra.record_id
            WHEN 'rfi_item' THEN (SELECT cri.bundle_id FROM client_request_items cri WHERE cri.id = ra.record_id)
            ELSE NULL
          END AS rfi_bundle_id
        FROM attachments a
        JOIN record_attachments ra ON ra.attachment_id = a.id
        WHERE (ra.record_type = 'client' AND ra.record_id = :cid)
           OR (ra.record_type = 'policy' AND ra.record_id IN (SELECT id FROM policies WHERE client_id = :cid AND archived = 0))
           OR (ra.record_type = 'activity' AND ra.record_id IN (SELECT id FROM activity_log WHERE client_id = :cid))
           OR (ra.record_type = 'project' AND ra.record_id IN (SELECT id FROM projects WHERE client_id = :cid))
           OR (ra.record_type = 'rfi_bundle' AND ra.record_id IN (SELECT id FROM client_request_bundles WHERE client_id = :cid))
           OR (ra.record_type = 'rfi_item' AND ra.record_id IN (
               SELECT cri.id FROM client_request_items cri
               JOIN client_request_bundles crb ON crb.id = cri.bundle_id
               WHERE crb.client_id = :cid
           ))
        ORDER BY ra.record_type, source_label, ra.sort_order
        LIMIT 500
    """, {"cid": client_id}).fetchall()

    all_atts = [dict(r) for r in rows]

    # Search filter
    if q:
        q_lower = q.lower()
        all_atts = [a for a in all_atts if q_lower in (a.get("title") or "").lower()
                     or q_lower in (a.get("filename") or "").lower()
                     or q_lower in (a.get("description") or "").lower()]

    # Human-readable size
    def _size_display(size):
        if not size:
            return ""
        if size < 1024:
            return f"{size} B"
        if size < 1024 * 1024:
            return f"{size / 1024:.0f} KB"
        return f"{size / (1024 * 1024):.1f} MB"

    for a in all_atts:
        a["size_display"] = _size_display(a.get("file_size"))

    # Build link URLs for source records
    def _link_url(record_type, record_id, source_label, rfi_bundle_id):
        if record_type == "policy" and source_label:
            uid = source_label.split(" — ")[0].strip()
            return f"/policies/{uid}/edit"
        if record_type == "project":
            return f"/clients/{client_id}/projects/{record_id}"
        if record_type == "client":
            return f"/clients/{client_id}"
        if record_type == "rfi_bundle":
            return f"/clients/{client_id}/requests/{record_id}"
        if record_type == "rfi_item" and rfi_bundle_id:
            return f"/clients/{client_id}/requests/{rfi_bundle_id}#req-item-{record_id}"
        return ""

    for a in all_atts:
        a["link_url"] = _link_url(
            a["record_type"], a["record_id"], a.get("source_label", ""), a.get("rfi_bundle_id")
        )

    total_count = len(all_atts)
    record_types = sorted({a["record_type"] for a in all_atts})

    # Build grouped structure
    groups = []
    if view == "grouped":
        from collections import OrderedDict
        seen = OrderedDict()
        for a in all_atts:
            key = (a["record_type"], a["record_id"])
            if key not in seen:
                seen[key] = {
                    "record_type": a["record_type"],
                    "record_id": a["record_id"],
                    "label": a.get("source_label") or a["record_type"].replace("_", " ").title(),
                    "link_url": a.get("link_url", ""),
                    "attachments": [],
                }
            seen[key]["attachments"].append(a)
        groups = list(seen.values())

    # Flat view + record type filter
    flat_list = all_atts if view == "flat" else []
    rt_filter = ""
    if view == "flat" and rt:
        rt_filter = rt
        flat_list = [a for a in flat_list if a["record_type"] == rt]

    ctx = {
        "request": request,
        "client": client_dict,
        "view": view,
        "q": q,
        "total_count": total_count,
        "groups": groups,
        "flat_list": flat_list,
        "record_types": record_types,
        "rt_filter": rt_filter,
        "dt_available": is_devonthink_available(),
        "categories": cfg.get("attachment_categories", []),
    }
    # HTMX swaps target #client-files-content — return only the inner partial
    tpl = "clients/_files_content.html" if request.headers.get("HX-Request") else "clients/_tab_files.html"
    return templates.TemplateResponse(tpl, ctx)


@router.get("/{client_id}/quick-brief", response_class=HTMLResponse)
def client_quick_brief(request: Request, client_id: int, conn=Depends(get_db)):
    """Quick Brief slideover - 10-second client digest."""
    from datetime import date
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Client not found", status_code=404)

    today_str = date.today().isoformat()

    # Summary from view
    summary = conn.execute("SELECT * FROM v_client_summary WHERE id = ?", [client_id]).fetchone()

    # Primary contact
    primary_contact = conn.execute("""
        SELECT c.name, c.email, c.phone, c.mobile
        FROM contacts c
        JOIN contact_client_assignments cca ON c.id = cca.contact_id
        WHERE cca.client_id = ? AND cca.contact_type = 'client' AND cca.is_primary = 1
        LIMIT 1
    """, [client_id]).fetchone()

    # Open issues
    open_issues = conn.execute("""
        SELECT subject, issue_severity, issue_uid, issue_sla_days,
               CAST(julianday('now') - julianday(activity_date) AS INTEGER) AS age_days
        FROM activity_log
        WHERE client_id = ? AND item_kind = 'issue'
          AND issue_id IS NULL
          AND issue_status NOT IN ('Resolved', 'Closed')
        ORDER BY CASE issue_severity WHEN 'Critical' THEN 0 WHEN 'High' THEN 1 WHEN 'Normal' THEN 2 ELSE 3 END
    """, [client_id]).fetchall()

    # Recent activities (last 3)
    recent = conn.execute("""
        SELECT activity_type, subject, activity_date
        FROM activity_log
        WHERE client_id = ? AND (item_kind IS NULL OR item_kind = 'followup')
        ORDER BY activity_date DESC, id DESC
        LIMIT 3
    """, [client_id]).fetchall()

    # Upcoming follow-ups (next 3)
    upcoming = conn.execute("""
        SELECT subject, follow_up_date, activity_type, contact_person
        FROM activity_log
        WHERE client_id = ? AND follow_up_date >= ?
          AND (follow_up_done = 0 OR follow_up_done IS NULL)
          AND (item_kind IS NULL OR item_kind = 'followup')
        ORDER BY follow_up_date ASC
        LIMIT 3
    """, [client_id, today_str]).fetchall()

    # Next renewals (next 3)
    renewals = conn.execute("""
        SELECT policy_type, carrier, expiration_date, renewal_status,
               CAST(julianday(expiration_date) - julianday('now') AS INTEGER) AS days_to
        FROM policies
        WHERE client_id = ? AND archived = 0
          AND (is_opportunity = 0 OR is_opportunity IS NULL)
          AND expiration_date >= ?
        ORDER BY expiration_date ASC
        LIMIT 3
    """, [client_id, today_str]).fetchall()

    # Scratchpad
    scratchpad = conn.execute(
        "SELECT content FROM client_scratchpad WHERE client_id = ?",
        [client_id]
    ).fetchone()

    return templates.TemplateResponse("clients/_quick_brief_slideover.html", {
        "request": request,
        "client": dict(client),
        "summary": dict(summary) if summary else {},
        "primary_contact": dict(primary_contact) if primary_contact else None,
        "open_issues": [dict(r) for r in open_issues],
        "recent": [dict(r) for r in recent],
        "upcoming": [dict(r) for r in upcoming],
        "renewals": [dict(r) for r in renewals],
        "scratchpad": scratchpad["content"] if scratchpad else "",
    })


@router.get("/{client_id}", response_class=HTMLResponse)
def client_detail(request: Request, client_id: int, add_contact: str = "", conn=Depends(get_db)):
    from collections import defaultdict
    client = get_client_by_id(conn, client_id, include_archived=True)
    if not client:
        return HTMLResponse("Client not found", status_code=404)
    summary = get_client_summary(conn, client_id)
    all_policies = [dict(p) for p in get_policies_for_client(conn, client_id)]
    opportunities = [p for p in all_policies if p.get("is_opportunity")]
    policies = [p for p in all_policies if not p.get("is_opportunity")]

    # Attach milestone progress and readiness scores for the policy table
    from policydb.web.routes.policies import _attach_milestone_progress, _attach_readiness_score
    policies = _attach_readiness_score(conn, _attach_milestone_progress(conn, policies))

    # Attach full policy contacts list to each opportunity for per-contact email links
    if opportunities:
        opp_ids = [o["id"] for o in opportunities]
        _pc_placeholders = ",".join("?" * len(opp_ids))
        _opp_contacts = conn.execute(
            f"SELECT cpa.policy_id, co.name, co.email, co.phone, cpa.role, co.organization "  # noqa: S608
            f"FROM contact_policy_assignments cpa "
            f"JOIN contacts co ON cpa.contact_id = co.id "
            f"WHERE cpa.policy_id IN ({_pc_placeholders}) ORDER BY cpa.id",
            opp_ids,
        ).fetchall()
        _opp_contacts_map: dict[int, list] = {}
        for _c in _opp_contacts:
            _opp_contacts_map.setdefault(_c["policy_id"], []).append(dict(_c))
        for o in opportunities:
            o["team"] = _opp_contacts_map.get(o["id"], [])
    activities = [dict(a) for a in get_activities(conn, client_id=client_id, days=90)]
    from policydb.web.routes.activities import _attach_pc_emails
    _attach_pc_emails(conn, activities)
    activity_types = cfg.get("activity_types")

    # Group policies by project_name; blank → "Corporate / Standalone" (sorted last).
    # Normalize keys (strip + collapse whitespace + lowercase) so minor format
    # differences ("Main St " vs "main st") still land in the same group.
    def _proj_key(name: str | None) -> str:
        if not name:
            return ""
        return " ".join(name.strip().split()).lower()

    groups: dict[str, list] = defaultdict(list)
    group_display: dict[str, str] = {}  # canonical display name per key
    for p in policies:
        raw = (p.get("project_name") or "").strip()
        key = _proj_key(raw)
        groups[key].append(p)
        if key and key not in group_display:
            group_display[key] = raw

    policy_groups = sorted(
        [(group_display.get(k, ""), v) for k, v in groups.items()],
        key=lambda x: ("\xff" if not x[0] else x[0].lower()),
    )

    # Build tower groups: {project_name: {tower_group: [layers sorted by attachment_point]}}
    # Policies without a tower_group are excluded; blank project_name → "Corporate / Standalone"
    tower_by_project: dict = defaultdict(lambda: defaultdict(list))
    for p in policies:
        tg = p.get("tower_group")
        if tg:
            proj = (p.get("project_name") or "").strip() or "Corporate / Standalone"
            tower_by_project[proj][tg].append(p)

    def _tower_sort_key(lp):
        att = lp.get("attachment_point")
        if att is not None:
            return (float(att), 0)
        pos = lp.get("layer_position") or "Primary"
        try:
            return (-1, int(pos))
        except (ValueError, TypeError):
            return (-1, 0)

    def _attach_ground_up(layers):
        """Sort layers and compute ground-up running limit for each."""
        sorted_layers = sorted(layers, key=_tower_sort_key)
        running = 0.0
        for lp in sorted_layers:
            lim = float(lp.get("limit_amount") or 0)
            att = lp.get("attachment_point")
            part = lp.get("participation_of")
            if att is not None and float(att) >= 0:
                # Use participation_of as the full layer size when present
                layer_size = float(part) if part else lim
                lp["ground_up"] = float(att) + layer_size
            else:
                running += lim
                lp["ground_up"] = running
        return sorted_layers

    # Sort: named projects A-Z, "Corporate / Standalone" last; within each project sort tower groups A-Z
    tower_groups = {
        proj: {
            tg: _attach_ground_up(layers)
            for tg, layers in sorted(tgs.items())
        }
        for proj, tgs in sorted(
            tower_by_project.items(),
            key=lambda x: ("\xff" if x[0] == "Corporate / Standalone" else x[0].lower()),
        )
    }

    # Build proportional tower visuals (groups co-carriers, computes flex/fill)
    def _build_tower_visuals(tg_dict: dict) -> dict:
        from policydb.analysis import layer_notation as _ln
        visuals = {}
        for proj, tgs in tg_dict.items():
            visuals[proj] = {}
            for tg_name, layers in tgs.items():
                if not layers:
                    continue
                total_gu = max(float(l.get("ground_up") or 0) for l in layers)
                if total_gu == 0:
                    continue
                # Group co-carriers at same attachment point
                grouped: dict[str, list] = {}
                for l in layers:
                    att = l.get("attachment_point")
                    key = str(float(att)) if att is not None else f"pos-{l.get('layer_position', 'Primary')}"
                    grouped.setdefault(key, []).append(l)
                visual_layers = []
                for gkey, carriers in grouped.items():
                    parts = [float(c["participation_of"]) for c in carriers if c.get("participation_of")]
                    carrier_limits = [float(c.get("limit_amount") or 0) for c in carriers]
                    if parts:
                        full_limit = max(parts)
                    elif len(carriers) > 1:
                        full_limit = sum(carrier_limits)
                    else:
                        full_limit = carrier_limits[0] if carrier_limits else 0
                    flex = max(full_limit / 1_000_000, 0.5)
                    att_val = carriers[0].get("attachment_point")
                    gu_val = max(float(c.get("ground_up") or 0) for c in carriers)
                    carrier_data = []
                    for c in carriers:
                        climit = float(c.get("limit_amount") or 0)
                        fill_pct = round(climit / full_limit * 100, 1) if full_limit > 0 else 100
                        carrier_data.append({
                            "carrier": c.get("carrier", ""),
                            "limit": climit,
                            "fill_pct": fill_pct,
                            "policy_uid": c.get("policy_uid", ""),
                            "policy_type": c.get("policy_type", ""),
                            "premium": c.get("premium") or 0,
                            "notation": _ln(c.get("limit_amount"), c.get("attachment_point"), c.get("participation_of")) or "",
                        })
                    is_shared = len(carriers) > 1 or any(c.get("participation_of") for c in carriers)
                    total_fill = sum(cd["fill_pct"] for cd in carrier_data)
                    open_pct = round(100 - total_fill, 1) if is_shared and total_fill < 100 else 0
                    open_amount = full_limit - sum(cd["limit"] for cd in carrier_data) if open_pct > 0 else 0
                    visual_layers.append({
                        "attachment": float(att_val) if att_val is not None else None,
                        "full_limit": full_limit,
                        "flex": flex,
                        "ground_up": gu_val,
                        "carriers": carrier_data,
                        "total_premium": sum(cd["premium"] for cd in carrier_data),
                        "is_shared": is_shared,
                        "open_pct": open_pct,
                        "open_amount": open_amount,
                    })
                visual_layers.sort(key=lambda vl: vl["attachment"] if vl["attachment"] is not None else -1)
                visuals[proj][tg_name] = {
                    "layers": visual_layers,
                    "total_ground_up": total_gu,
                    "total_premium": sum(vl["total_premium"] for vl in visual_layers),
                    "carrier_count": sum(len(vl["carriers"]) for vl in visual_layers),
                }
        return visuals

    tower_visuals = _build_tower_visuals(tower_groups) if tower_groups else {}

    # Archived policies for this client (for the collapsed audit section)
    archived_policies = [dict(r) for r in conn.execute(
        """SELECT policy_uid, policy_type, carrier, effective_date, expiration_date,
                  premium, policy_number, project_name
           FROM policies WHERE client_id = ? AND archived = 1
           ORDER BY expiration_date DESC""",
        (client_id,),
    ).fetchall()]

    # Legacy programs removed — programs now come from standalone programs table
    programs = []
    _program_linked_ids = set()

    # Load project notes keyed by normalized project name (from projects table)
    notes_rows = conn.execute(
        "SELECT id, LOWER(TRIM(name)) AS key, name, notes FROM projects WHERE client_id = ?",
        (client_id,),
    ).fetchall()
    project_notes = {r["key"]: r["notes"] for r in notes_rows}
    project_ids = {r["key"]: r["id"] for r in notes_rows}

    # Build project address dict from most recent policy per project
    project_addresses: dict = {}
    for p in sorted(policies, key=lambda x: x.get("id", 0), reverse=True):
        key = _proj_key(p.get("project_name"))
        if key not in project_addresses:
            project_addresses[key] = {
                "exposure_address": p.get("exposure_address") or "",
                "exposure_city":    p.get("exposure_city") or "",
                "exposure_state":   p.get("exposure_state") or "",
                "exposure_zip":     p.get("exposure_zip") or "",
            }

    scratch_row = conn.execute(
        "SELECT content, updated_at FROM client_scratchpad WHERE client_id=?",
        (client_id,),
    ).fetchone()
    client_scratchpad = scratch_row["content"] if scratch_row else ""
    client_scratchpad_updated = scratch_row["updated_at"] if scratch_row else ""
    client_saved_notes = get_saved_notes(conn, "client", str(client_id))

    contacts = get_client_contacts(conn, client_id, contact_type='client')

    team_contacts = get_client_contacts(conn, client_id, contact_type='internal')
    external_contacts = get_client_contacts(conn, client_id, contact_type='external')

    from policydb.email_templates import client_context as _client_ctx, render_tokens as _render_tokens
    _mail_ctx = _client_ctx(conn, client_id)
    mailto_subject = _render_tokens(cfg.get("email_subject_client", "Re: {{client_name}}"), _mail_ctx)

    # Aggregate placement touchpoints from contact_policy_assignments
    # Include archived policies so lost-policy contacts still appear
    _all_pols_incl_archived = [dict(r) for r in conn.execute(
        "SELECT * FROM policies WHERE client_id = ?", (client_id,)
    ).fetchall()]
    _pol_map = {p["id"]: p for p in _all_pols_incl_archived}
    _pol_subj_tpl = cfg.get("email_subject_policy", "Re: {{client_name}} \u2014 {{policy_type}}")
    _colleagues: dict[str, dict] = {}

    def _add_colleague(name: str, email: str, policy_dict: dict, organization: str = "") -> None:
        name = name.strip()
        if not name:
            return
        if name not in _colleagues:
            _colleagues[name] = {"name": name, "email": email.strip(), "organization": organization, "policies": []}
        elif not _colleagues[name]["email"] and email:
            _colleagues[name]["email"] = email.strip()
        p = policy_dict
        _pol_ctx = {
            "client_name": client["name"],
            "policy_type": p.get("policy_type") or "",
            "carrier": p.get("carrier") or "",
            "policy_uid": p.get("policy_uid") or "",
            "effective_date": p.get("effective_date") or "",
            "expiration_date": p.get("expiration_date") or "",
            "project_name": (p.get("project_name") or "").strip(),
            "project_name_sep": f" \u2014 {p.get('project_name')}" if p.get("project_name") else "",
        }
        _colleagues[name]["policies"].append({
            "policy_uid": p.get("policy_uid"),
            "policy_type": p.get("policy_type"),
            "carrier": p.get("carrier"),
            "project_name": (p.get("project_name") or "").strip(),
            "expiration_date": p.get("expiration_date") or "",
            "mailto_subject": _render_tokens(_pol_subj_tpl, _pol_ctx),
            "is_archived": bool(p.get("archived")),
        })

    # Source: contact_policy_assignments + contacts tables
    _pc_rows = conn.execute(
        """SELECT co.name, co.email, co.organization, cpa.policy_id
           FROM contact_policy_assignments cpa
           JOIN contacts co ON cpa.contact_id = co.id
           JOIN policies p ON cpa.policy_id = p.id
           WHERE p.client_id = ?
           ORDER BY co.name, p.policy_type""",
        (client_id,),
    ).fetchall()
    for row in _pc_rows:
        p = _pol_map.get(row["policy_id"])
        if p:
            _add_colleague(row["name"] or "", row["email"] or "", p, row["organization"] or "")

    placement_colleagues = sorted(_colleagues.values(), key=lambda x: x["name"].lower())

    # Render per-opportunity mailto subjects now that _pol_subj_tpl is available
    for o in opportunities:
        _opp_ctx = {
            "client_name": client["name"],
            "policy_type": o.get("policy_type") or "",
            "carrier": o.get("carrier") or "",
            "policy_uid": o.get("policy_uid") or "",
            "effective_date": o.get("target_effective_date") or "",
            "expiration_date": "",
            "project_name": (o.get("project_name") or "").strip(),
            "project_name_sep": f" \u2014 {o.get('project_name')}" if o.get("project_name") else "",
        }
        o["mailto_subject"] = _render_tokens(_pol_subj_tpl, _opp_ctx)

    attach_open_issues(conn, opportunities)

    # All contacts JSON for the contacts card autocomplete
    import json as _json
    _ac_rows = conn.execute(
        """SELECT co.name,
                  MAX(co.email)  AS email,
                  MAX(co.phone)  AS phone,
                  MAX(co.mobile) AS mobile,
                  MAX(cca.title) AS title,
                  MAX(cca.role)  AS role
           FROM contacts co
           JOIN contact_client_assignments cca ON co.id = cca.contact_id
           WHERE cca.contact_type='client' AND co.name IS NOT NULL AND co.name != ''
           GROUP BY co.name ORDER BY co.name"""
    ).fetchall()
    all_contacts_json = _json.dumps({
        r["name"]: {"email": r["email"] or "", "phone": r["phone"] or "",
                    "mobile": r["mobile"] or "",
                    "title": r["title"] or "", "role": r["role"] or ""}
        for r in _ac_rows
    })

    # Risk / Exposure tracking
    risks = [dict(r) for r in conn.execute(
        """SELECT r.*, p.policy_type AS linked_policy_type, p.carrier AS linked_carrier
           FROM client_risks r
           LEFT JOIN policies p ON r.policy_uid = p.policy_uid
           WHERE r.client_id=?
           ORDER BY
             CASE r.severity WHEN 'Critical' THEN 0 WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 ELSE 3 END,
             r.category""",
        (client_id,),
    ).fetchall()]
    for risk in risks:
        risk["coverage_lines"] = [dict(cl) for cl in conn.execute(
            "SELECT * FROM risk_coverage_lines WHERE risk_id=? ORDER BY coverage_line",
            (risk["id"],),
        ).fetchall()]
        risk["controls"] = [dict(c) for c in conn.execute(
            "SELECT * FROM risk_controls WHERE risk_id=? ORDER BY created_at",
            (risk["id"],),
        ).fetchall()]
    risk_categories = cfg.get("risk_categories", [])
    risk_severities = cfg.get("risk_severities", [])
    # Build list of policy UIDs for the "link to policy" dropdown
    policy_uid_options = [{"uid": p["policy_uid"], "label": f"{p['policy_uid']} — {p['policy_type']}"} for p in all_policies if not p.get("archived")]

    client_total_hours = get_client_total_hours(conn, client_id)
    linked_group = get_linked_group_for_client(conn, client_id)

    # Account Pulse data
    from datetime import date as _date
    _today = _date.today().isoformat()
    pulse_overdue = conn.execute(
        """SELECT COUNT(*) AS n FROM activity_log
           WHERE client_id = ? AND follow_up_date < ? AND follow_up_done = 0""",
        (client_id, _today),
    ).fetchone()["n"]
    pulse_overdue += conn.execute(
        """SELECT COUNT(*) AS n FROM policies
           WHERE client_id = ? AND follow_up_date < ? AND follow_up_date IS NOT NULL AND archived = 0""",
        (client_id, _today),
    ).fetchone()["n"]
    # Milestone progress across all active policies
    _ms_rows = conn.execute(
        """SELECT pm.completed FROM policy_milestones pm
           JOIN policies p ON pm.policy_uid = p.policy_uid
           WHERE p.client_id = ? AND p.archived = 0""",
        (client_id,),
    ).fetchall()
    pulse_milestone_done = sum(1 for r in _ms_rows if r["completed"])
    pulse_milestone_total = len(_ms_rows)
    # High/critical risks
    pulse_high_risks = conn.execute(
        "SELECT COUNT(*) AS n FROM client_risks WHERE client_id = ? AND severity IN ('High', 'Critical')",
        (client_id,),
    ).fetchone()["n"]
    # Recent timeline (activities + saved notes interleaved)
    _recent_acts = [dict(r) | {"_type": "activity", "_sort_date": r["activity_date"]} for r in conn.execute(
        """SELECT activity_type, subject, activity_date, duration_hours
           FROM activity_log WHERE client_id = ?
           ORDER BY activity_date DESC, id DESC LIMIT 5""",
        (client_id,),
    ).fetchall()]
    _recent_notes = [n | {"_type": "note", "_sort_date": n["created_at"][:10]} for n in
                     get_saved_notes_for_client_timeline(conn, client_id, limit=5)]
    pulse_recent = sorted(_recent_acts + _recent_notes, key=lambda x: x["_sort_date"], reverse=True)[:5]

    # ── Sidebar enrichment data ────────────────────────────────────────────
    # Renewal calendar: policy count per expiration month
    _rm_rows = conn.execute(
        """SELECT CAST(strftime('%m', expiration_date) AS INTEGER) AS month,
                  COUNT(*) AS cnt
           FROM policies p
           WHERE client_id = ? AND archived = 0
             AND (is_opportunity = 0 OR is_opportunity IS NULL)
             AND expiration_date IS NOT NULL
           GROUP BY month ORDER BY month""",
        (client_id,),
    ).fetchall()
    renewal_month_counts = {r["month"]: r["cnt"] for r in _rm_rows}
    # Compute peak renewal month from data (replaces manual client.renewal_month)
    computed_renewal_month = max(renewal_month_counts, key=renewal_month_counts.get) if renewal_month_counts else None

    # Next follow-up date + days until
    _nf = conn.execute(
        """SELECT MIN(follow_up_date) AS dt FROM activity_log
           WHERE client_id = ? AND follow_up_done = 0
             AND follow_up_date >= date('now')""",
        (client_id,),
    ).fetchone()
    next_followup_date = _nf["dt"] if _nf else None
    next_followup_days = None
    if next_followup_date:
        try:
            import dateparser as _dp
            _nf_dt = _dp.parse(next_followup_date)
            if _nf_dt:
                next_followup_days = (_nf_dt.date() - _date.today()).days
        except Exception:
            pass

    # Last activity date (full history, not limited to 90-day window)
    _la = conn.execute(
        "SELECT MAX(activity_date) AS dt FROM activity_log WHERE client_id = ?",
        (client_id,),
    ).fetchone()
    last_activity_relative = None
    if _la and _la["dt"]:
        try:
            import humanize as _humanize
            import dateparser as _dp
            _la_dt = _dp.parse(_la["dt"])
            if _la_dt:
                last_activity_relative = _humanize.naturaltime(datetime.now() - _la_dt)
        except Exception:
            last_activity_relative = _la["dt"]

    # Attachment rollup count for Files tab badge
    _att_count = conn.execute("""
        SELECT COUNT(DISTINCT a.id) FROM attachments a
        JOIN record_attachments ra ON ra.attachment_id = a.id
        WHERE (ra.record_type = 'client' AND ra.record_id = :cid)
           OR (ra.record_type = 'policy' AND ra.record_id IN (SELECT id FROM policies WHERE client_id = :cid AND archived = 0))
           OR (ra.record_type = 'activity' AND ra.record_id IN (SELECT id FROM activity_log WHERE client_id = :cid))
           OR (ra.record_type = 'project' AND ra.record_id IN (SELECT id FROM projects WHERE client_id = :cid))
           OR (ra.record_type = 'rfi_bundle' AND ra.record_id IN (SELECT id FROM client_request_bundles WHERE client_id = :cid))
           OR (ra.record_type = 'rfi_item' AND ra.record_id IN (
               SELECT cri.id FROM client_request_items cri
               JOIN client_request_bundles crb ON crb.id = cri.bundle_id
               WHERE crb.client_id = :cid
           ))
    """, {"cid": client_id}).fetchone()[0]

    from policydb.queries import REVIEW_CYCLE_LABELS as _REVIEW_CYCLE_LABELS
    from policydb.data_health import score_client as _score_client
    _client_dict = dict(client)
    _score_client(conn, _client_dict, include_staleness=True)

    sidebar_issues = [dict(r) for r in conn.execute(
        """SELECT issue_uid, subject, issue_severity
           FROM activity_log
           WHERE client_id = ?
             AND item_kind = 'issue'
             AND issue_status NOT IN ('Resolved', 'Closed')
           ORDER BY CASE issue_severity
               WHEN 'Critical' THEN 1 WHEN 'High' THEN 2
               WHEN 'Normal' THEN 3 ELSE 4 END
           LIMIT 5""",
        (client_id,),
    ).fetchall()]

    return templates.TemplateResponse("clients/detail.html", {
        "request": request,
        "active": "clients",
        "client": _client_dict,
        "summary": dict(summary) if summary else {},
        "client_total_hours": client_total_hours,
        "policy_groups": policy_groups,
        "tower_groups": tower_groups,
        "tower_visuals": tower_visuals,
        "bundles": _get_request_bundles(conn, client_id),
        "activities": activities,
        "activity_types": activity_types,
        "renewal_statuses": cfg.get("renewal_statuses"),
        "project_notes": project_notes,
        "project_ids": project_ids,
        "project_addresses": project_addresses,
        "archived_policies": archived_policies,
        "programs": programs,
        "program_linked_uids": _program_linked_ids,
        "client_scratchpad": client_scratchpad,
        "client_scratchpad_updated": client_scratchpad_updated,
        "client_saved_notes": client_saved_notes,
        "contacts": contacts,
        "team_contacts": team_contacts,
        "external_contacts": external_contacts,
        "billing_accounts": [dict(r) for r in conn.execute(
            "SELECT * FROM billing_accounts WHERE client_id=? ORDER BY is_master DESC, billing_id",
            (client_id,),
        ).fetchall()],
        "opportunities": opportunities,
        "placement_colleagues": placement_colleagues,
        "mailto_subject": mailto_subject,
        "all_contacts_json": all_contacts_json,
        "add_contact": add_contact,
        "contact_roles": cfg.get("contact_roles", []),
        "all_orgs": _get_all_client_contact_orgs(conn),
        "all_internal_contacts_json": _json.dumps({
            r["name"]: {"title": r["title"] or "", "email": r["email"] or "",
                        "phone": r["phone"] or "", "mobile": r["mobile"] or "", "role": r["role"] or ""}
            for r in conn.execute(
                """SELECT co.name, MAX(cca.title) AS title, MAX(co.email) AS email,
                          MAX(co.phone) AS phone, MAX(co.mobile) AS mobile, MAX(cca.role) AS role
                   FROM contacts co
                   JOIN contact_client_assignments cca ON co.id = cca.contact_id
                   WHERE cca.contact_type='internal' AND co.name IS NOT NULL AND co.name != ''
                   GROUP BY LOWER(TRIM(co.name)) ORDER BY co.name"""
            ).fetchall()
        }),
        "cycle_labels": _REVIEW_CYCLE_LABELS,
        "risks": risks,
        "risk_summary": _compute_risk_summary(risks),
        "risk_categories": risk_categories,
        "risk_severities": risk_severities,
        "risk_sources": cfg.get("risk_sources", []),
        "risk_control_types": cfg.get("risk_control_types", []),
        "risk_control_statuses": cfg.get("risk_control_statuses", []),
        "risk_adequacy_levels": cfg.get("risk_adequacy_levels", []),
        "policy_types": cfg.get("policy_types", []),
        "policy_uid_options": policy_uid_options,
        "linked_group": linked_group,
        "linked_relationships": cfg.get("linked_account_relationships", []),
        "pulse_overdue": pulse_overdue,
        "pulse_next_followup": next_followup_date,
        "pulse_milestone_done": pulse_milestone_done,
        "pulse_milestone_total": pulse_milestone_total,
        "pulse_high_risks": pulse_high_risks,
        "pulse_recent": pulse_recent,
        "today": _today,
        "today_iso": _today,
        "health": _compute_client_health(conn, client_id),
        "renewal_month_counts": renewal_month_counts,
        "computed_renewal_month": computed_renewal_month,
        "next_followup_date": next_followup_date,
        "next_followup_days": next_followup_days,
        "last_activity_relative": last_activity_relative,
        "dispositions": cfg.get("follow_up_dispositions", []),
        "pipeline_projects": _get_project_pipeline(conn, client_id),
        "location_projects": _get_project_locations(conn, client_id),
        "project_stages": cfg.get("project_stages", []),
        "project_types": cfg.get("project_types", []),
        "timeline_data": _build_timeline_data(_get_project_pipeline(conn, client_id)),
        "open_issues": [dict(r) for r in conn.execute("""
            SELECT a.id, a.issue_uid, a.subject, a.issue_status, a.issue_severity,
                   a.activity_date, a.issue_sla_days,
                   julianday(date('now')) - julianday(a.activity_date) AS days_open,
                   (SELECT COUNT(*) FROM activity_log sub WHERE sub.issue_id = a.id) AS activity_count
            FROM activity_log a
            WHERE a.client_id = ? AND a.item_kind = 'issue' AND a.issue_id IS NULL
              AND a.merged_into_id IS NULL
              AND (a.issue_status IS NULL OR a.issue_status NOT IN ('Resolved', 'Closed'))
            ORDER BY CASE a.issue_severity
              WHEN 'Critical' THEN 0 WHEN 'High' THEN 1 WHEN 'Normal' THEN 2 ELSE 3
            END, a.activity_date ASC
        """, (client_id,)).fetchall()],
        "health_score": _client_dict.get("health_score", 100),
        "health_missing": _client_dict.get("health_missing", []),
        "health_threshold": cfg.get("data_health_threshold", 85),
        "sidebar_issues": sidebar_issues,
        "attachment_count": _att_count,
        "pinned_notes": _pinned_notes_for_page(conn, "client", client_id),
        "pinned_scope": "client",
        "pinned_scope_id": str(client_id),
        "pinned_client_id": "",
    })


def _contacts_response(request, conn, client_id: int, duplicate_warning=None):
    """Shared helper: return the client contacts card partial with fresh data."""
    import json as _json
    contacts = get_client_contacts(conn, client_id, contact_type='client')
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    from policydb.email_templates import client_context as _client_ctx, render_tokens as _render_tokens
    _mail_ctx = _client_ctx(conn, client_id)
    mailto_subject = _render_tokens(cfg.get("email_subject_client", "Re: {{client_name}}"), _mail_ctx)
    # All known contacts across all clients for autocomplete fill
    all_ac_rows = conn.execute(
        """SELECT co.name,
                  MAX(co.email)  AS email,
                  MAX(co.phone)  AS phone,
                  MAX(co.mobile) AS mobile,
                  MAX(cca.title) AS title,
                  MAX(cca.role)  AS role
           FROM contacts co
           JOIN contact_client_assignments cca ON co.id = cca.contact_id
           WHERE cca.contact_type='client' AND co.name IS NOT NULL AND co.name != ''
           GROUP BY co.name ORDER BY co.name"""
    ).fetchall()
    all_contacts_json = _json.dumps({
        r["name"]: {"email": r["email"] or "", "phone": r["phone"] or "",
                    "mobile": r["mobile"] or "",
                    "title": r["title"] or "", "role": r["role"] or ""}
        for r in all_ac_rows
    })
    return templates.TemplateResponse("clients/_contacts.html", {
        "request": request,
        "client": dict(client) if client else {},
        "contacts": contacts,
        "mailto_subject": mailto_subject,
        "all_contacts_json": all_contacts_json,
        "contact_roles": cfg.get("contact_roles", []),
        "all_orgs": _get_all_client_contact_orgs(conn),
        "duplicate_warning": duplicate_warning,
    })


def _internal_contacts_response(request, conn, client_id: int):
    """Shared helper: return the internal team contacts card partial with fresh data."""
    import json as _json
    team_contacts = get_client_contacts(conn, client_id, contact_type='internal')
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    from policydb.email_templates import client_context as _client_ctx, render_tokens as _render_tokens
    _mail_ctx = _client_ctx(conn, client_id)
    mailto_subject = _render_tokens(cfg.get("email_subject_client", "Re: {{client_name}}"), _mail_ctx)
    # Autocomplete: all internal contacts across all clients, deduped by name
    _ac_rows = conn.execute(
        """SELECT co.name,
                  MAX(cca.title) AS title,
                  MAX(co.email)  AS email,
                  MAX(co.phone)  AS phone,
                  MAX(co.mobile) AS mobile,
                  MAX(cca.role)  AS role
           FROM contacts co
           JOIN contact_client_assignments cca ON co.id = cca.contact_id
           WHERE cca.contact_type='internal' AND co.name IS NOT NULL AND co.name != ''
           GROUP BY LOWER(TRIM(co.name)) ORDER BY co.name"""
    ).fetchall()
    all_internal_contacts_json = _json.dumps({
        r["name"]: {"title": r["title"] or "", "email": r["email"] or "",
                    "phone": r["phone"] or "", "mobile": r["mobile"] or "", "role": r["role"] or ""}
        for r in _ac_rows
    })
    return templates.TemplateResponse("clients/_team_contacts.html", {
        "request": request,
        "client": dict(client) if client else {},
        "team_contacts": team_contacts,
        "mailto_subject": mailto_subject,
        "all_internal_contacts_json": all_internal_contacts_json,
        "contact_roles": cfg.get("contact_roles", []),
        "team_assignments": cfg.get("team_assignments", []),
        "add_contact": "",
    })


@router.patch("/{client_id}/contacts/{contact_id}/cell")
async def contact_cell(request: Request, client_id: int, contact_id: int, conn=Depends(get_db)):
    """Save a single cell value for a client contact (matrix edit).
    contact_id in URL = assignment_id in new schema."""
    body = await request.json()
    field, value = body.get("field", ""), body.get("value", "")
    allowed = {"name", "title", "role", "email", "phone", "mobile", "notes", "organization"}
    if field not in allowed:
        return JSONResponse({"ok": False, "error": "Invalid field"}, status_code=400)
    formatted = value.strip()
    if field in ("phone", "mobile"):
        formatted = format_phone(formatted) if formatted else ""
    elif field == "email":
        formatted = clean_email(formatted) or ""
    # Shared fields -> update contacts table; per-assignment fields -> update junction table
    if field in ("name", "email", "phone", "mobile", "organization"):
        row = conn.execute("SELECT contact_id FROM contact_client_assignments WHERE id=?", (contact_id,)).fetchone()
        if row:
            conn.execute(
                f"UPDATE contacts SET {field}=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (formatted or None, row["contact_id"]),
            )
    else:
        # Per-assignment fields: role, title, notes
        conn.execute(
            f"UPDATE contact_client_assignments SET {field}=? WHERE id=? AND client_id=?",
            (formatted or None, contact_id, client_id),
        )
    conn.commit()
    return JSONResponse({"ok": True, "formatted": formatted})


@router.post("/{client_id}/contacts/add-row", response_class=HTMLResponse)
def contact_add_row(request: Request, client_id: int, conn=Depends(get_db)):
    """Create blank client contact row and return matrix row HTML."""
    cid = get_or_create_contact(conn, 'New Contact')
    cur = conn.execute(
        "INSERT INTO contact_client_assignments (contact_id, client_id, contact_type) VALUES (?, ?, 'client')",
        (cid, client_id),
    )
    conn.commit()
    c = {"id": cur.lastrowid, "contact_id": cid, "name": "New Contact", "title": None, "role": None,
         "email": None, "phone": None, "mobile": None, "notes": None,
         "organization": None, "is_primary": 0}
    return templates.TemplateResponse("clients/_contact_matrix_row.html", {
        "request": request, "c": c, "client": {"id": client_id},
        "contact_roles": cfg.get("contact_roles", []),
        "all_orgs": _get_all_client_contact_orgs(conn),
    })


@router.post("/{client_id}/contacts/add", response_class=HTMLResponse)
def contact_add(
    request: Request,
    client_id: int,
    name: str = Form(...),
    title: str = Form(""),
    role: str = Form(""),
    email: str = Form(""),
    phone: str = Form(""),
    mobile: str = Form(""),
    notes: str = Form(""),
    conn=Depends(get_db),
):
    from policydb.web.routes.contacts import _find_similar_contacts
    # Run duplicate check before creating — warn but don't block
    dupes = _find_similar_contacts(conn, name.strip(), source="client") if name.strip() else []
    cid = get_or_create_contact(conn, name,
                                email=clean_email(email) or None,
                                phone=format_phone(phone) or None,
                                mobile=format_phone(mobile) or None)
    assign_contact_to_client(conn, cid, client_id, contact_type='client',
                             title=title or None, role=role or None, notes=notes or None)
    conn.commit()
    # Pass duplicate warning so user can see it even after creation
    return _contacts_response(request, conn, client_id, duplicate_warning=dupes or None)


@router.post("/{client_id}/contacts/{contact_id}/edit", response_class=HTMLResponse)
def contact_edit(
    request: Request,
    client_id: int,
    contact_id: int,
    name: str = Form(...),
    title: str = Form(""),
    role: str = Form(""),
    email: str = Form(""),
    phone: str = Form(""),
    mobile: str = Form(""),
    notes: str = Form(""),
    conn=Depends(get_db),
):
    # contact_id = assignment_id; look up the real contact_id
    assignment = conn.execute(
        "SELECT contact_id FROM contact_client_assignments WHERE id=?", (contact_id,)
    ).fetchone()
    if assignment:
        # Update shared fields on the contacts table
        conn.execute(
            "UPDATE contacts SET name=?, email=?, phone=?, mobile=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (name, clean_email(email) or None, format_phone(phone) or None,
             format_phone(mobile) or None, assignment["contact_id"]),
        )
    # Update per-assignment fields on the junction table
    conn.execute(
        "UPDATE contact_client_assignments SET title=?, role=?, notes=? WHERE id=? AND client_id=?",
        (title or None, role or None, notes or None, contact_id, client_id),
    )

    conn.commit()
    return _contacts_response(request, conn, client_id)


@router.post("/{client_id}/contacts/{contact_id}/delete", response_class=HTMLResponse)
def contact_delete(
    request: Request,
    client_id: int,
    contact_id: int,
    conn=Depends(get_db),
):
    remove_contact_from_client(conn, contact_id)
    conn.commit()
    return _contacts_response(request, conn, client_id)


@router.post("/{client_id}/contacts/{contact_id}/set-primary", response_class=HTMLResponse)
def contact_set_primary_route(
    request: Request,
    client_id: int,
    contact_id: int,
    conn=Depends(get_db),
):
    set_primary_contact(conn, client_id, contact_id)
    conn.commit()
    return _contacts_response(request, conn, client_id)


@router.patch("/{client_id}/team/{contact_id}/cell")
async def team_contact_cell(request: Request, client_id: int, contact_id: int, conn=Depends(get_db)):
    """Save a single cell value for an internal team contact (matrix edit).
    contact_id in URL = assignment_id in new schema."""
    body = await request.json()
    field, value = body.get("field", ""), body.get("value", "")
    allowed = {"name", "title", "role", "assignment", "email", "phone", "mobile"}
    if field not in allowed:
        return JSONResponse({"ok": False, "error": "Invalid field"}, status_code=400)
    formatted = value.strip()
    if field in ("phone", "mobile"):
        formatted = format_phone(formatted) if formatted else ""
    elif field == "email":
        formatted = clean_email(formatted) or ""
    # Shared fields -> update contacts table; per-assignment fields -> update junction table
    if field in ("name", "email", "phone", "mobile"):
        row = conn.execute("SELECT contact_id FROM contact_client_assignments WHERE id=?", (contact_id,)).fetchone()
        if row:
            conn.execute(
                f"UPDATE contacts SET {field}=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (formatted or None, row["contact_id"]),
            )
    else:
        # Per-assignment fields: role, title, assignment
        conn.execute(
            f"UPDATE contact_client_assignments SET {field}=? WHERE id=? AND client_id=?",
            (formatted or None, contact_id, client_id),
        )
    conn.commit()
    return JSONResponse({"ok": True, "formatted": formatted})


@router.post("/{client_id}/team/add-row", response_class=HTMLResponse)
def team_contact_add_row(request: Request, client_id: int, conn=Depends(get_db)):
    """Create blank internal team contact row and return matrix row HTML."""
    cid = get_or_create_contact(conn, 'New Contact')
    cur = conn.execute(
        "INSERT INTO contact_client_assignments (contact_id, client_id, contact_type) VALUES (?, ?, 'internal')",
        (cid, client_id),
    )
    conn.commit()
    c = {"id": cur.lastrowid, "contact_id": cid, "name": "New Contact", "title": None, "role": None,
         "assignment": None, "email": None, "phone": None, "mobile": None}
    return templates.TemplateResponse("clients/_team_matrix_row.html", {
        "request": request, "c": c, "client": {"id": client_id},
        "contact_roles": cfg.get("contact_roles", []),
        "team_assignments": cfg.get("team_assignments", []),
    })


@router.post("/{client_id}/team/add", response_class=HTMLResponse)
def team_contact_add(
    request: Request,
    client_id: int,
    name: str = Form(...),
    title: str = Form(""),
    role: str = Form(""),
    assignment: str = Form(""),
    email: str = Form(""),
    phone: str = Form(""),
    mobile: str = Form(""),
    conn=Depends(get_db),
):
    cid = get_or_create_contact(conn, name,
                                email=clean_email(email) or None,
                                phone=format_phone(phone) or None,
                                mobile=format_phone(mobile) or None)
    assign_contact_to_client(conn, cid, client_id, contact_type='internal',
                             title=title or None, role=role or None,
                             assignment=assignment or None)
    conn.commit()
    return _internal_contacts_response(request, conn, client_id)


@router.post("/{client_id}/team/{contact_id}/edit", response_class=HTMLResponse)
def team_contact_edit(
    request: Request,
    client_id: int,
    contact_id: int,
    name: str = Form(...),
    title: str = Form(""),
    role: str = Form(""),
    assignment: str = Form(""),
    email: str = Form(""),
    phone: str = Form(""),
    mobile: str = Form(""),
    conn=Depends(get_db),
):
    # contact_id = assignment_id; look up the real contact_id
    asgn = conn.execute(
        "SELECT contact_id FROM contact_client_assignments WHERE id=?", (contact_id,)
    ).fetchone()
    if asgn:
        # Update shared fields on the contacts table
        conn.execute(
            "UPDATE contacts SET name=?, email=?, phone=?, mobile=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (name, clean_email(email) or None, format_phone(phone) or None,
             format_phone(mobile) or None, asgn["contact_id"]),
        )
    # Update per-assignment fields on the junction table
    conn.execute(
        "UPDATE contact_client_assignments SET title=?, role=?, assignment=? WHERE id=? AND client_id=?",
        (title or None, role or None, assignment or None, contact_id, client_id),
    )

    conn.commit()
    return _internal_contacts_response(request, conn, client_id)


@router.post("/{client_id}/team/{contact_id}/delete", response_class=HTMLResponse)
def team_contact_delete(
    request: Request,
    client_id: int,
    contact_id: int,
    conn=Depends(get_db),
):
    remove_contact_from_client(conn, contact_id)
    conn.commit()
    return _internal_contacts_response(request, conn, client_id)


# ── External Stakeholder Contacts ─────────────────────────────────────────────

def _external_contacts_response(request, conn, client_id: int):
    """Return rendered _external_contacts.html partial."""
    external_contacts = get_client_contacts(conn, client_id, contact_type='external')
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    from policydb.email_templates import client_context as _client_ctx, render_tokens as _render_tokens
    _mail_ctx = _client_ctx(conn, client_id)
    mailto_subject = _render_tokens(cfg.get("email_subject_client", "Re: {{client_name}}"), _mail_ctx)
    return templates.TemplateResponse("clients/_external_contacts.html", {
        "request": request,
        "client": dict(client) if client else {},
        "external_contacts": external_contacts,
        "mailto_subject": mailto_subject,
        "contact_roles": cfg.get("contact_roles", []),
        "all_orgs": sorted({r["organization"] for r in conn.execute("SELECT DISTINCT organization FROM contacts WHERE organization IS NOT NULL AND organization != ''").fetchall()}),
    })


@router.post("/{client_id}/external/assign", response_class=HTMLResponse)
def external_contact_assign(
    request: Request, client_id: int,
    name: str = Form(...), email: str = Form(""), phone: str = Form(""),
    mobile: str = Form(""), role: str = Form(""), title: str = Form(""),
    organization: str = Form(""),
    conn=Depends(get_db),
):
    """Assign an existing contact as an external stakeholder."""
    cid = get_or_create_contact(conn, name,
                                email=clean_email(email) or None,
                                phone=format_phone(phone) or None,
                                mobile=format_phone(mobile) or None)
    if organization:
        conn.execute("UPDATE contacts SET organization=? WHERE id=? AND (organization IS NULL OR organization='')",
                     (organization, cid))
    assign_contact_to_client(conn, cid, client_id, contact_type='external',
                             title=title or None, role=role or None)
    conn.commit()
    return _external_contacts_response(request, conn, client_id)


@router.post("/{client_id}/external/add-row", response_class=HTMLResponse)
def external_contact_add_row(request: Request, client_id: int, conn=Depends(get_db)):
    """Create blank external stakeholder contact — returns single row for initMatrix()."""
    cid = get_or_create_contact(conn, 'New Contact')
    conn.execute(
        "INSERT INTO contact_client_assignments (contact_id, client_id, contact_type) VALUES (?, ?, 'external')",
        (cid, client_id),
    )
    conn.commit()
    # Get the new assignment row (initMatrix expects a single <tr>, not the full card)
    assignment = conn.execute(
        """SELECT ca.id, ca.contact_id, c.name, c.email, c.phone, c.mobile, c.organization,
                  ca.role, ca.notes
           FROM contact_client_assignments ca
           JOIN contacts c ON c.id = ca.contact_id
           WHERE ca.contact_id = ? AND ca.client_id = ? AND ca.contact_type = 'external'
           ORDER BY ca.id DESC LIMIT 1""",
        (cid, client_id),
    ).fetchone()
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    from policydb.email_templates import client_context as _client_ctx, render_tokens as _render_tokens
    _mail_ctx = _client_ctx(conn, client_id)
    mailto_subject = _render_tokens(cfg.get("email_subject_client", "Re: {{client_name}}"), _mail_ctx)
    return templates.TemplateResponse("clients/_external_contact_row.html", {
        "request": request,
        "c": dict(assignment),
        "client": dict(client) if client else {},
        "mailto_subject": mailto_subject,
        "contact_roles": cfg.get("contact_roles", []),
        "all_orgs": sorted({r["organization"] for r in conn.execute("SELECT DISTINCT organization FROM contacts WHERE organization IS NOT NULL AND organization != ''").fetchall()}),
    })


@router.patch("/{client_id}/external/{contact_id}/cell")
async def external_contact_cell(request: Request, client_id: int, contact_id: int, conn=Depends(get_db)):
    """Save a single cell value for an external stakeholder contact."""
    body = await request.json()
    field, value = body.get("field", ""), body.get("value", "")
    allowed = {"name", "organization", "role", "notes", "email", "phone", "mobile"}
    if field not in allowed:
        return JSONResponse({"ok": False, "error": "Invalid field"}, status_code=400)
    formatted = value.strip()
    if field in ("phone", "mobile"):
        formatted = format_phone(formatted) if formatted else ""
    elif field == "email":
        formatted = clean_email(formatted) or ""
    if field in ("name", "email", "phone", "mobile", "organization"):
        row = conn.execute("SELECT contact_id FROM contact_client_assignments WHERE id=?", (contact_id,)).fetchone()
        if row:
            conn.execute(
                f"UPDATE contacts SET {field}=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (formatted or None, row["contact_id"]),
            )
    else:
        conn.execute(
            f"UPDATE contact_client_assignments SET {field}=? WHERE id=? AND client_id=?",
            (formatted or None, contact_id, client_id),
        )
    conn.commit()
    return JSONResponse({"ok": True, "formatted": formatted})


@router.post("/{client_id}/external/{contact_id}/delete", response_class=HTMLResponse)
def external_contact_delete(request: Request, client_id: int, contact_id: int, conn=Depends(get_db)):
    """Remove external stakeholder contact."""
    remove_contact_from_client(conn, contact_id)
    conn.commit()
    return _external_contacts_response(request, conn, client_id)


# ── Risk / Exposure Tracking ──────────────────────────────────────────────────

def _update_has_coverage(conn, risk_id: int):
    """Auto-derive has_coverage: 1 if any coverage line is Adequate, else 0."""
    row = conn.execute(
        "SELECT COUNT(*) AS cnt FROM risk_coverage_lines WHERE risk_id=? AND adequacy='Adequate'",
        (risk_id,),
    ).fetchone()
    conn.execute(
        "UPDATE client_risks SET has_coverage=? WHERE id=?",
        (1 if row["cnt"] > 0 else 0, risk_id),
    )


def _policy_uid_options(conn, client_id: int):
    all_policies = [dict(r) for r in conn.execute(
        "SELECT policy_uid, policy_type FROM policies WHERE client_id=? AND archived=0 ORDER BY policy_type",
        (client_id,),
    ).fetchall()]
    return [{"uid": p["policy_uid"], "label": f"{p['policy_uid']} — {p['policy_type']}"} for p in all_policies]


def _compute_client_health(conn, client_id: int) -> dict:
    """Compute client health score (0-100) from weighted signals."""
    from datetime import date
    today = date.today()
    today_str = today.isoformat()
    score = 100
    factors: list[dict] = []

    # 1. Days since last activity (max deduction: 25)
    last = conn.execute(
        "SELECT MAX(activity_date) AS d FROM activity_log WHERE client_id = ?",
        [client_id],
    ).fetchone()
    last_date = last["d"] if last else None
    if last_date:
        days_since = (today - date.fromisoformat(last_date)).days
    else:
        days_since = 999

    if days_since > 90:
        score -= 25
        factors.append({"label": f"No activity in {days_since}d", "impact": -25, "color": "red"})
    elif days_since > 60:
        score -= 20
        factors.append({"label": f"Last activity {days_since}d ago", "impact": -20, "color": "red"})
    elif days_since > 30:
        score -= 15
        factors.append({"label": f"Last activity {days_since}d ago", "impact": -15, "color": "amber"})
    elif days_since > 14:
        score -= 10
        factors.append({"label": f"Last activity {days_since}d ago", "impact": -10, "color": "amber"})
    elif days_since > 7:
        score -= 5
        factors.append({"label": f"Last activity {days_since}d ago", "impact": -5, "color": "neutral"})

    # 2. Overdue follow-ups (max deduction: 25)
    overdue_count = conn.execute(
        """SELECT COUNT(*) AS n FROM activity_log
           WHERE client_id = ? AND follow_up_date < ?
             AND (follow_up_done = 0 OR follow_up_done IS NULL)
             AND (item_kind IS NULL OR item_kind = 'followup')""",
        [client_id, today_str],
    ).fetchone()["n"]

    if overdue_count >= 3:
        score -= 25
        factors.append({"label": f"{overdue_count} overdue follow-ups", "impact": -25, "color": "red"})
    elif overdue_count == 2:
        score -= 15
        factors.append({"label": "2 overdue follow-ups", "impact": -15, "color": "red"})
    elif overdue_count == 1:
        score -= 10
        factors.append({"label": "1 overdue follow-up", "impact": -10, "color": "amber"})

    # 3. Open issues (max deduction: 25)
    issues = conn.execute(
        """SELECT issue_severity, issue_sla_days, activity_date,
                  CAST(julianday('now') - julianday(activity_date) AS INTEGER) AS age_days
           FROM activity_log
           WHERE client_id = ? AND item_kind = 'issue'
             AND merged_into_id IS NULL
             AND (issue_status IS NULL OR issue_status NOT IN ('Resolved', 'Closed'))""",
        [client_id],
    ).fetchall()

    sla_breached = any(
        r["issue_sla_days"] and r["age_days"] > r["issue_sla_days"] for r in issues
    )
    critical_count = sum(1 for r in issues if r["issue_severity"] == "Critical")
    high_count = sum(1 for r in issues if r["issue_severity"] == "High")
    other_count = sum(1 for r in issues if r["issue_severity"] not in ("Critical", "High"))

    issue_deduction = 0
    if sla_breached:
        issue_deduction = 25
        factors.append({"label": "Issue SLA breached", "impact": -25, "color": "red"})
    else:
        issue_deduction += min(critical_count * 15, 20)
        issue_deduction += min(high_count * 10, 15)
        issue_deduction += min(other_count * 5, 10)
        issue_deduction = min(issue_deduction, 25)
        if critical_count:
            factors.append({"label": f"{critical_count} critical issue(s)", "impact": -min(critical_count * 15, 20), "color": "red"})
        if high_count:
            factors.append({"label": f"{high_count} high issue(s)", "impact": -min(high_count * 10, 15), "color": "amber"})
    score -= issue_deduction

    # 4. Renewal proximity without recent action (max deduction: 15)
    next_renewal = conn.execute(
        """SELECT MIN(CAST(julianday(expiration_date) - julianday('now') AS INTEGER)) AS days_to
           FROM policies WHERE client_id = ? AND archived = 0
             AND (is_opportunity = 0 OR is_opportunity IS NULL)
             AND expiration_date >= ?""",
        [client_id, today_str],
    ).fetchone()

    days_to_renewal = next_renewal["days_to"] if next_renewal and next_renewal["days_to"] is not None else 999
    if days_to_renewal < 30 and days_since > 7:
        score -= 15
        factors.append({"label": f"Renewal in {days_to_renewal}d, no recent activity", "impact": -15, "color": "red"})
    elif days_to_renewal < 60 and days_since > 14:
        score -= 10
        factors.append({"label": f"Renewal in {days_to_renewal}d, no recent activity", "impact": -10, "color": "amber"})

    # 5. Open high/critical risks (max deduction: 10)
    risk_count = conn.execute(
        "SELECT COUNT(*) AS n FROM client_risks WHERE client_id = ? AND severity IN ('High', 'Critical')",
        [client_id],
    ).fetchone()["n"]

    if risk_count >= 3:
        score -= 10
        factors.append({"label": f"{risk_count} high/critical risks", "impact": -10, "color": "amber"})
    elif risk_count >= 1:
        score -= 5
        factors.append({"label": f"{risk_count} high/critical risk(s)", "impact": -5, "color": "neutral"})

    score = max(score, 0)
    level = "green" if score >= 70 else "amber" if score >= 40 else "red"

    return {
        "score": score,
        "level": level,
        "label": "Healthy" if level == "green" else "Needs Attention" if level == "amber" else "At Risk",
        "factors": factors,
    }


def _compute_risk_summary(risks: list[dict]) -> dict:
    """Compute aggregate risk posture stats for visual widgets."""
    total = len(risks)
    covered = sum(1 for r in risks if r.get("has_coverage"))
    gap = total - covered
    needs_review = sum(1 for r in risks if any(
        cl.get("adequacy") == "Needs Review" for cl in r.get("coverage_lines", [])
    ))
    total_controls = sum(len(r.get("controls", [])) for r in risks)
    implemented_controls = sum(
        sum(1 for c in r.get("controls", []) if c.get("status") == "Implemented")
        for r in risks
    )
    by_severity = {}
    for r in risks:
        sev = r.get("severity", "Unknown")
        by_severity.setdefault(sev, 0)
        by_severity[sev] += 1
    # Score: weighted coverage ratio (0-100)
    score = int(covered / total * 100) if total else 0
    return {
        "score": score,
        "total": total,
        "covered": covered,
        "gap": gap,
        "needs_review": needs_review,
        "total_controls": total_controls,
        "implemented_controls": implemented_controls,
        "by_severity": by_severity,
    }


def _risks_response(request, conn, client_id: int):
    """Shared helper: return the risks card partial with fresh data."""
    risks = [dict(r) for r in conn.execute(
        """SELECT r.*, p.policy_type AS linked_policy_type, p.carrier AS linked_carrier
           FROM client_risks r
           LEFT JOIN policies p ON r.policy_uid = p.policy_uid
           WHERE r.client_id=?
           ORDER BY
             CASE r.severity WHEN 'Critical' THEN 0 WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 ELSE 3 END,
             r.category""",
        (client_id,),
    ).fetchall()]
    # Attach coverage lines and controls per risk
    for risk in risks:
        risk["coverage_lines"] = [dict(cl) for cl in conn.execute(
            "SELECT * FROM risk_coverage_lines WHERE risk_id=? ORDER BY coverage_line",
            (risk["id"],),
        ).fetchall()]
        risk["controls"] = [dict(c) for c in conn.execute(
            "SELECT * FROM risk_controls WHERE risk_id=? ORDER BY created_at",
            (risk["id"],),
        ).fetchall()]
    return templates.TemplateResponse("clients/_risks.html", {
        "request": request,
        "client": {"id": client_id},
        "risks": risks,
        "risk_summary": _compute_risk_summary(risks),
        "risk_categories": cfg.get("risk_categories", []),
        "risk_severities": cfg.get("risk_severities", []),
        "risk_sources": cfg.get("risk_sources", []),
        "risk_control_types": cfg.get("risk_control_types", []),
        "risk_control_statuses": cfg.get("risk_control_statuses", []),
        "risk_adequacy_levels": cfg.get("risk_adequacy_levels", []),
        "policy_types": cfg.get("policy_types", []),
        "policy_uid_options": _policy_uid_options(conn, client_id),
    })


@router.get("/{client_id}/risks", response_class=HTMLResponse)
def risks_card(request: Request, client_id: int, conn=Depends(get_db)):
    """Return the full risks card partial (used by Cancel buttons)."""
    return _risks_response(request, conn, client_id)


@router.post("/{client_id}/risks/add", response_class=HTMLResponse)
def risk_add(
    request: Request,
    client_id: int,
    category: str = Form(...),
    description: str = Form(""),
    severity: str = Form("Medium"),
    source: str = Form(""),
    review_date: str = Form(""),
    has_coverage: int = Form(0),
    policy_uid: str = Form(""),
    notes: str = Form(""),
    conn=Depends(get_db),
):
    conn.execute(
        """INSERT INTO client_risks (client_id, category, description, severity, has_coverage, policy_uid, notes, source, review_date)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (client_id, category.strip(), description.strip() or None,
         severity, has_coverage, policy_uid.strip().upper() or None, notes.strip() or None,
         source.strip() or None, review_date.strip() or None),
    )
    conn.commit()
    return _risks_response(request, conn, client_id)


@router.patch("/{client_id}/risks/{risk_id}/cell")
async def risk_cell_save(request: Request, client_id: int, risk_id: int, conn=Depends(get_db)):
    """Save a single cell value from the risk matrix."""
    body = await request.json()
    field = body.get("field", "")
    value = body.get("value", "")
    allowed = {"category", "description", "severity", "source", "notes"}
    if field not in allowed:
        return JSONResponse({"ok": False, "error": "Invalid field"}, status_code=400)
    # Category and severity require a non-empty value
    if field in ("category", "severity") and not value.strip():
        return JSONResponse({"ok": False, "error": f"{field} cannot be empty"}, status_code=400)
    conn.execute(
        f"UPDATE client_risks SET {field}=? WHERE id=? AND client_id=?",
        (value.strip() or None, risk_id, client_id),
    )
    conn.commit()
    return JSONResponse({"ok": True})


@router.post("/{client_id}/risks/add-row", response_class=HTMLResponse)
def risk_add_row(request: Request, client_id: int, conn=Depends(get_db)):
    """Create a new risk with defaults and return a single matrix row."""
    categories = cfg.get("risk_categories", [])
    default_cat = categories[0] if categories else "General"
    cur = conn.execute(
        """INSERT INTO client_risks (client_id, category, severity, has_coverage)
           VALUES (?,?,?,0)""",
        (client_id, default_cat, "Medium"),
    )
    conn.commit()
    risk_id = cur.lastrowid
    risk = dict(conn.execute("SELECT * FROM client_risks WHERE id=?", (risk_id,)).fetchone())
    risk["coverage_lines"] = []
    risk["controls"] = []
    return templates.TemplateResponse("clients/_risk_matrix_row.html", {
        "request": request,
        "client": {"id": client_id},
        "r": risk,
        "risk_categories": categories,
        "risk_severities": cfg.get("risk_severities", []),
        "risk_sources": cfg.get("risk_sources", []),
        "risk_control_types": cfg.get("risk_control_types", []),
        "risk_control_statuses": cfg.get("risk_control_statuses", []),
        "risk_adequacy_levels": cfg.get("risk_adequacy_levels", []),
        "policy_types": cfg.get("policy_types", []),
        "policy_uid_options": _policy_uid_options(conn, client_id),
    })


@router.get("/{client_id}/risks/{risk_id}/edit", response_class=HTMLResponse)
def risk_edit_form(request: Request, client_id: int, risk_id: int, conn=Depends(get_db)):
    risk = conn.execute("SELECT * FROM client_risks WHERE id=? AND client_id=?", (risk_id, client_id)).fetchone()
    if not risk:
        return HTMLResponse("", status_code=404)
    return templates.TemplateResponse("clients/_risk_row_edit.html", {
        "request": request,
        "client": {"id": client_id},
        "r": dict(risk),
        "risk_categories": cfg.get("risk_categories", []),
        "risk_severities": cfg.get("risk_severities", []),
        "risk_sources": cfg.get("risk_sources", []),
        "policy_uid_options": _policy_uid_options(conn, client_id),
    })


@router.post("/{client_id}/risks/{risk_id}/edit", response_class=HTMLResponse)
def risk_edit_save(
    request: Request,
    client_id: int,
    risk_id: int,
    category: str = Form(...),
    description: str = Form(""),
    severity: str = Form("Medium"),
    source: str = Form(""),
    review_date: str = Form(""),
    has_coverage: int = Form(0),
    policy_uid: str = Form(""),
    notes: str = Form(""),
    conn=Depends(get_db),
):
    conn.execute(
        """UPDATE client_risks SET category=?, description=?, severity=?, has_coverage=?, policy_uid=?, notes=?,
           source=?, review_date=?
           WHERE id=? AND client_id=?""",
        (category.strip(), description.strip() or None, severity,
         has_coverage, policy_uid.strip().upper() or None, notes.strip() or None,
         source.strip() or None, review_date.strip() or None,
         risk_id, client_id),
    )
    conn.commit()
    return _risks_response(request, conn, client_id)


@router.post("/{client_id}/risks/{risk_id}/toggle-coverage", response_class=HTMLResponse)
def risk_toggle_coverage(request: Request, client_id: int, risk_id: int, conn=Depends(get_db)):
    current = conn.execute(
        "SELECT has_coverage FROM client_risks WHERE id=? AND client_id=?", (risk_id, client_id)
    ).fetchone()
    if current:
        conn.execute(
            "UPDATE client_risks SET has_coverage=? WHERE id=? AND client_id=?",
            (0 if current["has_coverage"] else 1, risk_id, client_id),
        )
        conn.commit()
    return _risks_response(request, conn, client_id)


@router.post("/{client_id}/risks/{risk_id}/delete", response_class=HTMLResponse)
def risk_delete(request: Request, client_id: int, risk_id: int, conn=Depends(get_db)):
    conn.execute("DELETE FROM client_risks WHERE id=? AND client_id=?", (risk_id, client_id))
    conn.commit()
    return _risks_response(request, conn, client_id)


# ── Coverage Lines ────────────────────────────────────────────────────────────

@router.post("/{client_id}/risks/{risk_id}/lines/add", response_class=HTMLResponse)
def risk_line_add(
    request: Request,
    client_id: int,
    risk_id: int,
    coverage_line: str = Form(...),
    adequacy: str = Form("Needs Review"),
    policy_uid: str = Form(""),
    notes: str = Form(""),
    conn=Depends(get_db),
):
    conn.execute(
        """INSERT OR IGNORE INTO risk_coverage_lines (risk_id, coverage_line, adequacy, policy_uid, notes)
           VALUES (?,?,?,?,?)""",
        (risk_id, coverage_line.strip(), adequacy, policy_uid.strip().upper() or None, notes.strip() or None),
    )
    _update_has_coverage(conn, risk_id)
    conn.commit()
    return _risks_response(request, conn, client_id)


@router.post("/{client_id}/risks/{risk_id}/lines/{line_id}/edit", response_class=HTMLResponse)
def risk_line_edit(
    request: Request,
    client_id: int,
    risk_id: int,
    line_id: int,
    adequacy: str = Form("Needs Review"),
    policy_uid: str = Form(""),
    notes: str = Form(""),
    conn=Depends(get_db),
):
    conn.execute(
        "UPDATE risk_coverage_lines SET adequacy=?, policy_uid=?, notes=? WHERE id=? AND risk_id=?",
        (adequacy, policy_uid.strip().upper() or None, notes.strip() or None, line_id, risk_id),
    )
    _update_has_coverage(conn, risk_id)
    conn.commit()
    return _risks_response(request, conn, client_id)


@router.post("/{client_id}/risks/{risk_id}/lines/{line_id}/delete", response_class=HTMLResponse)
def risk_line_delete(request: Request, client_id: int, risk_id: int, line_id: int, conn=Depends(get_db)):
    conn.execute("DELETE FROM risk_coverage_lines WHERE id=? AND risk_id=?", (line_id, risk_id))
    _update_has_coverage(conn, risk_id)
    conn.commit()
    return _risks_response(request, conn, client_id)


# ── Risk Controls ─────────────────────────────────────────────────────────────

@router.post("/{client_id}/risks/{risk_id}/controls/add", response_class=HTMLResponse)
def risk_control_add(
    request: Request,
    client_id: int,
    risk_id: int,
    control_type: str = Form(...),
    description: str = Form(...),
    status: str = Form("Recommended"),
    responsible: str = Form(""),
    target_date: str = Form(""),
    conn=Depends(get_db),
):
    conn.execute(
        """INSERT INTO risk_controls (risk_id, control_type, description, status, responsible, target_date)
           VALUES (?,?,?,?,?,?)""",
        (risk_id, control_type.strip(), description.strip(),
         status, responsible.strip() or None, target_date.strip() or None),
    )
    conn.commit()
    return _risks_response(request, conn, client_id)


@router.post("/{client_id}/risks/{risk_id}/controls/{ctrl_id}/edit", response_class=HTMLResponse)
def risk_control_edit(
    request: Request,
    client_id: int,
    risk_id: int,
    ctrl_id: int,
    control_type: str = Form(...),
    description: str = Form(...),
    status: str = Form("Recommended"),
    responsible: str = Form(""),
    target_date: str = Form(""),
    conn=Depends(get_db),
):
    conn.execute(
        """UPDATE risk_controls SET control_type=?, description=?, status=?, responsible=?, target_date=?
           WHERE id=? AND risk_id=?""",
        (control_type.strip(), description.strip(), status,
         responsible.strip() or None, target_date.strip() or None,
         ctrl_id, risk_id),
    )
    conn.commit()
    return _risks_response(request, conn, client_id)


@router.post("/{client_id}/risks/{risk_id}/controls/{ctrl_id}/delete", response_class=HTMLResponse)
def risk_control_delete(request: Request, client_id: int, risk_id: int, ctrl_id: int, conn=Depends(get_db)):
    conn.execute("DELETE FROM risk_controls WHERE id=? AND risk_id=?", (ctrl_id, risk_id))
    conn.commit()
    return _risks_response(request, conn, client_id)


@router.get("/{client_id}/edit", response_class=HTMLResponse)
def client_edit_form(request: Request, client_id: int, conn=Depends(get_db)):
    client = get_client_by_id(conn, client_id, include_archived=True)
    if not client:
        return HTMLResponse("Client not found", status_code=404)
    from policydb.queries import REVIEW_CYCLE_LABELS as _REVIEW_CYCLE_LABELS

    scratch_row = conn.execute(
        "SELECT content, updated_at FROM client_scratchpad WHERE client_id=?",
        (client_id,),
    ).fetchone()
    client_scratchpad = scratch_row["content"] if scratch_row else ""
    client_scratchpad_updated = scratch_row["updated_at"] if scratch_row else ""

    open_issues = [dict(r) for r in conn.execute(
        """SELECT a.id, a.issue_uid, a.subject
           FROM activity_log a
           WHERE a.client_id = ? AND a.item_kind = 'issue' AND a.issue_id IS NULL
             AND a.merged_into_id IS NULL
             AND (a.issue_status IS NULL OR a.issue_status NOT IN ('Resolved', 'Closed'))
           ORDER BY a.activity_date DESC""",
        (client_id,),
    ).fetchall()]

    return templates.TemplateResponse("clients/edit.html", {
        "request": request,
        "active": "clients",
        "client": dict(client),
        "industry_segments": cfg.get("industry_segments"),
        "cycle_labels": _REVIEW_CYCLE_LABELS,
        "client_scratchpad": client_scratchpad,
        "client_scratchpad_updated": client_scratchpad_updated,
        "activity_types": cfg.get("activity_types"),
        "open_issues": open_issues,
    })


@router.post("/{client_id}/follow-up")
def client_followup_set(
    client_id: int,
    follow_up_date: str = Form(""),
    conn=Depends(get_db),
):
    """Set or clear the client-level follow-up date."""
    conn.execute(
        "UPDATE clients SET follow_up_date = ? WHERE id = ?",
        (follow_up_date.strip() or None, client_id),
    )
    conn.commit()
    return JSONResponse({"ok": True, "follow_up_date": follow_up_date.strip() or None})


_STRATEGY_FIELDS = {
    "account_priorities", "renewal_strategy", "growth_opportunities",
    "relationship_risk", "service_model", "stewardship_date",
}


@router.patch("/{client_id}/strategy")
async def client_strategy_patch(client_id: int, request: Request, conn=Depends(get_db)):
    """PATCH a single account strategy field on a client."""
    body = await request.json()
    field = body.get("field", "")
    value = body.get("value", "")
    if field not in _STRATEGY_FIELDS:
        return JSONResponse({"ok": False, "error": "Invalid field"}, status_code=400)
    conn.execute(
        f"UPDATE clients SET {field} = ? WHERE id = ?",  # noqa: S608  — field validated against allowlist
        (value.strip() if value else "", client_id),
    )
    conn.commit()
    return JSONResponse({"ok": True, "field": field, "formatted": value.strip() if value else ""})


@router.post("/{client_id}/edit")
def client_edit_post(
    request: Request,
    client_id: int,
    action: str = Form("save"),
    name: str = Form(...),
    industry_segment: str = Form(...),
    cn_number: str = Form(""),
    is_prospect: str = Form(""),
    primary_contact: str = Form(""),
    contact_email: str = Form(""),
    contact_phone: str = Form(""),
    contact_mobile: str = Form(""),
    address: str = Form(""),
    notes: str = Form(""),
    broker_fee: str = Form(""),
    business_description: str = Form(""),
    website: str = Form(""),
    renewal_month: str = Form(""),
    client_since: str = Form(""),
    preferred_contact_method: str = Form(""),
    referral_source: str = Form(""),
    fein: str = Form(""),
    hourly_rate: str = Form(""),
    latitude: str = Form(""),
    longitude: str = Form(""),
    conn=Depends(get_db),
):
    def _float(v):
        try:
            return float(v) if str(v).strip() else None
        except ValueError:
            return None

    def _int(v):
        try:
            return int(v) if str(v).strip() else None
        except ValueError:
            return None

    old_row = dict(conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone())

    name = normalize_client_name(name) if name else name
    conn.execute(
        """UPDATE clients SET name=?, industry_segment=?, cn_number=?, is_prospect=?, primary_contact=?,
           contact_email=?, contact_phone=?, contact_mobile=?, address=?, notes=?,
           broker_fee=?, business_description=?,
           website=?, renewal_month=?, client_since=?, preferred_contact_method=?, referral_source=?,
           fein=?, hourly_rate=?, latitude=?, longitude=?
           WHERE id=?""",
        (name, industry_segment, cn_number.strip() or None, 1 if is_prospect else 0,
         primary_contact or None, clean_email(contact_email) or None,
         format_phone(contact_phone) or None, format_phone(contact_mobile) or None,
         address or None, notes or None,
         _float(broker_fee), business_description or None,
         website or None, _int(renewal_month), client_since or None,
         preferred_contact_method or None, referral_source or None,
         format_fein(fein) or None, _float(hourly_rate),
         _float(latitude), _float(longitude),
         client_id),
    )
    conn.commit()

    if action == "autosave":
        return JSONResponse({"ok": True})
    return RedirectResponse(f"/clients/{client_id}", status_code=303)


@router.post("/{client_id}/archive")
def client_archive(client_id: int, conn=Depends(get_db)):
    """Archive a client (soft delete — hidden from lists, data preserved)."""
    conn.execute("UPDATE clients SET archived=1 WHERE id=?", (client_id,))
    conn.commit()
    logger.info("Client %d archived", client_id)
    return RedirectResponse(f"/clients/{client_id}", status_code=303)


@router.post("/{client_id}/unarchive")
def client_unarchive(client_id: int, conn=Depends(get_db)):
    """Restore an archived client."""
    conn.execute("UPDATE clients SET archived=0 WHERE id=?", (client_id,))
    conn.commit()
    return RedirectResponse(f"/clients/{client_id}", status_code=303)


# ── Billing Accounts ──────────────────────────────────────────────────────────

def _billing_accounts_response(request, conn, client_id: int):
    """Return rendered billing accounts partial."""
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM billing_accounts WHERE client_id=? ORDER BY is_master DESC, billing_id",
        (client_id,),
    ).fetchall()]
    client = conn.execute("SELECT id, name FROM clients WHERE id=?", (client_id,)).fetchone()
    return templates.TemplateResponse("clients/_billing_accounts.html", {
        "request": request,
        "client": dict(client) if client else {"id": client_id},
        "billing_accounts": rows,
    })


@router.patch("/{client_id}/billing/{billing_id_row}/cell")
async def billing_account_cell(request: Request, client_id: int, billing_id_row: int, conn=Depends(get_db)):
    """Save a single cell value for a billing account (matrix edit)."""
    body = await request.json()
    field, value = body.get("field", ""), body.get("value", "")
    allowed = {"billing_id", "description", "fein", "entity_name"}
    if field not in allowed:
        return JSONResponse({"ok": False, "error": "Invalid field"}, status_code=400)
    formatted = value.strip()
    if field == "fein":
        formatted = format_fein(formatted)
    conn.execute(
        f"UPDATE billing_accounts SET {field}=?, modified_at=datetime('now') WHERE id=? AND client_id=?",
        (formatted or None, billing_id_row, client_id),
    )
    conn.commit()
    return JSONResponse({"ok": True, "formatted": formatted})


@router.post("/{client_id}/billing/add-row", response_class=HTMLResponse)
def billing_account_add_row(request: Request, client_id: int, conn=Depends(get_db)):
    """Create a blank billing account row and return matrix row HTML."""
    # Check if there's already a master — new rows default to non-master
    cur = conn.execute(
        "INSERT INTO billing_accounts (client_id, billing_id, is_master) VALUES (?, '', 0)",
        (client_id,),
    )
    conn.commit()
    row = {"id": cur.lastrowid, "billing_id": "", "entity_name": None, "fein": None, "description": None, "is_master": 0}
    return templates.TemplateResponse("clients/_billing_row.html", {
        "request": request, "b": row, "client": {"id": client_id},
    })


@router.post("/{client_id}/billing/{billing_id_row}/toggle-master", response_class=HTMLResponse)
def billing_account_toggle_master(request: Request, client_id: int, billing_id_row: int, conn=Depends(get_db)):
    """Toggle is_master flag on a billing account."""
    existing = conn.execute(
        "SELECT is_master FROM billing_accounts WHERE id=? AND client_id=?",
        (billing_id_row, client_id),
    ).fetchone()
    if not existing:
        return HTMLResponse("", status_code=404)
    # Clear all masters for this client, then set this one if it wasn't already
    conn.execute("UPDATE billing_accounts SET is_master=0, modified_at=datetime('now') WHERE client_id=?", (client_id,))
    if not existing["is_master"]:
        conn.execute("UPDATE billing_accounts SET is_master=1, modified_at=datetime('now') WHERE id=? AND client_id=?", (billing_id_row, client_id))
    conn.commit()
    return _billing_accounts_response(request, conn, client_id)


@router.post("/{client_id}/billing/{billing_id_row}/delete", response_class=HTMLResponse)
def billing_account_delete(request: Request, client_id: int, billing_id_row: int, conn=Depends(get_db)):
    """Delete a billing account."""
    conn.execute("DELETE FROM billing_accounts WHERE id=? AND client_id=?", (billing_id_row, client_id))
    conn.commit()
    return _billing_accounts_response(request, conn, client_id)


def _scratchpad_ctx(request, conn, client_id: int, content: str | None = None) -> dict:
    """Build context dict for the _scratchpad.html partial."""
    if content is None:
        row = conn.execute(
            "SELECT content, updated_at FROM client_scratchpad WHERE client_id=?", (client_id,)
        ).fetchone()
        content = row["content"] if row else ""
        updated = row["updated_at"] if row else ""
    else:
        updated_row = conn.execute(
            "SELECT updated_at FROM client_scratchpad WHERE client_id=?", (client_id,)
        ).fetchone()
        updated = updated_row["updated_at"] if updated_row else ""
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    return {
        "request": request,
        "client": dict(client) if client else {},
        "client_scratchpad": content,
        "client_scratchpad_updated": updated,
        "client_saved_notes": get_saved_notes(conn, "client", str(client_id)),
    }


@router.post("/{client_id}/scratchpad")
def client_scratchpad_save(
    request: Request,
    client_id: int,
    content: str = Form(""),
    conn=Depends(get_db),
):
    """Auto-save per-client working notes. Returns JSON if Accept header requests it."""
    from datetime import datetime, timezone
    conn.execute(
        "INSERT INTO client_scratchpad (client_id, content) VALUES (?, ?) "
        "ON CONFLICT(client_id) DO UPDATE SET content=excluded.content",
        (client_id, content),
    )
    conn.commit()
    if "application/json" in (request.headers.get("accept") or ""):
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
        return JSONResponse({"ok": True, "saved_at": now})
    return templates.TemplateResponse(
        "clients/_scratchpad.html", _scratchpad_ctx(request, conn, client_id, content)
    )


@router.post("/{client_id}/notes/save", response_class=HTMLResponse)
def client_note_save(request: Request, client_id: int, conn=Depends(get_db)):
    """Pin current scratchpad content as a saved note, then clear the scratchpad."""
    from datetime import date as _date
    row = conn.execute(
        "SELECT content FROM client_scratchpad WHERE client_id=?", (client_id,)
    ).fetchone()
    content = (row["content"] if row else "").strip()
    new_activity_id = None
    if content:
        save_note(conn, "client", str(client_id), content)
        # Also log to activity_log for unified account history
        account_exec = cfg.get("default_account_exec", "Grant")
        subject = content[:120] + ("…" if len(content) > 120 else "")
        cursor = conn.execute(
            """INSERT INTO activity_log
               (activity_date, client_id, policy_id, activity_type, subject, details, account_exec)
               VALUES (?, ?, NULL, 'Note', ?, ?, ?)""",
            (_date.today().isoformat(), client_id, subject, content, account_exec),
        )
        new_activity_id = cursor.lastrowid
        conn.execute(
            "UPDATE client_scratchpad SET content = '' WHERE client_id = ?",
            (client_id,),
        )
        conn.commit()

    # Render scratchpad response
    scratchpad_resp = templates.TemplateResponse(
        "clients/_scratchpad.html", _scratchpad_ctx(request, conn, client_id)
    )

    # If we logged an activity, append an OOB swap to insert it at top of activity list
    if new_activity_id:
        a_row = conn.execute(
            """SELECT a.*, c.name AS client_name, c.cn_number, p.policy_uid, p.project_id
               FROM activity_log a
               JOIN clients c ON a.client_id = c.id
               LEFT JOIN policies p ON a.policy_id = p.id
               WHERE a.id = ?""",
            (new_activity_id,),
        ).fetchone()
        if a_row:
            _a_dict = dict(a_row)
            from policydb.web.routes.activities import _attach_pc_emails
            _attach_pc_emails(conn, [_a_dict])
            activity_html = templates.TemplateResponse(
                "activities/_activity_row.html", {"request": request, "a": _a_dict, "dispositions": cfg.get("follow_up_dispositions", [])}
            ).body.decode()
            # OOB swap: prepend to activity list
            oob_html = f'<li hx-swap-oob="afterbegin:#activity-list">{activity_html}</li>'
            scratchpad_html = scratchpad_resp.body.decode()
            return HTMLResponse(scratchpad_html + oob_html)

    return scratchpad_resp


@router.delete("/{client_id}/notes/{note_id}", response_class=HTMLResponse)
def client_note_delete(request: Request, client_id: int, note_id: int, conn=Depends(get_db)):
    """Delete a saved note."""
    delete_saved_note(conn, note_id)
    return templates.TemplateResponse(
        "clients/_scratchpad.html", _scratchpad_ctx(request, conn, client_id)
    )


@router.get("/{client_id}/dedup")
def dedup_page(request: Request, client_id: int, conn=Depends(get_db)):
    """Client-level policy deduplication tool."""
    from policydb.dedup import find_duplicate_candidates
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Client not found", status_code=404)
    candidates = find_duplicate_candidates(conn, client_id)
    likely = [c for c in candidates if c["recommendation"] == "likely_duplicate"]
    possible = [c for c in candidates if c["recommendation"] == "possible_duplicate"]
    review = [c for c in candidates if c["recommendation"] == "different_policies"]
    return templates.TemplateResponse("clients/dedup.html", {
        "request": request,
        "client": client,
        "candidates": candidates,
        "likely": likely,
        "possible": possible,
        "review": review,
    })


@router.post("/{client_id}/dedup/merge")
def dedup_merge(
    request: Request,
    client_id: int,
    keep_uid: str = Form(""),
    archive_uid: str = Form(""),
    cherry_pick: str = Form("{}"),
    conn=Depends(get_db),
):
    """Merge two duplicate policies."""
    import json
    from policydb.dedup import merge_policies
    try:
        cherry_pick_dict = json.loads(cherry_pick)
    except Exception:
        cherry_pick_dict = {}
    result = merge_policies(conn, keep_uid, archive_uid, cherry_pick_dict)
    if result.get("ok"):
        fields = result.get("fields_transferred", [])
        field_count = len(fields)
        s = "s" if field_count != 1 else ""
        return HTMLResponse(f'''
            <div id="pair-{keep_uid}-{archive_uid}" class="border border-green-200 bg-green-50 rounded-lg p-4 mb-3">
                <div class="flex items-center gap-2">
                    <span class="text-green-600 text-lg">&#10003;</span>
                    <span class="text-sm font-medium text-green-800">
                        Merged {archive_uid} into {keep_uid}. {field_count} field{s} transferred.
                    </span>
                </div>
            </div>
        ''')
    return HTMLResponse(f'<div class="text-red-600 text-sm">{result.get("error", "Merge failed")}</div>')


@router.post("/{client_id}/dedup/dismiss")
def dedup_dismiss(
    request: Request,
    client_id: int,
    uid_a: str = Form(""),
    uid_b: str = Form(""),
    conn=Depends(get_db),
):
    """Dismiss a pair as not duplicates."""
    from policydb.dedup import dismiss_pair
    dismiss_pair(conn, client_id, uid_a, uid_b)
    return HTMLResponse(f'''
        <div id="pair-{uid_a}-{uid_b}" class="hidden"></div>
    ''')


@router.get("/{client_id}/export/full")
def export_full(client_id: int, conn=Depends(get_db)):
    """Full internal data export (XLSX) — all fields including internal notes."""
    from fastapi.responses import Response
    from policydb.exporter import export_full_xlsx
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Client not found", status_code=404)
    content = export_full_xlsx(conn, client_id, client["name"])
    safe = client["name"].lower().replace(" ", "_")
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe}_full.xlsx"'},
    )


@router.get("/{client_id}/export/book-review")
def export_book_review(client_id: int, conn=Depends(get_db)):
    """Export multi-tab Client Book Review XLSX for team gap review."""
    from fastapi.responses import Response
    from policydb.exporter import export_book_review_xlsx
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Client not found", status_code=404)
    safe = client["name"].lower().replace(" ", "_")
    content = export_book_review_xlsx(conn, client_id, client["name"])
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe}_book_review.xlsx"'},
    )


@router.get("/{client_id}/export/schedule")
def export_schedule(client_id: int, fmt: str = "md", conn=Depends(get_db)):
    from fastapi.responses import Response
    from policydb.exporter import (
        export_schedule_csv, export_schedule_json, export_schedule_md, export_schedule_xlsx,
    )
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Client not found", status_code=404)
    safe = client["name"].lower().replace(" ", "_")
    if fmt == "xlsx":
        content = export_schedule_xlsx(conn, client_id, client["name"])
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe}_schedule.xlsx"'},
        )
    if fmt == "csv":
        content = export_schedule_csv(conn, client_id)
        media_type = "text/csv"
        ext = "csv"
    elif fmt == "json":
        content = export_schedule_json(conn, client_id, client["name"])
        media_type = "application/json"
        ext = "json"
    else:
        content = export_schedule_md(conn, client_id, client["name"])
        media_type = "text/markdown"
        ext = "md"
    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{safe}_schedule.{ext}"'},
    )


@router.get("/{client_id}/export/project")
def export_project(client_id: int, project: str = "", conn=Depends(get_db)):
    from fastapi.responses import Response
    from policydb.exporter import export_project_group_xlsx
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Client not found", status_code=404)
    content = export_project_group_xlsx(conn, client_id, project, client["name"])
    safe_client = client["name"].lower().replace(" ", "_")
    safe_project = (project or "corporate").lower().replace(" ", "_").replace("/", "_")[:30]
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_client}_{safe_project}_policies.xlsx"'},
    )


@router.get("/{client_id}/export/rfi-by-location")
def export_rfi_by_location(client_id: int, conn=Depends(get_db)):
    from fastapi.responses import Response
    from policydb.exporter import export_rfi_by_location_xlsx
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Client not found", status_code=404)
    content = export_rfi_by_location_xlsx(conn, client_id)
    safe = client["name"].lower().replace(" ", "_")[:30]
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe}_rfi_by_location.xlsx"'},
    )


@router.get("/{client_id}/export/programs")
def export_programs(client_id: int, conn=Depends(get_db)):
    from fastapi.responses import Response
    from policydb.exporter import export_programs_xlsx
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Client not found", status_code=404)
    content = export_programs_xlsx(conn, client_id)
    safe = client["name"].lower().replace(" ", "_")[:30]
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe}_programs.xlsx"'},
    )


@router.get("/{client_id}/policies-json")
def client_policies_json(client_id: int, conn=Depends(get_db)):
    """Return client's policies as JSON for reassignment dropdowns."""
    from fastapi.responses import JSONResponse
    rows = conn.execute(
        "SELECT id, policy_uid, policy_type FROM policies WHERE client_id=? AND archived=0 ORDER BY policy_type",
        (client_id,),
    ).fetchall()
    return JSONResponse([dict(r) for r in rows])


@router.get("/{client_id}/copy-table")
def copy_table(client_id: int, project: str | None = None, conn=Depends(get_db)):
    """Return HTML + plain-text policy table for clipboard copy."""
    from fastapi.responses import JSONResponse
    from policydb.email_templates import build_policy_table
    result = build_policy_table(conn, client_id, project_name=project or None)
    return JSONResponse(result)


@router.get("/{client_id}/copy-table/opportunities")
def copy_table_opportunities(client_id: int, conn=Depends(get_db)):
    """Return HTML + plain-text opportunities table for clipboard copy."""
    from fastapi.responses import JSONResponse
    from policydb.email_templates import build_generic_table
    from policydb.queries import get_policies_for_client
    all_policies = [dict(p) for p in get_policies_for_client(conn, client_id)]
    opps = [p for p in all_policies if p.get("is_opportunity")]
    columns = [
        ("policy_type", "Line of Business", False),
        ("carrier", "Carrier (Target)", False),
        ("opportunity_status", "Status", False),
        ("target_effective_date", "Target Effective", False),
        ("premium", "Est. Premium", True),
        ("commission_amount", "Est. Revenue", True),
        ("placement_colleague", "Placement Team", False),
    ]
    return JSONResponse(build_generic_table(opps, columns))


@router.get("/{client_id}/copy-table/schedule")
def copy_table_schedule(client_id: int, conn=Depends(get_db)):
    """Return HTML + plain-text schedule of insurance for clipboard copy."""
    from fastapi.responses import JSONResponse
    from policydb.email_templates import build_generic_table
    from policydb.exporter import _schedule_rows_for_client
    rows = [dict(r) for r in _schedule_rows_for_client(conn, client_id)]
    # Strip client_name — already known from context
    for r in rows:
        r.pop("client_name", None)
    columns = [
        ("First Named Insured", "First Named Insured", False),
        ("Line of Business", "Line of Business", False),
        ("Carrier", "Carrier", False),
        ("Policy Number", "Policy #", False),
        ("Effective", "Effective", False),
        ("Expiration", "Expiration", False),
        ("Premium", "Premium", True),
        ("Limit", "Limit", True),
        ("Deductible", "Deductible", True),
        ("Form", "Form", False),
        ("Layer", "Layer", False),
        ("Project", "Project", False),
        ("Comments", "Comments", False),
    ]
    return JSONResponse(build_generic_table(rows, columns))


@router.get("/{client_id}/requests/{bundle_id}/copy-table")
def copy_table_request_bundle(client_id: int, bundle_id: int, conn=Depends(get_db)):
    """Return HTML + plain-text request bundle items for clipboard copy."""
    from fastapi.responses import JSONResponse
    from policydb.email_templates import build_generic_table
    items = _enrich_request_items(conn, [dict(r) for r in conn.execute(
        "SELECT * FROM client_request_items WHERE bundle_id=? ORDER BY received ASC, sort_order ASC, id ASC",
        (bundle_id,),
    ).fetchall()])
    # Build a coverage/location column
    for item in items:
        parts = []
        if item.get("policy_type"):
            parts.append(item["policy_type"])
        if item.get("carrier"):
            parts.append(item["carrier"])
        if item.get("project_name"):
            parts.append(item["project_name"])
        item["coverage_location"] = " — ".join(parts) if parts else ""
        item["status_label"] = "Received" if item.get("received") else "Outstanding"
    columns = [
        ("description", "Item", False),
        ("coverage_location", "Coverage / Location", False),
        ("category", "Category", False),
        ("status_label", "Status", False),
        ("notes", "Notes / Response", False),
    ]
    return JSONResponse(build_generic_table(items, columns))


# ─── Quick CSV exports per section ────────────────────────────────────────────

def _safe_filename(client_name: str, section: str) -> str:
    from datetime import date as _d
    safe = client_name.replace(" ", "_").replace("/", "-")[:30]
    return f"{safe}_{section}_{_d.today().isoformat()}.csv"


@router.get("/{client_id}/export/activities.csv")
def export_activities_csv(client_id: int, conn=Depends(get_db)):
    from policydb.utils import csv_response
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Not found", status_code=404)
    rows = [dict(r) for r in get_activities(conn, client_id=client_id)]
    cols = ["activity_date", "activity_type", "subject", "details", "contact_person",
            "duration_hours", "follow_up_date", "follow_up_done", "account_exec"]
    return csv_response(rows, _safe_filename(client["name"], "activities"), cols)


@router.get("/{client_id}/export/contacts.csv")
def export_contacts_csv(client_id: int, conn=Depends(get_db)):
    from policydb.utils import csv_response
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Not found", status_code=404)
    rows = [dict(r) for r in conn.execute(
        """SELECT co.name, cca.title, cca.role, cca.assignment, cca.contact_type,
                  co.email, co.phone, co.mobile, co.organization, cca.notes
           FROM contact_client_assignments cca
           JOIN contacts co ON cca.contact_id = co.id
           WHERE cca.client_id = ?
           ORDER BY cca.contact_type, co.name""",
        (client_id,),
    ).fetchall()]
    cols = ["name", "title", "role", "assignment", "contact_type", "email", "phone", "mobile", "organization", "notes"]
    return csv_response(rows, _safe_filename(client["name"], "contacts"), cols)


@router.get("/{client_id}/export/policies.csv")
def export_policies_csv(client_id: int, conn=Depends(get_db)):
    from policydb.utils import csv_response
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Not found", status_code=404)
    rows = [dict(r) for r in conn.execute(
        """SELECT COALESCE(project_name, 'Corporate / Standalone') AS project_location,
                  policy_type,
                  carrier, policy_number, effective_date, expiration_date,
                  premium, limit_amount, deductible, renewal_status, coverage_form,
                  layer_position, description
           FROM policies WHERE client_id = ? AND archived = 0
             AND (is_opportunity = 0 OR is_opportunity IS NULL)
           ORDER BY project_name, policy_type, layer_position""",
        (client_id,),
    ).fetchall()]
    cols = ["project_location", "policy_type", "carrier", "policy_number",
            "effective_date", "expiration_date", "premium", "limit_amount",
            "deductible", "renewal_status", "coverage_form", "layer_position", "description"]
    return csv_response(rows, _safe_filename(client["name"], "policies"), cols)


@router.get("/{client_id}/export/risks.csv")
def export_risks_csv(client_id: int, conn=Depends(get_db)):
    from policydb.utils import csv_response
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Not found", status_code=404)
    rows = [dict(r) for r in conn.execute(
        """SELECT category, description, severity, has_coverage, policy_uid,
                  notes, source, review_date, identified_date
           FROM client_risks WHERE client_id = ?
           ORDER BY CASE severity WHEN 'Critical' THEN 0 WHEN 'High' THEN 1
                    WHEN 'Medium' THEN 2 ELSE 3 END, category""",
        (client_id,),
    ).fetchall()]
    cols = ["category", "description", "severity", "has_coverage", "policy_uid",
            "notes", "source", "review_date", "identified_date"]
    return csv_response(rows, _safe_filename(client["name"], "risks"), cols)


@router.get("/{client_id}/export/followups.csv")
def export_followups_csv(client_id: int, conn=Depends(get_db)):
    from policydb.utils import csv_response
    from policydb.queries import get_all_followups
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Not found", status_code=404)
    overdue, upcoming = get_all_followups(conn, window=365, client_ids=[client_id])
    all_fu = overdue + upcoming
    cols = ["source", "subject", "follow_up_date", "days_overdue", "activity_type",
            "contact_person", "client_name", "policy_uid", "policy_type"]
    return csv_response(all_fu, _safe_filename(client["name"], "followups"), cols)


def _project_note_ctx(conn, client_id: int, project_name: str) -> dict:
    """Shared context builder for project note partials."""
    row = conn.execute(
        "SELECT id, notes FROM projects WHERE client_id = ? AND LOWER(TRIM(name)) = LOWER(TRIM(?))",
        (client_id, project_name),
    ).fetchone()
    policy_count = conn.execute(
        "SELECT COUNT(*) FROM policies WHERE client_id = ? AND LOWER(TRIM(COALESCE(project_name,''))) = LOWER(TRIM(?)) AND archived = 0",
        (client_id, project_name),
    ).fetchone()[0]
    client = conn.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    # Pull address from the most recent policy in this project
    addr_row = conn.execute(
        """SELECT exposure_address, exposure_city, exposure_state, exposure_zip
           FROM policies
           WHERE client_id = ? AND LOWER(TRIM(COALESCE(project_name,''))) = LOWER(TRIM(?))
             AND archived = 0
           ORDER BY id DESC LIMIT 1""",
        (client_id, project_name),
    ).fetchone()
    return {
        "project_name": project_name,
        "project_id": row["id"] if row else None,
        "note": row["notes"] if row else "",
        "policy_count": policy_count,
        "client": dict(client) if client else {},
        "exposure_address": addr_row["exposure_address"] if addr_row else "",
        "exposure_city": addr_row["exposure_city"] if addr_row else "",
        "exposure_state": addr_row["exposure_state"] if addr_row else "",
        "exposure_zip": addr_row["exposure_zip"] if addr_row else "",
    }


@router.get("/{client_id}/project-note", response_class=HTMLResponse)
def project_note_row(request: Request, client_id: int, project: str, conn=Depends(get_db)):
    """HTMX partial: display project header with note (used by Cancel)."""
    ctx = _project_note_ctx(conn, client_id, project)
    return templates.TemplateResponse("clients/_project_header.html", {"request": request, **ctx})


@router.get("/{client_id}/project-note/edit", response_class=HTMLResponse)
def project_note_edit(request: Request, client_id: int, project: str, conn=Depends(get_db)):
    """HTMX partial: edit form for a project note."""
    from policydb.web.routes.policies import US_STATES
    ctx = _project_note_ctx(conn, client_id, project)
    ctx["us_states"] = US_STATES
    return templates.TemplateResponse("clients/_project_header_edit.html", {"request": request, **ctx})


@router.post("/{client_id}/project-note", response_class=HTMLResponse)
def project_note_save(
    request: Request,
    client_id: int,
    project_name: str = Form(...),
    notes: str = Form(""),
    exposure_address: str = Form(""),
    exposure_city: str = Form(""),
    exposure_state: str = Form(""),
    exposure_zip: str = Form(""),
    conn=Depends(get_db),
):
    """HTMX: upsert project note and bulk-update location address on all policies in the project."""
    exposure_address = exposure_address.strip() if exposure_address else ""
    exposure_city = format_city(exposure_city) if exposure_city else ""
    exposure_state = format_state(exposure_state) if exposure_state else ""
    exposure_zip = format_zip(exposure_zip) if exposure_zip else ""
    conn.execute(
        "INSERT INTO projects (client_id, name, notes) VALUES (?, ?, ?) "
        "ON CONFLICT(client_id, name) DO UPDATE SET notes=excluded.notes",
        (client_id, project_name, notes.strip()),
    )
    conn.execute(
        """UPDATE policies SET
               exposure_address = ?,
               exposure_city    = ?,
               exposure_state   = ?,
               exposure_zip     = ?
           WHERE client_id = ?
             AND LOWER(TRIM(COALESCE(project_name,''))) = LOWER(TRIM(?))
             AND archived = 0""",
        (
            exposure_address or None,
            exposure_city or None,
            exposure_state or None,
            exposure_zip or None,
            client_id, project_name,
        ),
    )
    conn.commit()
    ctx = _project_note_ctx(conn, client_id, project_name)
    return templates.TemplateResponse("clients/_project_header.html", {"request": request, **ctx})


@router.post("/{client_id}/projects/{project_id}/notes")
def project_notes_autosave(
    client_id: int,
    project_id: int,
    content: str = Form(""),
    conn=Depends(get_db),
):
    from fastapi import HTTPException
    cur = conn.execute(
        "UPDATE projects SET notes = ? WHERE id = ? AND client_id = ?",
        (content, project_id, client_id),
    )
    conn.commit()
    if cur.rowcount == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    saved_at = babel_format_datetime(
        datetime.now(), "MMM d 'at' h:mma", locale="en_US"
    ).replace("AM", "am").replace("PM", "pm")
    return JSONResponse({"ok": True, "saved_at": saved_at})


@router.post("/{client_id}/projects/{project_id}/scratchpad")
def project_scratchpad_autosave(
    client_id: int,
    project_id: int,
    content: str = Form(""),
    conn=Depends(get_db),
):
    """Auto-save the project's working-notes scratchpad (upsert)."""
    from fastapi import HTTPException
    row = conn.execute(
        "SELECT 1 FROM projects WHERE id = ? AND client_id = ?",
        (project_id, client_id),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Project not found")
    conn.execute(
        """INSERT INTO project_scratchpad (project_id, content, updated_at)
           VALUES (?, ?, CURRENT_TIMESTAMP)
           ON CONFLICT(project_id) DO UPDATE SET
               content = excluded.content,
               updated_at = CURRENT_TIMESTAMP""",
        (project_id, content),
    )
    conn.commit()
    saved_at = babel_format_datetime(
        datetime.now(), "MMM d 'at' h:mma", locale="en_US"
    ).replace("AM", "am").replace("PM", "pm")
    return JSONResponse({"ok": True, "saved_at": saved_at})


@router.get("/{client_id}/projects/{project_id}", response_class=HTMLResponse)
def project_detail(
    client_id: int,
    project_id: int,
    request: Request,
    conn=Depends(get_db),
):
    from fastapi import HTTPException
    import dateparser
    project = conn.execute(
        "SELECT * FROM projects WHERE id = ? AND client_id = ?",
        (project_id, client_id),
    ).fetchone()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    client = conn.execute(
        "SELECT id, name FROM clients WHERE id = ?", (client_id,)
    ).fetchone()
    policies = conn.execute(
        """SELECT p.id, p.policy_uid, p.policy_type, p.carrier, p.renewal_status,
                  p.premium, p.expiration_date, p.effective_date, p.policy_number,
                  p.limit_amount, p.is_opportunity, p.client_id,
                  p.exposure_address, p.exposure_city, p.exposure_state, p.exposure_zip,
                  CAST(julianday(p.expiration_date) - julianday('now') AS INTEGER) AS days_to_renewal
           FROM policies p
           WHERE p.project_id = ? AND p.archived = 0
           ORDER BY p.policy_type""",
        (project_id,),
    ).fetchall()
    project = dict(project)
    # Pull address from first policy that has one (fallback when project has no own address)
    if not project.get("address"):
        for pol in policies:
            if pol["exposure_address"] or pol["exposure_city"]:
                project["exposure_address"] = pol["exposure_address"] or ""
                project["exposure_city"] = pol["exposure_city"] or ""
                project["exposure_state"] = pol["exposure_state"] or ""
                project["exposure_zip"] = pol["exposure_zip"] or ""
                break
    if project.get("updated_at"):
        try:
            dt = dateparser.parse(project["updated_at"])
            if dt:
                project["updated_at_fmt"] = babel_format_datetime(
                    dt, "MMM d 'at' h:mma", locale="en_US"
                ).replace("AM", "am").replace("PM", "pm")
        except Exception:
            project["updated_at_fmt"] = project["updated_at"][:16]

    # COPE data for this project/location
    cope_row = conn.execute(
        "SELECT * FROM cope_data WHERE project_id = ?", (project_id,)
    ).fetchone()
    cope = dict(cope_row) if cope_row else None

    # Programs linked to this location
    from policydb.queries import get_programs_for_project
    location_programs = get_programs_for_project(conn, project_id)

    # Project scratchpad (working notes with activity conversion)
    scratch_row = conn.execute(
        "SELECT content, updated_at FROM project_scratchpad WHERE project_id = ?",
        (project_id,),
    ).fetchone()
    project_scratchpad = scratch_row["content"] if scratch_row else ""
    project_scratchpad_updated = scratch_row["updated_at"] if scratch_row else ""

    # Open issues assigned to this location (direct project_id + linked policies'
    # issues), deduped by activity_log id, newest first.
    policy_ids = [p["id"] for p in policies]
    _placeholders = ",".join("?" * len(policy_ids)) if policy_ids else ""
    _where_parts = ["(a.project_id = ?)"]
    _params: list = [project_id]
    if policy_ids:
        _where_parts.append(f"(a.policy_id IN ({_placeholders}))")
        _params.extend(policy_ids)
    _where = " OR ".join(_where_parts)
    open_project_issues = [dict(r) for r in conn.execute(
        f"""SELECT a.id, a.issue_uid, a.subject, a.issue_status, a.issue_severity,
                   a.activity_date, a.policy_id, a.is_renewal_issue,
                   p.policy_uid, p.policy_type,
                   CAST(julianday(date('now')) - julianday(a.activity_date) AS INTEGER) AS days_open
            FROM activity_log a
            LEFT JOIN policies p ON p.id = a.policy_id
            WHERE a.item_kind = 'issue'
              AND a.client_id = ?
              AND ({_where})
              AND a.merged_into_id IS NULL
              AND (a.issue_status IS NULL OR a.issue_status NOT IN ('Resolved', 'Closed'))
            ORDER BY CASE a.issue_severity
                       WHEN 'Critical' THEN 1 WHEN 'High' THEN 2
                       WHEN 'Normal' THEN 3 ELSE 4 END,
                     a.activity_date DESC""",  # noqa: S608 — placeholders built from static counts
        [client_id, *_params],
    ).fetchall()]

    # Attachment count for the Files card header
    attachment_count = conn.execute(
        """SELECT COUNT(DISTINCT ra.attachment_id) FROM record_attachments ra
           WHERE ra.record_type = 'project' AND ra.record_id = ?""",
        (project_id,),
    ).fetchone()[0]

    return templates.TemplateResponse(
        "clients/project.html",
        {
            "request": request,
            "project": project,
            "client": dict(client),
            "policies": [dict(p) for p in policies],
            "cope": cope,
            "location_programs": location_programs,
            "renewal_statuses": cfg.get("renewal_statuses", []),
            "pinned_notes": _pinned_notes_for_page(conn, "project", project_id, client_id=client_id),
            "pinned_scope": "project",
            "pinned_scope_id": str(project_id),
            "pinned_client_id": str(client_id),
            "project_scratchpad": project_scratchpad,
            "project_scratchpad_updated": project_scratchpad_updated,
            "open_project_issues": open_project_issues,
            "attachment_count": attachment_count,
            "activity_types": cfg.get("activity_types", []),
        },
    )


@router.get("/{client_id}/projects/{project_id}/print", response_class=HTMLResponse)
def project_print(
    client_id: int,
    project_id: int,
    request: Request,
    conn=Depends(get_db),
):
    """Print-optimized view of project notes — use browser Print > Save as PDF."""
    from fastapi import HTTPException
    project = conn.execute(
        "SELECT * FROM projects WHERE id = ? AND client_id = ?",
        (project_id, client_id),
    ).fetchone()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    client = conn.execute(
        "SELECT id, name, cn_number FROM clients WHERE id = ?", (client_id,)
    ).fetchone()
    policies = conn.execute(
        """SELECT policy_uid, policy_type, carrier, premium, limit_amount,
                  effective_date, expiration_date, renewal_status,
                  exposure_address, exposure_city, exposure_state, exposure_zip
           FROM policies WHERE project_id = ? AND archived = 0
           ORDER BY policy_type""",
        (project_id,),
    ).fetchall()
    project = dict(project)
    for pol in policies:
        if pol["exposure_address"] or pol["exposure_city"]:
            project["exposure_address"] = pol["exposure_address"] or ""
            project["exposure_city"] = pol["exposure_city"] or ""
            project["exposure_state"] = pol["exposure_state"] or ""
            project["exposure_zip"] = pol["exposure_zip"] or ""
            break
    return templates.TemplateResponse(
        "clients/project_print.html",
        {
            "request": request,
            "project": project,
            "client": dict(client),
            "policies": [dict(p) for p in policies],
        },
    )


@router.get("/{client_id}/projects/{project_id}/pdf")
def project_pdf(
    client_id: int,
    project_id: int,
    conn=Depends(get_db),
):
    """Generate a PDF of project notes + policies via fpdf2."""
    from fastapi import HTTPException
    from fastapi.responses import Response
    from fpdf import FPDF

    project = conn.execute(
        "SELECT * FROM projects WHERE id = ? AND client_id = ?",
        (project_id, client_id),
    ).fetchone()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    client = conn.execute(
        "SELECT id, name, cn_number FROM clients WHERE id = ?", (client_id,)
    ).fetchone()
    policies = conn.execute(
        """SELECT policy_type, carrier, premium, limit_amount,
                  effective_date, expiration_date, renewal_status
           FROM policies WHERE project_id = ? AND archived = 0
           ORDER BY policy_type""",
        (project_id,),
    ).fetchall()

    project = dict(project)
    updated = project["updated_at"][:10] if project.get("updated_at") else ""
    cn = f" | {client['cn_number']}" if client["cn_number"] else ""
    pol_count = len(policies)

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()

    # Header
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 8, project["name"], new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(107, 114, 128)
    pdf.cell(0, 6, f"{client['name']}{cn}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(156, 163, 175)
    pdf.cell(0, 5, f"Updated {updated}  |  {pol_count} polic{'y' if pol_count == 1 else 'ies'}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_draw_color(229, 231, 235)
    pdf.line(10, pdf.get_y() + 3, 200, pdf.get_y() + 3)
    pdf.ln(8)

    # Notes section
    if project.get("notes"):
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(107, 114, 128)
        pdf.cell(0, 6, "PROJECT NOTES", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)
        pdf.set_font("Helvetica", "", 11)
        pdf.set_text_color(31, 41, 55)

        # Parse markdown line by line for clean rendering
        for line in project["notes"].split("\n"):
            stripped = line.strip()
            if not stripped:
                pdf.ln(3)
                continue
            if stripped.startswith("### "):
                pdf.ln(2)
                pdf.set_font("Helvetica", "B", 11)
                pdf.multi_cell(0, 5.5, stripped[4:])
                pdf.set_font("Helvetica", "", 11)
            elif stripped.startswith("## "):
                pdf.ln(3)
                pdf.set_font("Helvetica", "B", 12)
                pdf.multi_cell(0, 6, stripped[3:])
                pdf.set_font("Helvetica", "", 11)
            elif stripped.startswith("# "):
                pdf.ln(3)
                pdf.set_font("Helvetica", "B", 14)
                pdf.multi_cell(0, 7, stripped[2:])
                pdf.set_font("Helvetica", "", 11)
            elif stripped.startswith("- ") or stripped.startswith("* "):
                pdf.cell(6, 5.5, chr(8226))
                pdf.multi_cell(0, 5.5, stripped[2:].replace("**", ""))
            elif stripped.startswith("> "):
                pdf.set_text_color(107, 114, 128)
                pdf.cell(4, 5.5, "|")
                pdf.multi_cell(0, 5.5, stripped[2:].replace("**", ""))
                pdf.set_text_color(31, 41, 55)
            else:
                # Strip bold markers for clean text
                clean = stripped.replace("**", "")
                pdf.multi_cell(0, 5.5, clean)
        pdf.ln(4)

    # Policies table
    if policies:
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(107, 114, 128)
        pdf.cell(0, 6, "POLICIES", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        # Table header
        col_w = [42, 32, 26, 26, 22, 22, 20]
        headers = ["Line of Business", "Carrier", "Premium", "Limit", "Eff.", "Exp.", "Status"]
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(107, 114, 128)
        for i, h in enumerate(headers):
            align = "R" if i in (2, 3) else "L"
            pdf.cell(col_w[i], 5, h, align=align)
        pdf.ln()
        pdf.set_draw_color(209, 213, 219)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(1)

        # Table rows
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(55, 65, 81)
        total_premium = 0
        for p in policies:
            prem = p["premium"] or 0
            total_premium += prem
            prem_fmt = f"${prem:,.0f}" if prem else "—"
            lim_fmt = f"${p['limit_amount']:,.0f}" if p["limit_amount"] else "—"
            pdf.cell(col_w[0], 5, (p["policy_type"] or "")[:24])
            pdf.cell(col_w[1], 5, (p["carrier"] or "—")[:18])
            pdf.cell(col_w[2], 5, prem_fmt, align="R")
            pdf.cell(col_w[3], 5, lim_fmt, align="R")
            pdf.cell(col_w[4], 5, (p["effective_date"] or "—")[:10])
            pdf.cell(col_w[5], 5, (p["expiration_date"] or "—")[:10])
            pdf.cell(col_w[6], 5, (p["renewal_status"] or "—")[:10])
            pdf.ln()

        # Total row
        pdf.set_draw_color(209, 213, 219)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(1)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(col_w[0], 5, "Total")
        pdf.cell(col_w[1], 5, "")
        pdf.cell(col_w[2], 5, f"${total_premium:,.0f}", align="R")
        pdf.ln()

    # Footer
    pdf.ln(10)
    pdf.set_draw_color(229, 231, 235)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(3)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(176, 183, 192)
    pdf.cell(0, 4, f"{client['name']}  |  {project['name']}  |  Generated from PolicyDB")

    pdf_bytes = bytes(pdf.output())
    safe_name = project["name"].replace(" ", "_").replace("/", "-")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{client["name"]}_{safe_name}_Notes.pdf"'},
    )


# ── Project management: rename / merge / delete ──────────────────────────────


def _cleanup_orphan_projects(conn, client_id: int) -> None:
    """Delete project rows that have no policies linked by project_id."""
    conn.execute(
        """DELETE FROM projects
           WHERE client_id = ?
             AND id NOT IN (
               SELECT DISTINCT project_id FROM policies
               WHERE project_id IS NOT NULL AND client_id = ?
             )""",
        (client_id, client_id),
    )


@router.get("/{client_id}/project/rename-form", response_class=HTMLResponse)
def project_rename_form(request: Request, client_id: int, project: str, conn=Depends(get_db)):
    ctx = _project_note_ctx(conn, client_id, project)
    return templates.TemplateResponse("clients/_project_header_rename.html", {"request": request, **ctx})


@router.post("/{client_id}/project/rename")
def project_rename(
    request: Request,
    client_id: int,
    project_name: str = Form(...),
    new_name: str = Form(...),
    conn=Depends(get_db),
):
    new_name = new_name.strip()
    if not new_name:
        ctx = _project_note_ctx(conn, client_id, project_name)
        ctx["error"] = "Name cannot be empty."
        return templates.TemplateResponse("clients/_project_header_rename.html", {"request": request, **ctx})

    # Check for collision with a DIFFERENT project
    old_norm = project_name.strip().lower()
    new_norm = new_name.lower()
    if old_norm != new_norm:
        collision = conn.execute(
            "SELECT id FROM projects WHERE client_id = ? AND LOWER(TRIM(name)) = ?",
            (client_id, new_norm),
        ).fetchone()
        if collision:
            ctx = _project_note_ctx(conn, client_id, project_name)
            ctx["error"] = f"A project named '{new_name}' already exists. Use Merge instead."
            return templates.TemplateResponse("clients/_project_header_rename.html", {"request": request, **ctx})

    # Rename the project row
    conn.execute(
        "UPDATE projects SET name = ? WHERE client_id = ? AND LOWER(TRIM(name)) = LOWER(TRIM(?))",
        (new_name, client_id, project_name),
    )
    # Update project_name on all policies (active + archived)
    conn.execute(
        "UPDATE policies SET project_name = ? WHERE client_id = ? AND LOWER(TRIM(COALESCE(project_name,''))) = LOWER(TRIM(?))",
        (new_name, client_id, project_name),
    )
    conn.commit()
    return HTMLResponse("", headers={"HX-Redirect": f"/clients/{client_id}"})


@router.get("/{client_id}/project/merge-form", response_class=HTMLResponse)
def project_merge_form(request: Request, client_id: int, project: str, conn=Depends(get_db)):
    ctx = _project_note_ctx(conn, client_id, project)
    other = conn.execute(
        "SELECT name FROM projects WHERE client_id = ? AND LOWER(TRIM(name)) != LOWER(TRIM(?)) ORDER BY name",
        (client_id, project),
    ).fetchall()
    ctx["other_projects"] = [r["name"] for r in other]
    return templates.TemplateResponse("clients/_project_header_merge.html", {"request": request, **ctx})


@router.post("/{client_id}/project/merge")
def project_merge(
    request: Request,
    client_id: int,
    source_project: str = Form(...),
    target_project: str = Form(...),
    conn=Depends(get_db),
):
    source = conn.execute(
        "SELECT id, notes FROM projects WHERE client_id = ? AND LOWER(TRIM(name)) = LOWER(TRIM(?))",
        (client_id, source_project),
    ).fetchone()
    target = conn.execute(
        "SELECT id, notes FROM projects WHERE client_id = ? AND LOWER(TRIM(name)) = LOWER(TRIM(?))",
        (client_id, target_project),
    ).fetchone()
    if not source or not target:
        return HTMLResponse("Project not found", status_code=404)

    # Concatenate notes
    src_notes = (source["notes"] or "").strip()
    tgt_notes = (target["notes"] or "").strip()
    if src_notes:
        merged = f"{tgt_notes}\n\n[Merged from {source_project}]: {src_notes}".strip()
        conn.execute("UPDATE projects SET notes = ? WHERE id = ?", (merged, target["id"]))

    # Move policies by project_id
    conn.execute(
        "UPDATE policies SET project_name = ?, project_id = ? WHERE client_id = ? AND project_id = ?",
        (target_project, target["id"], client_id, source["id"]),
    )
    # Catch strays (archived or unlinked)
    conn.execute(
        """UPDATE policies SET project_name = ?, project_id = ?
           WHERE client_id = ? AND LOWER(TRIM(COALESCE(project_name,''))) = LOWER(TRIM(?))
             AND (project_id IS NULL OR project_id = ?)""",
        (target_project, target["id"], client_id, source_project, source["id"]),
    )
    _cleanup_orphan_projects(conn, client_id)
    conn.commit()
    return HTMLResponse("", headers={"HX-Redirect": f"/clients/{client_id}"})


@router.post("/{client_id}/project/delete")
def project_delete(
    request: Request,
    client_id: int,
    project_name: str = Form(...),
    conn=Depends(get_db),
):
    # Unassign all policies (moves to Corporate / Standalone)
    conn.execute(
        "UPDATE policies SET project_name = NULL, project_id = NULL WHERE client_id = ? AND LOWER(TRIM(COALESCE(project_name,''))) = LOWER(TRIM(?))",
        (client_id, project_name),
    )
    conn.execute(
        "DELETE FROM projects WHERE client_id = ? AND LOWER(TRIM(name)) = LOWER(TRIM(?))",
        (client_id, project_name),
    )
    conn.commit()
    return HTMLResponse("", headers={"HX-Redirect": f"/clients/{client_id}"})


def _project_contacts(conn, client_id: int, project: str) -> list[dict]:
    """Get all contacts relevant to a project: policy-assigned + client-level."""
    rows = conn.execute(
        """SELECT DISTINCT co.id, co.name FROM (
             SELECT co2.id, co2.name FROM contact_policy_assignments cpa
               JOIN contacts co2 ON cpa.contact_id = co2.id
               JOIN policies p ON cpa.policy_id = p.id
               WHERE p.client_id = ? AND LOWER(TRIM(COALESCE(p.project_name,''))) = LOWER(TRIM(?))
                 AND co2.name IS NOT NULL AND TRIM(co2.name) != '' AND p.archived = 0
             UNION
             SELECT co3.id, co3.name FROM contact_client_assignments cca
               JOIN contacts co3 ON cca.contact_id = co3.id
               WHERE cca.client_id = ? AND co3.name IS NOT NULL AND TRIM(co3.name) != ''
           ) co ORDER BY co.name""",
        (client_id, project, client_id),
    ).fetchall()
    return [dict(r) for r in rows]


@router.get("/{client_id}/project/log-form", response_class=HTMLResponse)
def project_log_form(request: Request, client_id: int, project: str, conn=Depends(get_db)):
    """HTMX partial: inline activity log form for a project."""
    ctx = _project_note_ctx(conn, client_id, project)
    ctx["activity_types"] = cfg.get("activity_types")
    ctx["contacts"] = _project_contacts(conn, client_id, project)
    # Policies list for "Specific policy" scope dropdown
    if ctx.get("project_id"):
        policies = conn.execute(
            """SELECT id, policy_uid, policy_type, carrier FROM policies
               WHERE project_id = ? AND archived = 0 ORDER BY policy_type""",
            (ctx["project_id"],),
        ).fetchall()
        ctx["policies"] = [dict(p) for p in policies]
    else:
        ctx["policies"] = []
    return templates.TemplateResponse("clients/_project_log_form.html", {"request": request, **ctx})


@router.post("/{client_id}/project/log", response_class=HTMLResponse)
def project_log_save(
    request: Request,
    client_id: int,
    project_name: str = Form(...),
    activity_type: str = Form(...),
    contact_person: str = Form(""),
    subject: str = Form(...),
    details: str = Form(""),
    follow_up_date: str = Form(""),
    scope: str = Form("project"),
    specific_policy_id: int = Form(0),
    duration_hours: str = Form(""),
    conn=Depends(get_db),
):
    """Create a single activity log entry at project or specific-policy level."""
    from datetime import date
    account_exec = cfg.get("default_account_exec", "")
    from policydb.utils import round_duration
    dur = round_duration(duration_hours)
    today = date.today().isoformat()

    # Resolve contact_person → contact_id
    _contact_id = None
    if contact_person and contact_person.strip():
        _row = conn.execute(
            "SELECT id FROM contacts WHERE LOWER(TRIM(name))=LOWER(TRIM(?))",
            (contact_person.strip(),),
        ).fetchone()
        if _row:
            _contact_id = _row["id"]

    # Resolve project_id from projects table
    proj_row = conn.execute(
        "SELECT id FROM projects WHERE client_id = ? AND LOWER(TRIM(name)) = LOWER(TRIM(?))",
        (client_id, project_name),
    ).fetchone()
    _project_id = proj_row["id"] if proj_row else None

    _fu_date = follow_up_date if follow_up_date else None

    if scope == "policy" and specific_policy_id:
        # Specific policy scope: one row with policy_id set
        conn.execute(
            """INSERT INTO activity_log
               (activity_date, client_id, policy_id, project_id, activity_type, contact_person,
                contact_id, subject, details, follow_up_date, duration_hours, account_exec)
               VALUES (?, ?, ?, NULL, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (today, client_id, specific_policy_id, activity_type, contact_person or None,
             _contact_id, subject, details or None, _fu_date, dur, account_exec),
        )
        if _fu_date:
            from policydb.queries import supersede_followups
            supersede_followups(conn, specific_policy_id, _fu_date)
        # Look up policy_uid for success message
        pol = conn.execute("SELECT policy_uid FROM policies WHERE id=?", (specific_policy_id,)).fetchone()
        log_msg = f"Logged activity to {pol['policy_uid']}" if pol else "Logged activity to policy"
    else:
        # Project-level scope: one row with project_id set, policy_id NULL
        conn.execute(
            """INSERT INTO activity_log
               (activity_date, client_id, policy_id, project_id, activity_type, contact_person,
                contact_id, subject, details, follow_up_date, duration_hours, account_exec)
               VALUES (?, ?, NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (today, client_id, _project_id, activity_type, contact_person or None,
             _contact_id, subject, details or None, _fu_date, dur, account_exec),
        )
        # No supersede_followups for project-level activities (no policy_id to chain)
        log_msg = f"Logged project activity to {project_name}"

    conn.commit()
    # Return the log form again with a success banner
    ctx = _project_note_ctx(conn, client_id, project_name)
    ctx["activity_types"] = cfg.get("activity_types")
    ctx["contacts"] = _project_contacts(conn, client_id, project_name)
    if _project_id:
        policies = conn.execute(
            """SELECT id, policy_uid, policy_type, carrier FROM policies
               WHERE project_id = ? AND archived = 0 ORDER BY policy_type""",
            (_project_id,),
        ).fetchall()
        ctx["policies"] = [dict(p) for p in policies]
    else:
        ctx["policies"] = []
    ctx["log_success"] = log_msg
    return templates.TemplateResponse("clients/_project_log_form.html", {"request": request, **ctx})


@router.get("/{client_id}/export/llm")
def export_llm(client_id: int, conn=Depends(get_db)):
    from fastapi.responses import Response
    from policydb.exporter import export_llm_client_md
    client = get_client_by_id(conn, client_id)
    if not client:
        return HTMLResponse("Client not found", status_code=404)
    content = export_llm_client_md(conn, client_id)
    safe = client["name"].lower().replace(" ", "_")
    return Response(
        content=content,
        media_type="text/markdown",
        headers={"Content-Disposition": f'attachment; filename="{safe}_llm.md"'},
    )


# ─── LINKED ACCOUNTS ─────────────────────────────────────────────────────────


def _linked_accounts_ctx(request, conn, client_id: int) -> dict:
    """Build context dict for the _linked_accounts.html partial."""
    linked_group = get_linked_group_for_client(conn, client_id)
    client = get_client_by_id(conn, client_id, include_archived=True)
    return {
        "request": request,
        "client": dict(client) if client else {},
        "linked_group": linked_group,
        "linked_relationships": cfg.get("linked_account_relationships", []),
    }


@router.get("/{client_id}/linked", response_class=HTMLResponse)
def linked_accounts_partial(request: Request, client_id: int, conn=Depends(get_db)):
    return templates.TemplateResponse(
        "clients/_linked_accounts.html", _linked_accounts_ctx(request, conn, client_id)
    )


@router.get("/{client_id}/linked/overview", response_class=HTMLResponse)
def linked_overview_partial(request: Request, client_id: int, conn=Depends(get_db)):
    linked_group = get_linked_group_for_client(conn, client_id)
    if not linked_group:
        return HTMLResponse("")
    overview = get_linked_group_overview(conn, linked_group["group"]["id"])
    client = get_client_by_id(conn, client_id, include_archived=True)
    return templates.TemplateResponse("clients/_linked_overview.html", {
        "request": request,
        "client": dict(client) if client else {},
        "linked_group": linked_group,
        "overview": overview,
    })


@router.get("/{client_id}/linked/search", response_class=HTMLResponse)
def linked_search(request: Request, client_id: int, q: str = "", conn=Depends(get_db)):
    """Search for active clients to link. Allows merging across existing groups."""
    if not q or len(q) < 2:
        return HTMLResponse("")
    # Find which clients are already in THIS client's group (exclude them)
    my_group = get_linked_group_for_client(conn, client_id)
    my_group_ids = set()
    if my_group:
        my_group_ids = {m["client_id"] for m in my_group["members"]}
    all_clients = conn.execute(
        "SELECT id, name, cn_number, industry_segment FROM clients WHERE archived = 0 ORDER BY name"
    ).fetchall()
    from rapidfuzz import fuzz
    results = []
    for c in all_clients:
        if c["id"] == client_id or c["id"] in my_group_ids:
            continue
        score = fuzz.WRatio(q.lower(), c["name"].lower())
        if score >= 50 or q.lower() in c["name"].lower():
            results.append(dict(c) | {"score": score})
    results.sort(key=lambda x: x["score"], reverse=True)
    return templates.TemplateResponse("clients/_linked_search_results.html", {
        "request": request,
        "results": results[:10],
        "client_id": client_id,
        "linked_group": my_group,
    })


@router.post("/{client_id}/linked/create", response_class=HTMLResponse)
def linked_create(
    request: Request,
    client_id: int,
    target_client_id: int = Form(...),
    label: str = Form(""),
    relationship: str = Form("Related"),
    conn=Depends(get_db),
):
    my_group = get_linked_group_for_client(conn, client_id)
    target_group = get_linked_group_for_client(conn, target_client_id)
    if my_group and target_group and my_group["group"]["id"] != target_group["group"]["id"]:
        # Both in different groups — merge target's group into ours
        for m in target_group["members"]:
            add_client_to_group(conn, my_group["group"]["id"], m["client_id"])
        delete_linked_group(conn, target_group["group"]["id"])
    elif my_group:
        add_client_to_group(conn, my_group["group"]["id"], target_client_id)
    elif target_group:
        add_client_to_group(conn, target_group["group"]["id"], client_id)
    else:
        create_linked_group(conn, label, relationship, [client_id, target_client_id])
    return templates.TemplateResponse(
        "clients/_linked_accounts.html", _linked_accounts_ctx(request, conn, client_id)
    )


@router.post("/{client_id}/linked/add", response_class=HTMLResponse)
def linked_add(
    request: Request,
    client_id: int,
    target_client_id: int = Form(...),
    conn=Depends(get_db),
):
    linked_group = get_linked_group_for_client(conn, client_id)
    target_group = get_linked_group_for_client(conn, target_client_id)
    if linked_group and target_group and linked_group["group"]["id"] != target_group["group"]["id"]:
        # Merge: move all members from target's group into this client's group
        for m in target_group["members"]:
            add_client_to_group(conn, linked_group["group"]["id"], m["client_id"])
        delete_linked_group(conn, target_group["group"]["id"])
    elif linked_group:
        add_client_to_group(conn, linked_group["group"]["id"], target_client_id)
    return templates.TemplateResponse(
        "clients/_linked_accounts.html", _linked_accounts_ctx(request, conn, client_id)
    )


@router.post("/{client_id}/linked/{member_id}/remove", response_class=HTMLResponse)
def linked_remove(
    request: Request, client_id: int, member_id: int, conn=Depends(get_db)
):
    remove_client_from_group(conn, member_id)
    return templates.TemplateResponse(
        "clients/_linked_accounts.html", _linked_accounts_ctx(request, conn, client_id)
    )


@router.post("/{client_id}/linked/edit", response_class=HTMLResponse)
def linked_edit(
    request: Request,
    client_id: int,
    label: str = Form(""),
    relationship: str = Form("Related"),
    conn=Depends(get_db),
):
    linked_group = get_linked_group_for_client(conn, client_id)
    if linked_group:
        update_linked_group(conn, linked_group["group"]["id"], label, relationship)
    return templates.TemplateResponse(
        "clients/_linked_accounts.html", _linked_accounts_ctx(request, conn, client_id)
    )


@router.post("/{client_id}/linked/dissolve", response_class=HTMLResponse)
def linked_dissolve(
    request: Request, client_id: int, conn=Depends(get_db)
):
    linked_group = get_linked_group_for_client(conn, client_id)
    if linked_group:
        delete_linked_group(conn, linked_group["group"]["id"])
    return templates.TemplateResponse(
        "clients/_linked_accounts.html", _linked_accounts_ctx(request, conn, client_id)
    )


# ── Account Summary ──────────────────────────────────────────────────────────


@router.get("/{client_id}/summary", response_class=HTMLResponse)
def client_summary_panel(
    request: Request,
    client_id: int,
    include_linked: int = 0,
    days: int = 90,
    conn=Depends(get_db),
):
    """HTMX partial: account summary card."""
    from policydb.exporter import build_account_summary
    summary = build_account_summary(conn, client_id, days=days, include_linked=bool(include_linked))
    linked_group = None
    from policydb.queries import get_linked_group_for_client
    linked_group = get_linked_group_for_client(conn, client_id)
    return templates.TemplateResponse("clients/_account_summary.html", {
        "request": request,
        "s": summary,
        "client_id": client_id,
        "include_linked": bool(include_linked),
        "has_linked_group": linked_group is not None,
        "days": days,
    })


@router.get("/{client_id}/summary/text")
def client_summary_text(
    client_id: int,
    include_linked: int = 0,
    days: int = 90,
    conn=Depends(get_db),
):
    """Return plain text account summary for clipboard."""
    from policydb.exporter import build_account_summary, render_account_summary_text
    summary = build_account_summary(conn, client_id, days=days, include_linked=bool(include_linked))
    text = render_account_summary_text(summary)
    from fastapi.responses import PlainTextResponse
    return PlainTextResponse(text)


# ── Request Tracker ─────


def _get_request_bundles(conn, client_id: int) -> list[dict]:
    """Fetch all request bundles for a client with item counts."""
    bundles = [dict(r) for r in conn.execute(
        "SELECT * FROM client_request_bundles WHERE client_id=? ORDER BY CASE status WHEN 'open' THEN 0 WHEN 'sent' THEN 1 WHEN 'partial' THEN 2 ELSE 3 END, updated_at DESC",
        (client_id,),
    ).fetchall()]
    for b in bundles:
        items = conn.execute(
            "SELECT COUNT(*) AS total, SUM(CASE WHEN received=1 THEN 1 ELSE 0 END) AS done FROM client_request_items WHERE bundle_id=?",
            (b["id"],),
        ).fetchone()
        b["item_total"] = items["total"] or 0
        b["item_done"] = items["done"] or 0
    return bundles


def _requests_response(request, conn, client_id: int):
    from datetime import date as _date
    bundles = _get_request_bundles(conn, client_id)
    client = conn.execute("SELECT id, name FROM clients WHERE id=?", (client_id,)).fetchone()
    return templates.TemplateResponse("clients/_requests.html", {
        "request": request,
        "client": dict(client) if client else {"id": client_id, "name": ""},
        "bundles": bundles,
        "today_iso": _date.today().isoformat(),
    })


@router.get("/{client_id}/requests", response_class=HTMLResponse)
def get_requests(request: Request, client_id: int, conn=Depends(get_db)):
    return _requests_response(request, conn, client_id)


@router.post("/{client_id}/requests", response_class=HTMLResponse)
def create_request_bundle(
    request: Request,
    client_id: int,
    title: str = Form("Information Request"),
    send_by_date: str = Form(""),
    conn=Depends(get_db),
):
    from policydb.db import next_rfi_uid
    rfi_uid = next_rfi_uid(conn, client_id)
    conn.execute(
        "INSERT INTO client_request_bundles (client_id, title, send_by_date, rfi_uid) VALUES (?, ?, ?, ?)",
        (client_id, title, send_by_date.strip() or None, rfi_uid),
    )
    bundle_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    # If a send-by date is set, create an activity_log follow-up so it surfaces in follow-ups
    if send_by_date.strip():
        from datetime import date as _date
        account_exec = cfg.get("default_account_exec", "")
        conn.execute(
            """INSERT INTO activity_log
               (activity_date, client_id, activity_type, subject, follow_up_date, account_exec)
               VALUES (?, ?, 'Task', ?, ?, ?)""",
            (
                _date.today().isoformat(),
                client_id,
                f"Send RFI: {rfi_uid} {title}",
                send_by_date.strip(),
                account_exec,
            ),
        )
    conn.commit()
    return _requests_response(request, conn, client_id)


def _enrich_request_items(conn, items: list[dict]) -> list[dict]:
    """Add policy_type and carrier to request items from the policies table."""
    for item in items:
        if item.get("policy_uid"):
            pol = conn.execute(
                "SELECT policy_type, carrier, project_name FROM policies WHERE policy_uid=?",
                (item["policy_uid"],),
            ).fetchone()
            if pol:
                item["policy_type"] = pol["policy_type"]
                item["carrier"] = pol["carrier"]
                if not item.get("project_name"):
                    item["project_name"] = pol["project_name"]
    return items


def _attach_item_attachment_counts(conn, items: list[dict]) -> None:
    """Populate item['attachment_count'] for a list of RFI items via one batched query."""
    if not items:
        return
    item_ids = [i["id"] for i in items]
    placeholders = ",".join("?" * len(item_ids))
    rows = conn.execute(
        f"""SELECT record_id AS item_id, COUNT(*) AS n
            FROM record_attachments
            WHERE record_type='rfi_item' AND record_id IN ({placeholders})
            GROUP BY record_id""",
        item_ids,
    ).fetchall()
    counts = {r["item_id"]: r["n"] for r in rows}
    for item in items:
        item["attachment_count"] = counts.get(item["id"], 0)


def _bundle_response(request, conn, client_id: int, bundle_id: int):
    bundle = conn.execute(
        "SELECT * FROM client_request_bundles WHERE id=? AND client_id=?",
        (bundle_id, client_id),
    ).fetchone()
    if not bundle:
        return HTMLResponse("Bundle not found", status_code=404)
    items = _enrich_request_items(conn, [dict(r) for r in conn.execute(
        "SELECT * FROM client_request_items WHERE bundle_id=? ORDER BY received ASC, sort_order ASC, id ASC",
        (bundle_id,),
    ).fetchall()])
    _attach_item_attachment_counts(conn, items)
    # Total attachment count for the Download ZIP button — bundle-level + sum of item counts
    bundle_att_count = conn.execute(
        "SELECT COUNT(*) FROM record_attachments WHERE record_type='rfi_bundle' AND record_id=?",
        (bundle_id,),
    ).fetchone()[0]
    item_att_total = sum(i.get("attachment_count", 0) for i in items)
    bundle_dict = dict(bundle)
    bundle_dict["total_attachment_count"] = bundle_att_count + item_att_total
    client = conn.execute("SELECT id, name FROM clients WHERE id=?", (client_id,)).fetchone()
    # Get policies for this client (for linking items to policies)
    policies = [dict(r) for r in conn.execute(
        "SELECT policy_uid, policy_type, carrier, project_name, effective_date FROM policies WHERE client_id=? AND archived=0 ORDER BY policy_type, effective_date DESC",
        (client_id,),
    ).fetchall()]
    return templates.TemplateResponse("clients/_request_bundle.html", {
        "request": request,
        "client": dict(client) if client else {"id": client_id, "name": ""},
        "bundle": bundle_dict,
        "items": items,
        "policies": policies,
        "request_categories": cfg.get("request_categories", []),
    })


@router.get("/{client_id}/requests/policy-items")
def request_policy_items(client_id: int, policy_uid: str = "", conn=Depends(get_db)):
    """JSON: return all request items linked to a specific policy_uid for inline display."""
    if not policy_uid:
        return JSONResponse({"items": []})
    items = [dict(r) for r in conn.execute(
        """SELECT i.id, i.bundle_id, i.description, i.category, i.received, i.received_at, i.notes
           FROM client_request_items i
           JOIN client_request_bundles b ON i.bundle_id = b.id
           WHERE b.client_id = ? AND i.policy_uid = ?
           ORDER BY i.received ASC, i.id ASC""",
        (client_id, policy_uid),
    ).fetchall()]
    bundle_url = f"/clients/{client_id}"
    return JSONResponse({"items": items, "bundle_url": bundle_url})


@router.get("/{client_id}/requests/program-view", response_class=HTMLResponse)
def request_program_view(request: Request, client_id: int, program_uid: str = "", conn=Depends(get_db)):
    """HTMX partial: request items scoped to a program — items from all child policies."""
    from policydb.queries import get_program_by_uid
    program = get_program_by_uid(conn, program_uid) if program_uid else None

    # Find the active bundle for this client
    bundle = conn.execute(
        "SELECT * FROM client_request_bundles WHERE client_id=? AND status IN ('open','sent','partial') ORDER BY updated_at DESC LIMIT 1",
        (client_id,),
    ).fetchone()
    bundle_id = bundle["id"] if bundle else None

    items = []
    if bundle_id and program:
        # Get all child policy UIDs for this program
        child_uids = [r["policy_uid"] for r in conn.execute(
            "SELECT policy_uid FROM policies WHERE program_id=? AND archived=0",
            (program["id"],),
        ).fetchall()]
        if child_uids:
            placeholders = ",".join("?" * len(child_uids))
            items = _enrich_request_items(conn, [dict(r) for r in conn.execute(
                f"SELECT * FROM client_request_items WHERE bundle_id=? AND policy_uid IN ({placeholders}) ORDER BY received ASC, sort_order ASC, id ASC",
                [bundle_id] + child_uids,
            ).fetchall()])

    client = conn.execute("SELECT id, name FROM clients WHERE id=?", (client_id,)).fetchone()
    return templates.TemplateResponse("clients/_request_policy_view.html", {
        "request": request,
        "client": dict(client) if client else {"id": client_id, "name": ""},
        "bundle": dict(bundle) if bundle else None,
        "bundle_id": bundle_id,
        "items": items,
        "policy_uid": program_uid,
        "request_categories": cfg.get("request_categories", []),
    })


@router.get("/{client_id}/requests/policy-view", response_class=HTMLResponse)
def request_policy_view(request: Request, client_id: int, policy_uid: str = "", conn=Depends(get_db)):
    """HTMX partial: server-rendered request items for a policy, with full card layout."""
    # Find the active bundle (or latest)
    bundle = conn.execute(
        "SELECT * FROM client_request_bundles WHERE client_id=? AND status IN ('open','sent','partial') ORDER BY updated_at DESC LIMIT 1",
        (client_id,),
    ).fetchone()
    bundle_id = bundle["id"] if bundle else None

    items = []
    if bundle_id:
        items = _enrich_request_items(conn, [dict(r) for r in conn.execute(
            "SELECT * FROM client_request_items WHERE bundle_id=? AND policy_uid=? ORDER BY received ASC, sort_order ASC, id ASC",
            (bundle_id, policy_uid),
        ).fetchall()])

    client = conn.execute("SELECT id, name FROM clients WHERE id=?", (client_id,)).fetchone()
    return templates.TemplateResponse("clients/_request_policy_view.html", {
        "request": request,
        "client": dict(client) if client else {"id": client_id, "name": ""},
        "bundle": dict(bundle) if bundle else None,
        "bundle_id": bundle_id,
        "items": items,
        "policy_uid": policy_uid,
        "request_categories": cfg.get("request_categories", []),
    })


@router.get("/{client_id}/requests/export-all")
def request_export_all(client_id: int, conn=Depends(get_db)):
    """Export all open request bundles as a multi-sheet XLSX."""
    from fastapi.responses import Response
    from policydb.exporter import export_client_requests_xlsx

    content = export_client_requests_xlsx(conn, client_id)
    client = conn.execute("SELECT name FROM clients WHERE id=?", (client_id,)).fetchone()
    client_name = client["name"] if client else "Client"
    from datetime import date
    filename = f"{client_name} - Outstanding Requests - {date.today().isoformat()}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{client_id}/requests/{bundle_id}", response_class=HTMLResponse)
def get_request_bundle(
    request: Request, client_id: int, bundle_id: int, conn=Depends(get_db)
):
    return _bundle_response(request, conn, client_id, bundle_id)


@router.post("/{client_id}/requests/{bundle_id}/items", response_class=HTMLResponse)
def add_request_item(
    request: Request,
    client_id: int,
    bundle_id: int,
    description: str = Form(...),
    policy_uid: str = Form(""),
    project_name: str = Form(""),
    category: str = Form(""),
    conn=Depends(get_db),
):
    conn.execute(
        "INSERT INTO client_request_items (bundle_id, description, policy_uid, project_name, category) VALUES (?, ?, ?, ?, ?)",
        (bundle_id, description, policy_uid, project_name, category),
    )
    conn.commit()
    return _bundle_response(request, conn, client_id, bundle_id)


@router.post(
    "/{client_id}/requests/{bundle_id}/items/{item_id}/toggle",
    response_class=HTMLResponse,
)
def toggle_request_item(
    request: Request,
    client_id: int,
    bundle_id: int,
    item_id: int,
    conn=Depends(get_db),
):
    item = conn.execute(
        "SELECT * FROM client_request_items WHERE id=? AND bundle_id=?",
        (item_id, bundle_id),
    ).fetchone()
    if not item:
        return HTMLResponse("Item not found", status_code=404)
    new_received = 0 if item["received"] else 1
    if new_received:
        conn.execute(
            "UPDATE client_request_items SET received=1, received_at=CURRENT_TIMESTAMP WHERE id=?",
            (item_id,),
        )
    else:
        conn.execute(
            "UPDATE client_request_items SET received=0, received_at=NULL WHERE id=?",
            (item_id,),
        )
    # Auto-update bundle status based on item completion
    counts = conn.execute(
        "SELECT COUNT(*) AS total, SUM(CASE WHEN received=1 THEN 1 ELSE 0 END) AS done FROM client_request_items WHERE bundle_id=?",
        (bundle_id,),
    ).fetchone()
    total = counts["total"] or 0
    done = counts["done"] or 0
    bundle = conn.execute(
        "SELECT * FROM client_request_bundles WHERE id=? AND client_id=?",
        (bundle_id, client_id),
    ).fetchone()
    if total > 0 and done == total:
        conn.execute(
            "UPDATE client_request_bundles SET status='complete', updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (bundle_id,),
        )
    elif bundle and bundle["status"] == "complete":
        conn.execute(
            "UPDATE client_request_bundles SET status='partial', updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (bundle_id,),
        )
    conn.commit()
    # Re-fetch updated item and bundle
    updated_item = dict(conn.execute(
        "SELECT * FROM client_request_items WHERE id=?", (item_id,)
    ).fetchone())
    _enrich_request_items(conn, [updated_item])
    _attach_item_attachment_counts(conn, [updated_item])
    bundle = conn.execute(
        "SELECT * FROM client_request_bundles WHERE id=? AND client_id=?",
        (bundle_id, client_id),
    ).fetchone()
    client = conn.execute("SELECT id, name FROM clients WHERE id=?", (client_id,)).fetchone()
    return templates.TemplateResponse("clients/_request_item.html", {
        "request": request,
        "item": updated_item,
        "bundle": dict(bundle),
        "client": dict(client) if client else {"id": client_id, "name": ""},
        "request_categories": cfg.get("request_categories", []),
    })


@router.patch("/{client_id}/requests/{bundle_id}/items/{item_id}")
async def edit_request_item(
    request: Request,
    client_id: int,
    bundle_id: int,
    item_id: int,
    conn=Depends(get_db),
):
    body = await request.json()
    field = body.get("field", "")
    value = body.get("value", "")
    allowed = {"description", "notes", "category", "policy_uid", "project_name"}
    if field not in allowed:
        return JSONResponse({"ok": False, "error": "Invalid field"}, status_code=400)
    conn.execute(
        f"UPDATE client_request_items SET {field}=? WHERE id=? AND bundle_id=?",
        (value, item_id, bundle_id),
    )
    conn.commit()
    return JSONResponse({"ok": True, "formatted": value})


@router.delete(
    "/{client_id}/requests/{bundle_id}/items/{item_id}",
    response_class=HTMLResponse,
)
def delete_request_item(
    request: Request,
    client_id: int,
    bundle_id: int,
    item_id: int,
    conn=Depends(get_db),
):
    # Re-parent any item-level attachments up to the bundle so history is
    # preserved ("we sent this type of application for this RFI before").
    # INSERT OR IGNORE protects the UNIQUE(attachment_id, record_type, record_id)
    # constraint in the case where the same file is already linked at the bundle
    # level — the existing bundle link wins, the item link is dropped.
    conn.execute(
        """INSERT OR IGNORE INTO record_attachments (attachment_id, record_type, record_id, sort_order, created_at)
           SELECT attachment_id, 'rfi_bundle', ?, sort_order, created_at
           FROM record_attachments
           WHERE record_type='rfi_item' AND record_id=?""",
        (bundle_id, item_id),
    )
    conn.execute(
        "DELETE FROM record_attachments WHERE record_type='rfi_item' AND record_id=?",
        (item_id,),
    )
    conn.execute(
        "DELETE FROM client_request_items WHERE id=? AND bundle_id=?",
        (item_id, bundle_id),
    )
    conn.commit()
    # Trigger a refresh of the bundle-level attachment panel so any promoted
    # files appear immediately. The listener lives in _request_bundle.html.
    return HTMLResponse("", headers={"HX-Trigger": "refreshBundleAttachments"})


@router.post(
    "/{client_id}/requests/{bundle_id}/status", response_class=HTMLResponse
)
def update_request_bundle_status(
    request: Request,
    client_id: int,
    bundle_id: int,
    status: str = Form(...),
    follow_up_date: str = Form(""),
    conn=Depends(get_db),
):
    from datetime import date as _date
    from policydb.utils import round_duration

    allowed_statuses = {"open", "sent", "partial", "complete"}
    if status not in allowed_statuses:
        return HTMLResponse("Invalid status", status_code=400)
    if status == "sent":
        conn.execute(
            "UPDATE client_request_bundles SET status=?, sent_at=CURRENT_TIMESTAMP, updated_at=CURRENT_TIMESTAMP WHERE id=? AND client_id=?",
            (status, bundle_id, client_id),
        )
        # Create an activity log entry so this shows in follow-ups
        bundle = conn.execute(
            "SELECT title, rfi_uid FROM client_request_bundles WHERE id=?", (bundle_id,)
        ).fetchone()
        counts = conn.execute(
            "SELECT COUNT(*) AS total, SUM(CASE WHEN received=0 THEN 1 ELSE 0 END) AS outstanding FROM client_request_items WHERE bundle_id=?",
            (bundle_id,),
        ).fetchone()
        outstanding = counts["outstanding"] or 0
        total = counts["total"] or 0
        _rfi_tag = f"{bundle['rfi_uid']} " if bundle and bundle["rfi_uid"] else ""
        subject = f"Sent {_rfi_tag}{bundle['title'] if bundle else 'information request'} — {outstanding} of {total} items outstanding"
        account_exec = cfg.get("default_account_exec", "Grant")
        fu = follow_up_date.strip() or None
        conn.execute(
            """INSERT INTO activity_log
               (activity_date, client_id, policy_id, activity_type, subject, details, follow_up_date, account_exec)
               VALUES (?, ?, NULL, 'Email', ?, ?, ?, ?)""",
            (_date.today().isoformat(), client_id, subject, None, fu, account_exec),
        )
    else:
        conn.execute(
            "UPDATE client_request_bundles SET status=?, updated_at=CURRENT_TIMESTAMP WHERE id=? AND client_id=?",
            (status, bundle_id, client_id),
        )
    conn.commit()
    return _requests_response(request, conn, client_id)


@router.post("/{client_id}/requests/quick-add", response_class=HTMLResponse)
def quick_add_request_item(
    request: Request,
    client_id: int,
    description: str = Form(...),
    policy_uid: str = Form(""),
    project_name: str = Form(""),
    category: str = Form(""),
    send_by_date: str = Form(""),
    conn=Depends(get_db),
):
    # Find the latest open bundle for this client, or create one
    bundle = conn.execute(
        "SELECT id FROM client_request_bundles WHERE client_id=? AND status='open' ORDER BY created_at DESC LIMIT 1",
        (client_id,),
    ).fetchone()
    if not bundle:
        from policydb.db import next_rfi_uid
        sbd = send_by_date.strip() or None
        rfi_uid = next_rfi_uid(conn, client_id)
        conn.execute(
            "INSERT INTO client_request_bundles (client_id, title, send_by_date, rfi_uid) VALUES (?, 'Information Request', ?, ?)",
            (client_id, sbd, rfi_uid),
        )
        bundle_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        # Create follow-up activity so the send deadline surfaces in follow-ups
        if sbd:
            from datetime import date as _date
            account_exec = cfg.get("default_account_exec", "")
            conn.execute(
                """INSERT INTO activity_log
                   (activity_date, client_id, activity_type, subject, follow_up_date, account_exec)
                   VALUES (?, ?, 'Task', ?, ?, ?)""",
                (_date.today().isoformat(), client_id, f"Send RFI: {rfi_uid}", sbd, account_exec),
            )
    else:
        bundle_id = bundle["id"]
    conn.execute(
        "INSERT INTO client_request_items (bundle_id, description, policy_uid, project_name, category) VALUES (?, ?, ?, ?, ?)",
        (bundle_id, description, policy_uid, project_name, category),
    )
    conn.commit()
    # Return the full policy-view partial so the list updates in place
    return request_policy_view(request, client_id, policy_uid, conn)


@router.post(
    "/{client_id}/requests/{bundle_id}/seed-from-checklist",
    response_class=HTMLResponse,
)
def seed_from_checklist(
    request: Request,
    client_id: int,
    bundle_id: int,
    conn=Depends(get_db),
):
    client_facing = cfg.get("client_facing_milestones", [])
    if not client_facing:
        return _bundle_response(request, conn, client_id, bundle_id)
    # Get all active policies and opportunities for this client
    policies = conn.execute(
        "SELECT policy_uid, policy_type, carrier, project_name FROM policies WHERE client_id=? AND archived=0",
        (client_id,),
    ).fetchall()
    for pol in policies:
        # For each client-facing milestone, check if it's incomplete (no row or completed=0)
        for ms_name in client_facing:
            existing_ms = conn.execute(
                "SELECT completed FROM policy_milestones WHERE policy_uid=? AND milestone=?",
                (pol["policy_uid"], ms_name),
            ).fetchone()
            # Skip if already completed
            if existing_ms and existing_ms["completed"]:
                continue
            # Build a descriptive item name
            desc = f"{ms_name} — {pol['policy_type']}"
            if pol["carrier"]:
                desc += f" ({pol['carrier']})"
            # Check if this item already exists in the bundle (avoid duplicates)
            existing_item = conn.execute(
                "SELECT id FROM client_request_items WHERE bundle_id=? AND description=? AND policy_uid=?",
                (bundle_id, desc, pol["policy_uid"]),
            ).fetchone()
            if existing_item:
                continue
            conn.execute(
                "INSERT INTO client_request_items (bundle_id, description, policy_uid, project_name, category) VALUES (?, ?, ?, ?, ?)",
                (bundle_id, desc, pol["policy_uid"], pol["project_name"] or "", ms_name.split()[0] if ms_name else "Other"),
            )
    conn.commit()
    return _bundle_response(request, conn, client_id, bundle_id)


@router.post("/{client_id}/requests/{bundle_id}/send-by", response_class=HTMLResponse)
def set_bundle_send_by(
    request: Request,
    client_id: int,
    bundle_id: int,
    send_by_date: str = Form(...),
    policy_uid: str = Form(""),
    conn=Depends(get_db),
):
    """Set or update send_by_date on a bundle. Creates an activity follow-up."""
    sbd = send_by_date.strip() or None
    conn.execute(
        "UPDATE client_request_bundles SET send_by_date=?, updated_at=CURRENT_TIMESTAMP WHERE id=? AND client_id=?",
        (sbd, bundle_id, client_id),
    )
    if sbd:
        from datetime import date as _date
        bundle_row = conn.execute("SELECT title, rfi_uid FROM client_request_bundles WHERE id=?", (bundle_id,)).fetchone()
        title = bundle_row["title"] if bundle_row else "Information Request"
        rfi_uid = bundle_row["rfi_uid"] if bundle_row else ""
        account_exec = cfg.get("default_account_exec", "")
        conn.execute(
            """INSERT INTO activity_log
               (activity_date, client_id, activity_type, subject, follow_up_date, account_exec)
               VALUES (?, ?, 'Task', ?, ?, ?)""",
            (_date.today().isoformat(), client_id, f"Send RFI: {rfi_uid} {title}", sbd, account_exec),
        )
    conn.commit()
    if policy_uid:
        return request_policy_view(request, client_id, policy_uid, conn)
    return _bundle_response(request, conn, client_id, bundle_id)


@router.get("/{client_id}/requests/{bundle_id}/export")
def request_bundle_export(client_id: int, bundle_id: int, conn=Depends(get_db)):
    """Export a request bundle as XLSX."""
    from fastapi.responses import Response
    from policydb.exporter import export_request_bundle_xlsx

    bundle = conn.execute(
        "SELECT title FROM client_request_bundles WHERE id=? AND client_id=?",
        (bundle_id, client_id),
    ).fetchone()
    if not bundle:
        return HTMLResponse("Bundle not found", status_code=404)

    content = export_request_bundle_xlsx(conn, bundle_id)
    client = conn.execute("SELECT name FROM clients WHERE id=?", (client_id,)).fetchone()
    client_name = client["name"] if client else "Client"
    from datetime import date
    filename = f"{client_name} - {bundle['title']} - {date.today().isoformat()}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Client-level field PATCH (e.g. caching geocoded lat/lng from sidebar map)
# ---------------------------------------------------------------------------

@router.patch("/{client_id}/field")
async def client_field_patch(
    request: Request,
    client_id: int,
    conn=Depends(get_db),
):
    """Update a single field on a client record."""
    body = await request.json()
    field = body.get("field", "")
    value = body.get("value", "")

    allowed = {
        "name", "cn_number", "industry_segment", "account_exec",
        "date_onboarded", "website", "fein", "broker_fee", "hourly_rate",
        "follow_up_date", "relationship_risk", "service_model",
        "business_description", "notes", "stewardship_date",
        "renewal_strategy", "growth_opportunities", "account_priorities",
        "latitude", "longitude",
    }
    if field not in allowed:
        return JSONResponse({"ok": False, "error": f"Invalid field: {field}"}, status_code=400)

    client = conn.execute("SELECT id FROM clients WHERE id = ?", (client_id,)).fetchone()
    if not client:
        return JSONResponse({"ok": False, "error": "Not found"}, status_code=404)

    formatted = value

    # Numeric fields
    if field in ("latitude", "longitude", "hourly_rate"):
        try:
            num = float(value) if str(value).strip() else None
        except ValueError:
            num = None
        conn.execute(f"UPDATE clients SET {field} = ? WHERE id = ?", (num, client_id))
        conn.commit()
        return JSONResponse({"ok": True, "formatted": str(num) if num is not None else ""})

    # Currency fields
    if field in ("broker_fee",):
        from policydb.utils import parse_currency_with_magnitude
        parsed = parse_currency_with_magnitude(value)
        conn.execute(f"UPDATE clients SET {field} = ? WHERE id = ?", (parsed, client_id))
        conn.commit()
        if parsed is not None:
            formatted = f"${parsed:,.0f}" if parsed == int(parsed) else f"${parsed:,.2f}"
        else:
            formatted = ""
        return JSONResponse({"ok": True, "formatted": formatted})

    conn.execute(f"UPDATE clients SET {field} = ? WHERE id = ?", (value or None, client_id))
    conn.commit()
    return JSONResponse({"ok": True, "formatted": formatted})


# ---------------------------------------------------------------------------
# Project Pipeline endpoints
# ---------------------------------------------------------------------------

@router.patch("/{client_id}/projects/{project_id}/field")
async def project_pipeline_field(
    request: Request,
    client_id: int,
    project_id: int,
    conn=Depends(get_db),
):
    """Update a single field on a pipeline project (contenteditable cell save)."""
    body = await request.json()
    field = body.get("field", "")
    value = body.get("value", "")

    allowed = {"project_type", "status", "name", "project_value", "start_date",
               "target_completion", "insurance_needed_by", "scope_description",
               "general_contractor", "owner_name", "address", "city", "state", "zip",
               "latitude", "longitude"}
    if field not in allowed:
        return JSONResponse({"ok": False, "error": f"Invalid field: {field}"}, status_code=400)

    project = conn.execute(
        "SELECT * FROM projects WHERE id = ? AND client_id = ?",
        (project_id, client_id),
    ).fetchone()
    if not project:
        return JSONResponse({"ok": False, "error": "Not found"}, status_code=404)

    formatted = value
    if field == "project_value":
        from policydb.utils import parse_currency_with_magnitude
        num = parse_currency_with_magnitude(value)
        conn.execute("UPDATE projects SET project_value = ? WHERE id = ?", (num, project_id))
        formatted = f"${num:,.0f}"
    elif field in ("latitude", "longitude"):
        try:
            num = float(value) if str(value).strip() else None
        except ValueError:
            num = None
        conn.execute(f"UPDATE projects SET {field} = ? WHERE id = ?", (num, project_id))
        formatted = str(num) if num is not None else ""
    elif field in ("start_date", "target_completion", "insurance_needed_by"):
        conn.execute(f"UPDATE projects SET {field} = ? WHERE id = ?",
                     (value.strip() or None, project_id))
        formatted = value.strip()
    elif field == "name":
        existing = conn.execute(
            "SELECT id FROM projects WHERE client_id = ? AND LOWER(TRIM(name)) = LOWER(TRIM(?)) AND id != ?",
            (client_id, value.strip(), project_id),
        ).fetchone()
        if existing:
            return JSONResponse({"ok": False, "error": "Project name already exists"}, status_code=400)
        conn.execute("UPDATE projects SET name = ? WHERE id = ?", (value.strip(), project_id))
        conn.execute("UPDATE policies SET project_name = ? WHERE project_id = ?", (value.strip(), project_id))
        formatted = value.strip()
    else:
        clean_value = value.strip() or None
        conn.execute(f"UPDATE projects SET {field} = ? WHERE id = ?",
                     (clean_value, project_id))
        formatted = value.strip()
        # Sync address fields to linked policies
        _address_to_exposure = {"address": "exposure_address", "city": "exposure_city",
                                "state": "exposure_state", "zip": "exposure_zip"}
        if field in _address_to_exposure:
            exposure_field = _address_to_exposure[field]
            conn.execute(f"UPDATE policies SET {exposure_field} = ? WHERE project_id = ? AND archived = 0",
                         (clean_value, project_id))

    conn.commit()
    return JSONResponse({"ok": True, "formatted": formatted})


@router.post("/{client_id}/projects/{project_id}/status", response_class=HTMLResponse)
def project_pipeline_status(
    request: Request,
    client_id: int,
    project_id: int,
    status: str = Form(...),
    conn=Depends(get_db),
):
    """HTMX endpoint: update project status, return updated badge partial."""
    stages = cfg.get("project_stages", ["Upcoming", "Quoting", "Bound", "Active", "Complete"])
    if status not in stages:
        status = stages[0]
    conn.execute("UPDATE projects SET status = ? WHERE id = ? AND client_id = ?",
                 (status, project_id, client_id))
    conn.commit()
    project = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
    if not project:
        return HTMLResponse("", status_code=404)

    # Check for active programs when marking Complete
    active_program_count = 0
    if status == "Complete":
        active_program_count = conn.execute(
            "SELECT COUNT(*) FROM programs WHERE project_id = ? AND archived = 0",
            (project_id,),
        ).fetchone()[0]

    client = get_client_by_id(conn, client_id)
    return templates.TemplateResponse("clients/_project_status_badge.html", {
        "request": request,
        "p": dict(project),
        "client": dict(client),
        "project_stages": stages,
        "active_program_count": active_program_count,
    })


@router.post("/{client_id}/projects/{project_id}/archive-programs")
def archive_project_programs(
    client_id: int,
    project_id: int,
    conn=Depends(get_db),
):
    """Archive all active programs linked to a project (used when project completes)."""
    conn.execute(
        "UPDATE programs SET archived = 1, updated_at = CURRENT_TIMESTAMP WHERE project_id = ? AND archived = 0",
        (project_id,),
    )
    conn.commit()
    return JSONResponse({"ok": True})


@router.post("/{client_id}/projects/{project_id}/type", response_class=HTMLResponse)
def project_pipeline_type(
    request: Request,
    client_id: int,
    project_id: int,
    project_type: str = Form(...),
    conn=Depends(get_db),
):
    """HTMX endpoint: update project type, return updated badge partial."""
    types = cfg.get("project_types", ["Location", "Construction", "Development", "Renovation"])
    if project_type not in types:
        project_type = types[0]
    conn.execute("UPDATE projects SET project_type = ? WHERE id = ? AND client_id = ?",
                 (project_type, project_id, client_id))
    conn.commit()
    project = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
    if not project:
        return HTMLResponse("", status_code=404)
    client = get_client_by_id(conn, client_id)
    return templates.TemplateResponse("clients/_project_type_badge.html", {
        "request": request,
        "p": dict(project),
        "client": dict(client),
        "project_types": types,
    })


@router.post("/{client_id}/projects/pipeline", response_class=HTMLResponse)
def project_pipeline_add(
    request: Request,
    client_id: int,
    conn=Depends(get_db),
):
    """Create a new pipeline project with default values."""
    base = "New Project"
    name = base
    counter = 2
    while conn.execute(
        "SELECT id FROM projects WHERE client_id = ? AND LOWER(TRIM(name)) = LOWER(TRIM(?))",
        (client_id, name),
    ).fetchone():
        name = f"{base} {counter}"
        counter += 1

    conn.execute(
        """INSERT INTO projects (client_id, name, project_type, status)
           VALUES (?, ?, 'Construction', 'Upcoming')""",
        (client_id, name),
    )
    project_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.commit()

    project = dict(conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone())
    project["total_coverages"] = 0
    project["placed_coverages"] = 0
    project["total_premium"] = 0
    project["total_revenue"] = 0

    return templates.TemplateResponse("clients/_project_pipeline_row.html", {
        "request": request,
        "p": project,
        "client": {"id": client_id},
        "project_stages": cfg.get("project_stages", []),
        "project_types": cfg.get("project_types", []),
    })


@router.post("/{client_id}/projects/location")
def project_location_add(
    client_id: int,
    conn=Depends(get_db),
):
    """Create a new location project."""
    base = "New Location"
    name = base
    counter = 2
    while conn.execute(
        "SELECT id FROM projects WHERE client_id = ? AND LOWER(TRIM(name)) = LOWER(TRIM(?))",
        (client_id, name),
    ).fetchone():
        name = f"{base} {counter}"
        counter += 1
    conn.execute(
        "INSERT INTO projects (client_id, name, project_type) VALUES (?, ?, 'Location')",
        (client_id, name),
    )
    conn.commit()
    return JSONResponse({"ok": True, "reload": True})


@router.get("/{client_id}/projects/{project_id}/coverage", response_class=HTMLResponse)
def project_coverage_detail(
    request: Request,
    client_id: int,
    project_id: int,
    conn=Depends(get_db),
):
    """Return coverage detail expansion for a pipeline project."""
    policies = [dict(r) for r in conn.execute("""
        SELECT policy_uid, policy_type, carrier, premium, renewal_status,
               is_opportunity, opportunity_status
        FROM policies
        WHERE project_id = ? AND archived = 0
        ORDER BY is_opportunity, policy_type
    """, (project_id,)).fetchall()]

    return templates.TemplateResponse("clients/_project_coverage_detail.html", {
        "request": request,
        "policies": policies,
    })


@router.delete("/{client_id}/projects/{project_id}/pipeline")
def project_pipeline_delete(
    client_id: int,
    project_id: int,
    conn=Depends(get_db),
):
    """Delete a pipeline project, unlinking its policies and cleaning up programs."""
    # Cascade safety: null out program_id on child policies of programs linked to this project,
    # and clean up tower data, before the CASCADE delete removes the programs.
    linked_programs = conn.execute(
        "SELECT id FROM programs WHERE project_id = ?", (project_id,)
    ).fetchall()
    for pgm in linked_programs:
        # Clean tower coverage/line refs for child policies
        child_ids = [r[0] for r in conn.execute(
            "SELECT id FROM policies WHERE program_id = ?", (pgm["id"],)
        ).fetchall()]
        if child_ids:
            id_list = ",".join(str(cid) for cid in child_ids)
            conn.execute(f"DELETE FROM program_tower_coverage WHERE excess_policy_id IN ({id_list}) OR underlying_policy_id IN ({id_list})")
            conn.execute(f"DELETE FROM program_tower_lines WHERE source_policy_id IN ({id_list})")
        # Null out program_id so policies survive the program deletion
        conn.execute("UPDATE policies SET program_id = NULL, tower_group = NULL WHERE program_id = ?", (pgm["id"],))

    conn.execute("UPDATE policies SET project_id = NULL, project_name = NULL WHERE project_id = ?", (project_id,))
    conn.execute("DELETE FROM projects WHERE id = ? AND client_id = ?", (project_id, client_id))
    conn.commit()
    return JSONResponse({"ok": True})


# ── Project-Level Exposure Routes ──────────────────────────────────────────────


@router.get("/{client_id}/projects/{project_id}/exposures", response_class=HTMLResponse)
def project_exposures(request: Request, client_id: int, project_id: int, year: int = 0, conn=Depends(get_db)):
    """Load project-level exposure matrix (HTMX partial)."""
    from datetime import date as _date
    if year == 0:
        year = _date.today().year
    ctx = _exposure_tab_context(conn, client_id, year, project_id)
    return templates.TemplateResponse("clients/_exposure_matrix.html", {"request": request, **ctx})


@router.get("/{client_id}/projects/{project_id}/exposures/types", response_class=HTMLResponse)
def project_exposure_types_dropdown(request: Request, client_id: int, project_id: int, year: int = 0, conn=Depends(get_db)):
    """Return dropdown HTML for the project exposure type picker."""
    from datetime import date as _date
    if year == 0:
        year = _date.today().year
    return HTMLResponse(_build_exposure_types_dropdown(client_id, year, project_id, conn))


@router.post("/{client_id}/projects/{project_id}/exposures/add-row", response_class=HTMLResponse)
async def project_exposure_add_row(request: Request, client_id: int, project_id: int, conn=Depends(get_db)):
    """Create a new project-level exposure row."""
    form = await request.form()
    return _exposure_add_row_handler(request, client_id, form, conn, project_id)


@router.patch("/{client_id}/projects/{project_id}/exposures/{exposure_id}/cell")
async def project_exposure_cell(request: Request, client_id: int, project_id: int, exposure_id: int, conn=Depends(get_db)):
    """Save a single cell value for a project-level exposure row."""
    return await exposure_cell(request, client_id, exposure_id, conn)


@router.patch("/{client_id}/projects/{project_id}/exposures/{exposure_id}/toggle-primary")
async def project_exposure_toggle_primary(request: Request, client_id: int, project_id: int, exposure_id: int, conn=Depends(get_db)):
    """Toggle primary status for a project-level policy-exposure link."""
    return await exposure_toggle_primary(request, client_id, exposure_id, conn, project_id=project_id)


@router.delete("/{client_id}/projects/{project_id}/exposures/{exposure_id}", response_class=HTMLResponse)
def project_exposure_delete(request: Request, client_id: int, project_id: int, exposure_id: int, conn=Depends(get_db)):
    """Delete a project-level exposure row."""
    conn.execute("DELETE FROM client_exposures WHERE id=? AND client_id=?", (exposure_id, client_id))
    conn.commit()
    return HTMLResponse("")


@router.post("/{client_id}/projects/{project_id}/exposures/copy-forward", response_class=HTMLResponse)
async def project_exposure_copy_forward(request: Request, client_id: int, project_id: int, conn=Depends(get_db)):
    """Copy project-level exposure types from one year to another."""
    form = await request.form()
    return _exposure_copy_forward_handler(request, client_id, form, conn, project_id)


@router.get("/{client_id}/projects/pipeline/export")
def project_pipeline_export(
    client_id: int,
    format: str = "xlsx",
    conn=Depends(get_db),
):
    """Export project pipeline as CSV or XLSX."""
    import io, re
    client = conn.execute("SELECT name FROM clients WHERE id = ?", (client_id,)).fetchone()
    if not client:
        return HTMLResponse("Not found", status_code=404)

    projects = _get_project_pipeline(conn, client_id)

    # Attach coverage list per project
    for p in projects:
        pols = conn.execute("""
            SELECT policy_type, is_opportunity, renewal_status
            FROM policies WHERE project_id = ? AND archived = 0
            ORDER BY is_opportunity, policy_type
        """, (p["id"],)).fetchall()
        coverages = []
        for pol in pols:
            status = "Opp" if pol["is_opportunity"] else (pol["renewal_status"] or "Placed")
            coverages.append(f"{pol['policy_type']} ({status})")
        p["coverage_list"] = ", ".join(coverages) if coverages else ""

    cols = ["name", "project_type", "status", "address", "city", "state", "zip",
            "insurance_needed_by", "start_date", "target_completion",
            "project_value", "total_premium", "total_revenue",
            "general_contractor", "owner_name", "coverage_list", "scope_description"]
    headers = ["Project", "Type", "Status", "Address", "City", "State", "ZIP",
               "Insurance Needed By", "Start Date", "Target Completion",
               "Project Value", "Total Premium", "Total Revenue",
               "General Contractor", "Owner", "Coverages", "Scope"]

    safe_name = re.sub(r'[^\w\s-]', '', client["name"]).strip().replace(' ', '_')

    if format == "csv":
        import csv as _csv
        output = io.StringIO()
        writer = _csv.writer(output)
        writer.writerow(headers)
        for p in projects:
            writer.writerow([p.get(c, "") or "" for c in cols])
        from starlette.responses import Response
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_pipeline.csv"'},
        )

    # XLSX via exporter shared styling
    from openpyxl import Workbook
    from policydb.exporter import _write_sheet, _wb_to_bytes
    sheet_rows = [{h: p.get(c, "") or "" for h, c in zip(headers, cols)} for p in projects]
    wb = Workbook()
    _write_sheet(wb, "Pipeline", sheet_rows)
    if wb.sheetnames and wb.sheetnames[0] == "Sheet":
        del wb["Sheet"]
    from starlette.responses import Response
    return Response(
        content=_wb_to_bytes(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_pipeline.xlsx"'},
    )


@router.get("/{client_id}/projects/pipeline/copy-table")
def project_pipeline_copy_table(client_id: int, conn=Depends(get_db)):
    """Return HTML + plain-text pipeline table for clipboard copy."""
    from fastapi.responses import JSONResponse
    from policydb.email_templates import build_generic_table
    projects = _get_project_pipeline(conn, client_id)
    for p in projects:
        pols = conn.execute("""
            SELECT policy_type, is_opportunity, renewal_status
            FROM policies WHERE project_id = ? AND archived = 0
            ORDER BY is_opportunity, policy_type
        """, (p["id"],)).fetchall()
        coverages = []
        for pol in pols:
            status = "Opp" if pol["is_opportunity"] else (pol["renewal_status"] or "Placed")
            coverages.append(f"{pol['policy_type']} ({status})")
        p["coverage_list"] = ", ".join(coverages) if coverages else ""
    columns = [
        ("name", "Project", False),
        ("project_type", "Type", False),
        ("status", "Status", False),
        ("address", "Address", False),
        ("city", "City", False),
        ("state", "State", False),
        ("insurance_needed_by", "Insurance Needed By", False),
        ("project_value", "Project Value", True),
        ("total_premium", "Total Premium", True),
        ("total_revenue", "Total Revenue", True),
        ("general_contractor", "General Contractor", False),
        ("coverage_list", "Coverages", False),
    ]
    return JSONResponse(build_generic_table(projects, columns))


@router.get("/{client_id}/projects/locations/export")
def project_locations_export(
    client_id: int,
    format: str = "xlsx",
    conn=Depends(get_db),
):
    """Export locations as CSV or XLSX."""
    import io, re
    client = conn.execute("SELECT name FROM clients WHERE id = ?", (client_id,)).fetchone()
    if not client:
        return HTMLResponse("Not found", status_code=404)

    locations = _get_project_locations(conn, client_id)

    for loc in locations:
        pols = conn.execute("""
            SELECT policy_type, is_opportunity, renewal_status
            FROM policies WHERE project_id = ? AND archived = 0
            ORDER BY is_opportunity, policy_type
        """, (loc["id"],)).fetchall()
        coverages = []
        for pol in pols:
            status = "Opp" if pol["is_opportunity"] else (pol["renewal_status"] or "Placed")
            coverages.append(f"{pol['policy_type']} ({status})")
        loc["coverage_list"] = ", ".join(coverages) if coverages else ""

    cols = ["name", "address", "city", "state", "zip",
            "total_premium", "total_revenue", "coverage_list"]
    headers = ["Location", "Address", "City", "State", "ZIP",
               "Total Premium", "Total Revenue", "Coverages"]

    safe_name = re.sub(r'[^\w\s-]', '', client["name"]).strip().replace(' ', '_')

    if format == "csv":
        import csv as _csv
        output = io.StringIO()
        writer = _csv.writer(output)
        writer.writerow(headers)
        for loc in locations:
            writer.writerow([loc.get(c, "") or "" for c in cols])
        from starlette.responses import Response
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_locations.csv"'},
        )

    from openpyxl import Workbook
    from policydb.exporter import _write_sheet, _wb_to_bytes
    sheet_rows = [{h: loc.get(c, "") or "" for h, c in zip(headers, cols)} for loc in locations]
    wb = Workbook()
    _write_sheet(wb, "Locations", sheet_rows)
    if wb.sheetnames and wb.sheetnames[0] == "Sheet":
        del wb["Sheet"]
    from starlette.responses import Response
    return Response(
        content=_wb_to_bytes(wb),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_locations.xlsx"'},
    )


@router.get("/{client_id}/projects/locations/copy-table")
def project_locations_copy_table(client_id: int, conn=Depends(get_db)):
    """Return HTML + plain-text locations table for clipboard copy."""
    from fastapi.responses import JSONResponse
    from policydb.email_templates import build_generic_table
    locations = _get_project_locations(conn, client_id)
    for loc in locations:
        pols = conn.execute("""
            SELECT policy_type, is_opportunity, renewal_status
            FROM policies WHERE project_id = ? AND archived = 0
            ORDER BY is_opportunity, policy_type
        """, (loc["id"],)).fetchall()
        coverages = []
        for pol in pols:
            status = "Opp" if pol["is_opportunity"] else (pol["renewal_status"] or "Placed")
            coverages.append(f"{pol['policy_type']} ({status})")
        loc["coverage_list"] = ", ".join(coverages) if coverages else ""
    columns = [
        ("name", "Location", False),
        ("address", "Address", False),
        ("city", "City", False),
        ("state", "State", False),
        ("zip", "ZIP", False),
        ("total_premium", "Total Premium", True),
        ("total_revenue", "Total Revenue", True),
        ("coverage_list", "Coverages", False),
    ]
    return JSONResponse(build_generic_table(locations, columns))


@router.get("/{client_id}/projects/pipeline/timeline")
def project_timeline_export(
    client_id: int,
    format: str = "pdf",
    conn=Depends(get_db),
):
    """Export project timeline as PDF."""
    from fpdf import FPDF
    from datetime import date as _date

    client = conn.execute("SELECT name FROM clients WHERE id = ?", (client_id,)).fetchone()
    if not client:
        return HTMLResponse("Not found", status_code=404)

    projects = _get_project_pipeline(conn, client_id)
    dated = [p for p in projects if p.get("start_date") or p.get("target_completion")]

    if not dated:
        return HTMLResponse("No projects with dates to render", status_code=400)

    pdf = FPDF()
    pdf.add_page("L")  # landscape
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, f"{client['name']} - Project Pipeline Timeline", ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 5, f"Generated {_date.today().strftime('%B %d, %Y')}", ln=True)
    pdf.ln(5)

    # Compute date range
    all_dates = []
    for p in dated:
        if p.get("start_date"): all_dates.append(p["start_date"])
        if p.get("target_completion"): all_dates.append(p["target_completion"])
        if p.get("insurance_needed_by"): all_dates.append(p["insurance_needed_by"])
    min_date = min(all_dates)
    max_date = max(all_dates)

    d_min = _date.fromisoformat(min_date)
    d_max = _date.fromisoformat(max_date)
    total_days = max((d_max - d_min).days, 1)

    chart_x = 60
    chart_w = 210  # landscape width minus margins
    bar_h = 8
    gap = 3

    # Status colors
    colors = {
        "Upcoming": (180, 180, 180),
        "Quoting": (59, 130, 246),
        "Bound": (34, 197, 94),
        "Active": (34, 197, 94),
        "Complete": (156, 163, 175),
    }

    pdf.set_font("Helvetica", "", 8)
    for p in dated:
        s = p.get("start_date") or p.get("target_completion")
        e = p.get("target_completion") or p.get("start_date")
        ds = _date.fromisoformat(s)
        de = _date.fromisoformat(e)

        x_start = chart_x + ((ds - d_min).days / total_days) * chart_w
        x_width = max(((de - ds).days / total_days) * chart_w, 3)

        # Label
        pdf.set_xy(5, pdf.get_y())
        pdf.cell(55, bar_h, p["name"][:25], 0, 0)

        # Bar
        r, g, b = colors.get(p.get("status", ""), (180, 180, 180))
        pdf.set_fill_color(r, g, b)
        pdf.rect(x_start, pdf.get_y(), x_width, bar_h, "F")

        # Insurance needed marker
        if p.get("insurance_needed_by"):
            di = _date.fromisoformat(p["insurance_needed_by"])
            x_ins = chart_x + ((di - d_min).days / total_days) * chart_w
            pdf.set_draw_color(220, 38, 38)
            pdf.line(x_ins, pdf.get_y(), x_ins, pdf.get_y() + bar_h)

        pdf.ln(bar_h + gap)

    content = pdf.output()
    from starlette.responses import Response
    return Response(
        content=bytes(content),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{client["name"]}_timeline.pdf"'},
    )


# ── Location Assignment Board ────────────────────────────────────────────────


@router.get("/{client_id}/locations", response_class=HTMLResponse)
def client_locations(request: Request, client_id: int, conn=Depends(get_db)):
    """Location assignment board for a client."""
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    if not client:
        return HTMLResponse("Client not found", status_code=404)

    policies = [dict(r) for r in conn.execute("""
        SELECT *
        FROM policies
        WHERE client_id = ? AND archived = 0
        ORDER BY policy_type
    """, (client_id,)).fetchall()]

    # Annotate each policy with whether it has any exposure links
    for p in policies:
        p["has_exposure_link"] = conn.execute(
            "SELECT 1 FROM policy_exposure_links WHERE policy_uid=? LIMIT 1",
            (p["policy_uid"],),
        ).fetchone() is not None

    # Group by project assignment
    unassigned = [p for p in policies if not p.get("project_name")]

    # Build location groups from projects table
    projects = [dict(r) for r in conn.execute(
        "SELECT * FROM projects WHERE client_id=? ORDER BY name", (client_id,)
    ).fetchall()]

    # Color palette for location groups
    colors = [
        ("blue-100", "blue-700", "blue-50"),
        ("green-100", "green-700", "green-50"),
        ("purple-100", "purple-700", "purple-50"),
        ("teal-100", "teal-700", "teal-50"),
        ("amber-100", "amber-700", "amber-50"),
        ("pink-100", "pink-700", "pink-50"),
    ]

    locations = []
    for i, proj in enumerate(projects):
        proj_policies = [p for p in policies if p.get("project_id") == proj["id"]]
        bg, text, light = colors[i % len(colors)]
        has_exposures = conn.execute(
            "SELECT 1 FROM client_exposures WHERE project_id=? LIMIT 1",
            (proj["id"],),
        ).fetchone() is not None
        program_count = conn.execute(
            "SELECT COUNT(*) FROM programs WHERE project_id = ? AND archived = 0",
            (proj["id"],),
        ).fetchone()[0]
        locations.append({
            "id": proj["id"], "name": proj["name"],
            "address": " ".join(filter(None, [
                proj.get("address"), proj.get("city"),
                proj.get("state"), proj.get("zip"),
            ])),
            "policies": proj_policies,
            "total_premium": sum(p.get("premium") or 0 for p in proj_policies),
            "color_bg": bg, "color_text": text, "color_light": light,
            "has_exposures": has_exposures,
            "program_count": program_count,
        })

    # Smart suggestions: group unassigned by shared exposure_address
    suggestions = []
    from collections import defaultdict
    addr_groups: dict[str, list] = defaultdict(list)
    for p in unassigned:
        addr = (p.get("exposure_address") or "").strip()
        if addr:
            addr_groups[addr].append(p)
    for addr, pols in addr_groups.items():
        if len(pols) >= 2:
            matching_loc = next(
                (loc for loc in locations if addr.lower() in loc["address"].lower()),
                None,
            )
            suggestions.append({
                "address": addr, "policies": pols, "count": len(pols),
                "matching_location": matching_loc,
            })

    return templates.TemplateResponse("clients/_location_board.html", {
        "request": request, "client_id": client_id, "client_name": client["name"],
        "unassigned": unassigned, "locations": locations, "suggestions": suggestions,
    })


@router.patch("/{client_id}/locations/assign", response_class=HTMLResponse)
def location_assign(
    request: Request,
    client_id: int,
    policy_uid: str = Form(...),
    project_id: int = Form(...),
    conn=Depends(get_db),
):
    """Assign a policy to a location/project."""
    project = conn.execute("SELECT name FROM projects WHERE id=?", (project_id,)).fetchone()
    if project:
        conn.execute(
            "UPDATE policies SET project_id=?, project_name=? WHERE policy_uid=? AND client_id=?",
            (project_id, project["name"], policy_uid, client_id),
        )
        loc = conn.execute(
            "SELECT address, city, state, zip FROM projects WHERE id=?",
            (project_id,),
        ).fetchone()
        if loc:
            conn.execute(
                """UPDATE policies SET exposure_address=?, exposure_city=?,
                   exposure_state=?, exposure_zip=?
                   WHERE policy_uid=? AND client_id=?""",
                (loc["address"] or "", loc["city"] or "", loc["state"] or "", loc["zip"] or "",
                 policy_uid, client_id),
            )
        conn.commit()
    return HTMLResponse("", headers={"HX-Trigger": "locationChanged"})


@router.patch("/{client_id}/locations/unassign", response_class=HTMLResponse)
def location_unassign(
    request: Request,
    client_id: int,
    policy_uid: str = Form(...),
    conn=Depends(get_db),
):
    """Remove a policy from its location/project."""
    conn.execute(
        "UPDATE policies SET project_id=NULL, project_name=NULL WHERE policy_uid=? AND client_id=?",
        (policy_uid, client_id),
    )
    conn.commit()
    return HTMLResponse("", headers={"HX-Trigger": "locationChanged"})


@router.patch("/{client_id}/locations/bulk-assign", response_class=HTMLResponse)
def location_bulk_assign(
    request: Request,
    client_id: int,
    address: str = Form(...),
    project_id: int = Form(...),
    conn=Depends(get_db),
):
    """Bulk-assign all unassigned policies sharing an exposure_address to a location."""
    project = conn.execute("SELECT name FROM projects WHERE id=?", (project_id,)).fetchone()
    if project:
        conn.execute(
            """UPDATE policies SET project_id=?, project_name=?
               WHERE client_id=? AND archived=0
               AND TRIM(exposure_address)=TRIM(?)
               AND (project_id IS NULL OR project_id=0)""",
            (project_id, project["name"], client_id, address),
        )
        loc = conn.execute(
            "SELECT address, city, state, zip FROM projects WHERE id=?",
            (project_id,),
        ).fetchone()
        if loc:
            conn.execute(
                """UPDATE policies SET exposure_address=?, exposure_city=?,
                   exposure_state=?, exposure_zip=?
                   WHERE project_id=? AND client_id=? AND archived=0""",
                (loc["address"] or "", loc["city"] or "", loc["state"] or "", loc["zip"] or "",
                 project_id, client_id),
            )
        conn.commit()
    return HTMLResponse("", headers={"HX-Trigger": "locationChanged"})


@router.post("/{client_id}/locations/create", response_class=HTMLResponse)
def location_create(
    request: Request,
    client_id: int,
    name: str = Form(...),
    address: str = Form(""),
    city: str = Form(""),
    state: str = Form(""),
    zip: str = Form(""),
    latitude: str = Form(""),
    longitude: str = Form(""),
    conn=Depends(get_db),
):
    """Create a new location/project for a client."""
    def _fl(v):
        try:
            return float(v) if str(v).strip() else None
        except ValueError:
            return None
    conn.execute(
        "INSERT INTO projects (name, client_id, address, city, state, zip, latitude, longitude, project_type) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Location')",
        (name, client_id, address, city, state, zip, _fl(latitude), _fl(longitude)),
    )
    conn.commit()
    return HTMLResponse("", headers={"HX-Trigger": "locationChanged"})


# ─── Bulk AI Import (policy data from messy spreadsheets) ────────────────────

# In-memory cache for bulk import review → apply
_BULK_IMPORT_CACHE: dict[str, tuple[list[dict], int, float]] = {}
_CLIENT_CONTACT_IMPORT_CACHE: dict[str, tuple[list[dict], int, float]] = {}


@router.get("/{client_id}/ai-bulk-import/prompt", response_class=HTMLResponse)
def client_ai_bulk_import_prompt(request: Request, client_id: int, conn=Depends(get_db)):
    """Generate the AI bulk import prompt pre-loaded with client context."""
    from policydb.llm_schemas import generate_policy_bulk_prompt, POLICY_BULK_IMPORT_SCHEMA, generate_json_template

    client_row = conn.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    if not client_row:
        return HTMLResponse("Client not found", status_code=404)
    client = dict(client_row)

    prompt_text = generate_policy_bulk_prompt(conn, client_id)
    json_template = generate_json_template(POLICY_BULK_IMPORT_SCHEMA)

    context_display = {"Client": client["name"]}
    if client["industry_segment"]:
        context_display["Industry"] = client["industry_segment"]
    pol_count = conn.execute(
        "SELECT COUNT(*) as c FROM policies WHERE client_id = ? AND archived = 0", (client_id,)
    ).fetchone()["c"]
    loc_count = conn.execute(
        "SELECT COUNT(*) as c FROM projects WHERE client_id = ? AND (project_type = 'Location' OR project_type IS NULL)", (client_id,)
    ).fetchone()["c"]
    context_display["Existing Policies"] = str(pol_count)
    context_display["Known Locations"] = str(loc_count)

    return templates.TemplateResponse("_ai_import_panel.html", {
        "request": request,
        "import_type": "bulk_policy",
        "prompt_text": prompt_text,
        "json_template": json_template,
        "context_display": context_display,
        "parse_url": f"/clients/{client_id}/ai-bulk-import/parse",
        "import_target": "#ai-bulk-import-results",
    })


@router.post("/{client_id}/ai-bulk-import/parse", response_class=HTMLResponse)
def client_ai_bulk_import_parse(
    request: Request,
    client_id: int,
    json_text: str = Form(...),
    conn=Depends(get_db),
):
    """Parse LLM JSON for bulk policy import. Returns review table."""
    from policydb.llm_schemas import parse_policy_bulk_json
    from policydb.utils import normalize_policy_number_for_matching

    result = parse_policy_bulk_json(json_text)
    if not result.get("ok"):
        return HTMLResponse(
            f'<div class="p-4 text-red-600 text-sm bg-red-50 rounded-lg">'
            f'<strong>Parse error:</strong> {result.get("error", "Unknown error")}</div>'
        )

    policies = result["policies"]
    warnings = result.get("warnings", [])

    # Match extracted policies against existing DB policies
    db_policies = conn.execute(
        """SELECT p.id, p.policy_uid, p.policy_number, p.policy_type, p.carrier,
                  p.effective_date, p.expiration_date, p.premium, p.limit_amount,
                  p.deductible, p.project_id, p.project_name,
                  c.name as client_name
           FROM policies p JOIN clients c ON p.client_id = c.id
           WHERE p.client_id = ? AND p.archived = 0""",
        (client_id,),
    ).fetchall()
    db_by_polnum: dict[str, dict] = {}
    for r in db_policies:
        pn = normalize_policy_number_for_matching(r["policy_number"] or "")
        if pn:
            db_by_polnum[pn] = dict(r)

    # Known locations for matching project_name
    locations = conn.execute(
        "SELECT id, name FROM projects WHERE client_id = ? AND (project_type = 'Location' OR project_type IS NULL)",
        (client_id,),
    ).fetchall()
    loc_by_name: dict[str, int] = {r["name"].lower().strip(): r["id"] for r in locations if r["name"]}

    enriched: list[dict] = []
    for pol in policies:
        entry = {
            "data": pol,
            "match": None,
            "match_type": "new",
            "location_id": None,
            "location_name": None,
            "has_layers": bool(pol.get("program_layers")),
            "layer_count": len(pol.get("program_layers", [])),
            "has_sub_coverages": bool(pol.get("sub_coverages")),
            "sub_coverage_count": len(pol.get("sub_coverages", [])),
        }
        pn = normalize_policy_number_for_matching(pol.get("policy_number") or "")
        if pn and pn in db_by_polnum:
            entry["match"] = db_by_polnum[pn]
            entry["match_type"] = "update"

        proj = (pol.get("project_name") or "").lower().strip()
        if proj and proj in loc_by_name:
            entry["location_id"] = loc_by_name[proj]
            entry["location_name"] = pol.get("project_name")

        enriched.append(entry)

    # Cache for apply step
    import time as _time
    import uuid as _uuid
    token = str(_uuid.uuid4())
    _BULK_IMPORT_CACHE[token] = (enriched, client_id, _time.time())

    new_count = sum(1 for e in enriched if e["match_type"] == "new")
    update_count = sum(1 for e in enriched if e["match_type"] == "update")
    program_count = sum(1 for e in enriched if e["has_layers"])
    located_count = sum(1 for e in enriched if e["location_id"])

    html_parts = ['<div class="space-y-4">']

    # Summary badges
    html_parts.append('<div class="flex flex-wrap gap-2 mb-4">')
    html_parts.append(f'<span class="px-2.5 py-1 rounded-full text-xs font-medium bg-green-50 text-green-700">{new_count} New</span>')
    html_parts.append(f'<span class="px-2.5 py-1 rounded-full text-xs font-medium bg-blue-50 text-blue-700">{update_count} Updates</span>')
    if program_count:
        html_parts.append(f'<span class="px-2.5 py-1 rounded-full text-xs font-medium bg-purple-50 text-purple-700">{program_count} Programs</span>')
    html_parts.append(f'<span class="px-2.5 py-1 rounded-full text-xs font-medium bg-amber-50 text-amber-700">{located_count}/{len(enriched)} Located</span>')
    html_parts.append('</div>')

    if warnings:
        html_parts.append('<div class="p-3 bg-amber-50 border border-amber-200 rounded-lg text-xs text-amber-700 mb-3">')
        html_parts.append(f'<strong>{len(warnings)} warnings:</strong><ul class="mt-1 list-disc pl-4">')
        for w in warnings[:10]:
            html_parts.append(f'<li>{w}</li>')
        if len(warnings) > 10:
            html_parts.append(f'<li>...and {len(warnings) - 10} more</li>')
        html_parts.append('</ul></div>')

    html_parts.append('<div class="overflow-x-auto">')
    html_parts.append('<table class="w-full text-xs">')
    html_parts.append('<thead><tr class="border-b border-gray-200 text-gray-500 text-left">')
    for h in ["Status", "Type", "Carrier", "Policy #", "Dates", "Premium", "Location", "Layers"]:
        html_parts.append(f'<th class="py-2 px-2">{h}</th>')
    html_parts.append('</tr></thead><tbody>')

    for entry in enriched:
        d = entry["data"]
        status_cls = "text-green-700 bg-green-50" if entry["match_type"] == "new" else "text-blue-700 bg-blue-50"
        status_label = "New" if entry["match_type"] == "new" else "Update"
        eff = d.get("effective_date", "")
        exp = d.get("expiration_date", "")
        dates = f"{eff} → {exp}" if eff or exp else ""
        premium = f"${d['premium']:,.0f}" if d.get("premium") else ""
        loc = entry["location_name"] or d.get("project_name") or ""
        loc_cls = "text-green-600" if entry["location_id"] else "text-gray-400"
        layers = str(entry["layer_count"]) if entry["has_layers"] else ""

        html_parts.append(f'<tr class="border-b border-gray-100 hover:bg-gray-50">')
        html_parts.append(f'<td class="py-1.5 px-2"><span class="px-1.5 py-0.5 rounded text-[10px] font-medium {status_cls}">{status_label}</span></td>')
        html_parts.append(f'<td class="py-1.5 px-2">{d.get("policy_type", "")}</td>')
        html_parts.append(f'<td class="py-1.5 px-2">{d.get("carrier", "")}</td>')
        html_parts.append(f'<td class="py-1.5 px-2 font-mono">{d.get("policy_number", "")}</td>')
        html_parts.append(f'<td class="py-1.5 px-2">{dates}</td>')
        html_parts.append(f'<td class="py-1.5 px-2">{premium}</td>')
        html_parts.append(f'<td class="py-1.5 px-2 {loc_cls}">{loc}</td>')
        html_parts.append(f'<td class="py-1.5 px-2 text-center">{layers}</td>')
        html_parts.append('</tr>')

    html_parts.append('</tbody></table></div>')

    html_parts.append(f'''
    <div class="flex items-center gap-3 pt-3 border-t border-gray-200">
      <button hx-post="/clients/{client_id}/ai-bulk-import/apply"
              hx-vals='{{"token": "{token}"}}'
              hx-target="#ai-bulk-import-results"
              hx-swap="innerHTML"
              class="bg-marsh hover:bg-marsh-light text-white font-medium text-sm px-5 py-2 rounded-lg transition-colors">
        Apply {len(enriched)} Policies
      </button>
      <span class="text-xs text-gray-400">{new_count} new + {update_count} updates</span>
    </div>
    ''')

    html_parts.append('</div>')
    return HTMLResponse("\n".join(html_parts))


@router.post("/{client_id}/ai-bulk-import/apply", response_class=HTMLResponse)
def client_ai_bulk_import_apply(
    request: Request,
    client_id: int,
    token: str = Form(""),
    conn=Depends(get_db),
):
    """Apply parsed bulk import: create new policies, update existing ones."""
    cache = _BULK_IMPORT_CACHE.get(token)
    if not cache:
        return HTMLResponse('<div class="p-4 text-red-600 text-sm">Session expired — please re-parse.</div>')

    enriched, cached_client_id, ts = cache
    if cached_client_id != client_id:
        return HTMLResponse('<div class="p-4 text-red-600 text-sm">Client mismatch.</div>')

    created = 0
    updated = 0
    programs_created = 0
    errors: list[str] = []

    for entry in enriched:
        d = entry["data"]
        try:
            if entry["match_type"] == "update" and entry["match"]:
                db_pol = entry["match"]
                updates = []
                params = []
                for field in ["policy_type", "carrier", "premium", "limit_amount", "deductible",
                              "effective_date", "expiration_date", "description", "layer_position",
                              "tower_group", "attachment_point", "participation_of",
                              "first_named_insured",
                              "underwriter_name", "placement_colleague", "exposure_address",
                              "coverage_form", "notes"]:
                    val = d.get(field)
                    if val is not None and str(val).strip():
                        updates.append(f"{field} = ?")
                        params.append(val)
                if entry["location_id"] and not db_pol.get("project_id"):
                    updates.append("project_id = ?")
                    params.append(entry["location_id"])
                    pname = entry["location_name"] or d.get("project_name") or ""
                    if pname:
                        updates.append("project_name = ?")
                        params.append(pname)

                if updates:
                    params.append(db_pol["id"])
                    conn.execute(
                        f"UPDATE policies SET {', '.join(updates)}, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                        params,
                    )
                    updated += 1
                # Create contact records for underwriter and placement colleague
                _upd_pol_id = db_pol["id"]
                _pc_name = (d.get("placement_colleague") or "").strip()
                if _pc_name:
                    _pc_cid = get_or_create_contact(conn, _pc_name)
                    assign_contact_to_policy(conn, _pc_cid, _upd_pol_id, is_placement_colleague=1)
                _uw_name = (d.get("underwriter_name") or "").strip()
                if _uw_name:
                    _uw_email = (d.get("underwriter_contact") or "").strip() or None
                    _uw_cid = get_or_create_contact(conn, _uw_name, email=_uw_email)
                    assign_contact_to_policy(conn, _uw_cid, _upd_pol_id, role="Underwriter")
            else:
                from policydb.db import next_policy_uid
                uid = next_policy_uid(conn)
                conn.execute(
                    """INSERT INTO policies (
                        policy_uid, client_id, policy_type, carrier, policy_number,
                        effective_date, expiration_date, premium, limit_amount, deductible,
                        description, layer_position, tower_group, attachment_point,
                        participation_of, first_named_insured, underwriter_name,
                        placement_colleague, exposure_address, coverage_form, notes,
                        project_id, project_name, renewal_status
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Not Started')""",
                    (
                        uid, client_id,
                        d.get("policy_type", ""), d.get("carrier", ""), d.get("policy_number", ""),
                        d.get("effective_date"), d.get("expiration_date"),
                        d.get("premium", 0) or 0, d.get("limit_amount", 0) or 0,
                        d.get("deductible", 0) or 0,
                        d.get("description", ""), d.get("layer_position", "Primary"),
                        d.get("tower_group", ""), d.get("attachment_point"),
                        d.get("participation_of"),
                        d.get("first_named_insured", ""), d.get("underwriter_name", ""),
                        d.get("placement_colleague", ""), d.get("exposure_address", ""),
                        d.get("coverage_form", ""), d.get("notes", ""),
                        entry["location_id"], entry.get("location_name") or d.get("project_name", ""),
                    ),
                )
                created += 1

                # Fetch policy ID once for all post-insert operations
                _new_pol_id = conn.execute("SELECT id FROM policies WHERE policy_uid = ?", (uid,)).fetchone()["id"]

                # Create contact records for underwriter and placement colleague
                _pc_name = (d.get("placement_colleague") or "").strip()
                if _pc_name:
                    _pc_cid = get_or_create_contact(conn, _pc_name)
                    assign_contact_to_policy(conn, _pc_cid, _new_pol_id, is_placement_colleague=1)
                _uw_name = (d.get("underwriter_name") or "").strip()
                if _uw_name:
                    _uw_email = (d.get("underwriter_contact") or "").strip() or None
                    _uw_cid = get_or_create_contact(conn, _uw_name, email=_uw_email)
                    assign_contact_to_policy(conn, _uw_cid, _new_pol_id, role="Underwriter")

                # Insert sub-coverages if present
                if d.get("sub_coverages"):
                    for j, sc in enumerate(d["sub_coverages"]):
                        conn.execute(
                            """INSERT OR IGNORE INTO policy_sub_coverages
                               (policy_id, coverage_type, limit_amount, deductible, notes, sort_order)
                               VALUES (?, ?, ?, ?, ?, ?)""",
                            (_new_pol_id, sc.get("coverage_type", ""),
                             sc.get("limit_amount"), sc.get("deductible"),
                             sc.get("notes", ""), j),
                        )

                if d.get("policy_number"):
                    try:
                        from policydb.match_memory import learn
                        learn(conn, _new_pol_id, "LLM Import", d["policy_number"], "policy_number", "llm")
                    except Exception:
                        pass

        except Exception as e:
            errors.append(f"Row {entry['data'].get('_index', '?')}: {str(e)}")

    conn.commit()
    _BULK_IMPORT_CACHE.pop(token, None)

    parts = ['<div class="p-4 space-y-3">']
    parts.append('<div class="flex items-center gap-2">')
    parts.append('<svg class="w-5 h-5 text-green-600" fill="none" stroke="currentColor" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M5 13l4 4L19 7"/></svg>')
    parts.append('<span class="text-sm font-medium text-gray-900">Import complete</span>')
    parts.append('</div>')
    parts.append('<div class="flex flex-wrap gap-2">')
    if created:
        parts.append(f'<span class="px-2.5 py-1 rounded-full text-xs font-medium bg-green-50 text-green-700">{created} created</span>')
    if updated:
        parts.append(f'<span class="px-2.5 py-1 rounded-full text-xs font-medium bg-blue-50 text-blue-700">{updated} updated</span>')
    if programs_created:
        parts.append(f'<span class="px-2.5 py-1 rounded-full text-xs font-medium bg-purple-50 text-purple-700">{programs_created} programs</span>')
    parts.append('</div>')
    if errors:
        parts.append(f'<div class="p-3 bg-red-50 border border-red-200 rounded-lg text-xs text-red-700">')
        parts.append(f'<strong>{len(errors)} errors:</strong><ul class="mt-1 list-disc pl-4">')
        for e in errors[:5]:
            parts.append(f'<li>{e}</li>')
        parts.append('</ul></div>')
    parts.append(f'<a href="/clients/{client_id}" class="text-sm text-marsh hover:underline">← Back to client</a>')
    parts.append('</div>')
    return HTMLResponse("\n".join(parts))


# ---------------------------------------------------------------------------
# AI Contact Import — client-level bulk contact import
# ---------------------------------------------------------------------------

@router.get("/{client_id}/ai-contact-import/prompt", response_class=HTMLResponse)
def client_ai_contact_import_prompt(
    request: Request, client_id: int, conn=Depends(get_db)
):
    """Return the AI import panel with contact bulk import prompt."""
    import json
    from policydb.llm_schemas import (
        CONTACT_BULK_IMPORT_SCHEMA,
        generate_contact_bulk_import_prompt,
    )

    client = conn.execute(
        "SELECT * FROM clients WHERE id = ?", (client_id,)
    ).fetchone()
    if not client:
        return HTMLResponse("Client not found", status_code=404)

    prompt_text = generate_contact_bulk_import_prompt(conn, client_id)

    # Build JSON template from schema examples
    example = {}
    for f in CONTACT_BULK_IMPORT_SCHEMA["fields"]:
        if f.get("example"):
            example[f["key"]] = f["example"]
    json_template = json.dumps([example], indent=2)

    context_display = {"Client": client["name"]}
    if client["industry_segment"]:
        context_display["Industry"] = client["industry_segment"]

    return templates.TemplateResponse("_ai_import_panel.html", {
        "request": request,
        "import_type": "client_contacts",
        "prompt_text": prompt_text,
        "json_template": json_template,
        "context_display": context_display,
        "parse_url": f"/clients/{client_id}/ai-contact-import/parse",
        "import_target": "#ai-contact-import-result",
    })


@router.post("/{client_id}/ai-contact-import/parse", response_class=HTMLResponse)
def client_ai_contact_import_parse(
    request: Request,
    client_id: int,
    json_text: str = Form(...),
    conn=Depends(get_db),
):
    """Parse LLM contact JSON and return review panel."""
    import time
    import uuid
    from policydb.llm_schemas import parse_contact_bulk_import_json

    result = parse_contact_bulk_import_json(json_text)
    if not result["ok"]:
        return HTMLResponse(
            f'<div class="p-4 bg-red-50 border border-red-200 rounded-lg text-red-700 text-sm">'
            f'{result["error"]}</div>',
            status_code=422,
        )

    client = conn.execute(
        "SELECT * FROM clients WHERE id = ?", (client_id,)
    ).fetchone()
    contacts = result["contacts"]
    warnings = result.get("warnings", [])

    # Fetch ALL existing client contacts across all types for dedup
    existing_names: set[str] = set()
    for ctype in ("client", "internal", "external"):
        rows = get_client_contacts(conn, client_id, contact_type=ctype)
        for r in rows:
            if r.get("name"):
                existing_names.add(r["name"].lower().strip())

    # Annotate contacts
    for contact in contacts:
        name_lower = contact["name"].lower().strip()
        contact["already_assigned"] = name_lower in existing_names

        existing = conn.execute(
            "SELECT id, email, phone, organization FROM contacts "
            "WHERE LOWER(TRIM(name)) = LOWER(TRIM(?))",
            (contact["name"],),
        ).fetchone()
        contact["existing_contact"] = dict(existing) if existing else None

        # Default contact_type
        if not contact.get("contact_type"):
            contact["contact_type"] = "client"

    # Cache for apply step
    token = str(uuid.uuid4())
    _CLIENT_CONTACT_IMPORT_CACHE[token] = (
        contacts,
        client_id,
        time.time(),
    )

    # Purge stale cache entries (>30 min)
    now = time.time()
    stale = [k for k, v in _CLIENT_CONTACT_IMPORT_CACHE.items() if now - v[2] > 1800]
    for k in stale:
        _CLIENT_CONTACT_IMPORT_CACHE.pop(k, None)

    return templates.TemplateResponse("clients/_ai_contacts_review.html", {
        "request": request,
        "client": dict(client),
        "contacts": contacts,
        "warnings": warnings,
        "token": token,
        "client_id": client_id,
        "contact_roles": cfg.get("contact_roles", []),
    })


@router.post("/{client_id}/ai-contact-import/apply", response_class=HTMLResponse)
async def client_ai_contact_import_apply(
    request: Request,
    client_id: int,
    conn=Depends(get_db),
):
    """Apply selected contacts from AI import to the client."""
    form = await request.form()
    token = form.get("token", "")

    cache = _CLIENT_CONTACT_IMPORT_CACHE.get(token)
    if not cache:
        return HTMLResponse(
            '<div class="p-4 text-red-600 text-sm">Session expired — please re-parse.</div>'
        )

    contacts, cached_client_id, ts = cache
    if cached_client_id != client_id:
        return HTMLResponse(
            '<div class="p-4 text-red-600 text-sm">Client mismatch.</div>'
        )

    created = 0
    updated = 0
    errors: list[str] = []

    for i, contact in enumerate(contacts):
        if not form.get(f"select_{i}"):
            continue

        # Read form overrides (user may have edited in review step)
        name = form.get(f"name_{i}", contact.get("name", "")).strip()
        email = form.get(f"email_{i}", contact.get("email", "")).strip()
        phone = form.get(f"phone_{i}", contact.get("phone", "")).strip()
        mobile = form.get(f"mobile_{i}", contact.get("mobile", "")).strip()
        org = form.get(f"org_{i}", contact.get("organization", "")).strip()
        title = form.get(f"title_{i}", contact.get("title", "")).strip()
        role = form.get(f"role_{i}", contact.get("role", ""))
        contact_type = form.get(f"type_{i}", contact.get("contact_type", "client"))

        if not name:
            continue

        try:
            # Normalize phone/email (both return plain strings)
            if email:
                email = clean_email(email)
            if phone:
                phone = format_phone(phone)
            if mobile:
                mobile = format_phone(mobile)

            cid = get_or_create_contact(
                conn,
                name,
                email=email or None,
                phone=phone or None,
                mobile=mobile or None,
                organization=org or None,
            )

            assign_contact_to_client(
                conn,
                cid,
                client_id,
                contact_type=contact_type,
                role=role,
                title=title,
            )

            if contact.get("existing_contact"):
                updated += 1
            else:
                created += 1
        except Exception as e:
            errors.append(f"{name}: {e}")

    conn.commit()
    _CLIENT_CONTACT_IMPORT_CACHE.pop(token, None)

    total = created + updated
    parts = [
        '<div class="p-4 space-y-3">',
        '<div class="flex items-center gap-2">',
        '<svg class="w-5 h-5 text-green-600" fill="none" viewBox="0 0 24 24" stroke="currentColor">',
        '<path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M5 13l4 4L19 7"/>',
        "</svg>",
        f'<span class="text-sm font-medium text-gray-900">'
        f'{total} contact{"s" if total != 1 else ""} imported</span>',
        "</div>",
        '<div class="flex gap-2">',
    ]
    if created:
        parts.append(
            f'<span class="px-2 py-0.5 rounded-full text-xs bg-green-50 text-green-700">'
            f"{created} created</span>"
        )
    if updated:
        parts.append(
            f'<span class="px-2 py-0.5 rounded-full text-xs bg-blue-50 text-blue-700">'
            f"{updated} updated</span>"
        )
    if errors:
        parts.append(
            f'<span class="px-2 py-0.5 rounded-full text-xs bg-red-50 text-red-700">'
            f"{len(errors)} error(s)</span>"
        )
    parts.append("</div>")
    if errors:
        parts.append('<div class="text-xs text-red-600 mt-1">')
        for e in errors:
            parts.append(f"<p>{e}</p>")
        parts.append("</div>")
    parts.append("</div>")
    response = HTMLResponse("\n".join(parts))
    response.headers["HX-Trigger"] = "refreshContacts"
    return response


# ─── EXPOSURE TRACKING ───────────────────────────────────────────────────────

def _exposure_tab_context(conn, client_id: int, year: int, project_id=None) -> dict:
    """Build shared context for the exposures tab/matrix."""
    from datetime import date as _date
    current_year = _date.today().year
    data_years = get_exposure_years(conn, client_id, project_id)
    # Include last 10 years so users can enter historical exposure data
    historical_range = list(range(current_year, current_year - 11, -1))
    available_years = sorted(set(data_years + historical_range), reverse=True)
    if year not in available_years:
        available_years = sorted(set(available_years + [year]), reverse=True)
    exposures = get_client_exposures(conn, client_id, year, project_id)
    observations = get_exposure_observations(conn, client_id, year, project_id)
    prior_year_has_data = (year - 1) in data_years

    # Build policy options for the combo dropdown
    policies = [dict(p) for p in get_policies_for_client(conn, client_id)]
    policy_options = [
        {"value": str(p["id"]), "label": f"{p['policy_type']} — {p.get('carrier') or '?'}"}
        for p in policies if not p.get("is_opportunity")
    ]

    # Annotate exposures with policy labels and link data
    policy_map = {str(p["id"]): f"{p['policy_type']} — {p.get('carrier') or '?'}" for p in policies}
    for e in exposures:
        e["policy_label"] = policy_map.get(str(e.get("policy_id"))) if e.get("policy_id") else None
        # Attach junction-table link data (rate, primary, linked policy)
        link = conn.execute(
            """SELECT pel.rate, pel.is_primary, pel.policy_uid, p.policy_type, p.carrier
               FROM policy_exposure_links pel
               JOIN policies p ON p.policy_uid = pel.policy_uid
               WHERE pel.exposure_id=?
               ORDER BY pel.is_primary DESC LIMIT 1""",
            (e["id"],),
        ).fetchone()
        e["link_rate"] = link["rate"] if link else None
        e["link_is_primary"] = link["is_primary"] if link else None
        e["link_policy_uid"] = link["policy_uid"] if link else None

    denom_options = cfg.get("exposure_denominators", [1, 100, 1000])

    # Build URL prefix for template links (corporate vs project-level)
    if project_id:
        exposure_url_prefix = f"/clients/{client_id}/projects/{project_id}"
        tab_reload_url = f"/clients/{client_id}/projects/{project_id}/exposures"
    else:
        exposure_url_prefix = f"/clients/{client_id}"
        tab_reload_url = f"/clients/{client_id}/tab/exposures"

    return {
        "client_id": client_id,
        "project_id": project_id,
        "selected_year": year,
        "available_years": available_years,
        "exposures": exposures,
        "observations": observations,
        "prior_year_has_data": prior_year_has_data,
        "policy_options": policy_options,
        "denom_options": denom_options,
        "exposure_url_prefix": exposure_url_prefix,
        "tab_reload_url": tab_reload_url,
    }


@router.get("/{client_id}/tab/exposures", response_class=HTMLResponse)
def client_tab_exposures(request: Request, client_id: int, year: int = 0, conn=Depends(get_db)):
    """Load exposures tab content (HTMX partial)."""
    from datetime import date as _date
    if year == 0:
        year = _date.today().year
    client = get_client_by_id(conn, client_id, include_archived=True)
    if not client:
        return HTMLResponse("Not found", status_code=404)
    ctx = _exposure_tab_context(conn, client_id, year)
    return templates.TemplateResponse("clients/_tab_exposures.html", {"request": request, **ctx})


def _build_exposure_types_dropdown(client_id: int, year: int, project_id, conn) -> str:
    """Build the exposure type picker dropdown HTML (shared by client + project routes)."""
    url_prefix = f"/clients/{client_id}/projects/{project_id}" if project_id else f"/clients/{client_id}"
    standard = cfg.get("standard_exposure_types", {})
    custom_types = get_distinct_custom_exposure_types(conn)

    existing = get_client_exposures(conn, client_id, year, project_id)
    existing_types = {e["exposure_type"] for e in existing}

    parts = ['<div class="text-[10px] text-gray-400 uppercase tracking-wide px-3 py-1.5 border-b border-gray-100">Standard</div>']
    for name, unit in standard.items():
        disabled = "opacity-40 pointer-events-none" if name in existing_types else "hover:bg-gray-50 cursor-pointer"
        parts.append(
            f'<div class="px-3 py-2 text-sm {disabled}" '
            f'hx-post="{url_prefix}/exposures/add-row" '
            f'hx-vals=\'{{"exposure_type":"{name}","year":"{year}","is_custom":"0","unit":"{unit}"}}\' '
            f'hx-target="#exposures-card" hx-swap="outerHTML">'
            f'{name}</div>'
        )
    if custom_types:
        parts.append('<div class="text-[10px] text-gray-400 uppercase tracking-wide px-3 py-1.5 border-t border-b border-gray-100">Previously Used</div>')
        for ct in custom_types:
            if ct in standard:
                continue
            disabled = "opacity-40 pointer-events-none" if ct in existing_types else "hover:bg-gray-50 cursor-pointer"
            parts.append(
                f'<div class="px-3 py-2 text-sm text-purple-600 {disabled}" '
                f'hx-post="{url_prefix}/exposures/add-row" '
                f'hx-vals=\'{{"exposure_type":"{ct}","year":"{year}","is_custom":"1","unit":"number"}}\' '
                f'hx-target="#exposures-card" hx-swap="outerHTML">'
                f'{ct}</div>'
            )
    parts.append(
        '<div class="border-t border-gray-200 px-3 py-2 hover:bg-gray-50 cursor-pointer">'
        f'<input type="text" id="custom-exposure-input" placeholder="Custom type name..." '
        f'class="w-full text-sm border-0 outline-none bg-transparent" '
        f"onkeydown=\"if(event.key==='Enter'){{var v=this.value.trim();if(v)htmx.ajax('POST','{url_prefix}/exposures/add-row',"
        f"{{values:{{exposure_type:v,year:'{year}',is_custom:'1',unit:'number'}},target:'#exposures-card',swap:'outerHTML'}});}}\">"
        '</div>'
    )
    return "\n".join(parts)


@router.get("/{client_id}/exposures/types", response_class=HTMLResponse)
def exposure_types_dropdown(request: Request, client_id: int, year: int = 0, project_id: int = 0, conn=Depends(get_db)):
    """Return dropdown HTML for the exposure type picker."""
    from datetime import date as _date
    if year == 0:
        year = _date.today().year
    proj = project_id if project_id else None
    return HTMLResponse(_build_exposure_types_dropdown(client_id, year, proj, conn))


def _exposure_add_row_handler(request, client_id, form, conn, project_id=None):
    """Shared handler for adding an exposure row (corporate + project)."""
    exposure_type = form.get("exposure_type", "").strip()
    year = int(form.get("year", 0))
    is_custom = int(form.get("is_custom", 0))
    unit = form.get("unit", "number")
    if not exposure_type or not year:
        return HTMLResponse("Missing fields", status_code=400)
    try:
        conn.execute(
            "INSERT INTO client_exposures (client_id, project_id, exposure_type, is_custom, unit, year) VALUES (?, ?, ?, ?, ?, ?)",
            (client_id, project_id, exposure_type, is_custom, unit, year),
        )
        conn.commit()
    except Exception:
        pass  # UNIQUE violation = already exists, just refresh
    ctx = _exposure_tab_context(conn, client_id, year, project_id)
    return templates.TemplateResponse("clients/_exposure_matrix.html", {"request": request, **ctx})


@router.post("/{client_id}/exposures/add-row", response_class=HTMLResponse)
async def exposure_add_row(request: Request, client_id: int, conn=Depends(get_db)):
    """Create a new exposure row and return refreshed matrix."""
    form = await request.form()
    return _exposure_add_row_handler(request, client_id, form, conn)


@router.patch("/{client_id}/exposures/{exposure_id}/cell")
async def exposure_cell(request: Request, client_id: int, exposure_id: int, conn=Depends(get_db)):
    """Save a single cell value for an exposure row."""
    body = await request.json()
    field, value = body.get("field", ""), body.get("value", "")
    allowed = {"amount", "source_document", "notes", "policy_id", "denominator", "exposure_type"}
    if field not in allowed:
        return JSONResponse({"ok": False, "error": "Invalid field"}, status_code=400)

    formatted = value.strip() if isinstance(value, str) else value

    if field == "exposure_type":
        # Only allow editing custom exposure types
        row = conn.execute("SELECT is_custom FROM client_exposures WHERE id=? AND client_id=?", (exposure_id, client_id)).fetchone()
        if not row or not row["is_custom"]:
            return JSONResponse({"ok": False, "error": "Cannot rename standard exposure types"}, status_code=400)
        if not formatted:
            return JSONResponse({"ok": False, "error": "Exposure type cannot be empty"}, status_code=400)
        conn.execute(
            "UPDATE client_exposures SET exposure_type=?, updated_at=CURRENT_TIMESTAMP WHERE id=? AND client_id=?",
            (formatted, exposure_id, client_id),
        )
        conn.commit()
        return JSONResponse({"ok": True, "formatted": formatted})

    if field == "amount":
        # Strip currency symbols and commas, parse to float
        cleaned = str(formatted).replace("$", "").replace(",", "").strip()
        if cleaned == "" or cleaned == "—":
            conn.execute(
                "UPDATE client_exposures SET amount=NULL, updated_at=CURRENT_TIMESTAMP WHERE id=? AND client_id=?",
                (exposure_id, client_id),
            )
            conn.commit()
            return JSONResponse({"ok": True, "formatted": "", "yoy": None, "yoy_direction": None})
        try:
            amount = float(cleaned)
        except ValueError:
            return JSONResponse({"ok": False, "error": "Invalid number"}, status_code=400)
        conn.execute(
            "UPDATE client_exposures SET amount=?, updated_at=CURRENT_TIMESTAMP WHERE id=? AND client_id=?",
            (amount, exposure_id, client_id),
        )
        conn.commit()
        from policydb.exposures import recalc_exposure_rate
        recalc_exposure_rate(conn, exposure_id=exposure_id)
        # Format and calculate YoY
        row = get_exposure_by_id(conn, exposure_id)
        if row and row.get("unit") == "currency":
            formatted = "${:,.0f}".format(amount)
        else:
            formatted = "{:,.0f}".format(amount)
        # Calculate YoY
        prior = conn.execute(
            """SELECT amount FROM client_exposures
               WHERE client_id=? AND exposure_type=? AND year=?
               AND COALESCE(project_id,0)=COALESCE(?,0)""",
            (client_id, row["exposure_type"], row["year"] - 1, row.get("project_id")),
        ).fetchone()
        yoy = None
        yoy_direction = None
        if prior and prior["amount"] and prior["amount"] != 0:
            pct = ((amount - prior["amount"]) / prior["amount"]) * 100
            yoy = "{:+.1f}%".format(pct)
            yoy_direction = "up" if pct > 0 else "down"
        # Include rate data in response
        link = conn.execute(
            "SELECT rate, is_primary FROM policy_exposure_links WHERE exposure_id=?",
            (exposure_id,),
        ).fetchone()
        resp = {"ok": True, "formatted": formatted, "yoy": yoy, "yoy_direction": yoy_direction}
        if link:
            resp["rate"] = link["rate"]
            resp["is_primary"] = link["is_primary"]
        return JSONResponse(resp)
    elif field == "denominator":
        denom = int(formatted) if formatted and formatted not in ("", "—") else 1
        if denom <= 0:
            denom = 1
        conn.execute(
            "UPDATE client_exposures SET denominator=?, updated_at=CURRENT_TIMESTAMP WHERE id=? AND client_id=?",
            (denom, exposure_id, client_id),
        )
        conn.commit()
        from policydb.exposures import recalc_exposure_rate
        recalc_exposure_rate(conn, exposure_id=exposure_id)
        link = conn.execute(
            "SELECT rate, is_primary FROM policy_exposure_links WHERE exposure_id=?",
            (exposure_id,),
        ).fetchone()
        return JSONResponse({
            "ok": True, "formatted": str(denom),
            "rate": link["rate"] if link else None,
            "is_primary": link["is_primary"] if link else None,
        })
    elif field == "policy_id":
        from policydb.exposures import create_exposure_link, delete_exposure_link
        old_link = conn.execute(
            "SELECT policy_uid FROM policy_exposure_links WHERE exposure_id=?",
            (exposure_id,),
        ).fetchone()
        pid = int(formatted) if formatted and formatted not in ("", "—", "0") else None
        if old_link:
            delete_exposure_link(conn, old_link["policy_uid"], exposure_id)
        if pid:
            pol = conn.execute("SELECT policy_uid, policy_type, carrier FROM policies WHERE id=?", (pid,)).fetchone()
            if pol:
                link = create_exposure_link(conn, pol["policy_uid"], exposure_id, is_primary=True)
                label = f"{pol['policy_type']} — {pol['carrier'] or '?'}"
                return JSONResponse({
                    "ok": True, "formatted": label,
                    "rate": link.get("rate"),
                    "is_primary": link.get("is_primary"),
                })
        return JSONResponse({"ok": True, "formatted": ""})
    else:
        conn.execute(
            f"UPDATE client_exposures SET {field}=?, updated_at=CURRENT_TIMESTAMP WHERE id=? AND client_id=?",
            (formatted or None, exposure_id, client_id),
        )
        conn.commit()
        return JSONResponse({"ok": True, "formatted": formatted})


def _render_exposure_row(request, conn, client_id, exposure_id, project_id=None):
    """Render a single exposure matrix row partial."""
    e = dict(conn.execute("SELECT * FROM client_exposures WHERE id=?", (exposure_id,)).fetchone())
    # Attach policy label
    policies = [dict(p) for p in get_policies_for_client(conn, client_id)]
    policy_map = {str(p["id"]): f"{p['policy_type']} — {p.get('carrier') or '?'}" for p in policies}
    e["policy_label"] = policy_map.get(str(e.get("policy_id"))) if e.get("policy_id") else None
    # Attach link data
    link = conn.execute(
        """SELECT pel.rate, pel.is_primary, pel.policy_uid, p.policy_type, p.carrier
           FROM policy_exposure_links pel
           JOIN policies p ON p.policy_uid = pel.policy_uid
           WHERE pel.exposure_id=?
           ORDER BY pel.is_primary DESC LIMIT 1""",
        (exposure_id,),
    ).fetchone()
    e["link_rate"] = link["rate"] if link else None
    e["link_is_primary"] = link["is_primary"] if link else None
    e["link_policy_uid"] = link["policy_uid"] if link else None
    # Attach prior year amount
    prior = conn.execute(
        """SELECT amount FROM client_exposures
           WHERE client_id=? AND exposure_type=? AND year=?
           AND COALESCE(project_id,0)=COALESCE(?,0)""",
        (client_id, e["exposure_type"], e["year"] - 1, e.get("project_id")),
    ).fetchone()
    e["prior_amount"] = prior["amount"] if prior else None
    # Build context
    policy_options = [
        {"value": str(p["id"]), "label": f"{p['policy_type']} — {p.get('carrier') or '?'}"}
        for p in policies if not p.get("is_opportunity")
    ]
    if project_id:
        exposure_url_prefix = f"/clients/{client_id}/projects/{project_id}"
    else:
        exposure_url_prefix = f"/clients/{client_id}"
    denom_options = cfg.get("exposure_denominators", [1, 100, 1000])
    return templates.TemplateResponse("clients/_exposure_matrix_row.html", {
        "request": request, "e": e,
        "policy_options": policy_options,
        "denom_options": denom_options,
        "exposure_url_prefix": exposure_url_prefix,
        "client_id": client_id,
    })


@router.patch("/{client_id}/exposures/{exposure_id}/toggle-primary")
async def exposure_toggle_primary(request: Request, client_id: int, exposure_id: int, conn=Depends(get_db), project_id: int = None):
    """Toggle primary status for a policy-exposure link."""
    from policydb.exposures import set_primary_exposure
    body = await request.form()
    policy_uid = body.get("policy_uid", "")
    if not policy_uid:
        return HTMLResponse("")
    link = conn.execute(
        "SELECT is_primary FROM policy_exposure_links WHERE policy_uid=? AND exposure_id=?",
        (policy_uid, exposure_id),
    ).fetchone()
    if not link:
        return HTMLResponse("")
    if link["is_primary"]:
        conn.execute(
            "UPDATE policy_exposure_links SET is_primary=0 WHERE policy_uid=? AND exposure_id=?",
            (policy_uid, exposure_id),
        )
        conn.commit()
    else:
        set_primary_exposure(conn, policy_uid, exposure_id)
    return _render_exposure_row(request, conn, client_id, exposure_id, project_id=project_id)


@router.delete("/{client_id}/exposures/{exposure_id}", response_class=HTMLResponse)
def exposure_delete(request: Request, client_id: int, exposure_id: int, conn=Depends(get_db)):
    """Delete an exposure row."""
    conn.execute("DELETE FROM client_exposures WHERE id=? AND client_id=?", (exposure_id, client_id))
    conn.commit()
    return HTMLResponse("")


def _exposure_copy_forward_handler(request, client_id, form, conn, project_id=None):
    """Shared handler for copy-forward (corporate + project)."""
    from_year = int(form.get("from_year", 0))
    to_year = int(form.get("to_year", 0))
    if not from_year or not to_year:
        return HTMLResponse("Missing year parameters", status_code=400)

    if project_id:
        source_rows = conn.execute(
            "SELECT exposure_type, is_custom, unit FROM client_exposures WHERE client_id=? AND year=? AND project_id=?",
            (client_id, from_year, project_id),
        ).fetchall()
    else:
        source_rows = conn.execute(
            "SELECT exposure_type, is_custom, unit FROM client_exposures WHERE client_id=? AND year=? AND project_id IS NULL",
            (client_id, from_year),
        ).fetchall()

    if not source_rows:
        ctx = _exposure_tab_context(conn, client_id, to_year, project_id)
        return templates.TemplateResponse("clients/_exposure_matrix.html", {"request": request, **ctx})

    for row in source_rows:
        try:
            conn.execute(
                "INSERT OR IGNORE INTO client_exposures (client_id, project_id, exposure_type, is_custom, unit, year) VALUES (?, ?, ?, ?, ?, ?)",
                (client_id, project_id, row["exposure_type"], row["is_custom"], row["unit"], to_year),
            )
        except Exception:
            pass
    conn.commit()

    ctx = _exposure_tab_context(conn, client_id, to_year, project_id)
    return templates.TemplateResponse("clients/_exposure_matrix.html", {"request": request, **ctx})


@router.post("/{client_id}/exposures/copy-forward", response_class=HTMLResponse)
async def exposure_copy_forward(request: Request, client_id: int, conn=Depends(get_db)):
    """Copy exposure types from one year to another (INSERT OR IGNORE)."""
    form = await request.form()
    return _exposure_copy_forward_handler(request, client_id, form, conn)


@router.get("/{client_id}/edit-followup-slideover", response_class=HTMLResponse)
def client_edit_followup_slideover(client_id: int, request: Request, conn=Depends(get_db)):
    """Return the edit slideover partial for a client follow-up."""
    row = conn.execute(
        "SELECT id, name, follow_up_date, notes FROM clients WHERE id = ?",
        (client_id,),
    ).fetchone()
    if not row:
        return HTMLResponse("<p class='p-4 text-sm text-gray-400'>Not found.</p>", status_code=404)
    return templates.TemplateResponse("action_center/_edit_client_slideover.html", {
        "request": request,
        "c": dict(row),
    })


@router.patch("/{client_id}/followup-field")
def patch_client_followup_field(client_id: int, body: dict = None, conn=Depends(get_db)):
    """Update follow_up_date or notes on a client (slideover inline edit)."""
    if not body:
        return JSONResponse({"ok": False, "error": "No body"}, status_code=400)
    field = body.get("field", "")
    value = body.get("value", "")
    allowed = {"follow_up_date", "notes"}
    if field not in allowed:
        return JSONResponse({"ok": False, "error": f"Field '{field}' not editable"}, status_code=400)
    conn.execute(f"UPDATE clients SET {field} = ? WHERE id = ?", (value or None, client_id))
    conn.commit()
    return {"ok": True, "formatted": value}
