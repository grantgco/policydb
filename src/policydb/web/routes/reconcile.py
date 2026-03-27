"""Reconciliation routes — compare an uploaded CSV/XLSX against PolicyDB policies."""

from __future__ import annotations

import logging
logger = logging.getLogger("policydb.web.routes.reconcile")

import io
from datetime import date

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, Response

from policydb import config as cfg
from policydb.queries import get_client_by_name
from policydb.utils import (
    normalize_carrier, normalize_coverage_type, normalize_policy_number,
    parse_currency_with_magnitude,
    _COVERAGE_ALIASES, _BASE_COVERAGE_ALIASES, _BASE_CARRIER_ALIASES,
    rebuild_coverage_aliases, rebuild_carrier_aliases,
)
from policydb.reconciler import (
    ReconcileRow,
    _build_reconcile_row,
    _find_likely_pairs,
    _score_pair,
    build_reconcile_xlsx,
    parse_uploaded_file,
    program_reconcile_summary,
    reconcile,
    summarize,
)
from policydb.web.app import get_db, templates

router = APIRouter(prefix="/reconcile")

# In-memory cache for last reconciliation (single-user local app)
import time as _time
import uuid as _uuid
_RESULT_CACHE: dict[str, tuple[bytes, float]] = {}  # token → (xlsx_bytes, timestamp)
_MISSING_CACHE: dict[str, tuple[list[dict], float]] = {}  # token → (missing_ext_rows, timestamp)
_LAST_MISSING_TOKEN: str = ""  # most recent token for batch-create fallback

# Board state cache: token → (results list, extras list, db_rows, timestamp)
_BOARD_CACHE: dict[str, tuple[list, list, list, float]] = {}

# Pre-match validation cache: token → (parsed_rows, warnings, upload_params, timestamp)
# upload_params stores client_id, scope, date_priority, filename, source_name so the match can run later
_PARSED_CACHE: dict[str, tuple[list[dict], list[str], dict, float]] = {}

# Source name per board token (for match memory learning on confirm/create)
_SOURCE_NAME_CACHE: dict[str, str] = {}

# Import session ID per board token (for completing session with stats)
_SESSION_ID_CACHE: dict[str, int] = {}

def _cache_cleanup():
    """Remove cache entries older than 1 hour."""
    cutoff = _time.time() - 3600
    for k in list(_RESULT_CACHE):
        if _RESULT_CACHE[k][1] < cutoff:
            del _RESULT_CACHE[k]
    for k in list(_MISSING_CACHE):
        if _MISSING_CACHE[k][1] < cutoff:
            del _MISSING_CACHE[k]
    for k in list(_BOARD_CACHE):
        if _BOARD_CACHE[k][3] < cutoff:
            del _BOARD_CACHE[k]
    for k in list(_PARSED_CACHE):
        if _PARSED_CACHE[k][3] < cutoff:
            del _PARSED_CACHE[k]


def _get_program_summary(token: str) -> dict:
    """Reconstruct program_summary from cached board state."""
    cache = _BOARD_CACHE.get(token)
    if not cache:
        return {}
    results, extras, db_rows, ts = cache
    return program_reconcile_summary(results + extras)


def _board_context(request, token: str, results: list, extras: list, conn=None) -> dict:
    """Build the shared template context for rendering _pairing_board.html."""
    summary = summarize(results + extras)
    all_clients = []
    if conn is not None:
        all_clients = [dict(c) for c in conn.execute(
            "SELECT id, name FROM clients WHERE archived=0 ORDER BY name"
        ).fetchall()]
    return {
        "request": request,
        "results": results,
        "extras": extras,
        "token": token,
        "summary": summary,
        "today": date.today().isoformat(),
        "field_display": _FIELD_DISPLAY,
        "policy_types": cfg.get("policy_types", []),
        "program_summary": _get_program_summary(token),
        "program_candidates": _detect_program_candidates(results),
        "all_clients": all_clients,
    }


def _render_counters(summary: dict) -> str:
    """Build OOB counter HTML matching the board-counters div in _pairing_board.html."""
    return (
        '<div id="board-counters" hx-swap-oob="true" class="flex flex-wrap items-center gap-3 mb-4">'
        '  <span class="flex items-center gap-1.5 text-sm text-gray-600">'
        '    <span class="w-2.5 h-2.5 rounded-full bg-green-500 inline-block"></span>'
        f'    Paired <span class="font-bold">{summary["paired_clean"]}</span>'
        '  </span>'
        '  <span class="flex items-center gap-1.5 text-sm text-gray-600">'
        '    <span class="w-2.5 h-2.5 rounded-full bg-amber-500 inline-block"></span>'
        f'    Review <span class="font-bold">{summary["paired_diffs"]}</span>'
        '  </span>'
        '  <span class="flex items-center gap-1.5 text-sm text-gray-600">'
        '    <span class="w-2.5 h-2.5 rounded-full bg-red-500 inline-block"></span>'
        f'    Unmatched <span class="font-bold">{summary["unmatched"]}</span>'
        '  </span>'
        '  <span class="flex items-center gap-1.5 text-sm text-gray-600">'
        '    <span class="w-2.5 h-2.5 rounded-full bg-purple-500 inline-block"></span>'
        f'    Extra <span class="font-bold">{summary["extra"]}</span>'
        '  </span>'
        '  <span class="text-gray-300 mx-1">|</span>'
        f'  <span class="text-xs text-gray-400">{summary["total"]} total rows</span>'
        '</div>'
    )


def _detect_program_candidates(results: list) -> list[dict]:
    """Detect groups of unmatched rows that look like a program.

    Detection strategies:
    1. Same client + type + dates, 2+ different carriers (existing)
    2. Tower pattern: same client + type, different limits suggesting layers
    3. Shared policy number prefix (e.g. PROG-GL-001A, PROG-GL-001B)

    Returns a list of dicts, each with:
      client, type, effective_date, expiration_date, carriers (list),
      row_indices (list of int), carrier_details (list of dicts with premium/limit/policy_number),
      detection_method (str)
    """
    from collections import defaultdict

    # ── Strategy 1: Same client + type + dates, 2+ carriers ──
    # Use normalized coverage type so "General Liability" and "Commercial General Liability"
    # are recognized as the same type and group into one program candidate.
    from policydb.utils import normalize_coverage_type as _nct
    from rapidfuzz import fuzz as _fuzz

    groups: dict[tuple, list[tuple[int, str, dict]]] = defaultdict(list)
    for i, r in enumerate(results):
        if r.status not in ("UNMATCHED", "MISSING"):
            continue
        e = r.ext or {}
        client = e.get("client_name", "")
        ptype = _nct(e.get("policy_type", ""))  # normalize
        eff = e.get("effective_date", "")
        exp = e.get("expiration_date", "")
        carrier = e.get("carrier", "")
        if client and ptype and carrier:
            key = (client, ptype, eff, exp)
            groups[key].append((i, carrier, e))

    # Merge groups with fuzzy-matching type names (e.g. "General Liability" vs "Comm General Liability")
    merged_keys = list(groups.keys())
    merge_map: dict[tuple, tuple] = {}  # maps secondary key → primary key
    for ki in range(len(merged_keys)):
        if merged_keys[ki] in merge_map:
            continue
        for kj in range(ki + 1, len(merged_keys)):
            if merged_keys[kj] in merge_map:
                continue
            a_client, a_type, a_eff, a_exp = merged_keys[ki]
            b_client, b_type, b_eff, b_exp = merged_keys[kj]
            if a_client == b_client and a_eff == b_eff and a_exp == b_exp:
                if _fuzz.WRatio(a_type, b_type) >= 80:
                    merge_map[merged_keys[kj]] = merged_keys[ki]
    # Apply merges
    for secondary, primary in merge_map.items():
        groups[primary].extend(groups.pop(secondary))

    candidates = []
    claimed_indices: set[int] = set()

    for key, entries in groups.items():
        unique_carriers = list(dict.fromkeys(c for _, c, _ in entries))
        if len(unique_carriers) >= 2:
            indices = [idx for idx, _, _ in entries]
            claimed_indices.update(indices)
            candidates.append({
                "client": key[0],
                "type": key[1],
                "effective_date": key[2],
                "expiration_date": key[3],
                "carriers": unique_carriers,
                "row_indices": indices,
                "carrier_details": [
                    {
                        "carrier": c,
                        "premium": e.get("premium", ""),
                        "limit_amount": e.get("limit_amount", ""),
                        "policy_number": e.get("policy_number", ""),
                    }
                    for _, c, e in entries
                ],
                "detection_method": "same_dates_multiple_carriers",
            })

    # ── Strategy 2: Shared policy number prefix (3+ chars before final segment) ──
    import re
    prefix_groups: dict[tuple, list[tuple[int, str, dict]]] = defaultdict(list)
    for i, r in enumerate(results):
        if i in claimed_indices:
            continue
        if r.status not in ("UNMATCHED", "MISSING"):
            continue
        e = r.ext or {}
        pn = (e.get("policy_number") or "").strip()
        client = e.get("client_name", "")
        ptype = e.get("policy_type", "")
        carrier = e.get("carrier", "")
        if pn and client and ptype:
            # Extract prefix: everything before the last segment (letter, digit suffix)
            # e.g. "PROG-GL-001A" → "PROG-GL-001", "TC-GL-2026-001" → "TC-GL-2026"
            m = re.match(r'^(.{4,}?)[- ]?([A-Za-z]|\d{1,2})$', pn)
            if m:
                prefix = m.group(1).rstrip("-_ ")
                key = (client, ptype, prefix)
                prefix_groups[key].append((i, carrier, e))

    for key, entries in prefix_groups.items():
        if len(entries) >= 2:
            indices = [idx for idx, _, _ in entries]
            if any(idx in claimed_indices for idx in indices):
                continue
            claimed_indices.update(indices)
            # Use dates from first entry
            first_ext = entries[0][2]
            candidates.append({
                "client": key[0],
                "type": key[1],
                "effective_date": first_ext.get("effective_date", ""),
                "expiration_date": first_ext.get("expiration_date", ""),
                "carriers": list(dict.fromkeys(c for _, c, _ in entries)),
                "row_indices": indices,
                "carrier_details": [
                    {
                        "carrier": c,
                        "premium": e.get("premium", ""),
                        "limit_amount": e.get("limit_amount", ""),
                        "policy_number": e.get("policy_number", ""),
                    }
                    for _, c, e in entries
                ],
                "detection_method": "shared_policy_prefix",
            })

    return candidates


def _load_db_policies(conn, client_id: int, scope: str) -> list[dict]:
    """Query PolicyDB policies for the reconcile comparison side."""
    conditions = ["p.archived = 0"]
    params: list = []

    if client_id > 0:
        conditions.append("p.client_id = ?")
        params.append(client_id)

    if scope == "active":
        conditions.append("p.expiration_date >= date('now', '-365 days')")

    where = " AND ".join(conditions)
    rows = conn.execute(
        f"""SELECT p.id, p.policy_uid, c.name AS client_name, p.policy_type, p.carrier,
                   p.policy_number, p.effective_date, p.expiration_date,
                   p.premium, p.limit_amount, p.deductible, p.client_id,
                   p.first_named_insured, p.placement_colleague, p.underwriter_name,
                   p.exposure_address, p.project_name, p.project_id,
                   p.is_program, p.program_id,
                   pr.name AS location_name,
                   prog.policy_uid AS program_uid
            FROM policies p
            JOIN clients c ON p.client_id = c.id
            LEFT JOIN projects pr ON p.project_id = pr.id
            LEFT JOIN policies prog ON p.program_id = prog.id
            WHERE {where}
            ORDER BY c.name, p.expiration_date""",
        params,
    ).fetchall()
    result = [dict(r) for r in rows]
    # Ensure project_name is populated from location join if not set directly
    for r in result:
        if not r.get("project_name") and r.get("location_name"):
            r["project_name"] = r["location_name"]
    return result


@router.post("/preview-columns")
async def reconcile_preview_columns(file: UploadFile = File(...)):
    """Return column headers and sample data from an uploaded file for column mapping UI."""
    from fastapi.responses import JSONResponse
    content = await file.read()
    headers = []
    sample_rows: list[list[str]] = []  # first 3 data rows for preview
    filename = file.filename or ""
    if filename.lower().endswith(('.xlsx', '.xls')) or content[:2] == b'PK':
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            first_row = next(rows_iter, None)
            if first_row:
                headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(first_row)]
                for _, data_row in zip(range(3), rows_iter):
                    sample_rows.append([str(v).strip() if v is not None else "" for v in data_row])
            wb.close()
        except Exception:
            pass
    else:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        import csv as _csv
        reader = _csv.reader(io.StringIO(text))
        first_row = next(reader, None)
        if first_row:
            headers = [h.strip() for h in first_row]
            for _, data_row in zip(range(3), reader):
                sample_rows.append([v.strip() for v in data_row])
    return JSONResponse({"headers": headers, "sample_rows": sample_rows})


@router.get("", response_class=HTMLResponse)
def reconcile_index(request: Request, conn=Depends(get_db)):
    all_clients = conn.execute(
        "SELECT id, name FROM clients WHERE archived=0 ORDER BY name"
    ).fetchall()
    # Source names for match memory: config defaults + any learned sources
    source_names = list(cfg.get("import_source_names", []))
    try:
        learned = conn.execute(
            "SELECT DISTINCT source_name FROM import_match_memory ORDER BY source_name"
        ).fetchall()
        for r in learned:
            if r["source_name"] not in source_names:
                source_names.append(r["source_name"])
    except Exception:
        pass  # table may not exist yet during migration
    return templates.TemplateResponse("reconcile/index.html", {
        "request": request,
        "active": "reconcile",
        "all_clients": [dict(c) for c in all_clients],
        "results": None,
        "summary": None,
        "pairs": [],
        "warnings": [],
        "errors": [],
        "selected_client_id": 0,
        "selected_scope": "active",
        "selected_source_name": "",
        "source_names": source_names,
        "filename": "",
        "run_date": "",
    })


@router.get("/reference-guide", response_class=HTMLResponse)
def reconcile_reference_guide(request: Request):
    """Printable data prep reference — coverage types, aliases, column headers."""
    from policydb.utils import _COVERAGE_ALIASES
    from policydb.importer import PolicyImporter
    import policydb.config as cfg

    # Group coverage aliases by canonical name
    coverage_groups: dict[str, list[str]] = {}
    for alias, canonical in _COVERAGE_ALIASES.items():
        if canonical not in coverage_groups:
            coverage_groups[canonical] = []
        if alias != canonical.lower():
            coverage_groups[canonical].append(alias)
    # Sort canonicals alphabetically
    coverage_types = sorted(coverage_groups.items())

    # Carrier aliases from config
    carrier_aliases = cfg.get("carrier_aliases", {})

    # Column header aliases from importer
    column_aliases: dict[str, list[str]] = {}
    for alias, canonical in PolicyImporter.ALIASES.items():
        if canonical not in column_aliases:
            column_aliases[canonical] = []
        column_aliases[canonical].append(alias)
    column_groups = sorted(column_aliases.items())

    return templates.TemplateResponse("reconcile/_reference_guide.html", {
        "request": request,
        "active": "reconcile",
        "coverage_types": coverage_types,
        "carrier_aliases": sorted(carrier_aliases.items()) if isinstance(carrier_aliases, dict) else [],
        "column_groups": column_groups,
    })


@router.get("/template-csv/{template_type}")
def reconcile_template_csv(template_type: str):
    """Download a CSV template for reconcile import."""
    import csv as _csv

    if template_type == "standard":
        headers = ["client_name", "policy_type", "carrier", "policy_number",
                    "effective_date", "expiration_date", "premium"]
        example = ["Acme Construction", "General Liability", "Hartford",
                    "GL-2025-001", "04/01/2025", "04/01/2026", "12500"]
    else:  # full
        headers = ["client_name", "policy_type", "carrier", "policy_number",
                    "effective_date", "expiration_date", "premium", "limit_amount",
                    "deductible", "program", "layer_position", "first_named_insured",
                    "placement_colleague", "underwriter_name"]
        example = ["Acme Construction", "General Liability", "Hartford",
                    "GL-2025-001", "04/01/2025", "04/01/2026", "12500", "1000000",
                    "5000", "", "Primary", "", "", ""]

    output = io.StringIO()
    writer = _csv.writer(output)
    writer.writerow(headers)
    writer.writerow(example)

    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=policydb-{template_type}-template.csv"},
    )


@router.post("", response_class=HTMLResponse)
async def reconcile_run(
    request: Request,
    file: UploadFile = File(...),
    client_id: int = Form(0),
    scope: str = Form("active"),
    column_mapping_json: str = Form(""),
    date_priority: str = Form(""),
    source_name: str = Form(""),
    as_of_date: str = Form(""),
    conn=Depends(get_db),
):
    all_clients = conn.execute(
        "SELECT id, name FROM clients WHERE archived=0 ORDER BY name"
    ).fetchall()
    # Source names for match memory datalist
    _source_names = list(cfg.get("import_source_names", []))
    try:
        _learned_sources = conn.execute(
            "SELECT DISTINCT source_name FROM import_match_memory ORDER BY source_name"
        ).fetchall()
        for r in _learned_sources:
            if r["source_name"] not in _source_names:
                _source_names.append(r["source_name"])
    except Exception:
        pass

    ctx = {
        "request": request,
        "active": "reconcile",
        "all_clients": [dict(c) for c in all_clients],
        "selected_client_id": client_id,
        "selected_scope": scope,
        "selected_source_name": source_name,
        "source_names": _source_names,
        "filename": file.filename or "",
        "run_date": date.today().isoformat(),
        "results": None,
        "summary": None,
        "warnings": [],
        "errors": [],
    }

    import json as _json
    col_map = {}
    if column_mapping_json:
        try:
            col_map = {k: v for k, v in _json.loads(column_mapping_json).items() if v and v != "ignore"}
        except Exception:
            pass

    content = await file.read()
    if not content:
        ctx["errors"] = ["Uploaded file is empty."]
        return templates.TemplateResponse("reconcile/index.html", ctx)

    # Check for duplicate file upload
    dup_warning = ""
    if source_name:
        try:
            from policydb.import_ledger import check_duplicate_file
            dup = check_duplicate_file(conn, content, source_name)
            if dup:
                dup_date = dup.get("imported_at", "unknown")[:10]
                dup_warning = f"This file was previously uploaded on {dup_date} ({dup.get('status', '')}). Re-processing."
        except Exception:
            pass

    # Load saved column mapping from source profile if user didn't provide one
    if not col_map and source_name:
        try:
            from policydb.import_ledger import get_saved_column_map
            saved_map = get_saved_column_map(conn, source_name)
            if saved_map:
                col_map = saved_map
                ctx["saved_column_map"] = saved_map
        except Exception:
            pass

    ext_rows, warnings = parse_uploaded_file(content, column_mapping=col_map or None, filename=file.filename or "")
    if dup_warning:
        warnings.insert(0, dup_warning)
    ctx["warnings"] = warnings
    ctx["column_mapping_json"] = column_mapping_json

    if not ext_rows:
        ctx["errors"] = ["No usable rows found in the uploaded file. Ensure it has client_name (or insured) and policy columns."]
        return templates.TemplateResponse("reconcile/index.html", ctx)

    # Cache parsed rows for validation panel and subsequent match
    _cache_cleanup()
    token = str(_uuid.uuid4())
    upload_params = {
        "client_id": client_id,
        "scope": scope,
        "date_priority": date_priority,
        "filename": file.filename or "",
        "column_mapping_json": column_mapping_json,
        "source_name": source_name,
        "as_of_date": as_of_date,
        "file_content": content,  # kept for session creation in run-match
        "column_map_used": col_map,  # actual mapping applied (may be from saved profile)
    }
    _PARSED_CACHE[token] = (ext_rows, warnings, upload_params, _time.time())

    # Build validation data and show validation panel
    return _build_validation_response(request, token, ext_rows, warnings, conn, ctx)


def _build_validation_response(request, token, parsed_rows, warnings, conn, ctx):
    """Build and return the validation panel response within the index page."""
    from rapidfuzz import fuzz
    from collections import defaultdict

    # Validate coverage types
    coverage_results = []
    for raw_type in sorted(set(r.get("policy_type", "") for r in parsed_rows if r.get("policy_type"))):
        normalized = normalize_coverage_type(raw_type)
        is_alias = raw_type.strip().lower() in _COVERAGE_ALIASES
        coverage_results.append({
            "raw": raw_type, "normalized": normalized,
            "recognized": is_alias or raw_type.strip().lower() == normalized.lower(),
            "count": sum(1 for r in parsed_rows if r.get("policy_type") == raw_type),
        })

    # Validate carriers
    carrier_results = []
    for raw_carrier in sorted(set(r.get("carrier", "") for r in parsed_rows if r.get("carrier"))):
        normalized = normalize_carrier(raw_carrier)
        recognized = normalized != raw_carrier.strip()
        carrier_results.append({
            "raw": raw_carrier, "normalized": normalized,
            "recognized": recognized or raw_carrier.strip().lower() == normalized.lower(),
            "count": sum(1 for r in parsed_rows if r.get("carrier") == raw_carrier),
        })

    # Validate dates
    dates_parsed = sum(1 for r in parsed_rows if r.get("effective_date") or r.get("expiration_date"))
    dates_total = len(parsed_rows)
    all_dates = [r.get("effective_date") or r.get("expiration_date")
                 for r in parsed_rows
                 if r.get("effective_date") or r.get("expiration_date")]
    date_range = (min(all_dates), max(all_dates)) if all_dates else ("", "")

    # Client names — fuzzy match to DB
    client_matches = []
    unique_clients = sorted(set(r.get("client_name", "") for r in parsed_rows if r.get("client_name")))
    db_clients = conn.execute("SELECT id, name FROM clients WHERE archived=0 ORDER BY name").fetchall()
    for name in unique_clients:
        best_match = max(db_clients, key=lambda c: fuzz.WRatio(name, c["name"]), default=None) if db_clients else None
        score = fuzz.WRatio(name, best_match["name"]) if best_match else 0
        client_matches.append({
            "raw": name, "db_name": best_match["name"] if best_match else "",
            "db_id": best_match["id"] if best_match else 0, "score": round(score),
            "count": sum(1 for r in parsed_rows if r.get("client_name") == name),
        })

    # Policy numbers
    polnums_present = sum(1 for r in parsed_rows if r.get("policy_number", "").strip())

    # Program auto-detection: group by client + type + dates with different carriers
    programs = []
    groups = defaultdict(set)
    for r in parsed_rows:
        key = (r.get("client_name", ""), r.get("policy_type", ""),
               r.get("effective_date", ""), r.get("expiration_date", ""))
        if r.get("carrier"):
            groups[key].add(r.get("carrier"))
    for key, carriers in groups.items():
        if len(carriers) >= 2:
            programs.append({
                "client": key[0], "type": key[1],
                "dates": f"{key[2]} - {key[3]}",
                "carriers": sorted(carriers),
            })

    # Check for location columns
    has_location = any(
        r.get("location") or r.get("address") or r.get("project_name") or r.get("project")
        for r in parsed_rows
    )

    # Get known policy types and carriers for the "map to" dropdowns
    from policydb import config as cfg
    known_policy_types = cfg.get("policy_types", [])
    known_carriers = cfg.get("carriers", [])

    ctx["validation_panel"] = True
    ctx["token"] = token
    ctx["row_count"] = len(parsed_rows)
    ctx["coverage_results"] = coverage_results
    ctx["carrier_results"] = carrier_results
    ctx["dates_parsed"] = dates_parsed
    ctx["dates_total"] = dates_total
    ctx["date_range"] = date_range
    ctx["client_matches"] = client_matches
    ctx["polnums_present"] = polnums_present
    ctx["programs"] = programs
    ctx["has_location"] = has_location
    ctx["known_policy_types"] = known_policy_types
    ctx["known_carriers"] = known_carriers

    return templates.TemplateResponse("reconcile/index.html", ctx)


@router.post("/run-match", response_class=HTMLResponse)
def reconcile_run_match(
    request: Request,
    token: str = Form(""),
    conn=Depends(get_db),
):
    """Run the actual reconcile match using cached parsed rows from the validation step."""
    cache = _PARSED_CACHE.get(token)
    if not cache:
        return HTMLResponse("<div class='p-4 text-red-600'>Session expired — please re-upload.</div>")

    ext_rows, warnings, upload_params, _ = cache
    client_id = upload_params.get("client_id", 0)
    scope = upload_params.get("scope", "active")
    date_priority = upload_params.get("date_priority", "")
    filename = upload_params.get("filename", "")
    source_name = upload_params.get("source_name", "")

    all_clients = conn.execute(
        "SELECT id, name FROM clients WHERE archived=0 ORDER BY name"
    ).fetchall()
    # Source names for template datalist
    _source_names = list(cfg.get("import_source_names", []))
    try:
        _learned_sources = conn.execute(
            "SELECT DISTINCT source_name FROM import_match_memory ORDER BY source_name"
        ).fetchall()
        for r in _learned_sources:
            if r["source_name"] not in _source_names:
                _source_names.append(r["source_name"])
    except Exception:
        pass

    ctx = {
        "request": request,
        "active": "reconcile",
        "all_clients": [dict(c) for c in all_clients],
        "selected_client_id": client_id,
        "selected_scope": scope,
        "selected_source_name": source_name,
        "source_names": _source_names,
        "filename": filename,
        "run_date": date.today().isoformat(),
        "results": None,
        "summary": None,
        "warnings": warnings,
        "errors": [],
    }

    db_rows = _load_db_policies(conn, client_id, scope)

    all_results = reconcile(ext_rows, db_rows, date_priority=bool(date_priority),
                            single_client=bool(client_id), conn=conn, source_name=source_name)
    logger.info("Reconcile match started for %d rows", len(ext_rows))

    missing_rows = [r for r in all_results if r.status in ("MISSING", "UNMATCHED")]
    extra_rows = [r for r in all_results if r.status == "EXTRA"]

    # Separate into board results (non-EXTRA) and extras for the pairing board cache
    board_results = [r for r in all_results if r.status != "EXTRA"]

    # Generate download token
    download_token = str(_uuid.uuid4())

    # Cache MISSING rows for batch create (keyed by download token)
    global _LAST_MISSING_TOKEN
    _MISSING_CACHE[download_token] = ([r.ext for r in missing_rows if r.ext], _time.time())
    _LAST_MISSING_TOKEN = download_token

    # Cache XLSX for download-without-reupload
    xlsx_bytes = build_reconcile_xlsx(all_results, run_date=date.today().isoformat(), filename=filename)
    _RESULT_CACHE[download_token] = (xlsx_bytes, _time.time())

    # Cache board state for interactive pairing operations
    _BOARD_CACHE[download_token] = (board_results, extra_rows, db_rows, _time.time())

    # Cache source name for match memory learning on confirm/create
    if source_name:
        _SOURCE_NAME_CACHE[download_token] = source_name

    # Log match memory stats
    memory_count = sum(1 for r in all_results if r.match_method == "memory")
    if memory_count:
        logger.info("Match memory: %d auto-matches from source '%s'", memory_count, source_name)

    # Create import session and save source profile
    session_id = None
    if source_name:
        try:
            from policydb.import_ledger import create_session, save_source_profile
            as_of_date = upload_params.get("as_of_date", "")
            file_content = upload_params.get("file_content")
            col_map_used = upload_params.get("column_map_used", {})

            session_id = create_session(
                conn, source_name=source_name, source_type="csv",
                file_name=filename, file_content=file_content,
                as_of_date=as_of_date, client_id=client_id if client_id else None,
                column_mapping=col_map_used,
            )
            # Save/update source profile with column mapping
            if col_map_used:
                save_source_profile(conn, source_name, column_map=col_map_used)

            # Store session_id in source name cache for later completion
            _SESSION_ID_CACHE[download_token] = session_id
        except Exception:
            logger.exception("Failed to create import session")

    # Clean up parsed cache — no longer needed (drop file_content to free memory)
    _PARSED_CACHE.pop(token, None)

    # Pass all results to index.html (it filters extras via selectattr in template)
    ctx["results"] = all_results
    ctx["extras"] = extra_rows
    ctx["summary"] = summarize(all_results)
    # Build suggested pairs and annotate with board_results index for manual-pair
    suggested_pairs = _find_likely_pairs(missing_rows, extra_rows)
    # Map each unmatched ReconcileRow to its index in board_results
    _unmatched_idx_map = {}
    for bi, br in enumerate(board_results):
        if br.status in ("MISSING", "UNMATCHED") and br.ext:
            _unmatched_idx_map[id(br)] = bi
    for sp in suggested_pairs:
        sp["missing_idx"] = _unmatched_idx_map.get(id(sp["missing"]), -1)
    ctx["pairs"] = suggested_pairs
    ctx["download_token"] = download_token
    ctx["today"] = date.today().isoformat()
    ctx["policy_types"] = cfg.get("policy_types", [])
    ctx["field_display"] = _FIELD_DISPLAY
    ctx["program_summary"] = program_reconcile_summary(all_results, carrier_map=_carrier_map)
    ctx["program_candidates"] = _detect_program_candidates(board_results)
    return templates.TemplateResponse("reconcile/index.html", ctx)


# ── Auto-Learn Alias Endpoints ─────────────────────────────────────────────


@router.post("/learn-coverage-alias", response_class=HTMLResponse)
def learn_coverage_alias(raw: str = Form(...), canonical: str = Form(...)):
    """Save a new coverage type alias to config."""
    from policydb import config as cfg
    aliases = cfg.get("coverage_aliases", {})
    if canonical not in aliases:
        aliases[canonical] = []
    if raw.strip().lower() not in [a.lower() for a in aliases[canonical]]:
        aliases[canonical].append(raw.strip())
    full = dict(cfg.load_config())
    full["coverage_aliases"] = aliases
    cfg.save_config(full)
    cfg.reload_config()
    rebuild_coverage_aliases()
    return HTMLResponse(
        f'<span class="text-[10px] px-1.5 py-0.5 rounded bg-green-100 text-green-700">'
        f'{raw} &rarr; {canonical} (saved)</span>'
    )


@router.post("/learn-carrier-alias", response_class=HTMLResponse)
def learn_carrier_alias(raw: str = Form(...), canonical: str = Form(...)):
    """Save a new carrier alias to config."""
    from policydb import config as cfg
    aliases = cfg.get("carrier_aliases", {})
    if canonical not in aliases:
        aliases[canonical] = []
    if raw.strip().lower() not in [a.lower() for a in aliases[canonical]]:
        aliases[canonical].append(raw.strip())
    full = dict(cfg.load_config())
    full["carrier_aliases"] = aliases
    cfg.save_config(full)
    cfg.reload_config()
    rebuild_carrier_aliases()
    return HTMLResponse(
        f'<span class="text-[10px] px-1.5 py-0.5 rounded bg-green-100 text-green-700">'
        f'{raw} &rarr; {canonical} (saved)</span>'
    )


# ── Pairing Board Endpoints ─────────────────────────────────────────────────


@router.get("/search-coverage", response_class=HTMLResponse)
def reconcile_search_coverage(
    request: Request,
    q: str = "",
    idx: int = 0,
    token: str = "",
    client_name: str = "",
    client_id: int = 0,
    conn=Depends(get_db),
):
    """Search DB policies for manual pairing. Filtered to client if client_id specified."""
    if not q or len(q) < 2:
        return HTMLResponse('<p class="text-xs text-gray-400 p-2">Type at least 2 characters...</p>')

    like = f"%{q}%"
    if client_id:
        rows = conn.execute("""
            SELECT p.id, p.policy_uid, c.name AS client_name, p.policy_type, p.carrier,
                   p.policy_number, p.effective_date, p.expiration_date, p.premium
            FROM policies p JOIN clients c ON p.client_id = c.id
            WHERE p.archived = 0 AND p.client_id = ?
              AND (p.policy_type LIKE ? OR p.carrier LIKE ? OR p.policy_number LIKE ?)
            ORDER BY p.policy_type, p.carrier LIMIT 15
        """, (client_id, like, like, like)).fetchall()
    else:
        rows = conn.execute("""
            SELECT p.id, p.policy_uid, c.name AS client_name, p.policy_type, p.carrier,
                   p.policy_number, p.effective_date, p.expiration_date, p.premium
            FROM policies p JOIN clients c ON p.client_id = c.id
            WHERE p.archived = 0
              AND (p.policy_type LIKE ? OR p.carrier LIKE ? OR p.policy_number LIKE ?
                   OR c.name LIKE ?)
            ORDER BY c.name, p.policy_type LIMIT 15
        """, (like, like, like, like)).fetchall()

    if not rows:
        return HTMLResponse('<p class="text-xs text-gray-500 p-2">No policies found matching that search.</p>')

    html = '<div class="space-y-1 max-h-48 overflow-y-auto p-2">'
    for r in rows:
        premium_str = f"${r['premium']:,.0f}" if r['premium'] else "—"
        html += (
            f'<div class="flex items-center justify-between text-xs bg-white border border-gray-200 rounded px-2 py-1.5 hover:bg-gray-50">'
            f'<div class="min-w-0 flex-1">'
            f'<span class="font-medium text-gray-800">{r["policy_type"]}</span>'
            f' <span class="text-gray-400">&middot;</span> '
            f'<span class="text-gray-600">{r["carrier"] or "—"}</span>'
            f' <span class="text-gray-400">&middot;</span> '
            f'<span class="text-gray-500">{r["client_name"]}</span>'
            f'<br><span class="text-gray-400">{r["policy_number"] or "—"} &middot; '
            f'{r["effective_date"] or "—"} &rarr; {r["expiration_date"] or "—"} &middot; {premium_str}</span>'
            f'</div>'
            f'<button type="button" '
            f'hx-post="/reconcile/manual-pair" '
            f'hx-vals=\'{{"idx": "{idx}", "policy_uid": "{r["policy_uid"]}", "token": "{token}"}}\' '
            f'hx-target="#pair-{idx}" hx-swap="outerHTML" '
            f'class="text-xs bg-green-600 text-white px-2 py-0.5 rounded ml-2 flex-shrink-0">Pair</button>'
            f'</div>'
        )
    html += '</div>'
    return HTMLResponse(html)


def _auto_learn_aliases(row: ReconcileRow) -> list[str]:
    """Auto-learn coverage/carrier aliases from a confirmed pair.

    Returns a list of human-readable descriptions of what was learned
    (e.g., ['"work comp" → Workers Compensation']).
    Skips aliases already in the hardcoded base.
    """
    learned: list[str] = []

    # Coverage alias
    if row.coverage_alias_applied and row.ext_type_raw and row.ext_type_normalized:
        raw_key = row.ext_type_raw.strip().lower()
        if raw_key not in _BASE_COVERAGE_ALIASES:
            aliases = cfg.get("coverage_aliases", {})
            canonical = row.ext_type_normalized
            if canonical not in aliases:
                aliases[canonical] = []
            if raw_key not in [a.lower() for a in aliases[canonical]]:
                aliases[canonical].append(row.ext_type_raw.strip())
                full = dict(cfg.load_config())
                full["coverage_aliases"] = aliases
                cfg.save_config(full)
                cfg.reload_config()
                rebuild_coverage_aliases()
                learned.append(f'"{row.ext_type_raw.strip()}" &rarr; {canonical}')
                logger.info("Auto-learned coverage alias: %s → %s", row.ext_type_raw.strip(), canonical)

    # Carrier alias
    if row.carrier_alias_applied and row.ext_carrier_raw and row.ext_carrier_normalized:
        raw_key = row.ext_carrier_raw.strip().lower()
        if raw_key not in _BASE_CARRIER_ALIASES:
            aliases = cfg.get("carrier_aliases", {})
            canonical = row.ext_carrier_normalized
            if canonical not in aliases:
                aliases[canonical] = []
            if raw_key not in [a.lower() for a in aliases[canonical]]:
                aliases[canonical].append(row.ext_carrier_raw.strip())
                full = dict(cfg.load_config())
                full["carrier_aliases"] = aliases
                cfg.save_config(full)
                cfg.reload_config()
                rebuild_carrier_aliases()
                learned.append(f'"{row.ext_carrier_raw.strip()}" &rarr; {canonical}')
                logger.info("Auto-learned carrier alias: %s → %s", row.ext_carrier_raw.strip(), canonical)

    return learned


def _render_learn_toast(learned: list[str]) -> str:
    """Render OOB toast HTML for auto-learned aliases."""
    if not learned:
        return ""
    pills = " ".join(
        f'<span class="inline-block px-2 py-0.5 rounded bg-blue-100 text-blue-700 text-[10px]">{item}</span>'
        for item in learned
    )
    return (
        f'<div id="learn-toast" hx-swap-oob="innerHTML">'
        f'<div class="flex items-center gap-2 px-3 py-2 mb-2 rounded-lg bg-blue-50 border border-blue-200 text-xs text-blue-800 '
        f'transition-opacity duration-1000" '
        f'x-data x-init="setTimeout(() => $el.style.opacity=\'0\', 3000); setTimeout(() => $el.remove(), 4000)">'
        f'<span class="font-semibold">Learned:</span> {pills}'
        f'</div></div>'
    )


@router.post("/confirm/{idx}", response_class=HTMLResponse)
def reconcile_confirm(request: Request, idx: int, token: str = Form(""), conn=Depends(get_db)):
    """Mark a paired row as confirmed. Returns updated pair row + OOB counters.

    Auto-learns coverage/carrier aliases if the pair used normalization.
    Auto-learns match memory if a source_name was provided.
    """
    cache = _BOARD_CACHE.get(token)
    if not cache:
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Session expired. Please re-run reconciliation.</div>')
    results, extras, db_rows, _ = cache
    if idx < 0 or idx >= len(results):
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Invalid row index.</div>')
    results[idx].confirmed = True
    logger.info("Reconcile pair confirmed: idx=%d", idx)

    # Auto-learn aliases
    learned = _auto_learn_aliases(results[idx])
    toast_html = _render_learn_toast(learned)

    # Auto-learn match memory (cross-source identity)
    source_name = _SOURCE_NAME_CACHE.get(token, "")
    row = results[idx]
    if source_name and row.ext and row.db and row.db.get("id"):
        try:
            from policydb.match_memory import learn_from_reconcile_pair
            count = learn_from_reconcile_pair(conn, row.db["id"], source_name, row.ext)
            if count:
                logger.info("Match memory: learned %d identities for policy_id=%d from '%s'",
                            count, row.db["id"], source_name)
        except Exception:
            logger.exception("Match memory learning failed on confirm")

    summary = summarize(results + extras)
    row_html = templates.TemplateResponse("reconcile/_pair_row.html", {
        "request": request,
        "row": results[idx],
        "idx": idx,
        "token": token,
        "field_display": _FIELD_DISPLAY,
        "policy_types": cfg.get("policy_types", []),
    }).body.decode()
    counter_html = _render_counters(summary)
    return HTMLResponse(row_html + counter_html + toast_html)


@router.post("/break/{idx}", response_class=HTMLResponse)
def reconcile_break(request: Request, idx: int, token: str = Form("")):
    """Break a pair: move DB policy to extras, convert row to unmatched."""
    cache = _BOARD_CACHE.get(token)
    if not cache:
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Session expired. Please re-run reconciliation.</div>')
    results, extras, db_rows, ts = cache
    if idx < 0 or idx >= len(results):
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Invalid row index.</div>')
    row = results[idx]
    # Move DB side to extras pool
    new_extra_html = ""
    extras_existed_before = len(extras) > 0
    if row.db:
        extra_row = ReconcileRow(ext=None, db=row.db, status="EXTRA")
        extras.append(extra_row)
        extra_row_html = templates.TemplateResponse("reconcile/_extra_row.html", {
            "request": request,
            "row": extra_row,
            "token": token,
            "today": date.today().isoformat(),
        }).body.decode()
        if extras_existed_before:
            # Append to existing extras pool
            new_extra_html = (
                '<div id="extras-pool" hx-swap-oob="beforeend">'
                + extra_row_html
                + '</div>'
            )
        else:
            # Create the extras pool from scratch (it wasn't rendered initially)
            new_extra_html = (
                '<div id="extras-pool" hx-swap-oob="afterend:#board-rows" class="mt-6">'
                '  <div class="flex items-center gap-3 mb-3">'
                '    <h3 class="text-sm font-semibold text-purple-700">In Coverage, Not in Upload (1)</h3>'
                '    <span class="text-xs text-gray-400">Drag a row above to pair it with an unmatched upload row</span>'
                '  </div>'
                + extra_row_html
                + '</div>'
            )
    # Convert to unmatched
    row.db = None
    row.status = "UNMATCHED"
    row.match_score = 0
    row.confidence = "none"
    row.confirmed = False
    row.diff_fields = []
    row.cosmetic_diffs = []
    row.fillable_fields = []
    row.score_policy_number = 0.0
    row.score_dates = 0.0
    row.score_type = 0.0
    row.score_carrier = 0.0
    row.score_name = 0.0
    row.match_method = ""
    # Return unmatched row + OOB new extra + OOB counters
    summary = summarize(results + extras)
    unmatched_html = templates.TemplateResponse("reconcile/_unmatched_row.html", {
        "request": request,
        "row": row,
        "idx": idx,
        "token": token,
    }).body.decode()
    counter_html = _render_counters(summary)
    return HTMLResponse(unmatched_html + new_extra_html + counter_html)


@router.post("/manual-pair", response_class=HTMLResponse)
def reconcile_manual_pair(
    request: Request,
    idx: int = Form(...),
    policy_uid: str = Form(...),
    token: str = Form(""),
    conn=Depends(get_db),
):
    """Manual pair: pair an unmatched row with a specific DB policy (from drag or search)."""
    cache = _BOARD_CACHE.get(token)
    if not cache:
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Session expired. Please re-run reconciliation.</div>')
    results, extras, db_rows, ts = cache
    if idx < 0 or idx >= len(results):
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Invalid row index.</div>')

    row = results[idx]
    if not row.ext:
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Row has no upload data.</div>')

    # Find the DB policy — check extras first, then db_rows, then query DB
    target_db = None
    extra_idx_to_remove = None
    for ei, er in enumerate(extras):
        if er.db and er.db.get("policy_uid") == policy_uid.upper():
            target_db = er.db
            extra_idx_to_remove = ei
            break
    if target_db is None:
        for db in db_rows:
            if db.get("policy_uid") == policy_uid.upper():
                target_db = db
                break
    if target_db is None:
        # Fallback: query DB directly
        db_row_raw = conn.execute(
            """SELECT p.policy_uid, c.name AS client_name, p.policy_type, p.carrier,
                      p.policy_number, p.effective_date, p.expiration_date,
                      p.premium, p.limit_amount, p.deductible, p.client_id
               FROM policies p JOIN clients c ON p.client_id = c.id
               WHERE p.policy_uid = ?""",
            (policy_uid.upper(),),
        ).fetchone()
        if db_row_raw:
            target_db = dict(db_row_raw)
    if target_db is None:
        return HTMLResponse(f'<div class="text-xs text-red-500 p-2">Policy {policy_uid} not found.</div>')

    # Score the pair
    breakdown = _score_pair(row.ext, target_db)

    # Update the row in place
    row.db = target_db
    row.status = "PAIRED"
    row.match_score = breakdown.total
    row.confidence = breakdown.confidence
    row.match_method = "manual"
    row.confirmed = False
    row.score_policy_number = breakdown.score_policy_number
    row.score_dates = breakdown.score_dates
    row.score_type = breakdown.score_type
    row.score_carrier = breakdown.score_carrier
    row.score_name = breakdown.score_name
    row.diff_fields = list(breakdown.diff_fields)
    row.cosmetic_diffs = list(breakdown.cosmetic_diffs)
    row.fillable_fields = list(breakdown.fillable_fields)
    row.eff_delta_days = breakdown.eff_delta_days
    row.exp_delta_days = breakdown.exp_delta_days
    row.ext_type_raw = breakdown.ext_type_raw
    row.ext_type_normalized = breakdown.ext_type_normalized
    row.coverage_alias_applied = breakdown.coverage_alias_applied
    row.ext_carrier_raw = breakdown.ext_carrier_raw
    row.ext_carrier_normalized = breakdown.ext_carrier_normalized
    row.carrier_alias_applied = breakdown.carrier_alias_applied

    # Remove from extras if found there
    if extra_idx_to_remove is not None:
        extras.pop(extra_idx_to_remove)

    # Auto-learn match memory for manual pairs (highest value — user explicitly paired these)
    source_name = _SOURCE_NAME_CACHE.get(token, "")
    if source_name and row.ext and target_db.get("id"):
        try:
            from policydb.match_memory import learn_from_reconcile_pair
            learn_from_reconcile_pair(conn, target_db["id"], source_name, row.ext)
        except Exception:
            logger.exception("Match memory learning failed on manual-pair")

    # Return the updated pair row (the JS in _pairing_board.html handles extra removal client-side)
    summary = summarize(results + extras)
    pair_html = templates.TemplateResponse("reconcile/_pair_row.html", {
        "request": request,
        "row": row,
        "idx": idx,
        "token": token,
        "field_display": _FIELD_DISPLAY,
        "policy_types": cfg.get("policy_types", []),
    }).body.decode()
    counter_html = _render_counters(summary)
    return HTMLResponse(pair_html + counter_html)


@router.post("/confirm-all", response_class=HTMLResponse)
def reconcile_confirm_all(request: Request, token: str = Form(""), conn=Depends(get_db)):
    """Confirm all high-confidence paired rows (score >= 75). Re-renders entire board."""
    cache = _BOARD_CACHE.get(token)
    if not cache:
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Session expired. Please re-run reconciliation.</div>')
    results, extras, db_rows, ts = cache
    source_name = _SOURCE_NAME_CACHE.get(token, "")
    confirmed_count = 0
    for row in results:
        if row.status == "PAIRED" and not row.confirmed and row.match_score >= 75:
            row.confirmed = True
            confirmed_count += 1
            # Learn match memory
            if source_name and row.ext and row.db and row.db.get("id"):
                try:
                    from policydb.match_memory import learn_from_reconcile_pair
                    learn_from_reconcile_pair(conn, row.db["id"], source_name, row.ext)
                except Exception:
                    pass
    if confirmed_count and source_name:
        logger.info("Match memory: batch-learned from %d confirmed pairs (source='%s')", confirmed_count, source_name)
    return templates.TemplateResponse("reconcile/_pairing_board.html",
                                      _board_context(request, token, results, extras))


@router.post("/confirm-all-programs", response_class=HTMLResponse)
def reconcile_confirm_all_programs(request: Request, token: str = Form("")):
    """Confirm all program-matched paired rows regardless of score. Re-renders entire board."""
    cache = _BOARD_CACHE.get(token)
    if not cache:
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Session expired. Please re-run reconciliation.</div>')
    results, extras, db_rows, ts = cache
    confirmed_count = 0
    for row in results:
        if row.status == "PAIRED" and not row.confirmed and row.db and row.db.get("program_id"):
            row.confirmed = True
            confirmed_count += 1
    return templates.TemplateResponse("reconcile/_pairing_board.html",
                                      _board_context(request, token, results, extras))


@router.post("/fill/{policy_uid}", response_class=HTMLResponse)
async def reconcile_fill(
    request: Request,
    policy_uid: str,
    conn=Depends(get_db),
):
    """HTMX: fill empty fields on a matched policy from imported values."""
    from policydb.importer import _parse_currency
    form = await request.form()

    _CURRENCY = {"premium", "limit_amount", "deductible"}
    _TEXT = {"carrier", "policy_number", "first_named_insured", "placement_colleague",
             "underwriter_name", "exposure_address", "project_name"}
    _DATE = {"effective_date", "expiration_date"}
    _ALLOWED = _CURRENCY | _TEXT | _DATE

    updates = []
    params = []
    filled_names = []
    for field in _ALLOWED:
        val = form.get(field, "").strip()
        if not val:
            continue
        if field in _CURRENCY:
            parsed = _parse_currency(val)
            if parsed and parsed > 0:
                updates.append(f"{field} = ?")
                params.append(parsed)
                filled_names.append(field.replace("_", " ").title())
        elif field in _DATE:
            updates.append(f"{field} = ?")
            params.append(val)
            filled_names.append(field.replace("_", " ").title())
        elif field in _TEXT:
            updates.append(f"{field} = ?")
            params.append(val)
            filled_names.append(field.replace("_", " ").title())

    # If project_name is being set, try to auto-match to a known location
    project_name_val = form.get("project_name", "").strip()
    if project_name_val:
        policy_row = conn.execute(
            "SELECT client_id, project_id FROM policies WHERE policy_uid = ?",
            (policy_uid.upper(),),
        ).fetchone()
        if policy_row and not policy_row["project_id"]:
            loc = conn.execute(
                "SELECT id FROM projects WHERE client_id = ? AND LOWER(name) = LOWER(?) "
                "AND (project_type = 'Location' OR project_type IS NULL)",
                (policy_row["client_id"], project_name_val),
            ).fetchone()
            if loc:
                updates.append("project_id = ?")
                params.append(loc["id"])
                filled_names.append("Location Link")

    if updates:
        # Get current values for provenance tracking before update
        pol = conn.execute(
            "SELECT id, " + ", ".join(f for f in _ALLOWED if form.get(f, "").strip()) +
            " FROM policies WHERE policy_uid = ?",
            (policy_uid.upper(),),
        ).fetchone()
        prior_values = dict(pol) if pol else {}
        policy_id = prior_values.pop("id", None)

        params.append(policy_uid.upper())
        conn.execute(
            f"UPDATE policies SET {', '.join(updates)} WHERE policy_uid = ?",
            params,
        )

        # Record provenance for each updated field
        source_name = ""
        session_id = None
        as_of_date = ""
        # Try to get source context from the form (passed via hidden fields)
        _token = form.get("token", "")
        if _token:
            source_name = _SOURCE_NAME_CACHE.get(_token, "")
            session_id = _SESSION_ID_CACHE.get(_token)
            # as_of_date would come from the session
            if session_id:
                try:
                    sess = conn.execute("SELECT as_of_date FROM import_sessions WHERE id = ?", (session_id,)).fetchone()
                    if sess:
                        as_of_date = sess["as_of_date"] or ""
                except Exception:
                    pass

        if policy_id and source_name:
            try:
                from policydb.import_ledger import record_provenance
                for field in _ALLOWED:
                    val = form.get(field, "").strip()
                    if not val:
                        continue
                    prior = str(prior_values.get(field, "") or "")
                    was_conflict = bool(prior.strip() and prior.strip() != val.strip())
                    record_provenance(
                        conn, policy_id, field, val,
                        source_name=source_name, source_session_id=session_id,
                        as_of_date=as_of_date, prior_value=prior, was_conflict=was_conflict,
                    )
            except Exception:
                logger.exception("Provenance recording failed on fill")

        conn.commit()

    label = ", ".join(filled_names) if filled_names else "No fields"
    return HTMLResponse(
        f'<span class="text-xs text-green-600 font-medium">Updated {label}</span>'
    )


@router.post("/archive/{policy_uid}", response_class=HTMLResponse)
def reconcile_archive(request: Request, policy_uid: str, token: str = Form(""), conn=Depends(get_db)):
    """HTMX: archive an EXTRA policy. Remove from board cache, return OOB delete + counters."""
    conn.execute("UPDATE policies SET archived=1 WHERE policy_uid=?", (policy_uid.upper(),))
    conn.commit()

    # Update board cache if active
    counter_html = ""
    cache = _BOARD_CACHE.get(token) if token else None
    if cache:
        results, extras, db_rows, ts = cache
        # Remove from extras
        for i, er in enumerate(extras):
            if er.db and er.db.get("policy_uid") == policy_uid.upper():
                extras.pop(i)
                break
        summary = summarize(results + extras)
        counter_html = _render_counters(summary)

    # Return empty div (hx-swap="outerHTML" removes the extra row) + OOB counters
    return HTMLResponse(
        f'<div id="extra-{policy_uid.upper()}" class="text-xs text-gray-400 italic p-2 mb-2">'
        f'Policy {policy_uid.upper()} archived.</div>'
        + counter_html
    )


@router.get("/create-form", response_class=HTMLResponse)
def reconcile_create_form(
    request: Request,
    client_name: str = "",
    policy_type: str = "",
    carrier: str = "",
    effective_date: str = "",
    expiration_date: str = "",
    premium: str = "",
    limit_amount: str = "",
    deductible: str = "",
    policy_number: str = "",
    description: str = "",
    project_name: str = "",
    placement_colleague: str = "",
    underwriter_name: str = "",
    commission_rate: str = "",
    row_uid: str = "",
    token: str = "",
    conn=Depends(get_db),
):
    """HTMX: render an inline quick-create form for a MISSING row."""
    client_row = get_client_by_name(conn, client_name) if client_name else None
    all_clients = conn.execute(
        "SELECT id, name FROM clients WHERE archived=0 ORDER BY name"
    ).fetchall()
    from policydb import config as cfg
    return templates.TemplateResponse("reconcile/_create_form.html", {
        "request": request,
        "ext": {
            "client_name": client_name,
            "policy_type": policy_type,
            "carrier": carrier,
            "effective_date": effective_date,
            "expiration_date": expiration_date,
            "premium": premium,
            "limit_amount": limit_amount,
            "deductible": deductible,
            "policy_number": policy_number,
            "description": description,
            "project_name": project_name,
            "placement_colleague": placement_colleague,
            "underwriter_name": underwriter_name,
            "commission_rate": commission_rate,
        },
        "matched_client": dict(client_row) if client_row else None,
        "all_clients": [dict(c) for c in all_clients],
        "policy_types": cfg.get("policy_types", []),
        "row_uid": row_uid,
        "token": token,
    })


@router.post("/create", response_class=HTMLResponse)
def reconcile_create(
    request: Request,
    row_uid: str = Form(""),
    client_id: int = Form(...),
    policy_type: str = Form(...),
    carrier: str = Form(...),
    effective_date: str = Form(...),
    expiration_date: str = Form(...),
    premium: float = Form(0.0),
    limit_amount: str = Form(""),
    deductible: str = Form(""),
    policy_number: str = Form(""),
    description: str = Form(""),
    project_name: str = Form(""),
    placement_colleague: str = Form(""),
    underwriter_name: str = Form(""),
    commission_rate: str = Form(""),
    is_program: str = Form("0"),
    token: str = Form(""),
    conn=Depends(get_db),
):
    """HTMX: create a new policy from a MISSING reconcile row, return confirmation."""
    from policydb.db import next_policy_uid
    from policydb import config as cfg

    uid = next_policy_uid(conn)
    account_exec = cfg.get("default_account_exec", "Grant")
    pgm = 1 if is_program == "1" else 0
    policy_type = normalize_coverage_type(policy_type)
    carrier = normalize_carrier(carrier) if carrier else ""
    policy_number = normalize_policy_number(policy_number) if policy_number else ""

    def _f(v):
        try: return float(v) if v else 0.0
        except (ValueError, TypeError): return 0.0

    # Auto-match project_name to known location for project_id
    project_id = None
    if project_name:
        loc = conn.execute(
            "SELECT id FROM projects WHERE client_id = ? AND LOWER(name) = LOWER(?) "
            "AND (project_type = 'Location' OR project_type IS NULL)",
            (client_id, project_name.strip()),
        ).fetchone()
        if loc:
            project_id = loc["id"]

    conn.execute(
        """INSERT INTO policies
           (policy_uid, client_id, policy_type, carrier, policy_number,
            effective_date, expiration_date, premium, limit_amount, deductible,
            description, project_name, project_id, underwriter_name,
            commission_rate, account_exec,
            is_program)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            uid, client_id, policy_type, carrier, policy_number or None,
            effective_date, expiration_date, premium,
            _f(limit_amount), _f(deductible),
            description or None, project_name or None, project_id,
            underwriter_name or None,
            _f(commission_rate), account_exec,
            pgm,
        ),
    )
    conn.commit()

    # Create structured contact records for placement colleague and underwriter
    _policy_row = conn.execute("SELECT id FROM policies WHERE policy_uid=?", (uid,)).fetchone()
    if _policy_row:
        _pid = _policy_row["id"]
        _pc_name = (placement_colleague or "").strip()
        if _pc_name:
            from policydb.queries import get_or_create_contact, assign_contact_to_policy
            _pc_cid = get_or_create_contact(conn, _pc_name)
            assign_contact_to_policy(conn, _pc_cid, _pid, is_placement_colleague=1)
        _uw_name = (underwriter_name or "").strip()
        if _uw_name:
            from policydb.queries import get_or_create_contact, assign_contact_to_policy
            _uw_cid = get_or_create_contact(conn, _uw_name)
            assign_contact_to_policy(conn, _uw_cid, _pid, role="Underwriter")
        conn.commit()

    client_name = conn.execute("SELECT name FROM clients WHERE id=?", (client_id,)).fetchone()["name"]

    # Update board cache — mark the row as PAIRED after creation
    counter_html = ""
    if token:
        cache = _BOARD_CACHE.get(token)
        if cache:
            results, extras, db_rows_cached, ts = cache
            try:
                idx = int(row_uid)
                if 0 <= idx < len(results):
                    new_db = {
                        "policy_uid": uid,
                        "client_name": client_name,
                        "policy_type": policy_type,
                        "carrier": carrier,
                        "policy_number": policy_number,
                        "effective_date": effective_date,
                        "expiration_date": expiration_date,
                        "premium": premium,
                        "limit_amount": _f(limit_amount),
                        "deductible": _f(deductible),
                        "client_id": client_id,
                    }
                    results[idx].db = new_db
                    results[idx].status = "PAIRED"
                    results[idx].match_score = 100.0
                    results[idx].confidence = "high"
                    results[idx].match_method = "created"
                    results[idx].confirmed = True
                    summary = summarize(results + extras)
                    counter_html = _render_counters(summary)
            except (ValueError, IndexError):
                pass

    # Learn match memory + record provenance for the newly created policy
    source_name = _SOURCE_NAME_CACHE.get(token, "")
    new_pid = conn.execute("SELECT id FROM policies WHERE policy_uid = ?", (uid,)).fetchone()["id"]
    if source_name and policy_number:
        try:
            from policydb.match_memory import learn
            learn(conn, new_pid, source_name, policy_number, "policy_number", "reconcile")
        except Exception:
            logger.exception("Match memory learning failed on create")

    if source_name and new_pid:
        try:
            from policydb.import_ledger import record_provenance_batch
            session_id = _SESSION_ID_CACHE.get(token)
            as_of_date = ""
            if session_id:
                sess = conn.execute("SELECT as_of_date FROM import_sessions WHERE id = ?", (session_id,)).fetchone()
                if sess:
                    as_of_date = sess["as_of_date"] or ""
            created_fields = {
                "policy_type": policy_type, "carrier": carrier, "policy_number": policy_number,
                "effective_date": effective_date, "expiration_date": expiration_date,
                "premium": str(premium), "description": description, "project_name": project_name,
            }
            created_fields = {k: v for k, v in created_fields.items() if v}
            record_provenance_batch(
                conn, new_pid, created_fields,
                source_name=source_name, source_session_id=session_id, as_of_date=as_of_date,
            )
        except Exception:
            logger.exception("Provenance recording failed on create")
        conn.commit()

    created_html = (
        f'<div class="pair-row flex items-center rounded-lg border border-green-200 bg-green-50 mb-2 px-4 py-3" data-status="confirmed" data-confirmed="true">'
        f'<span class="text-green-500 text-lg mr-3">&#10003;</span>'
        f'<div class="flex-1 min-w-0">'
        f'<span class="text-xs font-semibold text-green-700 px-2 py-0.5 rounded bg-green-100">CREATED</span>'
        f' <span class="text-sm font-medium text-gray-800 ml-2">{client_name}</span>'
        f' <span class="text-xs text-gray-400 mx-1">&middot;</span>'
        f' <span class="text-xs text-gray-600">{policy_type}</span>'
        f' <span class="text-xs text-gray-400 mx-1">&middot;</span>'
        f' <span class="text-xs text-gray-600">{carrier}</span>'
        f' <span class="text-xs text-gray-400 mx-1">&middot;</span>'
        f' <span class="text-xs text-gray-600">{effective_date} &rarr; {expiration_date}</span>'
        f' <span class="text-xs text-gray-400 mx-1">&middot;</span>'
        f' <span class="text-xs text-gray-700 tabular-nums">${premium:,.0f}</span>'
        f'</div>'
        f'<a href="/policies/{uid}/edit" class="text-xs text-blue-500 hover:underline ml-3">{uid} &rarr;</a>'
        f'</div>'
    )
    # OOB: replace the original unmatched row with the created confirmation
    oob_replace = f'<div id="pair-{row_uid}" hx-swap-oob="outerHTML">{created_html}</div>'
    # Primary response is empty (removes the form wrapper); OOB replaces the unmatched row
    return HTMLResponse('<div style="display:none"></div>' + oob_replace + counter_html)


@router.post("/apply/{policy_uid}", response_class=HTMLResponse)
def reconcile_apply(
    request: Request,
    policy_uid: str,
    row_uid: str = Form(""),
    carrier: str = Form(""),
    policy_number: str = Form(""),
    effective_date: str = Form(""),
    expiration_date: str = Form(""),
    premium: str = Form(""),
    limit_amount: str = Form(""),
    deductible: str = Form(""),
    policy_type: str = Form(""),
    conn=Depends(get_db),
):
    """HTMX: apply ext upload values to a DIFF policy (one-click update)."""
    def _f(v):
        try: return float(v) if v else None
        except (ValueError, TypeError): return None

    # Build update for non-empty ext fields only
    updates: dict = {}
    if carrier: updates["carrier"] = carrier
    if policy_number: updates["policy_number"] = policy_number
    if effective_date: updates["effective_date"] = effective_date
    if expiration_date: updates["expiration_date"] = expiration_date
    if premium: updates["premium"] = _f(premium)
    if limit_amount: updates["limit_amount"] = _f(limit_amount)
    if deductible: updates["deductible"] = _f(deductible)
    if policy_type: updates["policy_type"] = policy_type

    if updates:
        set_clause = ", ".join(f"{k}=?" for k in updates)
        conn.execute(
            f"UPDATE policies SET {set_clause} WHERE policy_uid=?",
            list(updates.values()) + [policy_uid.upper()],
        )
        conn.commit()

    client_name = conn.execute(
        """SELECT c.name FROM policies p JOIN clients c ON p.client_id=c.id WHERE p.policy_uid=?""",
        (policy_uid.upper(),),
    ).fetchone()["name"]

    premium_fmt = f"${float(premium):,.0f}" if premium else "—"
    ptype = policy_type or "—"
    carr = carrier or "—"

    return HTMLResponse(
        f'<tr id="row-{row_uid}" class="bg-green-50">'
        f'<td class="px-3 py-3 text-gray-300 text-xs">✓</td>'
        f'<td class="px-4 py-3"><span class="text-xs font-semibold px-2 py-0.5 rounded bg-green-100 text-green-700">UPDATED</span></td>'
        f'<td class="px-4 py-3 font-medium text-gray-800 text-sm">{client_name}</td>'
        f'<td class="px-4 py-3 text-xs text-gray-600">{ptype}</td>'
        f'<td class="px-4 py-3 text-xs text-gray-600">{carr}</td>'
        f'<td class="px-4 py-3 text-xs font-mono text-gray-500">{policy_number or "—"}</td>'
        f'<td class="px-4 py-3 text-xs whitespace-nowrap text-gray-600">{effective_date or "—"}</td>'
        f'<td class="px-4 py-3 text-xs whitespace-nowrap text-gray-600">{expiration_date or "—"}</td>'
        f'<td class="px-4 py-3 text-right tabular-nums text-gray-700">{premium_fmt}</td>'
        f'<td class="px-4 py-3"><a href="/policies/{policy_uid.upper()}/edit" class="text-xs text-marsh hover:underline">{policy_uid.upper()} →</a></td>'
        f'</tr>'
    )


_ALLOWED_FIELDS = {
    "policy_type", "carrier", "policy_number",
    "effective_date", "expiration_date",
    "premium", "limit_amount", "deductible",
    "first_named_insured", "placement_colleague",
    "underwriter_name", "exposure_address",
}

_CURRENCY_FIELDS = {"premium", "limit_amount", "deductible"}

_FIELD_DISPLAY = {
    "policy_type": "Coverage Type",
    "carrier": "Carrier",
    "policy_number": "Policy Number",
    "effective_date": "Effective Date",
    "expiration_date": "Expiration Date",
    "premium": "Premium",
    "limit_amount": "Limit",
    "deductible": "Deductible",
    "first_named_insured": "First Named Insured",
    "placement_colleague": "Placement Colleague",
    "underwriter_name": "Underwriter",
    "exposure_address": "Address",
    "project_name": "Location / Project",
}


def _parse_field_value(field_name: str, value: str):
    """Parse a field value for DB storage. Uses parse_currency_with_magnitude for money fields."""
    if field_name in _CURRENCY_FIELDS:
        parsed = parse_currency_with_magnitude(value)
        return parsed if parsed is not None else None
    return value


def _format_field_display(field_name: str, value: str) -> str:
    """Format a field value for display after apply."""
    if field_name in _CURRENCY_FIELDS and value:
        try:
            return "${:,.0f}".format(float(parse_currency_with_magnitude(value) or 0))
        except (ValueError, TypeError):
            return value or "—"
    return value or "—"


@router.patch("/apply-field/{policy_uid}", response_class=HTMLResponse)
async def reconcile_apply_field(
    request: Request,
    policy_uid: str,
    conn=Depends(get_db),
):
    """HTMX: apply a single field update from uploaded value to a DIFF policy."""
    form = await request.form()
    field_name = form.get("field_name", "")
    value = form.get("value", "")

    if field_name not in _ALLOWED_FIELDS:
        return HTMLResponse(
            f'<td colspan="4" class="text-xs text-red-500">Invalid field: {field_name}</td>',
            status_code=400,
        )

    db_value = _parse_field_value(field_name, value)

    conn.execute(
        f"UPDATE policies SET {field_name}=? WHERE policy_uid=?",
        (db_value, policy_uid.upper()),
    )
    conn.commit()

    display_label = _FIELD_DISPLAY.get(field_name, field_name)
    display_value = _format_field_display(field_name, value)
    return HTMLResponse(
        f'<div class="flex items-center gap-2 text-[10px]">'
        f'<span class="font-semibold text-green-600 w-24 flex-shrink-0">{display_label}</span>'
        f'<span class="text-green-700 font-semibold">{display_value}</span>'
        f'<span class="text-green-500">applied &#10003;</span>'
        f'</div>'
    )


def _apply_all_fields_for_row(row: ReconcileRow, conn) -> int:
    """Apply all diff_fields and fillable_fields from ext to DB for one pair.

    Returns count of fields applied.
    """
    if not row.ext or not row.db:
        return 0
    policy_uid = row.db.get("policy_uid", "")
    if not policy_uid:
        return 0

    applied = 0
    for f in list(row.diff_fields) + list(row.fillable_fields):
        if f not in _ALLOWED_FIELDS:
            continue
        value = str(row.ext.get(f) or "").strip()
        if not value:
            continue
        db_value = _parse_field_value(f, value)
        try:
            conn.execute(
                f"UPDATE policies SET {f}=? WHERE policy_uid=?",
                (db_value, policy_uid.upper()),
            )
            applied += 1
        except Exception:
            logger.warning("Failed to apply field %s for %s", f, policy_uid)
    return applied


@router.patch("/accept-all-fields/{idx}", response_class=HTMLResponse)
async def reconcile_accept_all_fields(
    request: Request,
    idx: int,
    conn=Depends(get_db),
):
    """Accept all diff + fillable fields for a single pair at once."""
    form = await request.form()
    token = form.get("token", "")
    cache = _BOARD_CACHE.get(token)
    if not cache:
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Session expired. Please re-run reconciliation.</div>')
    results, extras, db_rows, _ = cache
    if idx < 0 or idx >= len(results):
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Invalid row index.</div>')

    row = results[idx]
    applied = _apply_all_fields_for_row(row, conn)
    conn.commit()
    logger.info("Accept all fields: idx=%d, applied=%d fields", idx, applied)

    # Build response showing all fields as applied
    html_parts = []
    for f in list(row.diff_fields) + list(row.fillable_fields):
        if f not in _ALLOWED_FIELDS:
            continue
        value = str(row.ext.get(f) or "").strip()
        display_label = _FIELD_DISPLAY.get(f, f)
        display_value = _format_field_display(f, value)
        html_parts.append(
            f'<div class="flex items-center gap-2 text-[10px]">'
            f'<span class="font-semibold text-green-600 w-24 flex-shrink-0">{display_label}</span>'
            f'<span class="text-green-700 font-semibold">{display_value}</span>'
            f'<span class="text-green-500">applied &#10003;</span>'
            f'</div>'
        )
    return HTMLResponse("\n".join(html_parts) if html_parts else '<div class="text-[10px] text-gray-400">No fields to apply.</div>')


@router.get("/reconcile-all-preview", response_class=HTMLResponse)
def reconcile_all_preview(request: Request, token: str = ""):
    """Preview what reconcile-all would do: count pairs, fields, aliases to learn."""
    cache = _BOARD_CACHE.get(token)
    if not cache:
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Session expired. Please re-run reconciliation.</div>')
    results, extras, db_rows, _ = cache

    total_pairs = 0
    total_fields = 0
    aliases_to_learn: list[str] = []
    seen_aliases: set[str] = set()

    for row in results:
        if row.status == "PAIRED" and not row.confirmed:
            total_pairs += 1
            total_fields += len([f for f in list(row.diff_fields) + list(row.fillable_fields) if f in _ALLOWED_FIELDS])
            # Check for aliases to learn
            if row.coverage_alias_applied and row.ext_type_raw:
                key = row.ext_type_raw.strip().lower()
                if key not in _BASE_COVERAGE_ALIASES and key not in seen_aliases:
                    seen_aliases.add(key)
                    aliases_to_learn.append(f'"{row.ext_type_raw.strip()}" &rarr; {row.ext_type_normalized}')
            if row.carrier_alias_applied and row.ext_carrier_raw:
                key = row.ext_carrier_raw.strip().lower()
                if key not in _BASE_CARRIER_ALIASES and key not in seen_aliases:
                    seen_aliases.add(key)
                    aliases_to_learn.append(f'"{row.ext_carrier_raw.strip()}" &rarr; {row.ext_carrier_normalized}')

    alias_html = ""
    if aliases_to_learn:
        pills = " ".join(
            f'<span class="inline-block px-2 py-0.5 rounded bg-blue-100 text-blue-700 text-[10px]">{a}</span>'
            for a in aliases_to_learn
        )
        alias_html = f'<div class="mt-2 text-[10px] text-blue-700">Will learn {len(aliases_to_learn)} alias(es): {pills}</div>'

    return HTMLResponse(
        f'<div class="p-3 bg-amber-50 border border-amber-200 rounded-lg">'
        f'<div class="text-sm font-semibold text-amber-800 mb-1">Reconcile All</div>'
        f'<div class="text-xs text-amber-700">'
        f'Confirm <strong>{total_pairs}</strong> pairs and apply <strong>{total_fields}</strong> field updates.'
        f'</div>'
        f'{alias_html}'
        f'<div class="flex gap-2 mt-3">'
        f'<button type="button" hx-post="/reconcile/reconcile-all" '
        f'hx-vals=\'{{"token": "{token}"}}\' '
        f'hx-target="#pairing-board" hx-swap="innerHTML" '
        f'class="text-xs bg-green-600 text-white px-4 py-1.5 rounded hover:bg-green-700 font-medium">'
        f'Confirm &amp; Apply All</button>'
        f'<button type="button" onclick="this.closest(\'.p-3\').remove()" '
        f'class="text-xs text-gray-500 hover:text-gray-700 px-3 py-1.5">Cancel</button>'
        f'</div></div>'
    )


@router.post("/reconcile-all", response_class=HTMLResponse)
def reconcile_all(request: Request, token: str = Form(""), conn=Depends(get_db)):
    """Bulk confirm all pairs, apply all diffs, auto-learn all aliases."""
    cache = _BOARD_CACHE.get(token)
    if not cache:
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Session expired. Please re-run reconciliation.</div>')
    results, extras, db_rows, _ = cache

    confirmed_count = 0
    total_applied = 0
    all_learned: list[str] = []
    seen_coverage: set[str] = set()
    seen_carrier: set[str] = set()

    # Collect unique aliases first for deduplication
    coverage_to_learn: dict[str, str] = {}  # raw → canonical
    carrier_to_learn: dict[str, str] = {}

    for row in results:
        if row.status != "PAIRED" or row.confirmed:
            continue

        # Confirm
        row.confirmed = True
        confirmed_count += 1

        # Apply all fields
        applied = _apply_all_fields_for_row(row, conn)
        total_applied += applied

        # Collect aliases to learn (deduplicated)
        if row.coverage_alias_applied and row.ext_type_raw and row.ext_type_normalized:
            raw_key = row.ext_type_raw.strip().lower()
            if raw_key not in _BASE_COVERAGE_ALIASES and raw_key not in seen_coverage:
                seen_coverage.add(raw_key)
                coverage_to_learn[row.ext_type_raw.strip()] = row.ext_type_normalized
        if row.carrier_alias_applied and row.ext_carrier_raw and row.ext_carrier_normalized:
            raw_key = row.ext_carrier_raw.strip().lower()
            if raw_key not in _BASE_CARRIER_ALIASES and raw_key not in seen_carrier:
                seen_carrier.add(raw_key)
                carrier_to_learn[row.ext_carrier_raw.strip()] = row.ext_carrier_normalized

    conn.commit()

    # Write learned aliases once (not per-pair)
    if coverage_to_learn:
        aliases = cfg.get("coverage_aliases", {})
        for raw, canonical in coverage_to_learn.items():
            if canonical not in aliases:
                aliases[canonical] = []
            if raw.lower() not in [a.lower() for a in aliases[canonical]]:
                aliases[canonical].append(raw)
                all_learned.append(f'"{raw}" &rarr; {canonical}')
        full = dict(cfg.load_config())
        full["coverage_aliases"] = aliases
        cfg.save_config(full)
        cfg.reload_config()
        rebuild_coverage_aliases()

    if carrier_to_learn:
        aliases = cfg.get("carrier_aliases", {})
        for raw, canonical in carrier_to_learn.items():
            if canonical not in aliases:
                aliases[canonical] = []
            if raw.lower() not in [a.lower() for a in aliases[canonical]]:
                aliases[canonical].append(raw)
                all_learned.append(f'"{raw}" &rarr; {canonical}')
        full = dict(cfg.load_config())
        full["carrier_aliases"] = aliases
        cfg.save_config(full)
        cfg.reload_config()
        rebuild_carrier_aliases()

    logger.info("Reconcile all: confirmed=%d, applied=%d fields, learned=%d aliases",
                confirmed_count, total_applied, len(all_learned))

    # Build learning summary banner
    learn_banner = ""
    if all_learned:
        pills = " ".join(
            f'<span class="inline-block px-2 py-0.5 rounded bg-blue-100 text-blue-700 text-[10px]">{item}</span>'
            for item in all_learned
        )
        learn_banner = (
            f'<div class="mb-4 px-4 py-3 rounded-lg bg-blue-50 border border-blue-200 text-xs text-blue-800">'
            f'<span class="font-semibold">Reconciled {confirmed_count} pairs</span> · '
            f'Applied {total_applied} field updates · '
            f'Learned {len(all_learned)} new alias(es): {pills}'
            f'<button onclick="this.parentElement.remove()" class="ml-2 text-blue-400 hover:text-blue-600">&times;</button>'
            f'</div>'
        )
    elif confirmed_count:
        learn_banner = (
            f'<div class="mb-4 px-4 py-3 rounded-lg bg-green-50 border border-green-200 text-xs text-green-800">'
            f'<span class="font-semibold">Reconciled {confirmed_count} pairs</span> · '
            f'Applied {total_applied} field updates'
            f'<button onclick="this.parentElement.remove()" class="ml-2 text-green-400 hover:text-green-600">&times;</button>'
            f'</div>'
        )

    board_html = templates.TemplateResponse("reconcile/_pairing_board.html",
                                              _board_context(request, token, results, extras)).body.decode()

    return HTMLResponse(learn_banner + board_html)


# ── Program Management Endpoints ─────────────────────────────────────────────


@router.get("/program-group-form", response_class=HTMLResponse)
def program_group_form(
    request: Request,
    token: str = "",
    cand_idx: int = 0,
    indices: str = "",
    conn=Depends(get_db),
):
    """Render the inline form for creating a program from a detected candidate group or manual selection."""
    cache = _BOARD_CACHE.get(token)
    if not cache:
        return HTMLResponse('<div class="text-xs text-red-500 p-2">Session expired.</div>')
    results, extras, db_rows, ts = cache

    # Build candidate from explicit indices (manual multi-select) or from auto-detected candidates
    if indices:
        idx_list = [int(i) for i in indices.split(",") if i.strip().isdigit()]
        carriers = []
        carrier_details = []
        client_name = ""
        ptype = ""
        eff = ""
        exp = ""
        for i in idx_list:
            if 0 <= i < len(results):
                e = results[i].ext or {}
                d = results[i].db or {}
                carrier = e.get("carrier", "") or d.get("carrier", "")
                carriers.append(carrier)
                carrier_details.append({
                    "carrier": carrier,
                    "premium": e.get("premium", "") or d.get("premium", ""),
                    "limit_amount": e.get("limit_amount", "") or d.get("limit_amount", ""),
                    "policy_number": e.get("policy_number", "") or d.get("policy_number", ""),
                })
                if not client_name:
                    client_name = e.get("client_name", "") or d.get("client_name", "")
                if not ptype:
                    ptype = e.get("policy_type", "") or d.get("policy_type", "")
                if not eff:
                    eff = e.get("effective_date", "") or d.get("effective_date", "")
                if not exp:
                    exp = e.get("expiration_date", "") or d.get("expiration_date", "")
        cand = {
            "client": client_name,
            "type": ptype,
            "effective_date": eff,
            "expiration_date": exp,
            "carriers": [c for c in carriers if c],
            "row_indices": idx_list,
            "carrier_details": carrier_details,
            "detection_method": "manual_selection",
        }
    else:
        candidates = _detect_program_candidates(results)
        if cand_idx < 0 or cand_idx >= len(candidates):
            return HTMLResponse('<div class="text-xs text-red-500 p-2">Candidate not found.</div>')
        cand = candidates[cand_idx]

    matched_client = get_client_by_name(conn, cand["client"]) if cand["client"] else None
    all_clients = [dict(c) for c in conn.execute(
        "SELECT id, name FROM clients WHERE archived=0 ORDER BY name"
    ).fetchall()]

    return templates.TemplateResponse("reconcile/_program_group_form.html", {
        "request": request,
        "cand": cand,
        "token": token,
        "all_clients": all_clients,
        "matched_client": dict(matched_client) if matched_client else None,
        "policy_types": cfg.get("policy_types", []),
        "cand_idx": cand_idx,
    })


@router.post("/create-program-group", response_class=HTMLResponse)
def create_program_group(
    request: Request,
    token: str = Form(""),
    conn=Depends(get_db),
):
    """Placeholder — program creation from reconcile removed (use Programs UI instead)."""
    return HTMLResponse(
        '<div class="text-xs text-amber-600 p-2">Program creation from reconcile is no longer supported. '
        'Use the Programs page to create programs and assign child policies.</div>'
    )


@router.get("/search-programs", response_class=HTMLResponse)
def search_programs(request: Request):
    """Placeholder — program search from reconcile removed (use Programs UI instead)."""
    return HTMLResponse('<p class="text-xs text-gray-400 p-2">Program management moved to Programs page.</p>')


@router.post("/add-to-program", response_class=HTMLResponse)
def add_to_program(request: Request):
    """Placeholder — add-to-program from reconcile removed (use Programs UI instead)."""
    return HTMLResponse(
        '<div class="text-xs text-amber-600 p-2">Adding to programs from reconcile is no longer supported. '
        'Use the Programs page to manage program membership.</div>'
    )


@router.get("/download/{token}")
def reconcile_download_cached(token: str):
    """Download cached XLSX report. Regenerates from board cache if available (reflects mutations)."""
    _cache_cleanup()
    # Prefer board cache (reflects confirm/break/pair changes)
    board = _BOARD_CACHE.get(token)
    if board:
        results, extras, db_rows, _ = board
        xlsx_bytes = build_reconcile_xlsx(results + extras, run_date=date.today().isoformat())
    else:
        entry = _RESULT_CACHE.get(token)
        if not entry:
            return HTMLResponse("Report expired. Please re-run reconciliation.", status_code=404)
        xlsx_bytes, _ = entry
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="reconcile_{date.today()}.xlsx"'},
    )


@router.post("/download")
async def reconcile_download(
    file: UploadFile = File(...),
    client_id: int = Form(0),
    scope: str = Form("active"),
    column_mapping_json: str = Form(""),
    conn=Depends(get_db),
):
    import json as _json
    col_map = {}
    if column_mapping_json:
        try:
            col_map = {k: v for k, v in _json.loads(column_mapping_json).items() if v and v != "ignore"}
        except Exception:
            pass
    content = await file.read()
    ext_rows, _ = parse_uploaded_file(content, column_mapping=col_map or None, filename=file.filename or "")
    db_rows = _load_db_policies(conn, client_id, scope)
    results = reconcile(ext_rows, db_rows)
    xlsx = build_reconcile_xlsx(results, run_date=date.today().isoformat(), filename=file.filename or "")
    return Response(
        content=xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="reconcile_{date.today()}.xlsx"'},
    )
