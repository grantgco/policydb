"""Export system: schedule, client, book, renewals, LLM context dumps."""

from __future__ import annotations

import csv
import io
import json
import sqlite3
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from policydb import config as cfg
from policydb.analysis import build_program_audit
from policydb.display import fmt_currency, fmt_limit, fmt_pct, fmt_days

# Fields excluded from all client-facing (schedule) exports
_INTERNAL_FIELDS = {
    "commission_rate", "commission_amount", "prior_premium", "rate_change",
    "renewal_status", "placement_colleague", "underwriter_name",
    "underwriter_contact", "notes", "account_exec",
}

TODAY = date.today().isoformat()


def _schedule_rows_for_client(conn: sqlite3.Connection, client_id: int) -> list[sqlite3.Row]:
    return conn.execute(
        """SELECT * FROM v_schedule WHERE client_name = (SELECT name FROM clients WHERE id = ?)""",
        (client_id,),
    ).fetchall()


def _row_to_dict(row: sqlite3.Row, exclude: set[str] | None = None) -> dict:
    d = dict(row)
    if exclude:
        for k in exclude:
            d.pop(k, None)
    return d


# ─── SCHEDULE OF INSURANCE ───────────────────────────────────────────────────

def export_schedule_md(conn: sqlite3.Connection, client_id: int, client_name: str) -> str:
    rows = _schedule_rows_for_client(conn, client_id)
    account_exec = cfg.get("default_account_exec", "Grant")
    total = sum(r["Premium"] or 0 for r in rows)

    lines = [
        "# Schedule of Insurance",
        "",
        f"**Insured:** {client_name}",
        f"**Prepared:** {TODAY}",
        f"**Prepared by:** {account_exec}, Marsh",
        "",
        "| First Named Insured | Line of Business | Carrier | Policy # | Effective | Expiration | Premium | Limit | Deductible | Form | Layer | Comments |",
        "|---------------------|------------------|---------|----------|-----------|------------|---------|-------|------------|------|-------|----------|",
    ]

    for r in rows:
        pnum = r["Policy Number"] or ""
        form = r["Form"] or ""
        layer = r["Layer"] or "Primary"
        comments = (r["Comments"] or "").replace("|", "\\|")
        named = (r["First Named Insured"] or "").replace("|", "\\|")
        lines.append(
            f"| {named} | {r['Line of Business']} | {r['Carrier']} | {pnum} | {r['Effective']} | {r['Expiration']}"
            f" | {fmt_currency(r['Premium'])} | {fmt_limit(r['Limit'])} | {fmt_limit(r['Deductible'])}"
            f" | {form} | {layer} | {comments} |"
        )

    lines += [
        "",
        f"**Total Annual Premium:** {fmt_currency(total)}",
        f"**Policies:** {len(rows)}",
    ]
    return "\n".join(lines)


def export_schedule_csv(conn: sqlite3.Connection, client_id: int) -> str:
    rows = _schedule_rows_for_client(conn, client_id)
    if not rows:
        return ""
    buf = io.StringIO()
    cols = [k for k in rows[0].keys() if k != "client_name"]
    writer = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow({k: r[k] for k in cols})
    return buf.getvalue()


def export_schedule_json(conn: sqlite3.Connection, client_id: int, client_name: str) -> str:
    rows = _schedule_rows_for_client(conn, client_id)
    policies = []
    for r in rows:
        d = dict(r)
        d.pop("client_name", None)
        policies.append(d)
    return json.dumps({
        "client": client_name,
        "prepared": TODAY,
        "total_premium": sum(r["Premium"] or 0 for r in rows),
        "policies": policies,
    }, indent=2, default=str)


# ─── LLM CLIENT EXPORT ───────────────────────────────────────────────────────

def export_llm_client_md(conn: sqlite3.Connection, client_id: int) -> str:
    from collections import defaultdict
    from policydb.analysis import layer_notation as _ln

    client = conn.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    summary = conn.execute("SELECT * FROM v_client_summary WHERE id = ?", (client_id,)).fetchone()
    policies = conn.execute(
        "SELECT * FROM v_policy_status WHERE client_id = ? ORDER BY policy_type, layer_position",
        (client_id,),
    ).fetchall()
    pipeline = conn.execute(
        "SELECT * FROM v_renewal_pipeline WHERE client_name = ? ORDER BY expiration_date",
        (client["name"],),
    ).fetchall()
    activities = conn.execute(
        """SELECT a.*, c.name AS client_name, p.policy_uid, p.policy_type AS policy_type_ref
           FROM activity_log a
           JOIN clients c ON a.client_id = c.id
           LEFT JOIN policies p ON a.policy_id = p.id
           WHERE a.client_id = ? AND a.activity_date >= date('now', '-180 days')
           ORDER BY a.activity_date DESC""",
        (client_id,),
    ).fetchall()
    overdue = conn.execute(
        "SELECT * FROM v_overdue_followups WHERE client_name = ?",
        (client["name"],),
    ).fetchall()
    upcoming_followups = conn.execute(
        """SELECT a.id, a.activity_date, a.activity_type, a.subject, a.follow_up_date,
                  a.contact_person, p.policy_uid
           FROM activity_log a
           LEFT JOIN policies p ON a.policy_id = p.id
           WHERE a.client_id = ? AND a.follow_up_date >= date('now') AND a.follow_up_done = 0
           ORDER BY a.follow_up_date ASC""",
        (client_id,),
    ).fetchall()
    history = conn.execute(
        "SELECT * FROM premium_history WHERE client_id = ? ORDER BY policy_type, term_effective DESC",
        (client_id,),
    ).fetchall()
    project_notes_rows = conn.execute(
        "SELECT name AS project_name, notes FROM projects WHERE client_id = ? AND notes != '' ORDER BY name",
        (client_id,),
    ).fetchall()
    contacts = conn.execute(
        """SELECT co.name, co.email, co.phone, co.mobile, co.organization,
                  cca.title, cca.role, cca.assignment, cca.contact_type, cca.notes, cca.is_primary, cca.created_at
           FROM contact_client_assignments cca
           JOIN contacts co ON cca.contact_id = co.id
           WHERE cca.client_id = ?
           ORDER BY cca.is_primary DESC, co.name""",
        (client_id,),
    ).fetchall()
    scratchpad_row = conn.execute(
        "SELECT content FROM client_scratchpad WHERE client_id = ?", (client_id,)
    ).fetchone()
    audit = build_program_audit(conn, client_id)

    policy_contacts_rows = conn.execute(
        """SELECT cpa.policy_id, co.name, co.email, co.phone, cpa.role, p.policy_uid
           FROM contact_policy_assignments cpa
           JOIN contacts co ON cpa.contact_id = co.id
           JOIN policies p ON cpa.policy_id = p.id
           WHERE p.client_id = ? ORDER BY co.name""",
        (client_id,),
    ).fetchall()

    client_dict = dict(client)
    s = summary
    total_premium = s["total_premium"] if s else 0
    total_commission = s["total_commission"] if s else 0
    total_fees = s["total_fees"] if s else 0
    total_revenue = s["total_revenue"] if s else 0
    next_pol = pipeline[0] if pipeline else None

    # Build lookup maps
    project_notes_map = {r["project_name"]: r["notes"] for r in project_notes_rows}

    # Policy contacts grouped by policy_id
    from collections import defaultdict as _dd
    policy_contacts_map: dict = _dd(list)
    for pc in policy_contacts_rows:
        policy_contacts_map[pc["policy_id"]].append(dict(pc))

    # Primary location address per project (from most recent policy)
    project_addresses: dict[str, str] = {}
    for p in sorted([dict(r) for r in policies], key=lambda x: x.get("id", 0)):
        proj = (p.get("project_name") or "").strip()
        if proj not in project_addresses:
            parts = [p.get("exposure_address"), p.get("exposure_city"),
                     p.get("exposure_state"), p.get("exposure_zip")]
            addr = ", ".join(x for x in parts if x)
            if addr:
                project_addresses[proj] = addr

    # Tower layers grouped by project then tower_group
    tower_by_project: dict = defaultdict(lambda: defaultdict(list))
    for p_dict in [dict(p) for p in policies]:
        tg = p_dict.get("tower_group")
        if tg:
            proj = (p_dict.get("project_name") or "").strip()
            tower_by_project[proj][tg].append(p_dict)

    # Group policies by project
    policy_groups: dict[str, list] = defaultdict(list)
    for p in policies:
        proj = (dict(p).get("project_name") or "").strip()
        policy_groups[proj].append(dict(p))

    # Sort projects: named A-Z, blank ("Corporate / Standalone") last
    sorted_projects = sorted(policy_groups.keys(), key=lambda x: "\xff" if not x else x.lower())

    # ─── YAML frontmatter ────────────────────────────────────────────────────
    lines = [
        "---",
        "export_type: client_program_summary",
        f'client: "{client["name"]}"',
        f'industry: "{client["industry_segment"]}"',
        f'account_executive: "{client["account_exec"]}"',
        f'export_date: "{TODAY}"',
        f"total_policies: {len(policies)}",
        f"total_annual_premium: {fmt_currency(total_premium)}",
        f"estimated_annual_revenue: {fmt_currency(total_revenue)}",
    ]
    if next_pol:
        lines.append(
            f'next_renewal: "{next_pol["expiration_date"]} ({next_pol["days_to_renewal"]}d'
            f' — {next_pol["policy_uid"]}, {next_pol["policy_type"]})"'
        )
    lines += ["---", ""]

    # ─── Title + business description ────────────────────────────────────────
    lines += [f"# Client Program Summary: {client['name']}", ""]

    if client_dict.get("business_description"):
        lines += [client_dict["business_description"], ""]

    meta = f"**Industry:** {client['industry_segment']}"
    if client_dict.get("cn_number"):
        meta += f" | **CN:** {client['cn_number']}"
    meta += f" | **Onboarded:** {client['date_onboarded']} | **Account Executive:** {client['account_exec']}"
    if client_dict.get("client_since"):
        meta += f" | **Client Since:** {client_dict['client_since']}"
    if client_dict.get("website"):
        meta += f" | **Web:** {client_dict['website']}"
    if client_dict.get("preferred_contact_method"):
        meta += f" | **Preferred Contact:** {client_dict['preferred_contact_method']}"
    if client_dict.get("referral_source"):
        meta += f" | **Source:** {client_dict['referral_source']}"
    lines += [meta, ""]

    if client_dict.get("renewal_month"):
        import calendar
        month_name = calendar.month_name[client_dict["renewal_month"]]
        lines += [f"*Typical renewal month: {month_name}*", ""]

    # Contacts
    if contacts:
        lines += ["### Contacts", ""]
        for c in contacts:
            c = dict(c)
            primary_marker = " ★" if c.get("is_primary") else ""
            entry = f"- **{c['name']}{primary_marker}**"
            if c.get("title"):
                entry += f", {c['title']}"
            if c.get("role"):
                entry += f" ({c['role']})"
            if c.get("contact_type") == "internal":
                if c.get("assignment"):
                    entry += f" — {c['assignment']}"
                else:
                    entry += " [internal]"
            if c.get("email"):
                entry += f" — {c['email']}"
            if c.get("phone"):
                entry += f" · {c['phone']}"
            lines.append(entry)
        lines.append("")
    elif client_dict.get("primary_contact"):
        contact_line = f"- **{client['primary_contact']}**"
        if client_dict.get("contact_email"):
            contact_line += f" — {client['contact_email']}"
        if client_dict.get("contact_phone"):
            contact_line += f" · {client['contact_phone']}"
        lines += ["### Contacts", "", contact_line, ""]

    if client_dict.get("notes"):
        lines += ["### Internal Notes", "", client_dict["notes"], ""]

    scratchpad_content = (scratchpad_row["content"] if scratchpad_row else "").strip()
    if scratchpad_content:
        lines += ["### Working Notes", "", scratchpad_content, ""]

    # Risk / Exposure Profile
    risk_rows = conn.execute(
        """SELECT r.id, r.category, r.description, r.severity, r.has_coverage, r.policy_uid, r.notes,
                  r.source, r.review_date, r.identified_date
           FROM client_risks r WHERE r.client_id=?
           ORDER BY CASE r.severity WHEN 'Critical' THEN 0 WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 ELSE 3 END,
                    r.category""",
        (client_id,),
    ).fetchall()
    if risk_rows:
        lines += ["### Risk & Exposure Profile", ""]
        lines.append("| Category | Severity | Coverage | Description | Source | Notes |")
        lines.append("|----------|----------|----------|-------------|--------|-------|")
        for rr in risk_rows:
            rr = dict(rr)
            cov = f"Covered ({rr['policy_uid']})" if rr["has_coverage"] and rr["policy_uid"] else "Covered" if rr["has_coverage"] else "**GAP**"
            lines.append(f"| {rr['category']} | {rr['severity']} | {cov} | {rr['description'] or '—'} | {rr['source'] or '—'} | {rr['notes'] or '—'} |")
        lines.append("")
        # Coverage lines per risk
        for rr in risk_rows:
            rr = dict(rr)
            cl_rows = conn.execute(
                "SELECT coverage_line, adequacy, policy_uid, notes FROM risk_coverage_lines WHERE risk_id=? ORDER BY coverage_line",
                (rr["id"],),
            ).fetchall()
            if cl_rows:
                lines.append(f"**{rr['category']} — Coverage Lines:**")
                for cl in cl_rows:
                    cl = dict(cl)
                    pol = f" ({cl['policy_uid']})" if cl["policy_uid"] else ""
                    lines.append(f"- {cl['coverage_line']} — {cl['adequacy']}{pol}{' — ' + cl['notes'] if cl['notes'] else ''}")
                lines.append("")
            ctrl_rows = conn.execute(
                "SELECT control_type, description, status, responsible, target_date FROM risk_controls WHERE risk_id=? ORDER BY created_at",
                (rr["id"],),
            ).fetchall()
            if ctrl_rows:
                lines.append(f"**{rr['category']} — Controls:**")
                for ct in ctrl_rows:
                    ct = dict(ct)
                    resp = f" ({ct['responsible']})" if ct["responsible"] else ""
                    tgt = f" by {ct['target_date']}" if ct["target_date"] else ""
                    lines.append(f"- [{ct['status']}] {ct['description']} — {ct['control_type']}{resp}{tgt}")
                lines.append("")
        gaps = [dict(rr)["category"] for rr in risk_rows if not dict(rr).get("has_coverage")]
        if gaps:
            lines += [f"**Coverage gaps identified:** {', '.join(gaps)}", ""]

    # ─── Program Overview ────────────────────────────────────────────────────
    rev_detail = ""
    if total_commission and total_fees:
        rev_detail = f" ({fmt_currency(total_commission)} commission + {fmt_currency(total_fees)} flat fee)"
    elif total_commission:
        rev_detail = " (commission)"
    elif total_fees:
        rev_detail = " (flat fee)"

    lines += [
        "## Program Overview",
        "",
        f"- **Total annual premium:** {fmt_currency(total_premium)}",
        f"- **Estimated annual revenue:** {fmt_currency(total_revenue)}{rev_detail}",
        f"- **Active policies:** {len(policies)} ({audit['standalone_count']} standalone)",
        f"- **Coverage lines:** {', '.join(audit['coverage_lines'])}",
        f"- **Carriers:** {', '.join(audit['carriers'])}",
        "",
    ]
    if next_pol:
        lines += [
            f"Next renewal: **{next_pol['policy_type']}** ({next_pol['policy_uid']}) with "
            f"{next_pol['carrier']} — expires {next_pol['expiration_date']}, "
            f"{next_pol['days_to_renewal']} days out.",
            "",
        ]

    # ─── Renewal Pipeline ────────────────────────────────────────────────────
    if pipeline:
        lines += [
            "## Renewal Pipeline (Next 180 Days)",
            "",
            "| Policy | Line of Business | Carrier | Expires | Days | Premium | Urgency | Status | Colleague |",
            "|--------|------------------|---------|---------|------|---------|---------|--------|-----------|",
        ]
        for r in pipeline:
            lines.append(
                f"| {r['policy_uid']} | {r['policy_type']} | {r['carrier']}"
                f" | {r['expiration_date']} | {r['days_to_renewal']}"
                f" | {fmt_currency(r['premium'])} | {r['urgency']}"
                f" | {r['renewal_status']} | {r['placement_colleague'] or '—'} |"
            )
        lines.append("")

    # ─── Insurance Program (grouped by project) ───────────────────────────
    lines += ["## Insurance Program", ""]

    for proj in sorted_projects:
        group = policy_groups[proj]
        display_name = proj if proj else "Corporate / Standalone"
        lines += [f"### {display_name}", ""]

        addr = project_addresses.get(proj, "")
        if addr:
            lines += [f"**Location:** {addr}", ""]

        note = project_notes_map.get(proj, "")
        if note:
            lines += [note, ""]

        # Policy summary table
        lines += [
            "| UID | Line of Business | Carrier | Policy # | Effective | Expires | Premium | Limit | Ded. | Status |",
            "|-----|------------------|---------|----------|-----------|---------|---------|-------|------|--------|",
        ]
        for p in group:
            lines.append(
                f"| {p['policy_uid']} | {p['policy_type']} | {p['carrier']}"
                f" | {p.get('policy_number') or '—'} | {p['effective_date']} | {p['expiration_date']}"
                f" | {fmt_currency(p['premium'])} | {fmt_limit(p.get('limit_amount'))}"
                f" | {fmt_limit(p.get('deductible'))} | {p['renewal_status']} |"
            )
        lines.append("")

        # Per-policy narrative detail (description, placement, team, internal notes)
        for p in group:
            details = []
            if p.get("first_named_insured") and p["first_named_insured"] != client["name"]:
                details.append(f"First Named Insured: {p['first_named_insured']}")
            if p.get("access_point"):
                details.append(f"Access Point: {p['access_point']}")
            if p.get("description"):
                details.append(p["description"])
            # Policy team contacts (new system)
            team = policy_contacts_map.get(p.get("id", 0), [])
            if team:
                team_parts = []
                for tm in team:
                    part = tm["name"]
                    if tm.get("role"):
                        part += f" ({tm['role']})"
                    if tm.get("email"):
                        part += f" <{tm['email']}>"
                    team_parts.append(part)
                details.append("Team: " + "; ".join(team_parts))
            elif p.get("placement_colleague"):
                col = f"Placement: {p['placement_colleague']}"
                if p.get("placement_colleague_email"):
                    col += f" ({p['placement_colleague_email']})"
                details.append(col)
            if p.get("underwriter_name"):
                uw = f"Underwriter: {p['underwriter_name']}"
                if p.get("underwriter_contact"):
                    uw += f" ({p['underwriter_contact']})"
                details.append(uw)
            if p.get("notes"):
                details.append(f"Internal notes: {p['notes']}")
            if details:
                lines.append(f"**{p['policy_uid']} — {p['policy_type']}:** " + " | ".join(details))
        has_details = any(
            p.get("description") or p.get("placement_colleague") or p.get("notes")
            or p.get("underwriter_name") or p.get("first_named_insured")
            or p.get("access_point") or policy_contacts_map.get(p.get("id", 0))
            for p in group
        )
        if has_details:
            lines.append("")

        # Tower structure for this project (inline)
        proj_towers = tower_by_project.get(proj, {})
        if proj_towers:
            for tg_name in sorted(proj_towers.keys()):
                layers = sorted(
                    proj_towers[tg_name],
                    key=lambda lp: lp.get("attachment_point") or 0,
                    reverse=True,
                )
                total_tower_premium = sum(lp.get("premium") or 0 for lp in layers)
                lines += [
                    f"**Tower: {tg_name}** ({fmt_currency(total_tower_premium)} total, {len(layers)} carrier{'s' if len(layers) != 1 else ''})",
                    "",
                    "| Layer | Carrier | Limit | Premium | Colleague |",
                    "|-------|---------|-------|---------|-----------|",
                ]
                for lp in layers:
                    notation = _ln(lp.get("limit_amount"), lp.get("attachment_point"), lp.get("participation_of"))
                    lines.append(
                        f"| {notation or lp.get('layer_position', 'Primary')}"
                        f" | {lp['carrier']} | {fmt_limit(lp.get('limit_amount'))}"
                        f" | {fmt_currency(lp['premium'])} | {lp.get('placement_colleague') or '—'} |"
                    )
                lines.append("")

    # ─── Premium History ──────────────────────────────────────────────────────
    if history:
        lines += ["## Premium History", ""]
        current_type = None
        for r in history:
            if r["policy_type"] != current_type:
                current_type = r["policy_type"]
                lines += [
                    f"### {current_type}",
                    "",
                    "| Term | Carrier | Premium | Limit | Notes |",
                    "|------|---------|---------|-------|-------|",
                ]
            lines.append(
                f"| {r['term_effective']} → {r['term_expiration']}"
                f" | {r['carrier'] or '—'} | {fmt_currency(r['premium'])}"
                f" | {fmt_limit(r['limit_amount'])} | {r['notes'] or '—'} |"
            )
        lines.append("")

    # ─── Account Activity ────────────────────────────────────────────────────
    if activities:
        lines += [
            "## Account Activity (Last 180 Days)",
            "",
            "| Date | Type | Policy | Contact | Subject | Follow-Up |",
            "|------|------|--------|---------|---------|-----------|",
        ]
        detail_blocks = []
        for a in activities:
            subject = (a["subject"] or "").replace("|", "\\|")
            policy_ref = (a["policy_uid"] or "")
            lines.append(
                f"| {a['activity_date']} | {a['activity_type']}"
                f" | {policy_ref or '—'} | {a['contact_person'] or '—'} | {subject}"
                f" | {a['follow_up_date'] or '—'} |"
            )
            if a["details"] and a["details"].strip():
                detail_blocks.append((a["activity_date"], a["subject"], a["details"].strip(), policy_ref))
        lines.append("")
        if detail_blocks:
            lines += ["### Activity Notes", ""]
            for date, subject, details, policy_ref in detail_blocks:
                header = f"**{date} — {subject}**"
                if policy_ref:
                    header += f" [{policy_ref}]"
                lines.append(header)
                lines.append(f"{details}")
                lines.append("")

    # ─── Opportunities ────────────────────────────────────────────────────────
    opportunities = [p for p in [dict(r) for r in policies] if p.get("is_opportunity")]
    if opportunities:
        lines += ["## Opportunities / New Business", "", "| UID | Line | Status | Target Effective | Premium |",
                  "|-----|------|--------|-----------------|---------|"]
        for op in opportunities:
            lines.append(
                f"| {op['policy_uid']} | {op['policy_type']} | {op.get('opportunity_status') or '—'}"
                f" | {op.get('target_effective_date') or '—'} | {fmt_currency(op['premium'])} |"
            )
        lines.append("")

    # ─── Open Follow-Ups ─────────────────────────────────────────────────────
    if overdue or upcoming_followups:
        lines += ["## Open Follow-Ups", ""]
        if overdue:
            lines += ["**Overdue:**", ""]
            for o in overdue:
                lines.append(
                    f"- {o['follow_up_date']} ({o['days_overdue']}d overdue) — "
                    f"{o['activity_type']}: {o['subject']}"
                )
            lines.append("")
        if upcoming_followups:
            lines += ["**Upcoming:**", ""]
            for u in upcoming_followups:
                u = dict(u)
                entry = f"- {u['follow_up_date']} — {u['activity_type']}: {u['subject']}"
                if u.get("contact_person"):
                    entry += f" (w/ {u['contact_person']})"
                if u.get("policy_uid"):
                    entry += f" [{u['policy_uid']}]"
                lines.append(entry)
            lines.append("")

    return "\n".join(lines)


def export_llm_client_json(conn: sqlite3.Connection, client_id: int) -> str:
    client = conn.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    summary = conn.execute("SELECT * FROM v_client_summary WHERE id = ?", (client_id,)).fetchone()
    policies = conn.execute(
        "SELECT * FROM v_policy_status WHERE client_id = ?", (client_id,)
    ).fetchall()
    activities = conn.execute(
        """SELECT a.* FROM activity_log a WHERE a.client_id = ?
           AND a.activity_date >= date('now', '-90 days')
           ORDER BY a.activity_date DESC""",
        (client_id,),
    ).fetchall()
    overdue = conn.execute(
        "SELECT * FROM v_overdue_followups WHERE client_name = ?", (client["name"],)
    ).fetchall()
    upcoming_fups = conn.execute(
        """SELECT a.id, a.activity_date, a.activity_type, a.subject, a.follow_up_date,
                  a.contact_person, p.policy_uid
           FROM activity_log a
           LEFT JOIN policies p ON a.policy_id = p.id
           WHERE a.client_id = ? AND a.follow_up_date >= date('now') AND a.follow_up_done = 0
           ORDER BY a.follow_up_date ASC""",
        (client_id,),
    ).fetchall()
    history = conn.execute(
        "SELECT * FROM premium_history WHERE client_id = ? ORDER BY policy_type, term_effective DESC",
        (client_id,),
    ).fetchall()
    proj_notes = conn.execute(
        "SELECT name AS project_name, notes FROM projects WHERE client_id = ? AND notes != '' ORDER BY name",
        (client_id,),
    ).fetchall()
    contacts = conn.execute(
        """SELECT co.name, co.email, co.phone, co.mobile, co.organization,
                  cca.title, cca.role, cca.assignment, cca.contact_type, cca.notes, cca.is_primary, cca.created_at
           FROM contact_client_assignments cca
           JOIN contacts co ON cca.contact_id = co.id
           WHERE cca.client_id = ?
           ORDER BY cca.is_primary DESC, co.name""",
        (client_id,),
    ).fetchall()
    policy_contacts = conn.execute(
        """SELECT cpa.policy_id, co.name, co.email, co.phone, cpa.role, p.policy_uid
           FROM contact_policy_assignments cpa
           JOIN contacts co ON cpa.contact_id = co.id
           JOIN policies p ON cpa.policy_id = p.id
           WHERE p.client_id = ? ORDER BY p.policy_uid, co.name""",
        (client_id,),
    ).fetchall()
    scratchpad = conn.execute(
        "SELECT content, updated_at FROM client_scratchpad WHERE client_id = ?", (client_id,)
    ).fetchone()
    risks = conn.execute(
        """SELECT id, category, description, severity, has_coverage, policy_uid, notes,
                  source, review_date, identified_date
           FROM client_risks WHERE client_id=?
           ORDER BY CASE severity WHEN 'Critical' THEN 0 WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 ELSE 3 END,
                    category""",
        (client_id,),
    ).fetchall()
    audit = build_program_audit(conn, client_id)

    # Group policy_contacts by policy_uid
    from collections import defaultdict as _dd2
    pc_by_uid: dict = _dd2(list)
    for pc in policy_contacts:
        pc_by_uid[pc["policy_uid"]].append(dict(pc))

    data = {
        "metadata": {
            "export_type": "client_program",
            "date": TODAY,
            "client": client["name"],
        },
        "client": dict(client),
        "summary": dict(summary) if summary else {},
        "contacts": [dict(c) for c in contacts],
        "working_notes": scratchpad["content"] if scratchpad else "",
        "project_notes": [dict(r) for r in proj_notes],
        "policies": [
            {
                **dict(p),
                "team": pc_by_uid.get(p["policy_uid"], []),
                "computed": {
                    "days_to_renewal": p["days_to_renewal"],
                    "urgency": p["urgency"],
                    "commission_amount": p["commission_amount"],
                    "rate_change": p["rate_change"],
                },
            }
            for p in policies
        ],
        "risk_profile": [
            {
                **dict(r),
                "coverage_lines": [dict(cl) for cl in conn.execute(
                    "SELECT coverage_line, adequacy, policy_uid, notes FROM risk_coverage_lines WHERE risk_id=?",
                    (r["id"],),
                ).fetchall()],
                "controls": [dict(ct) for ct in conn.execute(
                    "SELECT control_type, description, status, responsible, target_date FROM risk_controls WHERE risk_id=?",
                    (r["id"],),
                ).fetchall()],
            }
            for r in risks
        ],
        "coverage_analysis": {
            "gaps": audit["gap_observations"],
            "coverage_gaps_from_risks": [dict(r)["category"] for r in risks if not dict(r).get("has_coverage")],
            "tower_count": audit["tower_count"],
            "standalone_count": audit["standalone_count"],
            "duplicate_count": audit["duplicate_count"],
        },
        "premium_history": [dict(r) for r in history],
        "activities": [dict(a) for a in activities],
        "overdue_followups": [dict(o) for o in overdue],
        "upcoming_followups": [dict(u) for u in upcoming_fups],
    }
    return json.dumps(data, indent=2, default=str)


# ─── LLM BOOK EXPORT ─────────────────────────────────────────────────────────

def export_llm_book_md(conn: sqlite3.Connection) -> str:
    clients = conn.execute("SELECT * FROM v_client_summary ORDER BY name").fetchall()
    all_policies = conn.execute("SELECT * FROM v_policy_status").fetchall()
    pipeline = conn.execute(
        "SELECT * FROM v_renewal_pipeline ORDER BY expiration_date"
    ).fetchall()
    overdue = conn.execute("SELECT * FROM v_overdue_followups").fetchall()

    total_premium = sum(c["total_premium"] for c in clients)
    total_commission = sum(c["total_commission"] for c in clients)

    urgency_counts: dict[str, dict] = {}
    for p in all_policies:
        u = p["urgency"]
        if u not in urgency_counts:
            urgency_counts[u] = {"count": 0, "premium": 0}
        urgency_counts[u]["count"] += 1
        urgency_counts[u]["premium"] += p["premium"] or 0

    lines = [
        "---",
        "export_type: book_of_business",
        f'account_executive: "{cfg.get("default_account_exec", "Grant")}"',
        f'export_date: "{TODAY}"',
        f"total_clients: {len(clients)}",
        f"total_policies: {len(all_policies)}",
        f"total_premium: {fmt_currency(total_premium)}",
        f"total_commission: {fmt_currency(total_commission)}",
        f"urgent_renewals: {urgency_counts.get('URGENT', {}).get('count', 0)}",
        "---",
        "",
        "# Book of Business Summary",
        "",
        "## Key Metrics",
        f"- Clients: {len(clients)}",
        f"- Policies: {len(all_policies)}",
        f"- Total premium: {fmt_currency(total_premium)}",
        f"- Total commission: {fmt_currency(total_commission)}",
        "",
        "## Urgency Summary",
        "",
    ]
    for urgency_level in ["EXPIRED", "URGENT", "WARNING", "UPCOMING", "OK"]:
        data = urgency_counts.get(urgency_level, {"count": 0, "premium": 0})
        lines.append(f"- {urgency_level}: {data['count']} policies, {fmt_currency(data['premium'])}")
    lines.append("")

    # Client summaries
    lines += ["## Client Summaries", ""]
    lines += [
        "| Client | Segment | Policies | Premium | Next Renewal | Flags |",
        "|--------|---------|----------|---------|--------------|-------|",
    ]
    for c in clients:
        flags = []
        if c["next_renewal_days"] is not None and c["next_renewal_days"] <= 90:
            flags.append("URGENT RENEWAL")
        if c["activity_last_90d"] == 0:
            flags.append("NO RECENT ACTIVITY")
        lines.append(
            f"| {c['name']} | {c['industry_segment']}"
            f" | {c['total_policies']} | {fmt_currency(c['total_premium'])}"
            f" | {fmt_days(c['next_renewal_days'])} | {', '.join(flags) or '—'} |"
        )
    lines.append("")

    # Stale renewals
    stale = [
        p for p in pipeline
        if p["renewal_status"] == "Not Started"
    ]
    if stale:
        lines += ["## Stale Renewals (Not Started — Within 180d)", ""]
        for p in stale:
            lines.append(f"- {p['client_name']} / {p['policy_type']} ({p['policy_uid']}) — expires {p['expiration_date']}, {p['days_to_renewal']}d")
        lines.append("")

    # Overdue follow-ups
    if overdue:
        lines += ["## Overdue Follow-Ups", ""]
        for o in overdue:
            lines.append(f"- {o['client_name']}: {o['subject']} (due {o['follow_up_date']}, {o['days_overdue']}d overdue)")
        lines.append("")

    return "\n".join(lines)


def export_llm_book_json(conn: sqlite3.Connection) -> str:
    clients = conn.execute("SELECT * FROM v_client_summary ORDER BY name").fetchall()
    all_policies = conn.execute("SELECT * FROM v_policy_status").fetchall()
    pipeline = conn.execute("SELECT * FROM v_renewal_pipeline ORDER BY expiration_date").fetchall()
    overdue = conn.execute("SELECT * FROM v_overdue_followups").fetchall()

    return json.dumps({
        "metadata": {"export_type": "book_of_business", "date": TODAY},
        "clients": [dict(c) for c in clients],
        "policies": [dict(p) for p in all_policies],
        "renewal_pipeline": [dict(p) for p in pipeline],
        "overdue_followups": [dict(o) for o in overdue],
    }, indent=2, default=str)


# ─── CLIENT EXPORT ────────────────────────────────────────────────────────────

def export_client_md(conn: sqlite3.Connection, client_id: int) -> str:
    """Full client profile + policies in Markdown."""
    return export_llm_client_md(conn, client_id)


def export_client_json(conn: sqlite3.Connection, client_id: int) -> str:
    return export_llm_client_json(conn, client_id)


def export_client_csv(conn: sqlite3.Connection, client_id: int) -> str:
    rows = conn.execute(
        "SELECT * FROM v_policy_status WHERE client_id = ? ORDER BY policy_type",
        (client_id,),
    ).fetchall()
    if not rows:
        return ""
    buf = io.StringIO()
    cols = list(rows[0].keys())
    writer = csv.DictWriter(buf, fieldnames=cols, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow(dict(r))
    return buf.getvalue()


# ─── RENEWAL EXPORT ───────────────────────────────────────────────────────────

def export_renewals_md(conn: sqlite3.Connection, window_days: int = 180) -> str:
    rows = conn.execute(
        """SELECT * FROM v_renewal_pipeline WHERE days_to_renewal <= ?
           ORDER BY expiration_date""",
        (window_days,),
    ).fetchall()

    # Build policy_contacts map keyed by policy_uid
    pc_rows = conn.execute(
        """SELECT p.policy_uid, co.name, cpa.role, co.email
           FROM contact_policy_assignments cpa
           JOIN contacts co ON cpa.contact_id = co.id
           JOIN policies p ON cpa.policy_id = p.id
           WHERE p.policy_uid IN (SELECT policy_uid FROM v_renewal_pipeline WHERE days_to_renewal <= ?)
           ORDER BY co.name""",
        (window_days,),
    ).fetchall()
    from collections import defaultdict as _ddr
    pc_map: dict = _ddr(list)
    for pc in pc_rows:
        pc_map[pc["policy_uid"]].append(pc["name"])

    total = sum(r["premium"] or 0 for r in rows)
    lines = [
        f"# Renewal Pipeline — Next {window_days} Days",
        "",
        f"**As of:** {TODAY}  ",
        f"**Policies:** {len(rows)}  ",
        f"**Premium at Risk:** {fmt_currency(total)}",
        "",
        "| UID | Client | Line | Carrier | Expires | Days | Urgency | Premium | Status | Team |",
        "|-----|--------|------|---------|---------|------|---------|---------|--------|------|",
    ]
    for r in rows:
        team = pc_map.get(r["policy_uid"])
        team_str = "; ".join(team) if team else (r["placement_colleague"] or "—")
        lines.append(
            f"| {r['policy_uid']} | {r['client_name']} | {r['policy_type']}"
            f" | {r['carrier']} | {r['expiration_date']} | {r['days_to_renewal']}"
            f" | {r['urgency']} | {fmt_currency(r['premium'])}"
            f" | {r['renewal_status']} | {team_str} |"
        )
    return "\n".join(lines)


def export_renewals_json(conn: sqlite3.Connection, window_days: int = 180) -> str:
    rows = conn.execute(
        """SELECT * FROM v_renewal_pipeline WHERE days_to_renewal <= ?
           ORDER BY expiration_date""",
        (window_days,),
    ).fetchall()
    return json.dumps({
        "window_days": window_days,
        "export_date": TODAY,
        "renewals": [dict(r) for r in rows],
    }, indent=2, default=str)


def export_renewals_csv(conn: sqlite3.Connection, window_days: int = 180) -> str:
    rows = conn.execute(
        """SELECT * FROM v_renewal_pipeline WHERE days_to_renewal <= ?
           ORDER BY expiration_date""",
        (window_days,),
    ).fetchall()
    if not rows:
        return ""
    buf = io.StringIO()
    cols = list(rows[0].keys())
    writer = csv.DictWriter(buf, fieldnames=cols)
    writer.writeheader()
    for r in rows:
        writer.writerow(dict(r))
    return buf.getvalue()


# ─── XLSX HELPERS ────────────────────────────────────────────────────────────

# Marsh brand palette (matches HTML copy-table in email_templates.py)
_HEADER_FILL = PatternFill("solid", fgColor="003865")
_HEADER_FONT = Font(name="Noto Sans", bold=True, color="FFFFFF", size=11)
_DATA_FONT = Font(name="Noto Sans", size=11, color="3D3C37")
_ALT_ROW_FILL = PatternFill("solid", fgColor="F7F3EE")
_BORDER_COLOR = "B9B6B1"
_THIN_BORDER = Border(
    left=Side(style="thin", color=_BORDER_COLOR),
    right=Side(style="thin", color=_BORDER_COLOR),
    top=Side(style="thin", color=_BORDER_COLOR),
    bottom=Side(style="thin", color=_BORDER_COLOR),
)

_CURRENCY_FMT = '"$"#,##0.00'
_CURRENCY_COLS = {
    "Premium", "Limit", "Deductible", "premium", "limit_amount", "deductible",
    "prior_premium", "commission_amount", "exposure_amount",
    "Prior Premium", "Commission", "Exposure Amount", "Broker Fee",
}

# Friendly header names — maps snake_case DB column names to human-readable labels.
# Headers already in friendly format (e.g. v_schedule aliases) pass through unchanged.
_FRIENDLY_HEADERS: dict[str, str] = {
    "policy_uid": "Policy ID",
    "policy_type": "Line of Business",
    "policy_number": "Policy #",
    "effective_date": "Effective",
    "expiration_date": "Expiration",
    "limit_amount": "Limit",
    "deductible": "Deductible",
    "premium": "Premium",
    "prior_premium": "Prior Premium",
    "rate_change": "Rate Change",
    "commission_rate": "Commission %",
    "commission_amount": "Commission",
    "coverage_form": "Form",
    "layer_position": "Layer",
    "tower_group": "Tower Group",
    "is_standalone": "Standalone",
    "project_name": "Location / Project",
    "renewal_status": "Status",
    "placement_colleague": "Placement Colleague",
    "placement_colleague_email": "Colleague Email",
    "underwriter_name": "Underwriter",
    "underwriter_contact": "Underwriter Contact",
    "follow_up_date": "Follow-Up Date",
    "attachment_point": "Attachment Point",
    "participation_of": "Participation",
    "exposure_basis": "Exposure Basis",
    "exposure_amount": "Exposure Amount",
    "exposure_unit": "Exposure Unit",
    "exposure_address": "Address",
    "exposure_city": "City",
    "exposure_state": "State",
    "exposure_zip": "ZIP",
    "account_exec": "Account Executive",
    "days_to_renewal": "Days to Renewal",
    "first_named_insured": "First Named Insured",
    "access_point": "Access Point",
    "contact_person": "Contact",
    "activity_date": "Date",
    "activity_type": "Type",
    "follow_up_done": "Complete",
    "created_at": "Created",
    "updated_at": "Updated",
    "cn_number": "Account #",
    "industry_segment": "Industry",
    "date_onboarded": "Onboarded",
    "broker_fee": "Broker Fee",
    "is_primary": "Primary",
    "contact_type": "Contact Type",
    "client_name": "Client",
    "term_effective": "Term Start",
    "term_expiration": "Term End",
    "has_coverage": "Covered",
    "identified_date": "Identified",
    "review_date": "Review Date",
    "scope_id": "Scope ID",
    "billing_id": "Billing ID",
    "entity_name": "Entity Name",
    "is_master": "Master Account",
    "sort_order": "Order",
    "received_at": "Received Date",
    "send_by_date": "Send By",
    "rfi_uid": "RFI ID",
    "sent_at": "Sent",
    "line_of_business": "Line of Business",
    "lead_broker": "Lead Broker",
    "milestone_profile": "Milestone Profile",
    "urgency": "Urgency",
    "notes": "Notes",
    "description": "Description",
    "carrier": "Carrier",
    "name": "Name",
    "email": "Email",
    "phone": "Phone",
    "mobile": "Mobile",
    "organization": "Organization",
    "title": "Title",
    "role": "Role",
    "assignment": "Assignment",
    "subject": "Subject",
    "details": "Details",
    "category": "Category",
    "severity": "Severity",
    "source": "Source",
    "website": "Website",
    "fein": "FEIN",
    "content": "Content",
    "status": "Status",
    "is_opportunity": "Opportunity",
    "policy_scratchpad": "Working Notes",
}


def _friendly(col_name: str) -> str:
    """Return human-readable header for a column name."""
    return _FRIENDLY_HEADERS.get(col_name, col_name)


def _write_sheet(wb: Workbook, title: str, rows: list, *, col_widths: dict[str, int] | None = None, wrap_text: bool = True) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["No data"])
        return

    raw_headers = list(rows[0].keys())
    display_headers = [_friendly(h) for h in raw_headers]
    ws.append(display_headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=wrap_text)

    for row in rows:
        ws.append([row[k] for k in raw_headers])

    # Apply styling to data cells: font, borders, alternating fills, currency
    _wrap = Alignment(wrap_text=wrap_text)
    _wrap_right = Alignment(wrap_text=wrap_text, horizontal="right")
    for col_idx, display_name in enumerate(display_headers, 1):
        is_currency = display_name in _CURRENCY_COLS
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _wrap_right if is_currency else _wrap
            if is_currency:
                cell.number_format = _CURRENCY_FMT
            # Alternating row fill (0-indexed data row: even rows get fill)
            if (row_idx - 2) % 2 == 1:
                cell.fill = _ALT_ROW_FILL

    # Column widths — use explicit overrides (matching display name) or auto-size
    for col_idx, (raw_name, display_name) in enumerate(zip(raw_headers, display_headers), 1):
        col_letter = get_column_letter(col_idx)
        if col_widths and raw_name in col_widths:
            ws.column_dimensions[col_letter].width = col_widths[raw_name]
        elif col_widths and display_name in col_widths:
            ws.column_dimensions[col_letter].width = col_widths[display_name]
        else:
            max_len = max(
                len(str(display_name)),
                *(len(str(row[raw_name] or "")) for row in rows),
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_schedule_xlsx(conn: sqlite3.Connection, client_id: int, client_name: str) -> bytes:
    rows = _schedule_rows_for_client(conn, client_id)
    # Strip client_name column — already in the filename
    cleaned = [{k: r[k] for k in r.keys() if k != "client_name"} for r in rows]
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Schedule of Insurance", cleaned)
    # Summary row
    ws = wb["Schedule of Insurance"]
    ws.append([])
    total = sum(r["Premium"] or 0 for r in rows)
    ws.append(["Total Annual Premium", fmt_currency(total)])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    # Project Notes sheet (if any)
    notes_rows = conn.execute(
        "SELECT name AS \"Project / Location\", notes AS \"Notes\" FROM projects WHERE client_id = ? AND notes != '' ORDER BY name",
        (client_id,),
    ).fetchall()
    if notes_rows:
        _write_sheet(wb, "Project Notes", [dict(r) for r in notes_rows])

    return _wb_to_bytes(wb)


def export_client_xlsx(conn: sqlite3.Connection, client_id: int) -> bytes:
    policies = conn.execute(
        "SELECT * FROM v_policy_status WHERE client_id = ? ORDER BY policy_type, layer_position",
        (client_id,),
    ).fetchall()
    history = conn.execute(
        "SELECT * FROM premium_history WHERE client_id = ? ORDER BY policy_type, term_effective DESC",
        (client_id,),
    ).fetchall()
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Policies", [dict(r) for r in policies])
    _write_sheet(wb, "Premium History", [dict(r) for r in history])
    return _wb_to_bytes(wb)


def export_full_xlsx(conn: sqlite3.Connection, client_id: int, client_name: str) -> bytes:
    """Full internal data export: all policy fields + contacts + notes + activities."""
    policies = conn.execute(
        """SELECT policy_uid, policy_type, carrier, policy_number,
                  effective_date, expiration_date, premium, limit_amount, deductible,
                  description, coverage_form, layer_position, tower_group, is_standalone,
                  project_name, renewal_status, commission_rate, commission_amount,
                  prior_premium, rate_change,
                  placement_colleague, placement_colleague_email,
                  underwriter_name, underwriter_contact,
                  follow_up_date, attachment_point, participation_of,
                  exposure_basis, exposure_amount, exposure_unit,
                  exposure_address, exposure_city, exposure_state, exposure_zip,
                  notes, account_exec, urgency, days_to_renewal,
                  first_named_insured, access_point
           FROM v_policy_status WHERE client_id = ?
           ORDER BY project_name, policy_type, layer_position""",
        (client_id,),
    ).fetchall()

    contacts = conn.execute(
        """SELECT co.name, cca.title, cca.role, cca.assignment, cca.contact_type,
                  co.email, co.phone, co.mobile, co.organization,
                  cca.notes, cca.is_primary, cca.created_at
           FROM contact_client_assignments cca
           JOIN contacts co ON cca.contact_id = co.id
           WHERE cca.client_id = ?
           ORDER BY cca.is_primary DESC, co.name""",
        (client_id,),
    ).fetchall()

    policy_team = conn.execute(
        """SELECT p.policy_uid, p.policy_type, co.name, cpa.role, co.email, co.phone
           FROM contact_policy_assignments cpa
           JOIN contacts co ON cpa.contact_id = co.id
           JOIN policies p ON cpa.policy_id = p.id
           WHERE p.client_id = ? ORDER BY p.policy_uid, co.name""",
        (client_id,),
    ).fetchall()

    project_notes = conn.execute(
        """SELECT name AS project_name, notes, created_at, updated_at
           FROM projects WHERE client_id = ? ORDER BY name""",
        (client_id,),
    ).fetchall()

    activities = conn.execute(
        """SELECT activity_date, activity_type, contact_person, subject, details,
                  follow_up_date, follow_up_done, account_exec, created_at
           FROM activity_log WHERE client_id = ?
           ORDER BY activity_date DESC""",
        (client_id,),
    ).fetchall()

    history = conn.execute(
        "SELECT policy_type, term_effective, term_expiration, carrier, premium, limit_amount, deductible, notes FROM premium_history WHERE client_id = ? ORDER BY policy_type, term_effective DESC",
        (client_id,),
    ).fetchall()

    scratchpad = conn.execute(
        "SELECT content, updated_at FROM client_scratchpad WHERE client_id = ?",
        (client_id,),
    ).fetchone()

    # Policy-level scratchpad notes (separate table keyed by policy_uid)
    policy_scratchpad_rows = conn.execute(
        """SELECT ps.policy_uid, ps.content AS notes, ps.updated_at
           FROM policy_scratchpad ps
           JOIN policies p ON ps.policy_uid = p.policy_uid
           WHERE p.client_id = ? AND ps.content != ''
           ORDER BY ps.policy_uid""",
        (client_id,),
    ).fetchall()

    # Risks register
    risks = conn.execute(
        """SELECT r.category, r.description, r.severity, r.has_coverage,
                  r.policy_uid, r.notes, r.source, r.review_date, r.identified_date
           FROM client_risks r WHERE r.client_id = ?
           ORDER BY CASE r.severity WHEN 'Critical' THEN 0 WHEN 'High' THEN 1
                    WHEN 'Medium' THEN 2 ELSE 3 END, r.category""",
        (client_id,),
    ).fetchall()

    # Saved / pinned notes
    saved_notes = conn.execute(
        """SELECT scope, scope_id, content, created_at
           FROM saved_notes
           WHERE scope = 'client' AND scope_id = ?
           ORDER BY created_at DESC""",
        (str(client_id),),
    ).fetchall()

    # Billing accounts
    billing = conn.execute(
        """SELECT billing_id, entity_name, description, is_master
           FROM billing_accounts WHERE client_id = ? ORDER BY is_master DESC, billing_id""",
        (client_id,),
    ).fetchall()

    # Client profile row (FEIN, broker_fee, key metadata)
    client_profile = conn.execute(
        """SELECT name, cn_number, fein, broker_fee, industry_segment,
                  account_exec, date_onboarded, website
           FROM clients WHERE id = ?""",
        (client_id,),
    ).fetchone()

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Policies (Full)", [dict(r) for r in policies])
    _write_sheet(wb, "Contacts", [dict(r) for r in contacts])
    _write_sheet(wb, "Policy Team", [dict(r) for r in policy_team])
    _write_sheet(wb, "Project Notes", [dict(r) for r in project_notes])
    if policy_scratchpad_rows:
        _write_sheet(wb, "Policy Notes", [dict(r) for r in policy_scratchpad_rows])
    _write_sheet(wb, "Activities", [dict(r) for r in activities])
    _write_sheet(wb, "Premium History", [dict(r) for r in history])
    if risks:
        _write_sheet(wb, "Risks", [dict(r) for r in risks])
    if saved_notes:
        _write_sheet(wb, "Saved Notes", [dict(r) for r in saved_notes])
    if billing:
        _write_sheet(wb, "Billing Accounts", [dict(r) for r in billing])

    # Client Profile sheet (key metadata including FEIN and broker_fee)
    ws_profile = wb.create_sheet("Client Profile")
    ws_profile.append(["Field", "Value"])
    ws_profile["A1"].font = Font(bold=True)
    ws_profile["B1"].font = Font(bold=True)
    if client_profile:
        for field, val in dict(client_profile).items():
            ws_profile.append([field, val])
    ws_profile.column_dimensions["A"].width = 24
    ws_profile.column_dimensions["B"].width = 40

    # Internal Notes as a simple text sheet
    client_row = conn.execute("SELECT notes FROM clients WHERE id = ?", (client_id,)).fetchone()
    ws_notes = wb.create_sheet("Internal Notes")
    ws_notes.append(["Internal Notes"])
    ws_notes["A1"].font = Font(bold=True)
    if client_row and client_row["notes"]:
        for line in client_row["notes"].splitlines():
            ws_notes.append([line])
    else:
        ws_notes.append(["(no internal notes)"])
    ws_notes.column_dimensions["A"].width = 80

    # Working Notes as a simple text sheet
    ws = wb.create_sheet("Working Notes")
    ws.append(["Working Notes"])
    ws["A1"].font = Font(bold=True)
    if scratchpad and scratchpad["content"]:
        for line in scratchpad["content"].splitlines():
            ws.append([line])
        ws.append([])
        ws.append([f"Last updated: {scratchpad['updated_at']}"])
    else:
        ws.append(["(no working notes)"])
    ws.column_dimensions["A"].width = 80

    return _wb_to_bytes(wb)


def export_renewals_xlsx(conn: sqlite3.Connection, window_days: int = 180) -> bytes:
    rows = conn.execute(
        """SELECT * FROM v_renewal_pipeline WHERE days_to_renewal <= ?
           ORDER BY expiration_date""",
        (window_days,),
    ).fetchall()
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, f"Renewals — Next {window_days}d", [dict(r) for r in rows])
    return _wb_to_bytes(wb)


# ─── REQUEST BUNDLE EXPORT ────────────────────────────────────────────────────

_RFI_COL_WIDTHS = {
    "Item": 45,
    "Coverage / Location": 35,
    "Category": 18,
    "Status": 14,
    "Received Date": 16,
    "Notes / Response": 45,
}


def _bundle_request_date(bundle: dict) -> str:
    """Return a human-readable request date from a bundle row.

    Prefers ``sent_at`` over ``created_at``.  Returns empty string when
    neither is available.
    """
    raw = bundle.get("sent_at") or bundle.get("created_at")
    if not raw:
        return ""
    try:
        dt = datetime.fromisoformat(str(raw))
        return dt.strftime("%B %d, %Y")
    except (ValueError, TypeError):
        # Already a plain date string — return as-is
        return str(raw)[:10]


def _bundle_date_label(bundle: dict) -> str:
    """Return 'Sent: <date>' or 'Created: <date>' depending on which field is used."""
    fmt = _bundle_request_date(bundle)
    if not fmt:
        return ""
    if bundle.get("sent_at"):
        return f"Sent: {fmt}"
    return f"Created: {fmt}"


def export_request_bundle_xlsx(conn, bundle_id: int) -> bytes:
    """Export a request bundle as an XLSX spreadsheet for the client."""
    bundle = conn.execute(
        "SELECT b.*, c.name AS client_name FROM client_request_bundles b JOIN clients c ON b.client_id = c.id WHERE b.id = ?",
        (bundle_id,),
    ).fetchone()
    items = conn.execute(
        """SELECT cri.*, p.policy_type, p.carrier, p.project_name AS pol_project
           FROM client_request_items cri
           LEFT JOIN policies p ON cri.policy_uid = p.policy_uid
           WHERE cri.bundle_id = ?
           ORDER BY cri.received ASC, cri.sort_order ASC, cri.id ASC""",
        (bundle_id,),
    ).fetchall()

    rows = []
    for item in items:
        i = dict(item)
        # Build a clear coverage/location reference
        ref_parts = []
        if i.get("policy_type"):
            ref_parts.append(i["policy_type"])
        if i.get("carrier"):
            ref_parts.append(i["carrier"])
        if i.get("pol_project") or i.get("project_name"):
            proj = i.get("pol_project") or i.get("project_name")
            ref_parts.append(proj)
        rows.append({
            "Item": i["description"],
            "Coverage / Location": " — ".join(ref_parts) if ref_parts else "",
            "Category": i.get("category") or "",
            "Status": "Received" if i["received"] else "Outstanding",
            "Received Date": (i.get("received_at") or "")[:10] if i["received"] else "",
            "Notes / Response": i.get("notes") or "",
        })

    wb = Workbook()
    wb.remove(wb.active)
    bundle_dict = dict(bundle) if bundle else {}
    client_name = bundle_dict.get("client_name", "Client")
    title = bundle_dict.get("title", "Request")
    _write_sheet(wb, title[:31], rows, col_widths=_RFI_COL_WIDTHS)  # sheet name max 31 chars

    # Add request date metadata above the data table
    date_label = _bundle_date_label(bundle_dict)
    if date_label:
        ws = wb[title[:31]]
        ws.insert_rows(1)
        ws["A1"] = date_label
        ws["A1"].font = Font(bold=True)

    return _wb_to_bytes(wb)


def render_request_compose_text(conn, bundle_id: int) -> str:
    """Generate formatted email body listing outstanding and received items."""
    bundle = conn.execute(
        "SELECT * FROM client_request_bundles WHERE id = ?",
        (bundle_id,),
    ).fetchone()
    items = conn.execute(
        """SELECT cri.*, p.policy_type, p.carrier, p.project_name AS pol_project
           FROM client_request_items cri
           LEFT JOIN policies p ON cri.policy_uid = p.policy_uid
           WHERE cri.bundle_id = ?
           ORDER BY cri.received ASC, cri.sort_order ASC, cri.id ASC""",
        (bundle_id,),
    ).fetchall()

    outstanding = []
    received = []
    for item in items:
        i = dict(item)
        desc = i["description"]
        context_parts = []
        if i.get("policy_type"):
            context_parts.append(i["policy_type"])
        if i.get("carrier"):
            context_parts.append(i["carrier"])
        proj = i.get("pol_project") or i.get("project_name")
        if proj:
            context_parts.append(proj)
        suffix = f" — {', '.join(context_parts)}" if context_parts else ""

        if i["received"]:
            date_str = f" (received {i['received_at'][:10]})" if i.get("received_at") else ""
            received.append(f"  ☑ {desc}{suffix}{date_str}")
        else:
            outstanding.append(f"  □ {desc}{suffix}")

    lines = []

    # Add request date at the top
    if bundle:
        date_label = _bundle_date_label(dict(bundle))
        if date_label:
            lines.append(date_label)
            lines.append("")

    if outstanding:
        lines.append(f"OUTSTANDING ({len(outstanding)} item{'s' if len(outstanding) != 1 else ''}):")
        lines.extend(outstanding)
    if received:
        if outstanding:
            lines.append("")
        lines.append("RECEIVED — thank you:")
        lines.extend(received)

    return "\n".join(lines)


def export_client_requests_xlsx(conn, client_id: int) -> bytes:
    """Export all non-complete request bundles for a client as a multi-sheet XLSX."""
    bundles = conn.execute(
        "SELECT * FROM client_request_bundles WHERE client_id=? AND status != 'complete' ORDER BY updated_at DESC",
        (client_id,),
    ).fetchall()

    wb = Workbook()
    wb.remove(wb.active)

    any_items = False
    for b in bundles:
        b = dict(b)
        items = conn.execute(
            """SELECT cri.*, p.policy_type, p.carrier, p.project_name AS pol_project
               FROM client_request_items cri
               LEFT JOIN policies p ON cri.policy_uid = p.policy_uid
               WHERE cri.bundle_id = ?
               ORDER BY cri.received ASC, cri.sort_order ASC, cri.id ASC""",
            (b["id"],),
        ).fetchall()
        rows = []
        for item in items:
            i = dict(item)
            ref_parts = []
            if i.get("policy_type"):
                ref_parts.append(i["policy_type"])
            if i.get("carrier"):
                ref_parts.append(i["carrier"])
            if i.get("pol_project") or i.get("project_name"):
                proj = i.get("pol_project") or i.get("project_name")
                ref_parts.append(proj)
            rows.append({
                "Item": i["description"],
                "Coverage / Location": " — ".join(ref_parts) if ref_parts else "",
                "Category": i.get("category") or "",
                "Status": "Received" if i["received"] else "Outstanding",
                "Received Date": (i.get("received_at") or "")[:10] if i["received"] else "",
                "Notes / Response": i.get("notes") or "",
            })
        if rows:
            any_items = True
        sheet_name = (b.get("rfi_uid") or b["title"] or "Request")[:31]
        _write_sheet(wb, sheet_name, rows, col_widths=_RFI_COL_WIDTHS)

        # Add request date metadata above the data table
        date_label = _bundle_date_label(b)
        if date_label:
            ws = wb[sheet_name]
            ws.insert_rows(1)
            ws["A1"] = date_label
            ws["A1"].font = Font(bold=True)

    if not any_items and not bundles:
        _write_sheet(wb, "Requests", [{"Item": "No outstanding items"}])

    return _wb_to_bytes(wb)


def export_rfi_by_location_xlsx(conn, client_id: int) -> bytes:
    """Export all open RFI items for a client, grouped by location (one sheet per location)."""
    from collections import defaultdict

    items = conn.execute(
        """SELECT cri.*, p.policy_type, p.carrier,
                  COALESCE(p.project_name, cri.project_name, '') AS location,
                  b.rfi_uid, b.title AS bundle_title, b.sent_at, b.created_at AS bundle_created
           FROM client_request_items cri
           JOIN client_request_bundles b ON cri.bundle_id = b.id
           LEFT JOIN policies p ON cri.policy_uid = p.policy_uid
           WHERE b.client_id = ? AND b.status != 'complete'
           ORDER BY location, cri.received ASC, cri.sort_order ASC, cri.id ASC""",
        (client_id,),
    ).fetchall()

    groups: dict[str, list[dict]] = defaultdict(list)
    for item in items:
        i = dict(item)
        loc = (i.get("location") or "").strip() or "Unassigned"
        ref_parts = []
        if i.get("policy_type"):
            ref_parts.append(i["policy_type"])
        if i.get("carrier"):
            ref_parts.append(i["carrier"])
        groups[loc].append({
            "Item": i["description"],
            "Coverage": " — ".join(ref_parts) if ref_parts else "",
            "RFI": i.get("rfi_uid") or i.get("bundle_title") or "",
            "Category": i.get("category") or "",
            "Status": "Received" if i["received"] else "Outstanding",
            "Received Date": (i.get("received_at") or "")[:10] if i["received"] else "",
            "Notes / Response": i.get("notes") or "",
        })

    _LOC_COL_WIDTHS = {
        "Item": 45, "Coverage": 30, "RFI": 16, "Category": 18,
        "Status": 14, "Received Date": 16, "Notes / Response": 45,
    }

    wb = Workbook()
    wb.remove(wb.active)
    if not groups:
        _write_sheet(wb, "Requests", [{"Item": "No outstanding items"}])
    else:
        for loc_name in sorted(groups.keys(), key=lambda x: (x == "Unassigned", x)):
            sheet_name = loc_name[:31]
            _write_sheet(wb, sheet_name, groups[loc_name], col_widths=_LOC_COL_WIDTHS)
    return _wb_to_bytes(wb)


def render_client_requests_compose_text(conn, client_id: int) -> str:
    """Generate formatted email body listing all outstanding items across all open bundles."""
    bundles = conn.execute(
        "SELECT * FROM client_request_bundles WHERE client_id=? AND status != 'complete' ORDER BY updated_at DESC",
        (client_id,),
    ).fetchall()

    bundle_count = len(bundles)
    all_lines = []
    total_outstanding = 0

    for b in bundles:
        b = dict(b)
        items = conn.execute(
            """SELECT cri.*, p.policy_type, p.carrier, p.project_name AS pol_project
               FROM client_request_items cri
               LEFT JOIN policies p ON cri.policy_uid = p.policy_uid
               WHERE cri.bundle_id = ?
               ORDER BY cri.received ASC, cri.sort_order ASC, cri.id ASC""",
            (b["id"],),
        ).fetchall()

        outstanding = []
        received = []
        for item in items:
            i = dict(item)
            desc = i["description"]
            context_parts = []
            if i.get("policy_type"):
                context_parts.append(i["policy_type"])
            if i.get("carrier"):
                context_parts.append(i["carrier"])
            proj = i.get("pol_project") or i.get("project_name")
            if proj:
                context_parts.append(proj)
            suffix = f" — {', '.join(context_parts)}" if context_parts else ""

            if i["received"]:
                date_str = f" (received {i['received_at'][:10]})" if i.get("received_at") else ""
                received.append(f"  \u2611 {desc}{suffix}{date_str}")
            else:
                outstanding.append(f"  \u25a1 {desc}{suffix}")

        total_outstanding += len(outstanding)
        rfi_label = b.get("rfi_uid") or b["title"] or "Request"
        title_label = b["title"] or "Information Request"
        date_suffix = ""
        date_label = _bundle_date_label(b)
        if date_label:
            date_suffix = f" ({date_label})"
        all_lines.append(f"\u2500\u2500\u2500 {rfi_label} \u2014 {title_label}{date_suffix} \u2500\u2500\u2500")
        if outstanding:
            all_lines.append(f"OUTSTANDING ({len(outstanding)} item{'s' if len(outstanding) != 1 else ''}):")
            all_lines.extend(outstanding)
        if received:
            if outstanding:
                all_lines.append("")
            all_lines.append("RECEIVED \u2014 thank you:")
            all_lines.extend(received)
        all_lines.append("")

    header = f"Outstanding items across {bundle_count} open request bundle{'s' if bundle_count != 1 else ''}:"
    return header + "\n\n" + "\n".join(all_lines).rstrip()


# ─── ACCOUNT SUMMARY ─────────────────────────────────────────────────────────


def build_account_summary(conn: sqlite3.Connection, client_id: int, days: int = 90, include_linked: bool = False) -> dict:
    """Build a structured account summary dict for a client."""
    from policydb.queries import (
        get_client_by_id, get_client_summary, get_activities,
        get_time_summary, get_policies_for_client,
        get_linked_group_for_client, get_all_followups,
    )
    from datetime import datetime

    client = get_client_by_id(conn, client_id)
    if not client:
        return {}
    client = dict(client)

    summary = get_client_summary(conn, client_id)
    summary = dict(summary) if summary else {}

    # Determine which client_ids to include
    client_ids = [client_id]
    linked_members = []
    linked_group = None
    if include_linked:
        linked_group = get_linked_group_for_client(conn, client_id)
        if linked_group:
            for m in linked_group.get("members", []):
                if m["client_id"] != client_id:
                    client_ids.append(m["client_id"])
                    linked_members.append({
                        "name": m["name"],
                        "total_premium": m.get("total_premium") or 0,
                        "total_policies": m.get("total_policies") or 0,
                        "next_renewal_days": m.get("next_renewal_days"),
                    })

    # Renewals (within 180 days)
    renewals = []
    for cid in client_ids:
        rows = conn.execute(
            """SELECT p.policy_uid, p.policy_type, p.carrier, p.expiration_date,
                      p.renewal_status, p.project_name,
                      CAST(julianday(p.expiration_date) - julianday('now') AS INTEGER) AS days_to_renewal
               FROM policies p
               WHERE p.client_id = ? AND p.archived = 0
                 AND (p.is_opportunity = 0 OR p.is_opportunity IS NULL)
                 AND p.expiration_date IS NOT NULL
                 AND julianday(p.expiration_date) - julianday('now') <= 180
               ORDER BY p.expiration_date""",
            (cid,),
        ).fetchall()
        # Attach milestone progress
        for r in rows:
            rd = dict(r)
            ms = conn.execute(
                "SELECT COUNT(*) AS total, SUM(CASE WHEN completed=1 THEN 1 ELSE 0 END) AS done FROM policy_milestones WHERE policy_uid=?",
                (rd["policy_uid"],),
            ).fetchone()
            rd["milestone_done"] = ms["done"] or 0
            rd["milestone_total"] = ms["total"] or 0
            if cid != client_id:
                c_row = conn.execute("SELECT name FROM clients WHERE id=?", (cid,)).fetchone()
                rd["client_name"] = c_row["name"] if c_row else ""
            renewals.append(rd)
    renewals.sort(key=lambda r: r.get("expiration_date") or "9999")

    # Follow-ups (overdue + upcoming 30d)
    overdue_all, upcoming_all = get_all_followups(conn, window=30)
    # Filter to our client_ids
    overdue_followups = [dict(r) for r in overdue_all if r.get("client_id") in client_ids]
    upcoming_followups = [dict(r) for r in upcoming_all if r.get("client_id") in client_ids]

    # Recent activity
    activities = []
    for cid in client_ids:
        acts = get_activities(conn, client_id=cid, days=days)
        for a in acts:
            ad = dict(a)
            if cid != client_id:
                ad["_from_linked"] = True
            activities.append(ad)
    activities.sort(key=lambda a: a.get("activity_date", ""), reverse=True)
    activities = activities[:20]  # cap at 20 most recent

    # Time summary
    time_data = get_time_summary(conn, client_id=client_id, days=days)

    # Coverage snapshot (all active policies)
    coverage = []
    for cid in client_ids:
        pols = get_policies_for_client(conn, cid)
        for p in pols:
            pd = dict(p)
            if not pd.get("is_opportunity"):
                coverage.append(pd)

    # High risks
    high_risks = [dict(r) for r in conn.execute(
        "SELECT * FROM client_risks WHERE client_id IN ({}) AND severity IN ('High', 'Critical') ORDER BY severity DESC".format(
            ",".join("?" * len(client_ids))
        ),
        client_ids,
    ).fetchall()]

    # Open request bundles
    open_requests = [dict(r) for r in conn.execute(
        """SELECT b.title, b.status,
                  (SELECT COUNT(*) FROM client_request_items WHERE bundle_id=b.id) AS total,
                  (SELECT COUNT(*) FROM client_request_items WHERE bundle_id=b.id AND received=0) AS outstanding
           FROM client_request_bundles b
           WHERE b.client_id IN ({}) AND b.status != 'complete'
           ORDER BY b.updated_at DESC""".format(",".join("?" * len(client_ids))),
        client_ids,
    ).fetchall()]

    # Renewal calendar — all active non-opp policies grouped by expiration month
    ph = ",".join("?" * len(client_ids))
    renewal_cal_rows = conn.execute(
        f"""SELECT strftime('%Y-%m', expiration_date) AS month_iso,
                   COUNT(*) AS count,
                   COALESCE(SUM(premium), 0) AS premium,
                   GROUP_CONCAT(policy_type, '|') AS types
            FROM policies
            WHERE client_id IN ({ph})
              AND archived = 0
              AND (is_opportunity = 0 OR is_opportunity IS NULL)
              AND expiration_date IS NOT NULL
            GROUP BY month_iso
            ORDER BY month_iso""",
        client_ids,
    ).fetchall()
    renewal_calendar = []
    for row in renewal_cal_rows:
        mi = row["month_iso"]
        try:
            from datetime import date as _d
            year, month = int(mi[:4]), int(mi[5:7])
            label = _d(year, month, 1).strftime("%b %Y")
        except (ValueError, TypeError):
            label = mi or ""
        type_counts: dict = {}
        for t in (row["types"] or "").split("|"):
            t = t.strip()
            if t:
                type_counts[t] = type_counts.get(t, 0) + 1
        renewal_calendar.append({
            "month_iso": mi,
            "month_label": label,
            "count": row["count"],
            "premium": row["premium"] or 0,
            "types": type_counts,
        })

    # Time by coverage line — hours logged per policy_type for this client
    time_by_policy = [dict(r) for r in conn.execute(
        f"""SELECT p.policy_type,
                   COALESCE(SUM(a.duration_hours), 0) AS hours,
                   COUNT(*) AS count
            FROM activity_log a
            JOIN policies p ON a.policy_id = p.id
            WHERE a.client_id IN ({ph})
              AND a.duration_hours IS NOT NULL AND a.duration_hours > 0
              AND a.activity_date >= date('now', ?)
            GROUP BY p.policy_type
            ORDER BY hours DESC""",
        client_ids + [f"-{days - 1} days"],
    ).fetchall()]

    # Notes snippet
    notes = (client.get("notes") or "")[:200]

    return {
        "client": {
            "name": client.get("name", ""),
            "cn_number": client.get("cn_number", ""),
            "industry": client.get("industry_segment", ""),
            "account_exec": cfg.get("default_account_exec", ""),
        },
        "financials": {
            "total_premium": summary.get("total_premium") or 0,
            "total_revenue": summary.get("total_revenue") or 0,
            "total_policies": summary.get("total_policies") or 0,
            "carrier_count": summary.get("carrier_count") or 0,
        },
        "renewals": renewals,
        "open_items": {
            "overdue": overdue_followups,
            "upcoming": upcoming_followups,
        },
        "recent_activity": activities,
        "time_summary": time_data,
        "coverage": coverage,
        "high_risks": high_risks,
        "open_requests": open_requests,
        "renewal_calendar": renewal_calendar,
        "time_by_policy": time_by_policy,
        "notes_snippet": notes,
        "linked_members": linked_members,
        "generated_at": datetime.now().strftime("%B %d, %Y %I:%M %p"),
        "days": days,
        "include_linked": include_linked,
    }


def render_account_summary_text(s: dict) -> str:
    """Render account summary dict as plain text for clipboard/email."""
    if not s:
        return ""

    c = s["client"]
    f = s["financials"]
    lines = [
        f"ACCOUNT SUMMARY: {c['name']}" + (f" [{c['cn_number']}]" if c['cn_number'] else ""),
        f"Generated: {s['generated_at']}",
        "\u2500" * 50,
    ]

    # Financials
    parts = []
    if f["total_premium"]:
        parts.append(f"${f['total_premium']:,.0f} Premium")
    if f["total_revenue"]:
        parts.append(f"${f['total_revenue']:,.0f} Revenue")
    if f["total_policies"]:
        parts.append(f"{f['total_policies']} Policies")
    if f["carrier_count"]:
        parts.append(f"{f['carrier_count']} Carriers")
    if parts:
        lines.append(" | ".join(parts))

    # Linked members
    if s.get("linked_members"):
        lines.append("")
        lines.append(f"LINKED ACCOUNTS ({len(s['linked_members'])})")
        for m in s["linked_members"]:
            lines.append(f"  {m['name']} \u2014 ${m['total_premium']:,.0f} premium, {m['total_policies']} policies")

    # Renewals
    if s["renewals"]:
        lines.append("")
        lines.append(f"RENEWAL STATUS (Next 180d) \u2014 {len(s['renewals'])} policies")
        for r in s["renewals"]:
            dtr = r.get("days_to_renewal")
            days_str = f"({dtr}d)" if dtr is not None else ""
            ms_str = f"[{r.get('milestone_done', 0)}/{r.get('milestone_total', 0)}]" if r.get("milestone_total") else ""
            client_prefix = f"{r['client_name']} \u2014 " if r.get("client_name") else ""
            exp = r.get("expiration_date", "")[:10] if r.get("expiration_date") else ""
            lines.append(f"  {client_prefix}{r.get('policy_type', '')} \u2014 {r.get('carrier', '')} \u2014 Exp {exp} {days_str} \u2014 {r.get('renewal_status', '')} {ms_str}")

    # Open items
    overdue = s["open_items"].get("overdue", [])
    upcoming = s["open_items"].get("upcoming", [])
    if overdue or upcoming:
        lines.append("")
        lines.append(f"OPEN ITEMS ({len(overdue) + len(upcoming)})")
        for o in overdue:
            lines.append(f"  \u26a0 {o.get('subject', '')} ({o.get('days_overdue', 0)}d overdue)")
        for u in upcoming:
            lines.append(f"  \u2192 {u.get('subject', '')} (due {u.get('follow_up_date', '')[:10]})")

    # Open requests
    if s.get("open_requests"):
        lines.append("")
        lines.append("OUTSTANDING REQUESTS")
        for req in s["open_requests"]:
            lines.append(f"  {req['title']} ({req['outstanding']} of {req['total']} items pending)")

    # Recent activity
    if s["recent_activity"]:
        total_h = s["time_summary"].get("total_hours", 0) if s.get("time_summary") else 0
        h_str = f" \u2014 {total_h:.1f}h logged" if total_h else ""
        lines.append("")
        lines.append(f"RECENT ACTIVITY (Last {s['days']}d){h_str}")
        for a in s["recent_activity"][:10]:
            dur = f" ({a['duration_hours']}h)" if a.get("duration_hours") else ""
            lines.append(f"  {a.get('activity_date', '')[:10]} {a.get('activity_type', '')} \u2014 {a.get('subject', '')}{dur}")

    # High risks
    if s.get("high_risks"):
        lines.append("")
        lines.append("HIGH RISKS")
        for r in s["high_risks"]:
            has_cov = "Covered" if r.get("has_coverage") else "No coverage"
            lines.append(f"  ! {r.get('category', '')} \u2014 {r.get('severity', '')} \u2014 {has_cov}")

    # Notes
    if s.get("notes_snippet"):
        lines.append("")
        lines.append("NOTES")
        lines.append(f"  {s['notes_snippet']}")

    return "\n".join(lines)


# ─── PROJECT GROUP EXPORT ────────────────────────────────────────────────────

def export_project_group_xlsx(
    conn: sqlite3.Connection,
    client_id: int,
    project_name: str,
    client_name: str,
) -> bytes:
    """Export policies for a single project/location as XLSX."""
    if project_name:
        policies = conn.execute(
            """SELECT policy_uid, policy_type, carrier, policy_number,
                      effective_date, expiration_date, premium, limit_amount, deductible,
                      description, coverage_form, layer_position,
                      project_name, renewal_status, commission_rate, commission_amount,
                      prior_premium, rate_change,
                      exposure_address, exposure_city, exposure_state, exposure_zip,
                      notes, urgency, days_to_renewal,
                      first_named_insured, access_point
               FROM v_policy_status
               WHERE client_id = ? AND LOWER(TRIM(COALESCE(project_name,''))) = LOWER(TRIM(?))
               ORDER BY policy_type, layer_position""",
            (client_id, project_name),
        ).fetchall()
    else:
        policies = conn.execute(
            """SELECT policy_uid, policy_type, carrier, policy_number,
                      effective_date, expiration_date, premium, limit_amount, deductible,
                      description, coverage_form, layer_position,
                      project_name, renewal_status, commission_rate, commission_amount,
                      prior_premium, rate_change,
                      exposure_address, exposure_city, exposure_state, exposure_zip,
                      notes, urgency, days_to_renewal,
                      first_named_insured, access_point
               FROM v_policy_status
               WHERE client_id = ? AND COALESCE(project_name, '') = ''
               ORDER BY policy_type, layer_position""",
            (client_id,),
        ).fetchall()

    policy_ids = [
        conn.execute("SELECT id FROM policies WHERE policy_uid = ?", (r["policy_uid"],)).fetchone()["id"]
        for r in policies
    ]

    policy_team = []
    if policy_ids:
        ph = ",".join("?" * len(policy_ids))
        policy_team = conn.execute(
            f"""SELECT p.policy_uid, p.policy_type, co.name, cpa.role, co.email, co.phone
                FROM contact_policy_assignments cpa
                JOIN contacts co ON cpa.contact_id = co.id
                JOIN policies p ON cpa.policy_id = p.id
                WHERE p.id IN ({ph}) ORDER BY p.policy_uid, co.name""",
            policy_ids,
        ).fetchall()

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Policies", [dict(r) for r in policies])

    # Total row
    ws = wb["Policies"]
    ws.append([])
    total = sum(r["premium"] or 0 for r in policies)
    ws.append(["Total Premium", fmt_currency(total)])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    if policy_team:
        _write_sheet(wb, "Policy Team", [dict(r) for r in policy_team])

    # Project note
    if project_name:
        note_row = conn.execute(
            "SELECT notes FROM projects WHERE client_id = ? AND LOWER(TRIM(name)) = LOWER(TRIM(?))",
            (client_id, project_name),
        ).fetchone()
        if note_row and note_row["notes"]:
            ws_n = wb.create_sheet("Project Notes")
            ws_n.append(["Project Notes"])
            ws_n["A1"].font = Font(bold=True)
            for line in (note_row["notes"] or "").split("\n"):
                ws_n.append([line])
            ws_n.column_dimensions["A"].width = 80

    return _wb_to_bytes(wb)


# ─── SINGLE POLICY EXPORT ───────────────────────────────────────────────────

def export_single_policy_xlsx(conn: sqlite3.Connection, policy_uid: str) -> bytes:
    """Export a single policy's full details as a multi-sheet XLSX."""
    row = conn.execute(
        "SELECT * FROM v_policy_status WHERE policy_uid = ?", (policy_uid,)
    ).fetchone()
    if not row:
        wb = Workbook()
        wb.active.append(["Policy not found"])
        return _wb_to_bytes(wb)

    d = dict(row)
    policy_id = conn.execute(
        "SELECT id FROM policies WHERE policy_uid = ?", (policy_uid,)
    ).fetchone()["id"]

    # Sheet 1: Policy Detail — transposed key/value
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Policy Detail")
    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for key, val in d.items():
        if key in ("id", "client_id"):
            continue
        ws.append([key, val])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    # Sheet 2: Policy Team
    team = conn.execute(
        """SELECT co.name, cpa.role, co.email, co.phone, co.mobile
           FROM contact_policy_assignments cpa
           JOIN contacts co ON cpa.contact_id = co.id
           WHERE cpa.policy_id = ? ORDER BY co.name""",
        (policy_id,),
    ).fetchall()
    if team:
        _write_sheet(wb, "Policy Team", [dict(r) for r in team])

    # Sheet 3: Milestones
    milestones = conn.execute(
        """SELECT milestone, completed, completed_at, is_critical
           FROM policy_milestones WHERE policy_uid = ? ORDER BY id""",
        (policy_uid,),
    ).fetchall()
    if milestones:
        _write_sheet(wb, "Milestones", [dict(r) for r in milestones])

    # Sheet 4: Activity Log
    activities = conn.execute(
        """SELECT activity_date, activity_type, contact_person, subject, details,
                  follow_up_date, follow_up_done, duration_hours
           FROM activity_log WHERE policy_id = ?
           ORDER BY activity_date DESC""",
        (policy_id,),
    ).fetchall()
    if activities:
        _write_sheet(wb, "Activity Log", [dict(r) for r in activities])

    # Sheet 5: Premium History
    history = conn.execute(
        """SELECT term_effective, term_expiration, carrier, premium,
                  limit_amount, deductible, notes
           FROM premium_history
           WHERE client_id = ? AND policy_type = ?
           ORDER BY term_effective DESC""",
        (d.get("client_id"), d.get("policy_type")),
    ).fetchall()
    if history:
        _write_sheet(wb, "Premium History", [dict(r) for r in history])

    return _wb_to_bytes(wb)


# ─── COMPLIANCE EXPORT ───────────────────────────────────────────────────────

# Status → fill colour mapping for Compliance Matrix cells
_COMPLIANCE_FILLS = {
    "compliant": PatternFill("solid", fgColor="C6EFCE"),
    "gap": PatternFill("solid", fgColor="FFC7CE"),
    "partial": PatternFill("solid", fgColor="FFEB9C"),
    "n/a": PatternFill("solid", fgColor="D9D2E9"),
    "na": PatternFill("solid", fgColor="D9D2E9"),
    "needs review": PatternFill("solid", fgColor="D9D9D9"),
    "waived": PatternFill("solid", fgColor="D9D9D9"),
}
_BOLD = Font(bold=True)
_BOLD_WHITE = Font(bold=True, color="FFFFFF")
_WRAP = Alignment(wrap_text=True, vertical="top")
_WRAP_CENTER = Alignment(wrap_text=True, horizontal="center", vertical="center")


def _filter_compliance_data(data: dict, project_ids: list[int]) -> dict:
    """Filter compliance data to only the specified location project_ids.

    Returns a new dict with filtered locations and recomputed overall_summary.
    """
    from policydb.compliance import compute_compliance_summary

    filtered_locs = [
        loc for loc in data["locations"]
        if loc["project"]["id"] in project_ids
    ]
    # Recompute overall summary from filtered set
    all_gov = {}
    for loc in filtered_locs:
        for line, g in loc.get("governing", {}).items():
            key = f"{loc['project']['id']}:{line}"
            all_gov[key] = g

    return {
        **data,
        "locations": filtered_locs,
        "overall_summary": compute_compliance_summary(all_gov),
    }


def export_compliance_xlsx(
    conn: sqlite3.Connection, client_id: int,
    project_ids: list[int] | None = None,
) -> tuple[bytes, str]:
    """Build a 5-sheet compliance workbook and return (bytes, filename)."""
    from policydb.compliance import get_client_compliance_data

    data = get_client_compliance_data(conn, client_id)
    if project_ids:
        data = _filter_compliance_data(data, project_ids)
    client_name = data.get("client_name") or ""
    if not client_name:
        row = conn.execute("SELECT name FROM clients WHERE id=?", (client_id,)).fetchone()
        client_name = row["name"] if row else f"Client_{client_id}"

    wb = Workbook()
    wb.remove(wb.active)

    _compliance_sheet_executive(wb, data, client_name)
    _compliance_sheet_matrix(wb, data)
    _compliance_sheet_gap_detail(wb, data)
    _compliance_sheet_all_requirements(wb, data)
    _compliance_sheet_cope(wb, conn, client_id, project_ids=project_ids)

    safe_name = client_name.replace(" ", "_").replace("/", "-")
    filename = f"Compliance_{safe_name}_{TODAY}.xlsx"
    return _wb_to_bytes(wb), filename


def _compliance_sheet_executive(wb: Workbook, data: dict, client_name: str) -> None:
    """Sheet 1 — Executive Summary."""
    ws = wb.create_sheet("Executive Summary")
    s = data["overall_summary"]

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    rows_to_write: list[tuple[str, Any]] = [
        ("Client", client_name),
        ("Report Date", TODAY),
        ("Overall Compliance", f"{s['compliance_pct']}%"),
        ("", ""),
        ("Total Requirements", s["total"]),
        ("Compliant", s["compliant"]),
        ("Gap", s["gap"]),
        ("Partial", s["partial"]),
        ("Waived", s["waived"]),
        ("N/A", s["na"]),
        ("Needs Review", s["needs_review"]),
        ("Locations", len(data["locations"])),
    ]

    for label, value in rows_to_write:
        ws.append([label, value])
        if label:
            ws.cell(row=ws.max_row, column=1).font = _BOLD

    # Key Findings — list every gap / partial with location and coverage line
    ws.append([])
    ws.append(["Key Findings"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

    findings: list[str] = []
    for loc in data["locations"]:
        loc_name = loc["project"].get("name", "Unknown")
        for line, gov in loc.get("governing", {}).items():
            status = (gov.get("compliance_status") or "Needs Review").lower()
            if status in ("gap", "partial"):
                findings.append(
                    f"{status.capitalize()}: {line} at {loc_name}"
                )

    if findings:
        for f in findings:
            ws.append(["", f])
    else:
        ws.append(["", "No gaps or partial compliance issues found."])


def _compliance_sheet_matrix(wb: Workbook, data: dict) -> None:
    """Sheet 2 — Compliance Matrix (coverage lines x locations)."""
    ws = wb.create_sheet("Compliance Matrix")

    locations = data["locations"]
    if not locations:
        ws.append(["No locations"])
        return

    # Collect all unique coverage lines across all locations
    all_lines: list[str] = []
    seen: set[str] = set()
    for loc in locations:
        for line in loc.get("governing", {}).keys():
            if line not in seen:
                all_lines.append(line)
                seen.add(line)

    if not all_lines:
        ws.append(["No coverage requirements"])
        return

    # Header row: "Coverage Line" + location names
    loc_names = [loc["project"].get("name", "?") for loc in locations]
    headers = ["Coverage Line"] + loc_names
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _BOLD_WHITE
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP_CENTER

    # Data rows
    for line in all_lines:
        row_vals: list[str] = [line]
        for loc in locations:
            gov = loc.get("governing", {}).get(line)
            status = (gov.get("compliance_status") or "Needs Review") if gov else ""
            row_vals.append(status)
        ws.append(row_vals)

        # Apply fill colours to status cells
        row_idx = ws.max_row
        for col_idx in range(2, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = _WRAP_CENTER
            status_key = (cell.value or "").lower()
            fill = _COMPLIANCE_FILLS.get(status_key)
            if fill:
                cell.fill = fill

    # Auto-size columns
    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def _compliance_sheet_gap_detail(wb: Workbook, data: dict) -> None:
    """Sheet 3 — Gap Detail (non-compliant rows only)."""
    gap_rows: list[dict] = []
    for loc in data["locations"]:
        loc_name = loc["project"].get("name", "Unknown")
        for line, gov in loc.get("governing", {}).items():
            status = (gov.get("compliance_status") or "Needs Review").lower()
            if status not in ("compliant", "waived", "n/a", "na"):
                sources = gov.get("source_requirements", [])
                source_names = ", ".join(
                    s.get("source_name", "") for s in sources if s.get("source_name")
                ) or gov.get("governing_source", "")
                clause_refs = ", ".join(
                    s.get("clause_ref", "") for s in sources if s.get("clause_ref")
                )
                req_limit = gov.get("required_limit") or 0
                source_notes = ", ".join(
                    s.get("source_notes", "") for s in sources if s.get("source_notes")
                )
                gap_rows.append({
                    "Location": loc_name,
                    "Coverage Line": line,
                    "Required Limit": fmt_limit(req_limit) if req_limit else "",
                    "In-Place Limit": "",
                    "Shortfall": "",
                    "Source Name": source_names,
                    "Source Clause Ref": clause_refs,
                    "Notes": gov.get("notes") or "",
                    "Source Notes": source_notes,
                })

    if gap_rows:
        _write_sheet(wb, "Gap Detail", gap_rows)
    else:
        ws = wb.create_sheet("Gap Detail")
        ws.append(["No gaps or issues found — all requirements are compliant."])


def _compliance_sheet_all_requirements(wb: Workbook, data: dict) -> None:
    """Sheet 4 — All Requirements with auto-filter."""
    all_rows: list[dict] = []
    for loc in data["locations"]:
        loc_name = loc["project"].get("name", "Unknown")
        for req in loc.get("requirements", []):
            endorsements = req.get("_endorsements_list") or []
            if isinstance(endorsements, str):
                try:
                    endorsements = json.loads(endorsements)
                except (ValueError, TypeError):
                    endorsements = [endorsements] if endorsements else []
            # Build linked policies string from junction table data
            policy_links = req.get("policy_links", [])
            if policy_links:
                linked_uids = ", ".join(lk.get("policy_uid", "") for lk in policy_links)
                link_types = ", ".join(lk.get("link_type", "direct") for lk in policy_links)
                primary_uid = next((lk["policy_uid"] for lk in policy_links if lk.get("is_primary")), "")
            else:
                linked_uids = req.get("linked_policy_uid") or ""
                link_types = "direct" if linked_uids else ""
                primary_uid = linked_uids

            all_rows.append({
                "Location": loc_name,
                "Coverage Line": req.get("coverage_line", ""),
                "Required Limit": fmt_limit(req.get("required_limit")) if req.get("required_limit") else "",
                "Max Deductible": fmt_limit(req.get("max_deductible")) if req.get("max_deductible") else "",
                "Deductible Type": req.get("deductible_type") or "",
                "Required Endorsements": ", ".join(endorsements),
                "Compliance Status": req.get("compliance_status") or "Needs Review",
                "Linked Policies": linked_uids,
                "Primary Policy": primary_uid,
                "Link Types": link_types,
                "Source Name": req.get("source_name") or "",
                "Source Clause Ref": req.get("clause_ref") or "",
                "Notes": req.get("notes") or "",
                "Source Notes": req.get("source_notes") or "",
            })

    if not all_rows:
        ws = wb.create_sheet("All Requirements")
        ws.append(["No requirements"])
        return

    _write_sheet(wb, "All Requirements", all_rows)

    # Apply auto-filter on the All Requirements sheet
    ws = wb["All Requirements"]
    ws.auto_filter.ref = ws.dimensions


def _compliance_sheet_cope(
    wb: Workbook, conn: sqlite3.Connection, client_id: int,
    project_ids: list[int] | None = None,
) -> None:
    """Sheet 5 — COPE Data (one row per location)."""
    sql = """SELECT p.name AS "Project Name",
                  COALESCE(p.address, '') || CASE WHEN p.city != '' THEN ', ' || p.city ELSE '' END
                    || CASE WHEN p.state != '' THEN ', ' || p.state ELSE '' END
                    || CASE WHEN p.zip != '' THEN ' ' || p.zip ELSE '' END AS "Address",
                  c.construction_type AS "Construction Type",
                  c.year_built AS "Year Built",
                  c.stories AS "Stories",
                  c.sq_footage AS "Sq Footage",
                  c.sprinklered AS "Sprinklered",
                  c.roof_type AS "Roof Type",
                  c.occupancy_description AS "Occupancy Description",
                  c.protection_class AS "Protection Class",
                  c.total_insurable_value AS "Total Insurable Value"
           FROM cope_data c
           JOIN projects p ON c.project_id = p.id
           WHERE p.client_id = ?"""
    params: list = [client_id]
    if project_ids:
        placeholders = ",".join("?" for _ in project_ids)
        sql += f" AND p.id IN ({placeholders})"
        params.extend(project_ids)
    sql += " ORDER BY p.name"
    cope_rows = conn.execute(sql, params).fetchall()

    if cope_rows:
        _write_sheet(wb, "COPE Data", [dict(r) for r in cope_rows])
    else:
        ws = wb.create_sheet("COPE Data")
        ws.append(["No COPE data available"])


# ─── COMPLIANCE MARKDOWN ─────────────────────────────────────────────────────


def export_compliance_md(
    conn: sqlite3.Connection, client_id: int,
    project_ids: list[int] | None = None,
) -> tuple[str, str]:
    """Build a Markdown compliance report and return (markdown_text, filename)."""
    from policydb.compliance import get_client_compliance_data

    data = get_client_compliance_data(conn, client_id)
    if project_ids:
        data = _filter_compliance_data(data, project_ids)
    client_name = data.get("client_name") or ""
    if not client_name:
        row = conn.execute("SELECT name FROM clients WHERE id=?", (client_id,)).fetchone()
        client_name = row["name"] if row else f"Client_{client_id}"

    safe_name = client_name.replace(" ", "_").replace("/", "-")
    filename = f"Compliance_{safe_name}_{TODAY}.md"

    locations = data["locations"]
    s = data["overall_summary"]
    lines: list[str] = []

    # ── Header ────────────────────────────────────────────────────────────────
    lines += [
        "# Insurance Compliance Review",
        "",
        f"**Client:** {client_name}  ",
        f"**Report Date:** {TODAY}  ",
        f"**Locations:** {len(locations)}",
        "",
        "---",
        "",
    ]

    # ── Executive Summary ─────────────────────────────────────────────────────
    lines += [
        "## Executive Summary",
        "",
        "| Metric | Value |",
        "|--------|-------|",
        f"| Overall Compliance | {s['compliance_pct']}% |",
        f"| Total Requirements | {s['total']} |",
        f"| Compliant | {s['compliant']} |",
        f"| Gaps | {s['gap']} |",
        f"| Partial | {s['partial']} |",
        f"| Waived | {s['waived']} |",
        f"| N/A | {s['na']} |",
        f"| Needs Review | {s['needs_review']} |",
        "",
    ]

    # ── Key Findings ──────────────────────────────────────────────────────────
    lines += ["## Key Findings", ""]
    findings: list[str] = []
    for loc in locations:
        loc_name = loc["project"].get("name", "Unknown")
        for line_name, gov in loc.get("governing", {}).items():
            status = (gov.get("compliance_status") or "Needs Review").lower()
            if status in ("gap", "partial"):
                findings.append(f"- **{status.capitalize()}:** {line_name} at {loc_name}")
    if findings:
        lines += findings
    else:
        lines.append("No gaps or partial compliance issues found.")
    lines.append("")

    # ── Compliance Matrix ─────────────────────────────────────────────────────
    if locations:
        all_cov_lines: list[str] = []
        seen: set[str] = set()
        for loc in locations:
            for line_name in loc.get("governing", {}).keys():
                if line_name not in seen:
                    all_cov_lines.append(line_name)
                    seen.add(line_name)

        if all_cov_lines:
            loc_names = [loc["project"].get("name", "?") for loc in locations]
            lines += ["## Compliance Matrix", ""]
            # Header row
            header = "| Coverage Line | " + " | ".join(loc_names) + " |"
            sep = "|" + "|".join(["---"] * (1 + len(loc_names))) + "|"
            lines += [header, sep]
            # Data rows
            for cov_line in all_cov_lines:
                cells = [cov_line]
                for loc in locations:
                    gov = loc.get("governing", {}).get(cov_line)
                    cells.append((gov.get("compliance_status") or "Needs Review") if gov else "")
                lines.append("| " + " | ".join(cells) + " |")
            lines.append("")

    # ── Gap Drill-Down ────────────────────────────────────────────────────────
    gap_rows: list[list[str]] = []
    for loc in locations:
        loc_name = loc["project"].get("name", "Unknown")
        for line_name, gov in loc.get("governing", {}).items():
            status = (gov.get("compliance_status") or "Needs Review").lower()
            if status not in ("compliant", "waived", "n/a", "na"):
                sources = gov.get("source_requirements", [])
                source_names = ", ".join(
                    src.get("source_name", "") for src in sources if src.get("source_name")
                ) or gov.get("governing_source", "")
                clause_refs = ", ".join(
                    src.get("clause_ref", "") for src in sources if src.get("clause_ref")
                )
                source_notes = ", ".join(
                    src.get("source_notes", "") for src in sources if src.get("source_notes")
                )
                req_limit = gov.get("required_limit") or 0
                gap_rows.append([
                    loc_name,
                    line_name,
                    fmt_limit(req_limit) if req_limit else "",
                    source_names,
                    clause_refs,
                    gov.get("notes") or "",
                    source_notes,
                ])

    if gap_rows:
        lines += [
            "## Gap Drill-Down",
            "",
            "| Location | Coverage Line | Required Limit | Source | Clause Ref | Notes | Source Notes |",
            "|----------|--------------|----------------|--------|------------|-------|-------------|",
        ]
        for row in gap_rows:
            lines.append("| " + " | ".join(row) + " |")
        lines.append("")

    # ── Per-Location Detail ───────────────────────────────────────────────────
    for loc in locations:
        proj = loc["project"]
        loc_name = proj.get("name", "Unknown")
        addr_parts = [proj.get("address", "")]
        if proj.get("city"):
            addr_parts.append(proj["city"])
        if proj.get("state"):
            addr_parts.append(proj["state"])
        if proj.get("zip"):
            addr_parts.append(proj["zip"])
        addr = ", ".join(p for p in addr_parts if p)

        loc_summary = loc.get("summary", {})
        pct = loc_summary.get("compliance_pct", 0)
        total = loc_summary.get("total", 0)
        compliant = loc_summary.get("compliant", 0)

        lines += [
            f"## Location: {loc_name}",
            "",
        ]
        if addr:
            lines.append(f"**Address:** {addr}  ")
        lines += [
            f"**Compliance:** {pct}% ({compliant}/{total})",
            "",
            "| Coverage | Req. Limit | Max Deduct. | Ded. Type | Endorsements | Status | Source | Clause Ref | Notes | Source Notes |",
            "|----------|-----------|-------------|-----------|-------------|--------|--------|------------|-------|-------------|",
        ]

        for req in loc.get("requirements", []):
            endorsements = req.get("_endorsements_list") or []
            if isinstance(endorsements, str):
                try:
                    endorsements = json.loads(endorsements)
                except (ValueError, TypeError):
                    endorsements = [endorsements] if endorsements else []
            cells = [
                req.get("coverage_line", ""),
                fmt_limit(req.get("required_limit")) if req.get("required_limit") else "",
                fmt_limit(req.get("max_deductible")) if req.get("max_deductible") else "",
                req.get("deductible_type") or "",
                ", ".join(endorsements),
                req.get("compliance_status") or "Needs Review",
                req.get("source_name") or "",
                req.get("clause_ref") or "",
                req.get("notes") or "",
                req.get("source_notes") or "",
            ]
            lines.append("| " + " | ".join(cells) + " |")
        lines.append("")

    # ── COPE Data ─────────────────────────────────────────────────────────────
    cope_sql = """SELECT p.name, p.address, p.city, p.state, p.zip,
                         c.construction_type, c.year_built, c.stories, c.sq_footage,
                         c.sprinklered, c.roof_type, c.occupancy_description,
                         c.protection_class, c.total_insurable_value
                  FROM cope_data c
                  JOIN projects p ON c.project_id = p.id
                  WHERE p.client_id = ?"""
    cope_params: list = [client_id]
    if project_ids:
        placeholders = ",".join("?" for _ in project_ids)
        cope_sql += f" AND p.id IN ({placeholders})"
        cope_params.extend(project_ids)
    cope_sql += " ORDER BY p.name"
    cope_rows_raw = conn.execute(cope_sql, cope_params).fetchall()

    if cope_rows_raw:
        lines += [
            "## COPE Data",
            "",
            "| Location | Address | Construction | Year | Stories | Sq Ft | Sprinkler | Roof | Occupancy | Prot. Class | TIV |",
            "|----------|---------|-------------|------|---------|-------|-----------|------|-----------|-------------|-----|",
        ]
        for cr in cope_rows_raw:
            cr = dict(cr)
            addr_parts = [cr.get("address") or ""]
            if cr.get("city"):
                addr_parts.append(cr["city"])
            if cr.get("state"):
                addr_parts.append(cr["state"])
            if cr.get("zip"):
                addr_parts.append(cr["zip"])
            addr = ", ".join(p for p in addr_parts if p)
            tiv = cr.get("total_insurable_value")
            cells = [
                cr.get("name") or "",
                addr,
                cr.get("construction_type") or "",
                str(cr.get("year_built") or ""),
                str(cr.get("stories") or ""),
                str(cr.get("sq_footage") or ""),
                cr.get("sprinklered") or "",
                cr.get("roof_type") or "",
                cr.get("occupancy_description") or "",
                cr.get("protection_class") or "",
                fmt_currency(tiv) if tiv else "",
            ]
            lines.append("| " + " | ".join(cells) + " |")
        lines.append("")

    return "\n".join(lines), filename


# ─── SAVE HELPER ─────────────────────────────────────────────────────────────

def save_export(content: str, filename: str) -> Path:
    exports_dir = Path(cfg.get("export_dir", str(Path.home() / ".policydb" / "exports")))
    exports_dir.mkdir(parents=True, exist_ok=True)
    out = exports_dir / filename
    out.write_text(content)
    return out


def save_export_bytes(content: bytes, filename: str) -> Path:
    exports_dir = Path(cfg.get("export_dir", str(Path.home() / ".policydb" / "exports")))
    exports_dir.mkdir(parents=True, exist_ok=True)
    out = exports_dir / filename
    out.write_bytes(content)
    return out


# ── Client Book Review XLSX ──────────────────────────────────────────────────

import re as _re

_PLACEHOLDER_RE = _re.compile(
    r"(?i)\b(tbd|tba|n/?a|pending|unknown|todo|xxx|placeholder|"
    r"see above|per above|same as|to be advised|to be determined)\b"
)

# Required fields for completeness scoring (field_name, label, is_numeric)
_COMPLETENESS_FIELDS = [
    ("carrier", "Carrier", False),
    ("policy_number", "Policy Number", False),
    ("effective_date", "Effective Date", False),
    ("expiration_date", "Expiration Date", False),
    ("premium", "Premium", True),
    ("limit_amount", "Limit", True),
    ("first_named_insured", "First Named Insured", False),
]


def _is_placeholder(val: str) -> bool:
    """Return True if value looks like a placeholder (TBD, N/A, etc.)."""
    if not val or not isinstance(val, str):
        return False
    return bool(_PLACEHOLDER_RE.search(val))


def _is_sketchy_short(val: str, field: str) -> bool:
    """Return True if a text value is suspiciously short (1-2 chars) for name-like fields."""
    if not val or not isinstance(val, str):
        return False
    name_fields = ("carrier", "first_named_insured", "placement_colleague", "underwriter_name")
    if field in name_fields and 0 < len(val.strip()) <= 2:
        return True
    return False


def _scan_sketchy_fields(policy: dict) -> list[tuple[str, str]]:
    """Scan a policy dict for placeholder/sketchy values. Returns [(field_label, value), ...]."""
    checks = [
        ("carrier", "Carrier"),
        ("policy_number", "Policy Number"),
        ("first_named_insured", "First Named Insured"),
        ("placement_colleague", "Placement Colleague"),
        ("underwriter_name", "Underwriter"),
        ("description", "Description"),
        ("coverage_form", "Coverage Form"),
    ]
    sketchy = []
    for field, label in checks:
        val = (policy.get(field) or "").strip()
        if not val:
            continue  # empty is caught by missing-fields logic
        if _is_placeholder(val) or _is_sketchy_short(val, field):
            sketchy.append((label, val))
    # Check $0 premium on non-program policies
    prem = policy.get("premium")
    if prem is not None and float(prem or 0) == 0 and not policy.get("program_id"):
        sketchy.append(("Premium", "$0"))
    return sketchy


def _compute_completeness(policy: dict) -> int:
    """Return 0-100 completeness score for a policy."""
    total = len(_COMPLETENESS_FIELDS)
    filled = 0
    for field, _label, is_numeric in _COMPLETENESS_FIELDS:
        val = policy.get(field)
        if is_numeric:
            if val is not None and float(val or 0) > 0:
                filled += 1
        else:
            sval = (str(val) if val else "").strip()
            if sval and not _is_placeholder(sval):
                filled += 1
    return round(filled / total * 100) if total else 100


def export_book_review_xlsx(conn: sqlite3.Connection, client_id: int, client_name: str) -> bytes:
    """Multi-tab XLSX workbook for team review of gaps and unknowns.

    Tabs: Instructions, Summary, All Policies, Suspected Duplicates,
    Unassigned Locations, Missing Fields, Program Review, Action Items.
    Uses friendly column labels for external team members.
    """
    from datetime import date
    from policydb.dedup import find_duplicate_candidates

    # ── Query all active policies for client ──
    policies = conn.execute(
        """SELECT p.policy_uid, p.policy_type, p.carrier, p.policy_number,
                  p.effective_date, p.expiration_date, p.premium, p.limit_amount,
                  p.deductible, p.description, p.coverage_form, p.layer_position,
                  p.tower_group, p.renewal_status, p.first_named_insured,
                  p.placement_colleague, p.underwriter_name,
                  p.exposure_address, p.project_name, p.project_id,
                  p.program_id, p.is_opportunity,
                  p.needs_investigation,
                  pr.name AS location_name,
                  pgm.name AS parent_program_name
           FROM policies p
           LEFT JOIN projects pr ON p.project_id = pr.id
           LEFT JOIN programs pgm ON p.program_id = pgm.id
           WHERE p.client_id = ? AND p.archived = 0
             AND (p.is_opportunity = 0 OR p.is_opportunity IS NULL)
           ORDER BY p.policy_type, p.carrier, p.effective_date""",
        (client_id,),
    ).fetchall()
    policies = [dict(r) for r in policies]

    # ── Query locations ──
    locations = conn.execute(
        "SELECT id, name, address, city, state FROM projects "
        "WHERE client_id = ? AND (project_type = 'Location' OR project_type IS NULL) ORDER BY name",
        (client_id,),
    ).fetchall()
    location_names = {r["id"]: r["name"] for r in locations}

    # ── Query programs from standalone programs table ──
    programs_db = conn.execute(
        "SELECT id, name, policy_type, effective_date, expiration_date FROM programs WHERE client_id = ? AND archived = 0 ORDER BY name",
        (client_id,),
    ).fetchall()
    programs_db = [dict(r) for r in programs_db]
    # Build carrier list per program from child policies
    program_child_carriers: dict[int, list[dict]] = {}
    for pgm in programs_db:
        children = conn.execute(
            "SELECT DISTINCT carrier, policy_number, premium, limit_amount FROM policies WHERE program_id = ? AND archived = 0 ORDER BY carrier",
            (pgm["id"],),
        ).fetchall()
        program_child_carriers[pgm["id"]] = [dict(c) for c in children]

    # ── Run dedup scan ──
    dedup_candidates = find_duplicate_candidates(conn, client_id)

    wb = Workbook()
    wb.remove(wb.active)

    # ════════════════════════════════════════════════════════════════════════
    # TAB 1: INSTRUCTIONS
    # ════════════════════════════════════════════════════════════════════════
    instructions = [
        {"#": 1, "Section": "HOW TO USE THIS WORKBOOK", "Details": ""},
        {"#": "", "Section": "", "Details": "This workbook was exported from PolicyDB to help the team review and complete the book of business."},
        {"#": "", "Section": "", "Details": "Each tab focuses on a different type of gap or issue. Work through them in order."},
        {"#": "", "Section": "", "Details": ""},
        {"#": 2, "Section": "TAB GUIDE", "Details": ""},
        {"#": "", "Section": "Summary", "Details": "High-level stats and gap counts. Review first to understand scope."},
        {"#": "", "Section": "All Policies", "Details": "Complete list. Use Has Location? / Has Carrier? / Has Policy# columns to spot gaps."},
        {"#": "", "Section": "Suspected Duplicates", "Details": "Policies that may be the same record imported from two sources. CONFIRM: are these the same policy?"},
        {"#": "", "Section": "Unassigned Locations", "Details": "Policies not assigned to a project/location. FIND: which project does each belong to?"},
        {"#": "", "Section": "Missing Fields", "Details": "Policies missing key data OR containing placeholder text (TBD, N/A, etc.). FIND: real values from source documents."},
        {"#": "", "Section": "Program Review", "Details": "Corporate programs — check carrier lists and identify unlinked policies."},
        {"#": "", "Section": "Action Items", "Details": "Prioritized checklist of everything that needs attention. Work top-down."},
        {"#": "", "Section": "", "Details": ""},
        {"#": 3, "Section": "HOW TO REPORT BACK", "Details": ""},
        {"#": "", "Section": "", "Details": "Fill in the 'Your Notes' column on the Action Items tab with what you find."},
        {"#": "", "Section": "", "Details": "For suspected duplicates: write SAME or DIFFERENT in the Verdict column."},
        {"#": "", "Section": "", "Details": "For missing fields: write the missing value directly in the Notes column."},
        {"#": "", "Section": "", "Details": "Return the completed workbook and we will update PolicyDB."},
    ]
    _write_sheet(wb, "Instructions", instructions, col_widths={"#": 5, "Section": 30, "Details": 80})

    # ════════════════════════════════════════════════════════════════════════
    # TAB 2: SUMMARY
    # ════════════════════════════════════════════════════════════════════════
    total_policies = len(policies)
    total_programs = len(programs_db)
    total_premium = sum(float(p.get("premium") or 0) for p in policies)
    unassigned = [p for p in policies if not p.get("project_id") and not p.get("program_id")]
    missing_carrier = [p for p in policies if not (p.get("carrier") or "").strip()]
    missing_premium = [p for p in policies if not p.get("premium")]
    missing_polnum = [p for p in policies if not (p.get("policy_number") or "").strip()]
    missing_dates = [p for p in policies if not p.get("effective_date") or not p.get("expiration_date")]

    # Pre-compute sketchy data and completeness for all policies
    policy_sketchy = {}  # policy_uid -> [(field_label, value), ...]
    policy_completeness = {}  # policy_uid -> int 0-100
    for p in policies:
        policy_sketchy[p["policy_uid"]] = _scan_sketchy_fields(p)
        policy_completeness[p["policy_uid"]] = _compute_completeness(p)

    sketchy_count = sum(1 for v in policy_sketchy.values() if v)
    avg_completeness = round(sum(policy_completeness[p["policy_uid"]] for p in policies) / len(policies)) if policies else 100

    summary_rows = [
        {"Item": "Client", "Value": client_name},
        {"Item": "Report Date", "Value": date.today().isoformat()},
        {"Item": "", "Value": ""},
        {"Item": "Total Policies", "Value": total_policies},
        {"Item": "Total Programs", "Value": total_programs},
        {"Item": "Total Premium", "Value": total_premium},
        {"Item": "Known Locations", "Value": len(locations)},
        {"Item": "Average Data Completeness", "Value": f"{avg_completeness}%"},
        {"Item": "", "Value": ""},
        {"Item": "GAPS IDENTIFIED", "Value": ""},
        {"Item": "Policies Without Location Assignment", "Value": len(unassigned)},
        {"Item": "Policies Missing Carrier", "Value": len(missing_carrier)},
        {"Item": "Policies Missing Premium", "Value": len(missing_premium)},
        {"Item": "Policies Missing Policy Number", "Value": len(missing_polnum)},
        {"Item": "Policies Missing Dates", "Value": len(missing_dates)},
        {"Item": "Policies with Placeholder/TBD Data", "Value": sketchy_count},
        {"Item": "Policies Flagged for Investigation", "Value": sum(1 for p in policies if p.get("needs_investigation"))},
        {"Item": "Suspected Duplicates", "Value": len(dedup_candidates)},
    ]
    _write_sheet(wb, "Summary", summary_rows, col_widths={"Item": 40, "Value": 25})

    # ════════════════════════════════════════════════════════════════════════
    # TAB 2: ALL POLICIES
    # ════════════════════════════════════════════════════════════════════════
    all_rows = []
    for p in policies:
        location = p.get("location_name") or p.get("project_name") or ""
        parent = p.get("parent_program_name") or ""
        sketchy_here = policy_sketchy.get(p["policy_uid"], [])
        completeness = policy_completeness.get(p["policy_uid"], 100)
        all_rows.append({
            "Policy ID": p["policy_uid"],
            "Coverage Type": p.get("policy_type", ""),
            "Carrier": p.get("carrier", ""),
            "Policy Number": p.get("policy_number", ""),
            "Effective Date": p.get("effective_date", ""),
            "Expiration Date": p.get("expiration_date", ""),
            "Premium": float(p.get("premium") or 0),
            "Limit": float(p.get("limit_amount") or 0),
            "Deductible": float(p.get("deductible") or 0),
            "Location / Project": location,
            "Layer": p.get("layer_position", ""),
            "Renewal Status": p.get("renewal_status", ""),
            "Program": parent,
            "First Named Insured": p.get("first_named_insured", ""),
            "Placement Colleague": p.get("placement_colleague", ""),
            "Underwriter": p.get("underwriter_name", ""),
            "Description": p.get("description", ""),
            "Data Completeness %": f"{completeness}%",
            "Has Location?": "Yes" if p.get("project_id") else "NO",
            "Has Carrier?": "Yes" if (p.get("carrier") or "").strip() else "NO",
            "Has Policy #?": "Yes" if (p.get("policy_number") or "").strip() else "NO",
            "Sketchy Data?": "YES" if sketchy_here else "",
            "Needs Investigation?": "YES" if p.get("needs_investigation") else "",
        })
    _write_sheet(wb, "All Policies", all_rows)

    # ════════════════════════════════════════════════════════════════════════
    # TAB 4: SUSPECTED DUPLICATES
    # ════════════════════════════════════════════════════════════════════════
    dup_rows = []
    for c in dedup_candidates:
        a = c["policy_a"]
        b = c["policy_b"]
        signals = ", ".join(s.replace("~", " (fuzzy)") for s in c["match_signals"])
        fills_a = ", ".join(c.get("fillable_a", []))
        fills_b = ", ".join(c.get("fillable_b", []))

        dup_rows.append({
            "Score": c["score"],
            "Confidence": c["recommendation"].replace("_", " ").title(),
            "Policy A": a.get("policy_uid", ""),
            "A — Type": a.get("policy_type", ""),
            "A — Carrier": a.get("carrier", ""),
            "A — Policy #": a.get("policy_number", ""),
            "A — Dates": f"{a.get('effective_date', '')} – {a.get('expiration_date', '')}",
            "A — Premium": float(a.get("premium") or 0),
            "A — Location": a.get("project_name", "") or a.get("location_name", "") or "",
            "Policy B": b.get("policy_uid", ""),
            "B — Type": b.get("policy_type", ""),
            "B — Carrier": b.get("carrier", ""),
            "B — Policy #": b.get("policy_number", ""),
            "B — Dates": f"{b.get('effective_date', '')} – {b.get('expiration_date', '')}",
            "B — Premium": float(b.get("premium") or 0),
            "B — Location": b.get("project_name", "") or b.get("location_name", "") or "",
            "Match Signals": signals,
            "A Has (B Missing)": fills_a,
            "B Has (A Missing)": fills_b,
            "Verdict (SAME/DIFFERENT)": "",
            "Notes": "",
        })

    if not dup_rows:
        dup_rows.append({
            "Score": "", "Confidence": "", "Policy A": "", "A — Type": "",
            "A — Carrier": "", "A — Policy #": "", "A — Dates": "",
            "A — Premium": "", "A — Location": "",
            "Policy B": "", "B — Type": "", "B — Carrier": "",
            "B — Policy #": "", "B — Dates": "", "B — Premium": "",
            "B — Location": "", "Match Signals": "",
            "A Has (B Missing)": "", "B Has (A Missing)": "",
            "Verdict (SAME/DIFFERENT)": "No suspected duplicates found",
            "Notes": "",
        })
    _write_sheet(wb, "Suspected Duplicates", dup_rows)

    # ════════════════════════════════════════════════════════════════════════
    # TAB 5: UNASSIGNED LOCATIONS
    # ════════════════════════════════════════════════════════════════════════
    unassigned_rows = []
    for p in unassigned:
        unassigned_rows.append({
            "Policy ID": p["policy_uid"],
            "Coverage Type": p.get("policy_type", ""),
            "Carrier": p.get("carrier", ""),
            "Policy Number": p.get("policy_number", ""),
            "Premium": float(p.get("premium") or 0),
            "Effective Date": p.get("effective_date", ""),
            "Expiration Date": p.get("expiration_date", ""),
            "Address on File": p.get("exposure_address", ""),
            "Notes": p.get("description", ""),
            "Action Needed": "FIND which project/location this policy covers. If corporate-wide, write 'CORPORATE'.",
        })
    _write_sheet(wb, "Unassigned Locations", unassigned_rows)

    # ════════════════════════════════════════════════════════════════════════
    # TAB 4: MISSING FIELDS
    # ════════════════════════════════════════════════════════════════════════
    missing_rows = []
    for p in policies:
        missing = []
        if not (p.get("carrier") or "").strip():
            missing.append("Carrier")
        if not p.get("premium"):
            missing.append("Premium")
        if not (p.get("policy_number") or "").strip():
            missing.append("Policy Number")
        if not p.get("effective_date"):
            missing.append("Effective Date")
        if not p.get("expiration_date"):
            missing.append("Expiration Date")
        if not p.get("limit_amount"):
            missing.append("Limit")
        if not (p.get("first_named_insured") or "").strip():
            missing.append("First Named Insured")
        if not (p.get("placement_colleague") or "").strip():
            missing.append("Placement Colleague")
        if not (p.get("underwriter_name") or "").strip():
            missing.append("Underwriter")

        sketchy = policy_sketchy.get(p["policy_uid"], [])
        completeness = policy_completeness.get(p["policy_uid"], 100)

        if missing or sketchy:
            has_critical = any(f in missing for f in ["Carrier", "Premium", "Effective Date", "Expiration Date"])
            issue_count = len(missing) + len(sketchy)
            missing_rows.append({
                "Policy ID": p["policy_uid"],
                "Coverage Type": p.get("policy_type", ""),
                "Carrier": p.get("carrier", ""),
                "Data Completeness %": f"{completeness}%",
                "Missing Fields": ", ".join(missing) if missing else "",
                "Sketchy Fields": ", ".join(f"{lbl} = '{val}'" for lbl, val in sketchy) if sketchy else "",
                "Total Issues": issue_count,
                "Priority": "High" if has_critical or len(sketchy) >= 3 else "Medium",
            })
    missing_rows.sort(key=lambda r: (-r["Total Issues"], r["Coverage Type"]))
    _write_sheet(wb, "Missing Fields", missing_rows)

    # ════════════════════════════════════════════════════════════════════════
    # TAB 5: PROGRAM REVIEW
    # ════════════════════════════════════════════════════════════════════════
    program_rows = []
    for pgm in programs_db:
        children = program_child_carriers.get(pgm["id"], [])
        carrier_list = ", ".join(c["carrier"] for c in children if c.get("carrier")) or ""
        carrier_count = len([c for c in children if c.get("carrier")])
        total_prog_premium = sum(float(c.get("premium") or 0) for c in children)

        child_count = conn.execute(
            "SELECT COUNT(*) as c FROM policies WHERE program_id = ? AND archived = 0",
            (pgm["id"],),
        ).fetchone()["c"]

        # Check for standalone policies with same type that might belong
        potential_members = [
            p for p in policies
            if not p.get("program_id")
            and p.get("policy_type") == pgm.get("policy_type")
        ]

        program_rows.append({
            "Program": pgm.get("name", ""),
            "Coverage Type": pgm.get("policy_type", ""),
            "Carriers": carrier_list,
            "Carrier Count": carrier_count,
            "Total Premium": total_prog_premium,
            "Effective Date": pgm.get("effective_date", ""),
            "Expiration Date": pgm.get("expiration_date", ""),
            "Child Policies Linked": child_count,
            "Potential Unlinked Policies": len(potential_members),
            "Review Note": f"{len(potential_members)} standalone {pgm.get('policy_type', '')} policies may belong to this program" if potential_members else "OK",
        })
    _write_sheet(wb, "Program Review", program_rows)

    # ════════════════════════════════════════════════════════════════════════
    # TAB 6: ACTION ITEMS
    # ════════════════════════════════════════════════════════════════════════
    actions = []
    action_num = 0

    # Dedup items first — highest priority
    for c in dedup_candidates:
        a = c["policy_a"]
        b = c["policy_b"]
        action_num += 1
        actions.append({
            "#": action_num,
            "Priority": "High" if c["confidence"] == "high" else "Medium",
            "Category": "Suspected Duplicate",
            "Policy ID": f"{a.get('policy_uid', '')} vs {b.get('policy_uid', '')}",
            "What To Do": f"CONFIRM: Are these the same policy? Score: {c['score']}. See Suspected Duplicates tab.",
            "Context": f"{a.get('policy_type', '')} · {a.get('carrier', '')} · {a.get('effective_date', '')}",
            "Your Notes": "",
        })

    for p in unassigned:
        action_num += 1
        actions.append({
            "#": action_num,
            "Priority": "Medium",
            "Category": "Location Assignment",
            "Policy ID": p["policy_uid"],
            "What To Do": f"FIND which project/location this {p.get('policy_type', '')} ({p.get('carrier', '')}) covers. Write project name or 'CORPORATE'.",
            "Context": p.get("exposure_address", "") or "No address on file",
            "Your Notes": "",
        })

    for row in missing_rows:
        if row["Priority"] == "High":
            action_num += 1
            parts = []
            if row.get("Missing Fields"):
                parts.append(f"MISSING: {row['Missing Fields']}")
            if row.get("Sketchy Fields"):
                parts.append(f"REPLACE: {row['Sketchy Fields']}")
            actions.append({
                "#": action_num,
                "Priority": "High",
                "Category": "Missing / Sketchy Data",
                "Policy ID": row["Policy ID"],
                "What To Do": ". ".join(parts) + ". Check AMS, carrier portal, or policy documents.",
                "Context": f"{row['Coverage Type']} / {row.get('Carrier', '')}",
                "Your Notes": "",
            })

    # Sketchy-only rows (have placeholder data but no missing critical fields)
    for row in missing_rows:
        if row["Priority"] != "High" and row.get("Sketchy Fields"):
            action_num += 1
            actions.append({
                "#": action_num,
                "Priority": "Medium",
                "Category": "Placeholder Data",
                "Policy ID": row["Policy ID"],
                "What To Do": f"REPLACE placeholder values: {row['Sketchy Fields']}. Provide real data.",
                "Context": f"{row['Coverage Type']} / {row.get('Carrier', '')}",
                "Your Notes": "",
            })

    for prog_row in program_rows:
        if prog_row["Potential Unlinked Policies"] > 0:
            action_num += 1
            actions.append({
                "#": action_num,
                "Priority": "Medium",
                "Category": "Program Membership",
                "Policy ID": prog_row["Program ID"],
                "What To Do": f"VERIFY: {prog_row['Review Note']}. Should these standalone policies be part of this program?",
                "Context": f"{prog_row['Carrier Count']} carriers, {prog_row['Child Policies Linked']} linked",
                "Your Notes": "",
            })

    if not actions:
        actions.append({
            "#": 1, "Priority": "", "Category": "",
            "Policy ID": "", "What To Do": "No action items — book looks complete!",
            "Context": "", "Your Notes": "",
        })

    _write_sheet(wb, "Action Items", actions)

    return _wb_to_bytes(wb)


# ─── PROGRAM EXPORT ──────────────────────────────────────────────────────────


def export_programs_xlsx(conn: sqlite3.Connection, client_id: int) -> bytes:
    """Export all programs for a client — one sheet per program with child policies."""
    programs = conn.execute(
        """SELECT id, program_uid, name, line_of_business, effective_date,
                  expiration_date, renewal_status, lead_broker, notes
           FROM programs WHERE client_id = ? AND (archived = 0 OR archived IS NULL)
           ORDER BY name""",
        (client_id,),
    ).fetchall()

    wb = Workbook()
    wb.remove(wb.active)

    if not programs:
        _write_sheet(wb, "Programs", [{"Program": "No programs found for this client"}])
        return _wb_to_bytes(wb)

    # Summary sheet
    summary_rows = []
    for prog in programs:
        p = dict(prog)
        child_count = conn.execute(
            "SELECT COUNT(*) FROM policies WHERE program_id = ?", (p["id"],)
        ).fetchone()[0]
        total_premium = conn.execute(
            "SELECT COALESCE(SUM(premium), 0) FROM policies WHERE program_id = ?", (p["id"],)
        ).fetchone()[0]
        summary_rows.append({
            "Program": p["name"] or p["program_uid"],
            "Line of Business": p.get("line_of_business") or "",
            "Effective": p.get("effective_date") or "",
            "Expiration": p.get("expiration_date") or "",
            "Status": p.get("renewal_status") or "",
            "Lead Broker": p.get("lead_broker") or "",
            "Policies": child_count,
            "Total Premium": total_premium,
        })
    _write_sheet(wb, "Summary", summary_rows)

    # One sheet per program with child policies
    for prog in programs:
        p = dict(prog)
        sheet_name = (p["name"] or p["program_uid"] or "Program")[:31]
        children = conn.execute(
            """SELECT policy_uid, policy_type, carrier, policy_number,
                      effective_date, expiration_date, premium, limit_amount,
                      deductible, description, layer_position, project_name,
                      access_point, renewal_status
               FROM policies WHERE program_id = ?
               ORDER BY layer_position, policy_type""",
            (p["id"],),
        ).fetchall()
        rows = [dict(r) for r in children]
        if not rows:
            rows = [{"policy_type": "(no child policies)"}]
        _write_sheet(wb, sheet_name, rows)

        # Add program metadata header row above the data
        ws = wb[sheet_name]
        meta_parts = []
        if p.get("line_of_business"):
            meta_parts.append(p["line_of_business"])
        if p.get("effective_date") and p.get("expiration_date"):
            meta_parts.append(f"{p['effective_date']} — {p['expiration_date']}")
        if p.get("lead_broker"):
            meta_parts.append(f"Lead: {p['lead_broker']}")
        if meta_parts:
            ws.insert_rows(1)
            ws["A1"] = " | ".join(meta_parts)
            ws["A1"].font = Font(name="Noto Sans", bold=True, size=11, color="003865")

    return _wb_to_bytes(wb)
