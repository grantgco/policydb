"""Read-only reconciliation logic: compare an uploaded CSV against PolicyDB policies."""

from __future__ import annotations

import csv
import io
from collections import namedtuple
from dataclasses import dataclass, field
from datetime import datetime
from rapidfuzz import fuzz

from policydb.importer import PolicyImporter, _parse_currency, _parse_date
from policydb.utils import (
    _COVERAGE_ALIASES,
    normalize_carrier,
    normalize_client_name_for_matching,
    normalize_coverage_type,
    normalize_policy_number_for_matching,
    parse_currency,
)


# ─── TYPES ────────────────────────────────────────────────────────────────────

COMPARE_FIELDS = [
    "client_name",
    "policy_type",
    "carrier",
    "policy_number",
    "effective_date",
    "expiration_date",
    "premium",
    "limit_amount",
    "deductible",
]

_STATUS_SORT = {"DIFF": 0, "PAIRED": 0, "MISSING": 1, "UNMATCHED": 1, "EXTRA": 2, "MATCH": 3}


@dataclass
class ReconcileRow:
    """One row in reconciliation results."""
    ext: dict | None = None         # uploaded record; None for EXTRA
    db: dict | None = None          # PolicyDB record; None for UNMATCHED/MISSING
    status: str = "PAIRED"          # "PAIRED" | "UNMATCHED" | "EXTRA" | legacy: "MATCH" | "DIFF" | "MISSING"
    match_score: float = 0.0       # 0-100 total
    confidence: str = "none"       # "high" | "medium" | "low" | "none"
    match_method: str = ""         # "policy_number" | "scored" | "manual" | legacy: "date_pair" | "fuzzy" | ""
    confirmed: bool = False        # user stamped

    # Per-field score breakdown
    score_policy_number: float = 0.0
    score_dates: float = 0.0
    score_type: float = 0.0
    score_carrier: float = 0.0
    score_name: float = 0.0

    # Diff tracking
    diff_fields: list[str] = field(default_factory=list)
    cosmetic_diffs: list[str] = field(default_factory=list)
    fillable_fields: list[str] = field(default_factory=list)

    # Metadata
    eff_delta_days: int | None = None
    exp_delta_days: int | None = None
    ext_type_raw: str = ""
    ext_type_normalized: str = ""
    coverage_alias_applied: bool = False
    ext_carrier_raw: str = ""
    ext_carrier_normalized: str = ""
    carrier_alias_applied: bool = False



# ─── FILE PARSING ─────────────────────────────────────────────────────────────

def _normalize_headers(raw_headers: list[str]) -> dict[str, str]:
    """Map raw CSV column names to canonical PolicyDB field names via PolicyImporter.ALIASES."""
    aliases = PolicyImporter.ALIASES
    mapping = {}
    for h in raw_headers:
        if h is None:
            continue
        key = h.strip().lower().replace(" ", "_").replace("-", "_")
        canonical = aliases.get(key, key)
        mapping[h] = canonical
    return mapping


def _process_raw_rows(
    raw_rows: list[dict],
    column_mapping: dict | None,
    warnings: list[str],
) -> list[dict]:
    """Normalize, map, and parse a list of raw row dicts (from CSV or XLSX)."""
    if not raw_rows:
        warnings.append("No data rows found in uploaded file.")
        return []

    if column_mapping:
        raw_headers = list(raw_rows[0].keys())
        header_map = {h: column_mapping.get(h, h) for h in raw_headers}
    else:
        header_map = _normalize_headers(list(raw_rows[0].keys()))

    rows: list[dict] = []
    for i, raw in enumerate(raw_rows, start=2):  # row 1 = header
        row: dict = {}
        for raw_key, value in raw.items():
            canonical = header_map.get(raw_key, raw_key)
            row[canonical] = (str(value).strip() if value is not None else "")

        if not row.get("client_name") and not row.get("policy_number"):
            warnings.append(f"Row {i}: skipped — no client name or policy number found.")
            continue

        for field_name in ("effective_date", "expiration_date"):
            if row.get(field_name):
                parsed = _parse_date(row[field_name])
                if parsed:
                    row[field_name] = parsed
                else:
                    warnings.append(f"Row {i}: could not parse date '{row[field_name]}' for {field_name}.")
                    row[field_name] = ""

        for field_name in ("premium", "limit_amount", "deductible"):
            if row.get(field_name):
                row[field_name] = _parse_currency(row[field_name])
            else:
                row[field_name] = 0.0

        rows.append(row)
    return rows


def _parse_csv_content(content: bytes, column_mapping: dict | None) -> tuple[list[dict], list[str]]:
    """Parse CSV bytes into normalized dicts."""
    warnings: list[str] = []
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
        warnings.append("File encoding detected as Latin-1 (not UTF-8). Special characters may be affected.")

    reader = csv.DictReader(io.StringIO(text))
    raw_rows = list(reader)
    rows = _process_raw_rows(raw_rows, column_mapping, warnings)
    return rows, warnings


def _parse_xlsx_content(content: bytes, column_mapping: dict | None) -> tuple[list[dict], list[str]]:
    """Parse XLSX bytes into normalized dicts using openpyxl."""
    from openpyxl import load_workbook

    warnings: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return [], [f"Could not read Excel file: {e}"]

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], ["Excel file has no data."]

    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header_row)]

    raw_rows: list[dict] = []
    for row_vals in rows_iter:
        if all(v is None for v in row_vals):
            continue  # skip blank rows
        raw = {}
        for h, v in zip(headers, row_vals):
            if v is None:
                raw[h] = ""
            elif isinstance(v, datetime):
                raw[h] = v.strftime("%Y-%m-%d")
            else:
                raw[h] = str(v).strip()
        raw_rows.append(raw)

    wb.close()
    rows = _process_raw_rows(raw_rows, column_mapping, warnings)
    return rows, warnings


def parse_uploaded_file(
    content: bytes,
    column_mapping: dict | None = None,
    filename: str = "",
) -> tuple[list[dict], list[str]]:
    """
    Parse uploaded CSV or XLSX bytes into normalized dicts.

    Auto-detects format by filename extension or content magic bytes.
    """
    if filename.lower().endswith(('.xlsx', '.xls')) or content[:2] == b'PK':
        return _parse_xlsx_content(content, column_mapping)
    return _parse_csv_content(content, column_mapping)


# Backward compatibility alias
parse_uploaded_csv = parse_uploaded_file


# ─── FIELD COMPARISON ─────────────────────────────────────────────────────────

def _date_delta_days(d1: str | None, d2: str | None) -> int | None:
    """Return absolute day difference between two YYYY-MM-DD strings, or None."""
    if not d1 or not d2:
        return None
    try:
        dt1 = datetime.strptime(d1, "%Y-%m-%d")
        dt2 = datetime.strptime(d2, "%Y-%m-%d")
        return abs((dt1 - dt2).days)
    except ValueError:
        return None


# ─── MATCHING ─────────────────────────────────────────────────────────────────

def _same_year(d1: str | None, d2: str | None) -> bool:
    """Return True if two YYYY-MM-DD strings share the same calendar year."""
    if not d1 or not d2 or len(d1) < 4 or len(d2) < 4:
        return False
    return d1[:4] == d2[:4]


# ─── ADDITIVE SCORING (new — replaces _fuzzy_match in Task 4) ────────────────

ScoreBreakdown = namedtuple(
    "ScoreBreakdown",
    [
        "score_policy_number",
        "score_dates",
        "score_type",
        "score_carrier",
        "score_name",
        "total",
        "confidence",
        "diff_fields",
        "cosmetic_diffs",
        "fillable_fields",
        "eff_delta_days",
        "exp_delta_days",
        "ext_type_raw",
        "ext_type_normalized",
        "coverage_alias_applied",
        "ext_carrier_raw",
        "ext_carrier_normalized",
        "carrier_alias_applied",
    ],
)


def _score_pair(
    ext: dict,
    db: dict,
    single_client: bool = False,
) -> ScoreBreakdown:
    """Additive per-field scoring for an (ext, db) pair.

    No hard gates — every signal contributes points independently.

    Signals and max points:
        Policy Number  40   Exact normalized=40, fuzzy>=90=32, >=75=20, missing=0
        Dates          30   Split 15+15 (eff+exp). Exact=15, <=14d=12, <=45d=8, same year=4, >1yr=0
        Policy Type    15   Normalized match=15, fuzzy>=85=12, >=70=8, <70=0
        Carrier        10   Normalized match=10, fuzzy>=80=7, >=60=4, <60=0
        Client Name     5   Normalized match=5, fuzzy>=80=4, >=60=2, <60=0

    Confidence tiers: high>=75, medium>=45, low<45.
    """

    # Diff tracking lists — populated by each scoring section below
    diff_fields: list[str] = []
    cosmetic_diffs: list[str] = []
    fillable_fields: list[str] = []

    # ── Policy Number (max 40) ────────────────────────────────────────────────
    ext_pn_raw = str(ext.get("policy_number") or "").strip()
    db_pn_raw = str(db.get("policy_number") or "").strip()
    ext_pn = normalize_policy_number_for_matching(ext_pn_raw)
    db_pn = normalize_policy_number_for_matching(db_pn_raw)

    score_policy_number = 0.0
    if ext_pn and db_pn:
        if ext_pn == db_pn:
            score_policy_number = 40.0
        else:
            pn_ratio = fuzz.ratio(ext_pn, db_pn)
            if pn_ratio >= 90:
                score_policy_number = 32.0
            elif pn_ratio >= 75:
                score_policy_number = 20.0
    # Missing on either side = 0 (no penalty, no reward)

    # ── Dates (max 30 = 15 eff + 15 exp) ─────────────────────────────────────
    ext_eff = str(ext.get("effective_date") or "").strip()
    db_eff = str(db.get("effective_date") or "").strip()
    ext_exp = str(ext.get("expiration_date") or "").strip()
    db_exp = str(db.get("expiration_date") or "").strip()

    eff_delta = _date_delta_days(ext_eff, db_eff)
    exp_delta = _date_delta_days(ext_exp, db_exp)

    def _date_score(delta: int | None, d1: str, d2: str) -> float:
        """Score a single date pair (max 15)."""
        if delta is None:
            return 0.0
        if delta == 0:
            return 15.0
        if delta <= 14:
            return 12.0
        if delta <= 45:
            return 8.0
        if _same_year(d1, d2):
            return 4.0
        return 0.0

    score_eff = _date_score(eff_delta, ext_eff, db_eff)
    score_exp = _date_score(exp_delta, ext_exp, db_exp)
    score_dates = score_eff + score_exp

    # ── Policy Type (max 15) ─────────────────────────────────────────────────
    ext_type_raw = str(ext.get("policy_type") or "").strip()
    db_type_raw = str(db.get("policy_type") or "").strip()
    ext_type_norm = normalize_coverage_type(ext_type_raw)
    db_type_norm = normalize_coverage_type(db_type_raw)
    coverage_alias_applied = (
        bool(ext_type_raw)
        and ext_type_raw.strip().lower() != ext_type_norm.lower()
    )

    score_type = 0.0
    if ext_type_norm and db_type_norm:
        if ext_type_norm.lower() == db_type_norm.lower():
            score_type = 15.0
            # Raw values differ but normalized match = cosmetic
            if ext_type_raw.lower() != db_type_raw.lower():
                cosmetic_diffs.append("policy_type")
        else:
            type_ratio = fuzz.WRatio(ext_type_norm, db_type_norm)
            if type_ratio >= 85:
                score_type = 12.0
                cosmetic_diffs.append("policy_type")
            elif type_ratio >= 70:
                score_type = 8.0
                diff_fields.append("policy_type")
            else:
                diff_fields.append("policy_type")

    # ── Carrier (max 10) ──────────────────────────────────────────────────────
    ext_carrier_raw = str(ext.get("carrier") or "").strip()
    db_carrier_raw = str(db.get("carrier") or "").strip()
    ext_carrier_norm = normalize_carrier(ext_carrier_raw)
    db_carrier_norm = normalize_carrier(db_carrier_raw)
    carrier_alias_applied = (
        bool(ext_carrier_raw)
        and ext_carrier_raw.strip().lower() != ext_carrier_norm.lower()
    )

    score_carrier = 0.0
    if ext_carrier_norm and db_carrier_norm:
        if ext_carrier_norm.lower() == db_carrier_norm.lower():
            score_carrier = 10.0
            if ext_carrier_raw.lower() != db_carrier_raw.lower():
                cosmetic_diffs.append("carrier")
        else:
            carrier_ratio = fuzz.WRatio(ext_carrier_norm, db_carrier_norm)
            if carrier_ratio >= 80:
                score_carrier = 7.0
                cosmetic_diffs.append("carrier")
            elif carrier_ratio >= 60:
                score_carrier = 4.0
                diff_fields.append("carrier")
            else:
                diff_fields.append("carrier")

    # ── Client Name / FNI (max 5) ────────────────────────────────────────────
    if single_client:
        score_name = 5.0
    else:
        ext_client = normalize_client_name_for_matching(ext.get("client_name"))
        db_client = normalize_client_name_for_matching(db.get("client_name"))
        ext_fni = normalize_client_name_for_matching(ext.get("first_named_insured"))
        db_fni = normalize_client_name_for_matching(db.get("first_named_insured"))

        # Cross-match: ext client vs db client, ext client vs db FNI,
        #              ext FNI vs db client, ext FNI vs db FNI — take best
        best_name_ratio = 0.0
        pairs_to_check = []
        if ext_client and db_client:
            pairs_to_check.append((ext_client, db_client))
        if ext_client and db_fni:
            pairs_to_check.append((ext_client, db_fni))
        if ext_fni and db_client:
            pairs_to_check.append((ext_fni, db_client))
        if ext_fni and db_fni:
            pairs_to_check.append((ext_fni, db_fni))

        for a, b in pairs_to_check:
            ratio = fuzz.WRatio(a, b)
            if ratio > best_name_ratio:
                best_name_ratio = ratio

        if best_name_ratio >= 95:
            score_name = 5.0
        elif best_name_ratio >= 80:
            score_name = 4.0
        elif best_name_ratio >= 60:
            score_name = 2.0
        else:
            score_name = 0.0

    # ── Total and confidence ──────────────────────────────────────────────────
    total = score_policy_number + score_dates + score_type + score_carrier + score_name

    if total >= 75:
        confidence = "high"
    elif total >= 45:
        confidence = "medium"
    else:
        confidence = "low"

    # Currency fields: premium, limit_amount, deductible
    for cf in ("premium", "limit_amount", "deductible"):
        ext_val = parse_currency(ext.get(cf, 0))
        db_val = parse_currency(db.get(cf, 0))
        if ext_val > 0 and db_val > 0:
            pct_diff = abs(ext_val - db_val) / max(ext_val, db_val)
            if pct_diff > 0.01:
                diff_fields.append(cf)
        elif ext_val > 0 and db_val == 0:
            fillable_fields.append(cf)

    # ── Policy Number diff tracking ────────────────────────────────────────────
    if ext_pn_raw and not db_pn_raw:
        fillable_fields.append("policy_number")
    elif ext_pn_raw and db_pn_raw:
        if ext_pn != db_pn:
            # Normalized forms differ — real diff
            diff_fields.append("policy_number")
        elif ext_pn_raw.strip().lower() != db_pn_raw.strip().lower():
            # Same after normalization but raw values differ — cosmetic
            cosmetic_diffs.append("policy_number")

    # ── Text field diff tracking (no scoring weight, just diff detection) ────
    for tf in ("first_named_insured", "placement_colleague",
               "underwriter_name", "project_name"):
        ext_val = str(ext.get(tf) or "").strip()
        db_val = str(db.get(tf) or "").strip()
        if ext_val and not db_val:
            fillable_fields.append(tf)
        elif ext_val and db_val and ext_val.lower() != db_val.lower():
            diff_fields.append(tf)

    return ScoreBreakdown(
        score_policy_number=score_policy_number,
        score_dates=score_dates,
        score_type=score_type,
        score_carrier=score_carrier,
        score_name=score_name,
        total=total,
        confidence=confidence,
        diff_fields=diff_fields,
        cosmetic_diffs=cosmetic_diffs,
        fillable_fields=fillable_fields,
        eff_delta_days=eff_delta,
        exp_delta_days=exp_delta,
        ext_type_raw=ext_type_raw,
        ext_type_normalized=ext_type_norm,
        coverage_alias_applied=coverage_alias_applied,
        ext_carrier_raw=ext_carrier_raw,
        ext_carrier_normalized=ext_carrier_norm,
        carrier_alias_applied=carrier_alias_applied,
    )


def find_candidates(ext_row: dict, db_rows: list[dict], limit: int = 8, single_client: bool = False) -> list[tuple[dict, float]]:
    """
    Return top candidate DB rows for a given ext_row, for manual match selection.
    Uses _score_pair() for consistent scoring. Returns top ``limit`` by score, no
    threshold — the UI shows all candidates and lets the user pick.
    Returns list of (db_row, score) sorted by score descending.
    """
    scored: list[tuple[dict, float]] = []

    for db in db_rows:
        breakdown = _score_pair(ext_row, db, single_client=single_client)
        scored.append((db, round(breakdown.total, 1)))

    scored.sort(key=lambda x: x[1], reverse=True)
    return scored[:limit]


# ─── MAIN RECONCILE ───────────────────────────────────────────────────────────

def _build_reconcile_row(ext: dict, db: dict, breakdown: ScoreBreakdown,
                         match_method: str) -> ReconcileRow:
    """Create a PAIRED ReconcileRow from a ScoreBreakdown."""
    return ReconcileRow(
        ext=ext, db=db, status="PAIRED",
        match_score=breakdown.total,
        confidence=breakdown.confidence,
        match_method=match_method,
        score_policy_number=breakdown.score_policy_number,
        score_dates=breakdown.score_dates,
        score_type=breakdown.score_type,
        score_carrier=breakdown.score_carrier,
        score_name=breakdown.score_name,
        diff_fields=list(breakdown.diff_fields),
        cosmetic_diffs=list(breakdown.cosmetic_diffs),
        fillable_fields=list(breakdown.fillable_fields),
        eff_delta_days=breakdown.eff_delta_days,
        exp_delta_days=breakdown.exp_delta_days,
        ext_type_raw=breakdown.ext_type_raw,
        ext_type_normalized=breakdown.ext_type_normalized,
        coverage_alias_applied=breakdown.coverage_alias_applied,
        ext_carrier_raw=breakdown.ext_carrier_raw,
        ext_carrier_normalized=breakdown.ext_carrier_normalized,
        carrier_alias_applied=breakdown.carrier_alias_applied,
    )



def reconcile(
    ext_rows: list[dict],
    db_rows: list[dict],
    date_priority: bool = False,
    single_client: bool = False,
    conn=None,
    source_name: str = "",
) -> list[ReconcileRow]:
    """
    Match ext_rows against db_rows using a 4-pass algorithm.

    Pass 0: Match memory lookup — if a source_name and conn are provided,
            look up prior identity pairs so previously matched policies
            auto-pair without scoring.
    Pass 1: Exact policy number match (normalized). Uses _score_pair()
            for full scoring.
    Pass 2: Scored match for remaining unmatched ext rows. Best match per ext
            row with score >= 45. Each DB row can only match once (1:1).
    Pass 3: Remaining ext rows → UNMATCHED, remaining DB rows → EXTRA.

    Sort order: Unconfirmed PAIRED sorted by amber (45-74) ascending first,
    then green (75+) ascending, then UNMATCHED, then confirmed, then EXTRA.
    """
    results: list[ReconcileRow] = []

    # Build DB indexes — use normalized policy numbers for flexible matching
    db_by_polnum: dict[str, dict] = {}
    db_by_id: dict[int, tuple[dict, int]] = {}  # policy_id → (db_row, index)
    for idx, db in enumerate(db_rows):
        db_by_id[db.get("id", 0)] = (db, idx)
        pn = normalize_policy_number_for_matching(db.get("policy_number") or "")
        if pn and pn not in db_by_polnum:
            db_by_polnum[pn] = db

    db_unmatched: set[int] = set(range(len(db_rows)))
    ext_matched: set[int] = set()

    # ── Pass 0: Match memory lookup ────────────────────────────────────────────
    _memory_match_count = 0
    if conn is not None and source_name:
        try:
            from policydb.match_memory import lookup_batch
            # Build external keys from all ext rows (raw + normalized policy numbers)
            ext_keys: dict[str, list[int]] = {}  # external_key → [ext_row_indices]
            for i, ext in enumerate(ext_rows):
                pn = (ext.get("policy_number") or "").strip()
                if pn:
                    ext_keys.setdefault(pn, []).append(i)
                    norm = normalize_policy_number_for_matching(pn)
                    if norm and norm != pn:
                        ext_keys.setdefault(norm, []).append(i)

            if ext_keys:
                memory = lookup_batch(conn, source_name, list(ext_keys.keys()))
                for ext_key, policy_id in memory.items():
                    if policy_id not in db_by_id:
                        continue
                    db, db_idx = db_by_id[policy_id]
                    if db_idx not in db_unmatched:
                        continue
                    for i in ext_keys.get(ext_key, []):
                        if i in ext_matched:
                            continue
                        db_unmatched.discard(db_idx)
                        ext_matched.add(i)
                        ext = ext_rows[i]

                        breakdown = _score_pair(ext, db, single_client=single_client)
                        row = _build_reconcile_row(ext, db, breakdown, "memory")
                        results.append(row)
                        _memory_match_count += 1
                        break  # one ext row per db match
        except Exception:
            import logging
            logging.getLogger(__name__).exception("Match memory Pass 0 failed — falling through to scoring")

    if _memory_match_count:
        import logging
        logging.getLogger(__name__).info("Pass 0 (match memory): %d auto-matches from source '%s'",
                                         _memory_match_count, source_name)

    # ── Pass 1: Exact policy number match ──────────────────────────────────────
    for i, ext in enumerate(ext_rows):
        ext_pn = normalize_policy_number_for_matching(ext.get("policy_number") or "")
        if not ext_pn:
            continue
        db = db_by_polnum.get(ext_pn)
        if db is None:
            continue
        db_idx = db_rows.index(db)
        if db_idx not in db_unmatched:
            continue  # already claimed — send to Pass 2
        db_unmatched.discard(db_idx)
        ext_matched.add(i)

        breakdown = _score_pair(ext, db, single_client=single_client)
        row = _build_reconcile_row(ext, db, breakdown, "policy_number")
        results.append(row)

    # ── Pass 2: Scored match for remaining unmatched ext rows ─────────────────
    for i, ext in enumerate(ext_rows):
        if i in ext_matched:
            continue

        best_db = None
        best_db_idx = -1
        best_breakdown = None
        best_score = 0.0

        for db_idx in list(db_unmatched):
            db = db_rows[db_idx]
            breakdown = _score_pair(ext, db, single_client=single_client)

            if breakdown.total > best_score:
                best_score = breakdown.total
                best_db = db
                best_db_idx = db_idx
                best_breakdown = breakdown

        if best_db is not None and best_score >= 45:
            db_unmatched.discard(best_db_idx)
            ext_matched.add(i)
            row = _build_reconcile_row(ext, best_db, best_breakdown, "scored")
            results.append(row)

    # ── Pass 3: Unmatched / Extra ─────────────────────────────────────────────
    # Remaining ext rows → UNMATCHED
    for i, ext in enumerate(ext_rows):
        if i in ext_matched:
            continue
        # Build metadata even for unmatched rows
        raw_type = ext.get("policy_type", "")
        norm_type = normalize_coverage_type(raw_type) if raw_type else ""
        row = ReconcileRow(
            ext=ext, db=None, status="UNMATCHED", match_score=0.0,
            ext_type_raw=raw_type, ext_type_normalized=norm_type,
            coverage_alias_applied=(bool(raw_type) and raw_type.strip().lower() != norm_type.lower()),
        )
        results.append(row)

    # Remaining DB rows → EXTRA
    for idx in sorted(db_unmatched):
        results.append(ReconcileRow(ext=None, db=db_rows[idx], status="EXTRA", match_score=0.0))

    # Sort: unconfirmed PAIRED amber (45-74) ascending, then green (75+) ascending,
    # then UNMATCHED, then confirmed PAIRED, then EXTRA
    def _sort_key(r: ReconcileRow):
        side = r.db if r.db else r.ext
        client = (side or {}).get("client_name", "") or ""
        exp = (side or {}).get("expiration_date", "") or ""

        if r.status == "PAIRED" and not r.confirmed:
            if r.match_score < 75:
                # Amber — sort first, ascending by score
                return (0, r.match_score, client.lower(), exp)
            else:
                # Green — sort second, ascending by score
                return (1, r.match_score, client.lower(), exp)
        elif r.status == "UNMATCHED":
            return (2, 0.0, client.lower(), exp)
        elif r.status == "PAIRED" and r.confirmed:
            return (3, r.match_score, client.lower(), exp)
        else:  # EXTRA
            return (4, 0.0, client.lower(), exp)

    results.sort(key=_sort_key)
    return results


def program_reconcile_summary(results: list[ReconcileRow]) -> dict[int, dict]:
    """Group reconcile results by program for summary display."""
    by_program: dict[int, dict] = {}
    for r in results:
        if r.db and r.db.get("program_id"):
            pid = r.db["program_id"]
            if pid not in by_program:
                by_program[pid] = {
                    "matched": 0, "total_premium": 0.0, "children": [],
                }
            if r.status == "PAIRED":
                by_program[pid]["matched"] += 1
                by_program[pid]["total_premium"] += float(
                    r.ext.get("premium") or 0
                ) if r.ext else 0
            by_program[pid]["children"].append(r)
    return by_program


# ─── CROSS-PAIR SCORING ───────────────────────────────────────────────────────

def _find_likely_pairs(
    unmatched_rows: list[ReconcileRow],
    extra_rows: list[ReconcileRow],
    threshold: float = 30.0,
) -> list[dict]:
    """Cross-match UNMATCHED ext rows against EXTRA db rows using _score_pair().

    Returns [{"id": int, "score": float, "missing": ReconcileRow, "extra": ReconcileRow}, ...]
    sorted by score descending, deduplicated (one pair per UNMATCHED, one per EXTRA).

    Note: The dict keys use "missing" and "extra" for backward compatibility with
    templates that reference these keys.
    """
    if not unmatched_rows or not extra_rows:
        return []

    # Score all combinations
    scored: list[tuple[float, int, int]] = []
    for mi, mr in enumerate(unmatched_rows):
        if not mr.ext:
            continue
        for ei, er in enumerate(extra_rows):
            if not er.db:
                continue
            breakdown = _score_pair(mr.ext, er.db)
            if breakdown.total >= threshold:
                scored.append((breakdown.total, mi, ei))

    # Sort by score descending, deduplicate greedily
    scored.sort(key=lambda x: x[0], reverse=True)
    used_unmatched: set[int] = set()
    used_extra: set[int] = set()
    pairs: list[dict] = []

    for score, mi, ei in scored:
        if mi in used_unmatched or ei in used_extra:
            continue
        used_unmatched.add(mi)
        used_extra.add(ei)
        pairs.append({
            "id": len(pairs),
            "score": round(score, 1),
            "missing": unmatched_rows[mi],
            "extra": extra_rows[ei],
        })

    return pairs


# ─── SUMMARY ──────────────────────────────────────────────────────────────────

def summarize(results: list[ReconcileRow]) -> dict:
    paired = [r for r in results if r.status == "PAIRED"]
    paired_with_diffs = [r for r in paired if r.diff_fields]
    confirmed = [r for r in paired if r.confirmed]
    review = [r for r in paired if not r.confirmed]
    return {
        "total": len(results),
        "paired": len(paired),
        "paired_clean": len(paired) - len(paired_with_diffs),
        "paired_diffs": len(paired_with_diffs),
        "confirmed": len(confirmed),
        "review": len(review),
        "unmatched": sum(1 for r in results if r.status == "UNMATCHED"),
        "extra": sum(1 for r in results if r.status == "EXTRA"),
        # Legacy compat
        "match": sum(1 for r in results if r.status in ("MATCH", "PAIRED") and not r.diff_fields),
        "diff": sum(1 for r in results if r.status in ("DIFF",) or (r.status == "PAIRED" and r.diff_fields)),
        "missing": sum(1 for r in results if r.status in ("MISSING", "UNMATCHED")),
    }


# ─── XLSX EXPORT ──────────────────────────────────────────────────────────────

def build_reconcile_xlsx(results: list[ReconcileRow], run_date: str = "", filename: str = "") -> bytes:
    """Build a 5-sheet XLSX reconciliation report."""
    from openpyxl import Workbook
    from policydb.exporter import _write_sheet, _wb_to_bytes

    wb = Workbook()
    wb.remove(wb.active)

    summary = summarize(results)

    # Sheet 1: Summary
    ws_sum = wb.create_sheet("Summary")
    from policydb.exporter import _HEADER_FILL, _HEADER_FONT
    from openpyxl.styles import Alignment
    ws_sum.append(["PolicyDB Reconciliation Report"])
    ws_sum.append(["Run Date", run_date or ""])
    ws_sum.append(["Source File", filename or ""])
    ws_sum.append([])
    ws_sum.append(["Status", "Count", "Description"])
    ws_sum["A5"].font = _HEADER_FONT
    ws_sum["B5"].font = _HEADER_FONT
    ws_sum["C5"].font = _HEADER_FONT
    ws_sum["A5"].fill = _HEADER_FILL
    ws_sum["B5"].fill = _HEADER_FILL
    ws_sum["C5"].fill = _HEADER_FILL
    ws_sum.append(["MATCH", summary["match"], "Aligned within tolerances"])
    ws_sum.append(["DIFF", summary["diff"], "Matched but field discrepancies found"])
    ws_sum.append(["MISSING", summary["missing"], "In uploaded file, not found in PolicyDB"])
    ws_sum.append(["EXTRA", summary["extra"], "In PolicyDB, not in uploaded file"])
    ws_sum.append(["TOTAL", summary["total"], ""])
    ws_sum.column_dimensions["A"].width = 12
    ws_sum.column_dimensions["B"].width = 10
    ws_sum.column_dimensions["C"].width = 45

    # Sheets 2-5: build row dicts (handle both PAIRED/UNMATCHED and legacy MATCH/DIFF/MISSING)
    def _diff_rows():
        return [_diff_dict(r) for r in results
                if r.status == "DIFF" or (r.status == "PAIRED" and r.diff_fields)]

    def _missing_rows():
        return [r.ext for r in results if r.status in ("MISSING", "UNMATCHED")]

    def _extra_rows():
        cols = ["policy_uid", "client_name", "policy_type", "carrier", "policy_number",
                "effective_date", "expiration_date", "premium", "limit_amount", "deductible",
                "location_name", "program_uid"]
        return [{c: (r.db or {}).get(c, "") for c in cols} for r in results if r.status == "EXTRA"]

    def _all_rows():
        rows = []
        for r in results:
            side = r.db if r.db else r.ext
            row = {"status": r.status, "match_score": round(r.match_score, 1),
                   "diff_fields": ", ".join(r.diff_fields)}
            for f in COMPARE_FIELDS:
                row[f"ext_{f}"] = (r.ext or {}).get(f, "")
                row[f"db_{f}"] = (r.db or {}).get(f, "")
            row["db_policy_uid"] = r.db.get("policy_uid", "") if r.db else ""
            row["db_location"] = r.db.get("location_name", "") if r.db else ""
            row["db_program"] = r.db.get("program_uid", "") if r.db else ""
            rows.append(row)
        return rows

    _write_sheet(wb, "Differences (DIFF)", _diff_rows())
    _write_sheet(wb, "Missing", _missing_rows())
    _write_sheet(wb, "Extra in PolicyDB", _extra_rows())
    _write_sheet(wb, "All Results", _all_rows())

    return _wb_to_bytes(wb)


def _diff_dict(r: ReconcileRow) -> dict:
    """Build a flat comparison dict for a DIFF row."""
    row = {
        "diff_fields": ", ".join(r.diff_fields),
        "match_score": round(r.match_score, 1),
        "db_policy_uid": (r.db or {}).get("policy_uid", ""),
        "db_location": (r.db or {}).get("location_name", ""),
        "db_program": (r.db or {}).get("program_uid", ""),
    }
    for f in COMPARE_FIELDS:
        row[f"ext_{f}"] = (r.ext or {}).get(f, "")
        row[f"db_{f}"] = (r.db or {}).get(f, "")
    return row
