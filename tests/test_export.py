"""Tests for export system."""

import io
import json
import sqlite3

import pytest

from policydb.db import get_connection, init_db
from policydb.seed import run_seed
from policydb.exporter import (
    export_schedule_md,
    export_schedule_csv,
    export_schedule_json,
    export_llm_client_md,
    export_llm_client_json,
    export_llm_book_md,
    export_request_bundle_xlsx,
    export_client_requests_xlsx,
)


@pytest.fixture
def seeded_db(tmp_path, monkeypatch):
    db_path = tmp_path / "test.sqlite"
    monkeypatch.setattr("policydb.db.DB_PATH", db_path)
    monkeypatch.setattr("policydb.db.DB_DIR", tmp_path)
    monkeypatch.setattr("policydb.db.EXPORTS_DIR", tmp_path / "exports")
    monkeypatch.setattr("policydb.db.CONFIG_PATH", tmp_path / "config.yaml")
    init_db(path=db_path)
    conn = get_connection(db_path)

    # Insert minimal test data without calling seed (avoids click prompts)
    conn.execute(
        "INSERT INTO clients (name, industry_segment, account_exec) VALUES ('Acme Corp', 'Technology', 'Grant')"
    )
    client_id = conn.execute("SELECT id FROM clients WHERE name='Acme Corp'").fetchone()["id"]
    conn.execute(
        """INSERT INTO policies
           (policy_uid, client_id, policy_type, carrier, policy_number,
            effective_date, expiration_date, premium, limit_amount, deductible,
            description, renewal_status, commission_rate, prior_premium, account_exec, notes)
           VALUES ('POL-001', ?, 'General Liability', 'Zurich', 'GL-12345',
                   '2025-01-01', '2026-01-01', 50000, 2000000, 25000,
                   'Covers GL for all operations.', 'In Progress', 0.12, 45000, 'Grant', 'Internal note.')""",
        (client_id,),
    )
    conn.execute(
        """INSERT INTO policies
           (policy_uid, client_id, policy_type, carrier, policy_number,
            effective_date, expiration_date, premium, renewal_status,
            commission_rate, account_exec)
           VALUES ('POL-002', ?, 'Cyber / Tech E&O', 'Coalition', 'CYBER-99',
                   '2025-01-01', '2026-01-01', 85000, 'Bound', 0.08, 'Grant')""",
        (client_id,),
    )
    conn.commit()
    yield db_path, client_id, conn
    conn.close()


def test_schedule_md_excludes_internal_fields(seeded_db):
    db_path, client_id, conn = seeded_db
    content = export_schedule_md(conn, client_id, "Acme Corp")
    # Internal fields must not appear
    assert "commission_rate" not in content
    assert "commission_amount" not in content
    assert "prior_premium" not in content
    assert "renewal_status" not in content
    assert "Internal note" not in content
    # Client-facing content must be present
    assert "Acme Corp" in content
    assert "General Liability" in content
    assert "Zurich" in content
    assert "Covers GL for all operations" in content


def test_schedule_csv_excludes_internal_fields(seeded_db):
    db_path, client_id, conn = seeded_db
    content = export_schedule_csv(conn, client_id)
    assert "commission_rate" not in content
    assert "prior_premium" not in content
    assert "renewal_status" not in content


def test_schedule_json_excludes_internal_fields(seeded_db):
    db_path, client_id, conn = seeded_db
    content = export_schedule_json(conn, client_id, "Acme Corp")
    data = json.loads(content)
    policies = data["policies"]
    assert len(policies) > 0
    for p in policies:
        assert "commission_rate" not in p
        assert "prior_premium" not in p
        assert "renewal_status" not in p


def test_llm_export_includes_internal_fields(seeded_db):
    db_path, client_id, conn = seeded_db
    content = export_llm_client_md(conn, client_id)
    # LLM export should include renewal status and internal data
    assert "In Progress" in content or "Bound" in content
    assert "Grant" in content  # account exec
    assert "Acme Corp" in content


def test_llm_export_json_structure(seeded_db):
    db_path, client_id, conn = seeded_db
    content = export_llm_client_json(conn, client_id)
    data = json.loads(content)
    assert "metadata" in data
    assert data["metadata"]["export_type"] == "client_program"
    assert "client" in data
    assert "policies" in data
    assert len(data["policies"]) == 2
    # Computed fields should be present
    for p in data["policies"]:
        assert "computed" in p
        assert "days_to_renewal" in p["computed"]
        assert "urgency" in p["computed"]


def test_schedule_md_total_premium(seeded_db):
    db_path, client_id, conn = seeded_db
    content = export_schedule_md(conn, client_id, "Acme Corp")
    # Total: 50000 + 85000 = 135000
    assert "135,000" in content or "135000" in content


def test_llm_book_md_structure(seeded_db):
    db_path, client_id, conn = seeded_db
    content = export_llm_book_md(conn)
    assert "Book of Business" in content
    assert "Acme Corp" in content
    assert "export_type: book_of_business" in content


# ─── RFI / Request Bundle Export Tests ────────────────────────────────────────


@pytest.fixture
def seeded_rfi_db(seeded_db):
    """Extend seeded_db with request bundle and items."""
    db_path, client_id, conn = seeded_db
    conn.execute(
        "INSERT INTO client_request_bundles (id, client_id, title, status) VALUES (1, ?, 'Q1 Renewal Info', 'open')",
        (client_id,),
    )
    conn.execute(
        """INSERT INTO client_request_items (bundle_id, description, policy_uid, category, received, notes, sort_order)
           VALUES (1, 'Updated loss runs for GL', 'POL-001', 'Loss Data', 0, 'Need 5-year history', 1)"""
    )
    conn.execute(
        """INSERT INTO client_request_items (bundle_id, description, policy_uid, category, received, received_at, notes, sort_order)
           VALUES (1, 'Signed application', 'POL-002', 'Applications', 1, '2025-12-15', 'Received via email', 2)"""
    )
    conn.commit()
    yield db_path, client_id, conn


def test_request_bundle_xlsx_valid(seeded_rfi_db):
    """export_request_bundle_xlsx returns valid XLSX bytes."""
    db_path, client_id, conn = seeded_rfi_db
    content = export_request_bundle_xlsx(conn, 1)
    assert isinstance(content, bytes)
    assert len(content) > 100
    # Verify it's a valid XLSX (ZIP magic number)
    assert content[:2] == b"PK"


def test_request_bundle_xlsx_word_wrap(seeded_rfi_db):
    """Data cells in request bundle export have word wrap enabled."""
    from openpyxl import load_workbook
    db_path, client_id, conn = seeded_rfi_db
    content = export_request_bundle_xlsx(conn, 1)
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # Check a data cell (row 2, col 1 = first Item cell)
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            assert cell.alignment.wrap_text, f"Cell ({row_idx},{col_idx}) missing wrap_text"


def test_request_bundle_xlsx_column_widths(seeded_rfi_db):
    """RFI export sheets have explicit column widths for readability."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    db_path, client_id, conn = seeded_rfi_db
    content = export_request_bundle_xlsx(conn, 1)
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # Column order is: Item, Coverage/Location, Category, Status,
    # Received Date, Attached File(s), Notes / Response
    assert ws.column_dimensions["A"].width == 45   # Item
    assert ws.column_dimensions["B"].width == 35   # Coverage / Location
    assert ws.column_dimensions["F"].width == 35   # Attached File(s)
    assert ws.column_dimensions["G"].width == 45   # Notes / Response


def test_request_bundle_xlsx_has_attached_files_column(seeded_rfi_db):
    """RFI export includes an 'Attached File(s)' column so the workbook
    doubles as a manifest when paired with a ZIP of files."""
    from openpyxl import load_workbook
    db_path, client_id, conn = seeded_rfi_db
    content = export_request_bundle_xlsx(conn, 1)
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # Header row — find the first row whose A cell contains "Item"
    header_row = None
    for r in range(1, ws.max_row + 1):
        if (ws.cell(row=r, column=1).value or "") == "Item":
            header_row = r
            break
    assert header_row is not None, "Item header not found"
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Attached File(s)" in headers
    # Comes before Notes / Response
    assert headers.index("Attached File(s)") < headers.index("Notes / Response")


def test_client_requests_xlsx_valid(seeded_rfi_db):
    """export_client_requests_xlsx returns valid XLSX bytes with word wrap."""
    from openpyxl import load_workbook
    db_path, client_id, conn = seeded_rfi_db
    content = export_client_requests_xlsx(conn, client_id)
    assert isinstance(content, bytes)
    assert content[:2] == b"PK"
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    # Check word wrap on data cells
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            assert cell.alignment.wrap_text, f"Cell ({row_idx},{col_idx}) missing wrap_text"


def test_client_requests_xlsx_empty(seeded_db):
    """export_client_requests_xlsx handles clients with no requests."""
    db_path, client_id, conn = seeded_db
    content = export_client_requests_xlsx(conn, client_id)
    assert isinstance(content, bytes)
    assert content[:2] == b"PK"


def test_client_requests_xlsx_groups_by_program_and_location(seeded_db):
    """Tabs are grouped by program (if the policy has one) or location,
    not by RFI bundle. Each row carries an 'RFI' column for traceability."""
    from openpyxl import load_workbook
    db_path, client_id, conn = seeded_db

    # Program: Corporate Property, linked to POL-001
    conn.execute(
        """INSERT INTO programs (program_uid, client_id, name, line_of_business)
           VALUES ('PRG-001', ?, 'Corporate Property', 'Property')""",
        (client_id,),
    )
    program_id = conn.execute("SELECT id FROM programs WHERE program_uid='PRG-001'").fetchone()["id"]
    conn.execute("UPDATE policies SET program_id=? WHERE policy_uid='POL-001'", (program_id,))
    # Location-only item: POL-002 stays unlinked to a program but gets a project_name
    conn.execute("UPDATE policies SET project_name='Downtown Office' WHERE policy_uid='POL-002'")

    # Two open bundles with one item each
    conn.execute(
        "INSERT INTO client_request_bundles (id, client_id, rfi_uid, title, status) "
        "VALUES (10, ?, 'RFI-010', 'Q1 Renewal', 'open')",
        (client_id,),
    )
    conn.execute(
        "INSERT INTO client_request_bundles (id, client_id, rfi_uid, title, status) "
        "VALUES (11, ?, 'RFI-011', 'Mid-term', 'open')",
        (client_id,),
    )
    conn.execute(
        """INSERT INTO client_request_items (bundle_id, description, policy_uid, category, received, sort_order)
           VALUES (10, 'Statement of Values', 'POL-001', 'Property', 0, 1)"""
    )
    conn.execute(
        """INSERT INTO client_request_items (bundle_id, description, policy_uid, category, received, sort_order)
           VALUES (11, 'Signed application', 'POL-002', 'Applications', 0, 1)"""
    )
    # Unassigned: no policy, no project_name
    conn.execute(
        """INSERT INTO client_request_items (bundle_id, description, category, received, sort_order)
           VALUES (10, 'Org chart', 'Corporate Docs', 0, 2)"""
    )
    conn.commit()

    wb = load_workbook(io.BytesIO(export_client_requests_xlsx(conn, client_id)))
    names = wb.sheetnames
    assert "Corporate Property" in names
    assert "Downtown Office" in names
    assert "Unassigned" in names
    # Program tab comes before location tab; Unassigned is last
    assert names.index("Corporate Property") < names.index("Downtown Office")
    assert names[-1] == "Unassigned"

    # Program sheet has the RFI column populated with the bundle uid
    ws = wb["Corporate Property"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "RFI" in headers
    assert "Attached File(s)" in headers
    rfi_col = headers.index("RFI") + 1
    assert ws.cell(row=2, column=rfi_col).value == "RFI-010"


def test_client_requests_xlsx_merges_multiple_bundles_into_one_tab(seeded_db):
    """Items from two separate bundles that share a program land in the same tab."""
    from openpyxl import load_workbook
    db_path, client_id, conn = seeded_db

    conn.execute(
        """INSERT INTO programs (program_uid, client_id, name) VALUES ('PRG-002', ?, 'Casualty Program')""",
        (client_id,),
    )
    prg_id = conn.execute("SELECT id FROM programs WHERE program_uid='PRG-002'").fetchone()["id"]
    conn.execute("UPDATE policies SET program_id=? WHERE policy_uid='POL-001'", (prg_id,))
    conn.execute("UPDATE policies SET program_id=? WHERE policy_uid='POL-002'", (prg_id,))

    conn.execute(
        "INSERT INTO client_request_bundles (id, client_id, rfi_uid, title, status) "
        "VALUES (20, ?, 'RFI-020', 'Bundle A', 'open')",
        (client_id,),
    )
    conn.execute(
        "INSERT INTO client_request_bundles (id, client_id, rfi_uid, title, status) "
        "VALUES (21, ?, 'RFI-021', 'Bundle B', 'open')",
        (client_id,),
    )
    conn.execute(
        "INSERT INTO client_request_items (bundle_id, description, policy_uid, received, sort_order) "
        "VALUES (20, 'Loss runs', 'POL-001', 0, 1)"
    )
    conn.execute(
        "INSERT INTO client_request_items (bundle_id, description, policy_uid, received, sort_order) "
        "VALUES (21, 'Exposure schedule', 'POL-002', 0, 1)"
    )
    conn.commit()

    wb = load_workbook(io.BytesIO(export_client_requests_xlsx(conn, client_id)))
    assert wb.sheetnames == ["Casualty Program"]
    ws = wb["Casualty Program"]
    # Header row + two data rows
    assert ws.max_row == 3
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    rfi_col = headers.index("RFI") + 1
    rfi_values = {ws.cell(row=r, column=rfi_col).value for r in range(2, ws.max_row + 1)}
    assert rfi_values == {"RFI-020", "RFI-021"}


def test_build_client_rfi_zip_layout(seeded_rfi_db, tmp_path, monkeypatch):
    """build_client_rfi_zip produces a client-friendly ZIP with the
    workbook at root, per-bundle top folders, and no MANIFEST.txt."""
    import zipfile
    from policydb.web.routes.attachments import build_client_rfi_zip

    db_path, client_id, conn = seeded_rfi_db

    # Attach one local file to item 1 so the ZIP has at least one file
    attachments_dir = tmp_path / ".policydb" / "files" / "attachments"
    attachments_dir.mkdir(parents=True)
    sample_path = attachments_dir / "loss_runs_2024.pdf"
    sample_path.write_bytes(b"%PDF-1.4 dummy")

    # Point the attachments dir override at our temp path
    monkeypatch.setattr(
        "policydb.web.routes.attachments._ATTACHMENTS_DIR",
        attachments_dir,
    )

    conn.execute(
        """INSERT INTO attachments (uid, title, source, file_path, filename, file_size, mime_type, category)
           VALUES ('ATT-001', 'Loss Runs 2024', 'local', ?, 'loss_runs_2024.pdf', 14, 'application/pdf', 'Loss Data')""",
        (str(sample_path),),
    )
    att_id = conn.execute("SELECT id FROM attachments WHERE uid='ATT-001'").fetchone()["id"]
    item_id = conn.execute(
        "SELECT id FROM client_request_items WHERE bundle_id=1 AND sort_order=1"
    ).fetchone()["id"]
    conn.execute(
        "INSERT INTO record_attachments (attachment_id, record_type, record_id) VALUES (?, 'rfi_item', ?)",
        (att_id, item_id),
    )
    conn.commit()

    data, download_name, total_files = build_client_rfi_zip(conn, client_id)

    assert download_name.endswith(".zip")
    assert "Outstanding Requests" in download_name
    assert total_files == 1

    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        names = zf.namelist()

    # No technical manifest
    assert "MANIFEST.txt" not in names
    # Client-friendly workbook at root
    assert "Outstanding Requests.xlsx" in names
    # Exactly one PDF, nested under a program/location tab folder that
    # mirrors the workbook. The seeded policy has no program and no
    # project_name, so the item lands in "Unassigned/".
    pdfs = [n for n in names if n.endswith(".pdf")]
    assert len(pdfs) == 1
    assert pdfs[0].startswith("Unassigned/")
    # Item folder carries the RFI label in brackets so the bundle is
    # still traceable from the file path (matches the workbook's RFI column).
    assert "[Q1 Renewal Info]" in pdfs[0]


def test_build_client_rfi_zip_raises_when_no_bundles(seeded_db):
    """Clients with no open bundles raise ValueError so the route can 404."""
    from policydb.web.routes.attachments import build_client_rfi_zip
    db_path, client_id, conn = seeded_db
    with pytest.raises(ValueError):
        build_client_rfi_zip(conn, client_id)
